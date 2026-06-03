"""Tests for Arc F S2.4: Sprint-1 fields surfaced in Excel and control center.

Covers:
- Security sheet GHAS + OSSF columns (present + backward-compat with missing data)
- All Repos sheet new flag columns (has_any_release, readme_stale)
- Repo detail hidden data Arc F columns
- Control center lane assignment for readme_stale and GHAS critical alerts
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from src.excel_all_repos_helpers import ALL_REPOS_HEADERS, build_all_repo_rows
from src.excel_export import _build_all_repos, _build_security
from src.excel_repo_data_helpers import _arc_f_detail_cols, repo_detail_rows
from src.excel_security_sheet_helpers import SECURITY_HEADERS, build_security_sheet_rows
from src.operator_control_center import build_operator_snapshot

# ─────────────────────────────────────────────────────────────
# Shared fixture helpers
# ─────────────────────────────────────────────────────────────


def _make_audit_with_arc_f(
    name: str = "repo-alpha",
    readme_stale: bool | None = None,
    has_any_release: bool = False,
    latest_release_age_days: int | None = None,
    ossf_score: float | None = None,
    ossf_available: bool = False,
    ossf_checks: list | None = None,
    dep_critical: int = 0,
    dep_available: bool = False,
    cs_critical: int = 0,
    cs_available: bool = False,
    ss_open: int = 0,
    ss_available: bool = False,
    html_url: str = "",
) -> dict:
    """Build a minimal audit dict with Arc F signals populated."""
    readme_details = {
        "exists": True,
        "readme_stale": readme_stale,
        "readme_staleness_ratio": 8.5 if readme_stale else None,
        "readme_last_touched_days": 200 if readme_stale else None,
        "code_last_touched_days": 10 if readme_stale else None,
    }
    activity_details = {
        "has_any_release": has_any_release,
        "latest_release_age_days": latest_release_age_days,
        "release_count": 1 if has_any_release else 0,
    }
    ossf_scorecard: dict = {}
    if ossf_available:
        ossf_scorecard = {
            "available": True,
            "score": ossf_score,
            "checks": ossf_checks or [],
        }
    return {
        "metadata": {"name": name, "html_url": html_url, "language": "Python"},
        "overall_score": 0.5,
        "interest_score": 0.3,
        "grade": "C",
        "completeness_tier": "functional",
        "interest_tier": "interesting",
        "interest_grade": "C",
        "badges": [],
        "next_badges": [],
        "flags": [],
        "hotspots": [],
        "implementation_hotspots": [],
        "action_candidates": [],
        "lenses": {},
        "security_posture": {
            "label": "healthy",
            "score": 0.6,
            "secrets_found": 0,
            "dangerous_files": [],
            "has_security_md": False,
            "has_dependabot": False,
            "evidence": [],
            "local": {},
            "github": {},
        },
        "score_explanation": {},
        "ossf_scorecard": ossf_scorecard,
        "analyzer_results": [
            {
                "dimension": "readme",
                "score": 0.5,
                "max_score": 1.0,
                "findings": [],
                "details": readme_details,
            },
            {
                "dimension": "activity",
                "score": 0.6,
                "max_score": 1.0,
                "findings": [],
                "details": activity_details,
            },
        ],
    }


def _make_ghas_entry(
    dep_critical: int = 0,
    dep_high: int = 0,
    dep_available: bool = True,
    cs_critical: int = 0,
    cs_high: int = 0,
    cs_available: bool = True,
    ss_open: int = 0,
    ss_available: bool = True,
) -> dict:
    return {
        "dependabot": {
            "critical": dep_critical,
            "high": dep_high,
            "available": dep_available,
        },
        "code_scanning": {
            "critical": cs_critical,
            "high": cs_high,
            "available": cs_available,
        },
        "secret_scanning": {
            "open": ss_open,
            "available": ss_available,
        },
    }


def _make_report_data(audits: list[dict], ghas_alerts: dict | None = None) -> dict:
    base: dict = {
        "username": "testuser",
        "generated_at": "2026-05-10T12:00:00+00:00",
        "audits": audits,
        "repos_audited": len(audits),
        "average_score": 0.5,
        "portfolio_grade": "C",
        "hotspots": [],
        "implementation_hotspots": [],
        "implementation_hotspots_summary": {},
        "collections": {},
        "profiles": {},
        "lenses": {},
        "scenario_summary": {},
        "preflight_summary": {"status": "ok", "checks": []},
        "review_summary": {
            "review_id": "test-1",
            "source_run_id": "testuser:2026-05-10",
            "status": "open",
        },
        "watch_state": {},
        "review_targets": [],
        "material_changes": [],
        "managed_state_drift": [],
        "governance_drift": [],
        "governance_preview": {},
        "campaign_summary": {},
        "writeback_preview": {},
        "rollback_preview": {},
    }
    if ghas_alerts:
        base["ghas_alerts"] = ghas_alerts
    return base


# ─────────────────────────────────────────────────────────────
# Security sheet
# ─────────────────────────────────────────────────────────────


class TestSecuritySheetGhasAndOssf:
    def test_headers_include_ghas_and_ossf_columns(self) -> None:
        assert "Dep Critical" in SECURITY_HEADERS
        assert "Dep High" in SECURITY_HEADERS
        assert "CodeQL Critical" in SECURITY_HEADERS
        assert "CodeQL High" in SECURITY_HEADERS
        assert "Secrets Open" in SECURITY_HEADERS
        assert "OSSF Score" in SECURITY_HEADERS

    def test_ghas_counts_populated_when_lookup_provided(self) -> None:
        audit = _make_audit_with_arc_f()
        ghas = {
            "repo-alpha": _make_ghas_entry(dep_critical=3, dep_high=5, cs_critical=1, ss_open=2)
        }
        rows = build_security_sheet_rows([audit], ghas_lookup=ghas)
        assert len(rows) == 1
        row = rows[0]
        # Dep Critical at index 8 (0-based)
        assert row[8] == 3
        assert row[9] == 5
        assert row[10] == 1
        assert row[12] == 2

    def test_ghas_unavailable_shows_zeros_not_dashes(self) -> None:
        """When available=False, counts should be 0 (not raw API values)."""
        audit = _make_audit_with_arc_f()
        ghas = {
            "repo-alpha": _make_ghas_entry(
                dep_critical=99,
                dep_available=False,
                cs_available=False,
                ss_available=False,
            )
        }
        rows = build_security_sheet_rows([audit], ghas_lookup=ghas)
        row = rows[0]
        assert row[8] == 0  # dep_critical zeroed because available=False
        assert row[10] == 0  # cs_critical

    def test_missing_ghas_lookup_shows_dashes(self) -> None:
        audit = _make_audit_with_arc_f()
        rows = build_security_sheet_rows([audit], ghas_lookup=None)
        row = rows[0]
        assert row[8] == "—"
        assert row[9] == "—"
        assert row[10] == "—"
        assert row[12] == "—"
        assert row[13] == "—"

    def test_ossf_score_shows_value_when_available(self) -> None:
        audit = _make_audit_with_arc_f(ossf_available=True, ossf_score=7.3)
        rows = build_security_sheet_rows([audit], ghas_lookup=None)
        assert rows[0][13] == 7.3

    def test_ossf_score_shows_dash_when_unavailable(self) -> None:
        audit = _make_audit_with_arc_f(ossf_available=False)
        rows = build_security_sheet_rows([audit], ghas_lookup=None)
        assert rows[0][13] == "—"

    def test_row_length_matches_headers(self) -> None:
        audit = _make_audit_with_arc_f()
        rows = build_security_sheet_rows([audit])
        assert len(rows[0]) == len(SECURITY_HEADERS)

    def test_security_sheet_builds_without_ghas_data(self, tmp_path: Path) -> None:
        """Full sheet build must not raise with missing GHAS data in report."""
        wb = Workbook()
        report = _make_report_data([_make_audit_with_arc_f()])
        _build_security(wb, report)
        assert "Security" in wb.sheetnames

    def test_security_sheet_builds_with_ghas_data(self, tmp_path: Path) -> None:
        wb = Workbook()
        report = _make_report_data(
            [_make_audit_with_arc_f()],
            ghas_alerts={"repo-alpha": _make_ghas_entry(dep_critical=2)},
        )
        _build_security(wb, report)
        ws = wb["Security"]
        dep_critical_col = SECURITY_HEADERS.index("Dep Critical") + 1
        assert ws.cell(row=2, column=dep_critical_col).value == 2


# ─────────────────────────────────────────────────────────────
# All Repos sheet
# ─────────────────────────────────────────────────────────────


class TestAllReposNewColumns:
    def test_has_release_column_present(self) -> None:
        assert "Has Release" in ALL_REPOS_HEADERS

    def test_readme_stale_column_present(self) -> None:
        assert "README Stale" in ALL_REPOS_HEADERS

    def test_has_release_yes_when_activity_flag_set(self) -> None:
        audit = _make_audit_with_arc_f(has_any_release=True)
        rows = build_all_repo_rows(
            [audit],
            score_history=None,
            risk_lookup=None,
            render_sparkline=lambda _: "",
        )
        has_release_idx = ALL_REPOS_HEADERS.index("Has Release")
        assert rows[0]["values"][has_release_idx] == "Yes"

    def test_has_release_no_when_flag_false(self) -> None:
        audit = _make_audit_with_arc_f(has_any_release=False)
        rows = build_all_repo_rows(
            [audit],
            score_history=None,
            risk_lookup=None,
            render_sparkline=lambda _: "",
        )
        has_release_idx = ALL_REPOS_HEADERS.index("Has Release")
        assert rows[0]["values"][has_release_idx] == "No"

    def test_readme_stale_yes_when_flag_set(self) -> None:
        audit = _make_audit_with_arc_f(readme_stale=True)
        rows = build_all_repo_rows(
            [audit],
            score_history=None,
            risk_lookup=None,
            render_sparkline=lambda _: "",
        )
        readme_stale_idx = ALL_REPOS_HEADERS.index("README Stale")
        assert rows[0]["values"][readme_stale_idx] == "Yes"

    def test_readme_stale_no_when_flag_false(self) -> None:
        audit = _make_audit_with_arc_f(readme_stale=False)
        rows = build_all_repo_rows(
            [audit],
            score_history=None,
            risk_lookup=None,
            render_sparkline=lambda _: "",
        )
        readme_stale_idx = ALL_REPOS_HEADERS.index("README Stale")
        assert rows[0]["values"][readme_stale_idx] == "No"

    def test_values_length_matches_headers(self) -> None:
        audit = _make_audit_with_arc_f()
        rows = build_all_repo_rows(
            [audit],
            score_history=None,
            risk_lookup=None,
            render_sparkline=lambda _: "",
        )
        assert len(rows[0]["values"]) == len(ALL_REPOS_HEADERS)

    def test_all_repos_sheet_builds_without_new_fields(self) -> None:
        """Backward compat: audit without readme/activity fields must not raise."""
        audit = {
            "metadata": {"name": "bare-repo", "html_url": ""},
            "overall_score": 0.4,
            "interest_score": 0.2,
            "grade": "D",
            "completeness_tier": "wip",
            "interest_tier": "mundane",
            "interest_grade": "D",
            "badges": [],
            "next_badges": [],
            "flags": [],
            "analyzer_results": [],
        }
        wb = Workbook()
        _build_all_repos(wb, _make_report_data([audit]), score_history=None, risk_lookup=None)
        assert "All Repos" in wb.sheetnames


# ─────────────────────────────────────────────────────────────
# Repo detail hidden data
# ─────────────────────────────────────────────────────────────


class TestArcFDetailCols:
    def test_readme_stale_yes(self) -> None:
        audit = _make_audit_with_arc_f(readme_stale=True)
        cols = _arc_f_detail_cols(audit, None)
        assert cols[0] == "Yes"

    def test_readme_stale_no(self) -> None:
        audit = _make_audit_with_arc_f(readme_stale=False)
        cols = _arc_f_detail_cols(audit, None)
        assert cols[0] == "No"

    def test_readme_stale_none_shows_dash(self) -> None:
        audit = _make_audit_with_arc_f(readme_stale=None)
        cols = _arc_f_detail_cols(audit, None)
        assert cols[0] == "—"

    def test_has_release_yes(self) -> None:
        audit = _make_audit_with_arc_f(has_any_release=True, latest_release_age_days=45)
        cols = _arc_f_detail_cols(audit, None)
        assert cols[4] == "Yes"
        assert cols[5] == 45

    def test_latest_release_age_none_shows_dash(self) -> None:
        audit = _make_audit_with_arc_f(has_any_release=False)
        cols = _arc_f_detail_cols(audit, None)
        assert cols[5] == "—"

    def test_ossf_score_and_failing_checks(self) -> None:
        audit = _make_audit_with_arc_f(
            ossf_available=True,
            ossf_score=3.2,
            ossf_checks=[
                {"name": "Branch-Protection", "score": 0},
                {"name": "Code-Review", "score": 0},
                {"name": "Vulnerabilities", "score": 1},
            ],
        )
        cols = _arc_f_detail_cols(audit, None)
        assert cols[6] == 3.2
        assert "Branch-Protection" in cols[7]
        assert "Code-Review" in cols[7]
        assert "Vulnerabilities" not in cols[7]  # score=1, not failing

    def test_ossf_unavailable_shows_dashes(self) -> None:
        audit = _make_audit_with_arc_f(ossf_available=False)
        cols = _arc_f_detail_cols(audit, None)
        assert cols[6] == "—"
        assert cols[7] == "—"

    def test_row_extended_by_8_columns(self) -> None:
        """repo_detail_rows returns 79 columns per row (70 base + 8 Arc F + 1 risk tier)."""
        data = _make_report_data([_make_audit_with_arc_f()])
        detail_rows, _, _ = repo_detail_rows(data, None)
        assert len(detail_rows) == 1
        assert len(detail_rows[0]) == 79

    def test_risk_tier_appended_as_column_79(self) -> None:
        data = _make_report_data([_make_audit_with_arc_f(name="repo-alpha")])
        detail_rows, _, _ = repo_detail_rows(data, None, risk_lookup={"repo-alpha": "elevated"})
        assert len(detail_rows[0]) == 79
        assert detail_rows[0][78] == "elevated"

    def test_risk_tier_dash_when_no_lookup(self) -> None:
        data = _make_report_data([_make_audit_with_arc_f()])
        detail_rows, _, _ = repo_detail_rows(data, None)
        assert detail_rows[0][78] == "—"

    def test_backward_compat_without_arc_f_fields(self) -> None:
        """Audit with no readme/activity Arc F fields must still produce 78 cols."""
        bare_audit = {
            "metadata": {"name": "bare", "html_url": "", "description": None},
            "overall_score": 0.5,
            "interest_score": 0.2,
            "grade": "C",
            "completeness_tier": "functional",
            "badges": [],
            "flags": [],
            "hotspots": [],
            "implementation_hotspots": [],
            "action_candidates": [],
            "lenses": {},
            "security_posture": {},
            "score_explanation": {},
            "ossf_scorecard": {},
            "analyzer_results": [],
        }
        data = _make_report_data([bare_audit])
        detail_rows, _, _ = repo_detail_rows(data, None)
        assert len(detail_rows[0]) == 79


# ─────────────────────────────────────────────────────────────
# Control center lane assignments
# ─────────────────────────────────────────────────────────────


class TestControlCenterArcFLanes:
    def _snapshot(self, report_data: dict, tmp_path: Path) -> dict:
        return build_operator_snapshot(report_data, output_dir=tmp_path)

    def test_readme_stale_lands_in_urgent_lane(self, tmp_path: Path) -> None:
        audit = _make_audit_with_arc_f(name="my-repo", readme_stale=True)
        report = _make_report_data([audit])
        snapshot = self._snapshot(report, tmp_path)
        queue = snapshot.get("operator_queue", [])
        arc_f_items = [i for i in queue if "arc-f:readme-stale:my-repo" in i.get("item_id", "")]
        assert arc_f_items, "Expected an arc-f:readme-stale item in the queue"
        assert arc_f_items[0]["lane"] == "urgent"

    def test_readme_not_stale_does_not_add_item(self, tmp_path: Path) -> None:
        audit = _make_audit_with_arc_f(name="clean-repo", readme_stale=False)
        report = _make_report_data([audit])
        snapshot = self._snapshot(report, tmp_path)
        queue = snapshot.get("operator_queue", [])
        arc_f_items = [i for i in queue if "arc-f:readme-stale" in i.get("item_id", "")]
        assert not arc_f_items

    def test_ghas_critical_dependabot_lands_in_blocked_lane(self, tmp_path: Path) -> None:
        audit = _make_audit_with_arc_f(name="vuln-repo")
        report = _make_report_data(
            [audit],
            ghas_alerts={"vuln-repo": _make_ghas_entry(dep_critical=5)},
        )
        snapshot = self._snapshot(report, tmp_path)
        queue = snapshot.get("operator_queue", [])
        arc_f_items = [i for i in queue if "arc-f:ghas-critical:vuln-repo" in i.get("item_id", "")]
        assert arc_f_items, "Expected an arc-f:ghas-critical item in the queue"
        assert arc_f_items[0]["lane"] == "blocked"

    def test_ghas_critical_codeql_lands_in_blocked_lane(self, tmp_path: Path) -> None:
        audit = _make_audit_with_arc_f(name="cs-repo")
        report = _make_report_data(
            [audit],
            ghas_alerts={"cs-repo": _make_ghas_entry(cs_critical=2)},
        )
        snapshot = self._snapshot(report, tmp_path)
        queue = snapshot.get("operator_queue", [])
        arc_f_items = [i for i in queue if "arc-f:ghas-critical:cs-repo" in i.get("item_id", "")]
        assert arc_f_items
        assert arc_f_items[0]["lane"] == "blocked"

    def test_ghas_zero_criticals_does_not_add_blocked_item(self, tmp_path: Path) -> None:
        audit = _make_audit_with_arc_f(name="safe-repo")
        report = _make_report_data(
            [audit],
            ghas_alerts={"safe-repo": _make_ghas_entry(dep_critical=0, cs_critical=0, ss_open=0)},
        )
        snapshot = self._snapshot(report, tmp_path)
        queue = snapshot.get("operator_queue", [])
        arc_f_items = [i for i in queue if "arc-f:ghas-critical" in i.get("item_id", "")]
        assert not arc_f_items

    def test_ossf_low_score_lands_in_ready_lane(self, tmp_path: Path) -> None:
        audit = _make_audit_with_arc_f(name="low-ossf-repo", ossf_available=True, ossf_score=3.1)
        report = _make_report_data([audit])
        snapshot = self._snapshot(report, tmp_path)
        queue = snapshot.get("operator_queue", [])
        arc_f_items = [i for i in queue if "arc-f:ossf-low:low-ossf-repo" in i.get("item_id", "")]
        assert arc_f_items, "Expected an arc-f:ossf-low item in the queue"
        assert arc_f_items[0]["lane"] == "ready"

    def test_ossf_adequate_score_does_not_add_item(self, tmp_path: Path) -> None:
        audit = _make_audit_with_arc_f(name="ok-ossf-repo", ossf_available=True, ossf_score=6.0)
        report = _make_report_data([audit])
        snapshot = self._snapshot(report, tmp_path)
        queue = snapshot.get("operator_queue", [])
        arc_f_items = [i for i in queue if "arc-f:ossf-low" in i.get("item_id", "")]
        assert not arc_f_items

    def test_no_arc_f_items_when_all_data_absent(self, tmp_path: Path) -> None:
        """Audit with no Arc F data must not add any arc-f items."""
        bare_audit = {
            "metadata": {"name": "bare", "html_url": ""},
            "overall_score": 0.5,
            "interest_score": 0.3,
            "grade": "C",
            "completeness_tier": "functional",
            "analyzer_results": [],
            "ossf_scorecard": {},
        }
        report = _make_report_data([bare_audit])
        snapshot = self._snapshot(report, tmp_path)
        queue = snapshot.get("operator_queue", [])
        arc_f_items = [i for i in queue if "arc-f:" in i.get("item_id", "")]
        assert not arc_f_items


# ---------------------------------------------------------------------------
# Arc F S3.4: control-center duplicate-group lane entries
# ---------------------------------------------------------------------------


class _FakeDuplicateGroup:
    """Minimal stand-in for DuplicateGroup dataclass."""

    def __init__(self, members, representative, min_pairwise_cosine):
        self.members = members
        self.representative = representative
        self.min_pairwise_cosine = min_pairwise_cosine


class _FakeSemanticIndex:
    """Stub SemanticIndex that returns a fixed list of DuplicateGroups."""

    def __init__(self, groups):
        self._groups = groups

    def find_duplicate_groups(self, **kwargs):
        return self._groups


class TestControlCenterDuplicateGroupLanes:
    def _snapshot_with_index(self, report_data: dict, tmp_path: Path, semantic_index) -> dict:
        return build_operator_snapshot(
            report_data, output_dir=tmp_path, semantic_index=semantic_index
        )

    def test_duplicate_group_lane_entry_appears(self, tmp_path: Path) -> None:
        """When semantic_index returns duplicate groups, lane entries appear in the queue."""
        group = _FakeDuplicateGroup(
            members=["ArgueMap", "ConvictionMapper", "TradeoffAtlas"],
            representative="TradeoffAtlas",
            min_pairwise_cosine=0.91,
        )
        fake_idx = _FakeSemanticIndex([group])

        report = _make_report_data([_make_audit_with_arc_f("ArgueMap")])
        snapshot = self._snapshot_with_index(report, tmp_path, fake_idx)

        queue = snapshot.get("operator_queue", [])
        dup_items = [i for i in queue if "arc-f:duplicate-group:" in i.get("item_id", "")]
        assert len(dup_items) == 1

        item = dup_items[0]
        assert item["lane"] == "ready"
        assert "ArgueMap" in item["summary"] or "ConvictionMapper" in item["summary"]
        assert item["priority"] == 35

    def test_duplicate_group_summary_contains_all_members(self, tmp_path: Path) -> None:
        """The lane entry summary mentions all member repo names."""
        group = _FakeDuplicateGroup(
            members=["RepoA", "RepoB", "RepoC"],
            representative="RepoC",
            min_pairwise_cosine=0.88,
        )
        fake_idx = _FakeSemanticIndex([group])

        report = _make_report_data([_make_audit_with_arc_f("RepoA")])
        snapshot = self._snapshot_with_index(report, tmp_path, fake_idx)

        queue = snapshot.get("operator_queue", [])
        dup_items = [i for i in queue if "arc-f:duplicate-group:" in i.get("item_id", "")]
        assert dup_items
        combined = dup_items[0]["title"] + dup_items[0]["summary"]
        for name in ["RepoA", "RepoB", "RepoC"]:
            assert name in combined

    def test_no_duplicate_entries_without_semantic_index(self, tmp_path: Path) -> None:
        """When semantic_index=None, no duplicate-group lane entries are generated."""
        report = _make_report_data([_make_audit_with_arc_f("SomeRepo")])
        snapshot = build_operator_snapshot(report, output_dir=tmp_path, semantic_index=None)

        queue = snapshot.get("operator_queue", [])
        dup_items = [i for i in queue if "arc-f:duplicate-group:" in i.get("item_id", "")]
        assert dup_items == []

    def test_no_duplicate_entries_when_groups_empty(self, tmp_path: Path) -> None:
        """When semantic_index returns no groups, no duplicate-group entries are added."""
        fake_idx = _FakeSemanticIndex([])
        report = _make_report_data([_make_audit_with_arc_f("SomeRepo")])
        snapshot = self._snapshot_with_index(report, tmp_path, fake_idx)

        queue = snapshot.get("operator_queue", [])
        dup_items = [i for i in queue if "arc-f:duplicate-group:" in i.get("item_id", "")]
        assert dup_items == []
