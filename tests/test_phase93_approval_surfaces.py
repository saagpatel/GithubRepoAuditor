from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from src.approval_ledger import render_approval_center_markdown
from src.excel_export import _build_approval_ledger, _build_hidden_data_sheets
from src.review_pack import export_review_pack
from src.scheduled_handoff import render_scheduled_handoff_markdown
from src.web_export import _render_html


def _report() -> dict:
    return {
        "username": "testuser",
        "generated_at": "2026-04-13T12:00:00+00:00",
        "repos_audited": 1,
        "average_score": 0.61,
        "portfolio_grade": "C",
        "audits": [],
        "operator_summary": {
            "headline": "Approval review is the next bounded operator step.",
            "counts": {"blocked": 0, "urgent": 1, "ready": 0, "deferred": 0},
            "approval_workflow_summary": {
                "summary": "Governance: all is the strongest approval review candidate right now.",
            },
            "next_approval_review": {
                "summary": "Review Governance: all next and decide whether to capture approval.",
            },
            "top_ready_for_review_approvals": [
                {
                    "label": "Governance: all",
                    "subject_key": "all",
                    "summary": "Governance scope all is ready for review.",
                    "approval_command": "audit testuser --approve-governance --governance-scope all",
                }
            ],
            "top_needs_reapproval_approvals": [],
            "top_overdue_approval_followups": [
                {
                    "label": "Governance: all",
                    "subject_key": "all",
                    "summary": "Governance: all is still approved, but its local follow-up review is overdue.",
                    "follow_up_state": "overdue-follow-up",
                    "follow_up_command": "audit testuser --review-governance --governance-scope all",
                    "last_reviewed_by": "owner",
                    "next_follow_up_due_at": "2026-04-12T12:00:00+00:00",
                }
            ],
            "top_due_soon_approval_followups": [],
            "top_approved_manual_approvals": [],
            "top_blocked_approvals": [],
        },
        "operator_queue": [
            {
                "repo": "RepoA",
                "title": "RepoA approval follow-up",
                "lane": "urgent",
                "lane_label": "Urgent",
                "lane_reason": "Approval review is now the clean next step.",
                "recommended_action": "Review the approval packet and capture local approval if it still matches.",
                "approval_line": "Approval Workflow: Governance: all is ready for review.",
            }
        ],
        "approval_ledger": [
            {
                "approval_id": "governance:all",
                "approval_subject_type": "governance",
                "subject_key": "all",
                "label": "Governance: all",
                "approval_state": "ready-for-review",
                "source_run_id": "testuser:2026-04-13T12:00:00+00:00",
                "fingerprint": "fingerprint-1",
                "approved_at": "",
                "approved_by": "",
                "approval_ready": True,
                "apply_ready_after_approval": True,
                "approval_command": "audit testuser --approve-governance --governance-scope all",
                "manual_apply_command": "",
                "summary": "Governance scope all is ready for review.",
            },
            {
                "approval_id": "governance:followup",
                "approval_subject_type": "governance",
                "subject_key": "all",
                "label": "Governance: all",
                "approval_state": "approved-manual",
                "follow_up_state": "overdue-follow-up",
                "source_run_id": "testuser:2026-04-13T12:00:00+00:00",
                "fingerprint": "fingerprint-1",
                "approved_at": "2026-04-01T12:00:00+00:00",
                "approved_by": "owner",
                "last_reviewed_at": "2026-04-06T12:00:00+00:00",
                "last_reviewed_by": "owner",
                "next_follow_up_due_at": "2026-04-12T12:00:00+00:00",
                "approval_ready": False,
                "apply_ready_after_approval": True,
                "approval_command": "audit testuser --approve-governance --governance-scope all",
                "follow_up_command": "audit testuser --review-governance --governance-scope all",
                "manual_apply_command": "audit testuser --writeback-apply --governance-scope all",
                "summary": "Governance: all is still approved, but its local follow-up review is overdue.",
            },
        ],
        "approval_workflow_summary": {
            "summary": "Governance: all is the strongest approval review candidate right now.",
        },
        "next_approval_review": {
            "summary": "Review Governance: all next and decide whether to capture approval.",
        },
    }


def test_review_pack_web_and_scheduled_handoff_include_approval_workflow(tmp_path: Path) -> None:
    report = _report()

    review_pack = export_review_pack(report, tmp_path)["review_pack_path"].read_text()
    html = _render_html(report)
    handoff = render_scheduled_handoff_markdown(
        {
            "username": report["username"],
            "generated_at": report["generated_at"],
            "operator_summary": report["operator_summary"],
            "operator_queue": report["operator_queue"],
            "operator_recent_changes": [],
            "campaign_summary": {},
            "writeback_preview": {},
            "managed_state_drift": [],
            "issue_candidate": {},
        }
    )
    approval_center = render_approval_center_markdown(
        {
            "username": report["username"],
            "generated_at": report["generated_at"],
            "approval_workflow_summary": report["approval_workflow_summary"],
            "next_approval_review": report["next_approval_review"],
            "approval_ledger": report["approval_ledger"],
        }
    )

    assert "Approval Workflow:" in review_pack
    assert "Next Approval Review:" in review_pack
    assert "Approval Workflow" in html
    assert "Next Approval Review" in html
    assert "Approval Workflow" in handoff
    assert "Next Approval Review" in handoff
    assert "Approval Workflow" in approval_center
    assert "Ready For Review" in approval_center
    assert "Overdue Follow-Up" in approval_center
    assert "--review-governance --governance-scope all" in approval_center


def test_excel_approval_ledger_sheet_and_hidden_data_exist() -> None:
    report = _report()
    workbook = Workbook()
    workbook.remove(workbook.active)

    _build_approval_ledger(workbook, report)
    _build_hidden_data_sheets(workbook, report, diff_data=None, score_history={})

    assert "Approval Ledger" in workbook.sheetnames
    assert "Data_ApprovalLedger" in workbook.sheetnames
    sheet = workbook["Approval Ledger"]
    assert sheet["A1"].value == "Approval Ledger"
    assert "strongest approval review candidate" in str(sheet["A2"].value)
    assert sheet["C5"].value == "Follow-Up"
    assert sheet["G5"].value == "Next Follow-Up Due"
