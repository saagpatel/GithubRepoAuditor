from __future__ import annotations

import json
import xml.etree.ElementTree as ET
import zipfile

import pytest
from openpyxl import Workbook, load_workbook

from src.excel_export import (
    RADAR_DIMS,
    RADAR_LABELS,
    _build_action_items,
    _build_all_repos,
    _build_by_collection,
    _build_by_lens,
    _build_campaigns,
    _build_changes,
    _build_compare_sheet,
    _build_dashboard,
    _build_executive_summary,
    _build_governance_audit,
    _build_governance_controls,
    _build_hidden_data_sheets,
    _build_historical_intelligence,
    _build_hotspots,
    _build_implementation_hotspots,
    _build_operator_outcomes,
    _build_portfolio_explorer,
    _build_print_pack,
    _build_repo_detail,
    _build_repo_profiles,
    _build_review_history_sheet,
    _build_review_queue,
    _build_run_changes,
    _build_scenario_planner,
    _build_score_explainer,
    _build_security,
    _build_security_debt,
    _build_trend_summary,
    _build_trends,
    _build_writeback_audit,
    export_excel,
)
from src.excel_template import DEFAULT_TEMPLATE_PATH, TEMPLATE_INFO_SHEET
from src.report_enrichment import (
    build_queue_pressure_summary,
    build_top_recommendation_summary,
    build_trust_actionability_summary,
    build_weekly_review_pack,
    no_baseline_summary,
    no_linked_artifact_summary,
)


def _make_audit(name: str, score: float, grade: str = "C", tier: str = "functional", **kwargs) -> dict:
    dims = kwargs.get("dims", {})
    results = [
        {"dimension": d, "score": dims.get(d, score), "max_score": 1.0, "findings": [], "details": {}}
        for d in RADAR_DIMS
    ]
    # Add security dimension
    sec_details = kwargs.get("security_details", {"secrets_found": 0, "dangerous_files": [], "has_security_md": False, "has_dependabot": False})
    results.append({"dimension": "security", "score": kwargs.get("security_score", 0.8), "max_score": 1.0, "findings": [], "details": sec_details})
    # Add interest
    results.append({"dimension": "interest", "score": 0.3, "max_score": 1.0, "findings": [], "details": {}})
    return {
        "metadata": {"name": name, "html_url": f"https://github.com/user/{name}", "language": "Python"},
        "overall_score": score, "interest_score": 0.3, "grade": grade,
        "completeness_tier": tier, "badges": [], "flags": [],
        "lenses": {
            "ship_readiness": {"score": score, "orientation": "higher-is-better", "summary": "Ready", "drivers": []},
            "maintenance_risk": {"score": round(1 - score, 3), "orientation": "higher-is-riskier", "summary": "Risk", "drivers": []},
            "showcase_value": {"score": min(1.0, score + 0.1), "orientation": "higher-is-better", "summary": "Story", "drivers": []},
            "security_posture": {"score": kwargs.get("security_score", 0.8), "orientation": "higher-is-better", "summary": "Security", "drivers": []},
            "momentum": {"score": 0.5, "orientation": "higher-is-better", "summary": "Momentum", "drivers": []},
            "portfolio_fit": {"score": score, "orientation": "higher-is-better", "summary": "Fit", "drivers": []},
        },
        "security_posture": {
            "label": "healthy" if kwargs.get("security_score", 0.8) >= 0.65 else "critical",
            "score": kwargs.get("security_score", 0.8),
            "secrets_found": sec_details.get("secrets_found", 0),
            "dangerous_files": sec_details.get("dangerous_files", []),
            "has_security_md": sec_details.get("has_security_md", False),
            "has_dependabot": sec_details.get("has_dependabot", False),
            "evidence": [],
        },
        "action_candidates": [
            {
                "key": "testing",
                "title": "Strengthen tests",
                "lens": "ship_readiness",
                "effort": "medium",
                "confidence": 0.8,
                "expected_lens_delta": 0.12,
                "expected_tier_movement": "Closer to shipped",
                "rationale": "testing is weak",
            }
        ],
        "hotspots": [
            {
                "category": "finish-line",
                "severity": 0.72,
                "title": "Promising but under-finished",
                "summary": "Worth finishing",
                "recommended_action": "Strengthen tests",
            }
        ],
        "implementation_hotspots": [
            {
                "scope": "file",
                "path": "src/core.py",
                "module": "src",
                "category": "code-complexity",
                "pressure_score": 0.74,
                "suggestion_type": "refactor",
                "why_it_matters": "src/core.py is carrying concentrated complexity.",
                "suggested_first_move": "Split the biggest function and add one regression test.",
                "signal_summary": "Complexity pressure 0.74 across 2 complex blocks.",
            }
        ],
        "analyzer_results": results,
    }


def _make_report(audits=None) -> dict:
    if audits is None:
        audits = [
            _make_audit("RepoA", 0.85, "A", "shipped"),
            _make_audit("RepoB", 0.60, "C", "functional"),
            _make_audit("RepoC", 0.40, "D", "wip", security_score=0.3, security_details={"secrets_found": 2, "dangerous_files": [".env"], "has_security_md": False, "has_dependabot": False}),
        ]
    return {
        "username": "user",
        "generated_at": "2026-03-29T10:00:00+00:00",
        "audits": audits,
        "repos_audited": len(audits),
        "average_score": 0.62,
        "portfolio_grade": "B",
        "hotspots": [
            {
                "repo": "RepoC",
                "category": "security-debt",
                "severity": 0.83,
                "title": "Security posture needs attention",
                "summary": "Needs attention",
                "recommended_action": "Improve security posture",
                "tier": "wip",
            }
        ],
        "implementation_hotspots": [
            {
                "repo": "RepoC",
                "scope": "file",
                "path": "src/core.py",
                "module": "src",
                "category": "code-complexity",
                "pressure_score": 0.74,
                "suggestion_type": "refactor",
                "why_it_matters": "src/core.py is carrying concentrated complexity.",
                "suggested_first_move": "Split the biggest function and add one regression test.",
                "signal_summary": "Complexity pressure 0.74 across 2 complex blocks.",
            }
        ],
        "implementation_hotspots_summary": {
            "summary": "1 repos have concrete implementation pressure. Start with RepoC in file src/core.py.",
        },
        "collections": {
            "showcase": {
                "description": "Worth showing",
                "repos": [{"name": "RepoA", "reason": "High value"}],
            }
        },
        "profiles": {
            "default": {"description": "Balanced"}
        },
        "lenses": {
            "ship_readiness": {"description": "Ready", "average_score": 0.7},
        },
        "scenario_summary": {
            "top_levers": [
                {
                    "key": "testing",
                    "title": "Strengthen tests",
                    "lens": "ship_readiness",
                    "repo_count": 2,
                    "average_expected_lens_delta": 0.11,
                    "projected_tier_promotions": 1,
                }
            ],
            "portfolio_projection": {
                "current_shipped": 1,
                "projected_shipped": 2,
                "projected_average_score_delta": 0.03,
            },
        },
        "security_posture": {
            "average_score": 0.61,
            "critical_repos": ["RepoC"],
            "repos_with_secrets": ["RepoC"],
            "provider_coverage": {
                "github": {"available_repos": 2, "total_repos": len(audits)},
                "scorecard": {"available_repos": 1, "total_repos": len(audits)},
            },
            "open_alerts": {"code_scanning": 2, "secret_scanning": 1},
        },
        "security_governance_preview": [
            {
                "repo": "RepoC",
                "key": "enable-codeql-default-setup",
                "priority": "high",
                "title": "Enable CodeQL default setup",
                "expected_posture_lift": 0.12,
                "effort": "medium",
                "source": "github",
                "why": "Code scanning is not configured",
            }
        ],
        "campaign_summary": {
            "campaign_type": "security-review",
            "label": "Security Review",
            "portfolio_profile": "default",
            "collection_name": None,
            "action_count": 1,
            "repo_count": 1,
        },
        "writeback_preview": {
            "repos": [
                {
                    "repo": "RepoC",
                    "topics": ["ghra-call-security-review"],
                    "issue_title": "[Repo Auditor] Security Review",
                    "notion_action_count": 1,
                    "action_ids": ["security-review-abc123"],
                }
            ]
        },
        "writeback_results": {
            "results": [
                {
                    "repo_full_name": "user/RepoC",
                    "target": "github-issue",
                    "status": "created",
                    "url": "https://github.com/user/RepoC/issues/1",
                }
            ]
        },
        "review_summary": {"review_id": "review-1", "status": "open"},
        "review_targets": [
            {
                "repo": "RepoC",
                "title": "Security posture needs attention",
                "severity": 0.83,
                "next_step": "Preview governance controls",
                "decision_hint": "ready-for-governance-approval",
                "safe_to_defer": False,
            }
        ],
        "review_history": [
            {
                "review_id": "review-1",
                "generated_at": "2026-03-29T10:00:00+00:00",
                "material_change_count": 2,
                "status": "open",
                "decision_state": "needs-review",
                "sync_state": "local-only",
                "emitted": True,
            }
        ],
        "material_changes": [
            {
                "change_type": "security",
                "repo": "RepoC",
                "severity": 0.83,
                "title": "Security posture needs attention",
            }
        ],
        "operator_summary": {
            "headline": "There is live drift or high-severity change that needs attention now.",
            "counts": {"blocked": 0, "urgent": 2, "ready": 1, "deferred": 0},
            "trend_status": "stable",
            "trend_summary": "The queue is stable but still sticky: 1 attention item is persisting from the last run.",
            "new_attention_count": 0,
            "resolved_attention_count": 0,
            "persisting_attention_count": 1,
            "reopened_attention_count": 0,
            "quiet_streak_runs": 0,
            "aging_status": "watch",
            "decision_memory_status": "attempted",
            "primary_target_last_seen_at": "2026-03-29T10:00:00+00:00",
            "primary_target_last_intervention": {
                "item_id": "review-target:RepoC",
                "repo": "RepoC",
                "title": "Security posture needs attention",
                "event_type": "drifted",
                "recorded_at": "2026-03-29T10:00:00+00:00",
                "outcome": "drifted",
            },
            "primary_target_last_outcome": "no-change",
            "primary_target_resolution_evidence": "The last intervention was drifted for RepoC: Security posture needs attention, but the item is still open.",
            "primary_target_confidence_score": 0.7,
            "primary_target_confidence_label": "medium",
            "primary_target_confidence_reasons": [
                "Urgent drift or regression needs attention before ready work.",
                "A prior intervention happened, but the item is still open.",
            ],
            "recent_interventions": [
                {
                    "item_id": "review-target:RepoC",
                    "repo": "RepoC",
                    "title": "Security posture needs attention",
                    "event_type": "drifted",
                    "recorded_at": "2026-03-29T10:00:00+00:00",
                    "outcome": "drifted",
                }
            ],
            "recently_quieted_count": 0,
            "confirmed_resolved_count": 0,
            "reopened_after_resolution_count": 0,
            "decision_memory_window_runs": 3,
            "resolution_evidence_summary": "The last intervention was drifted for RepoC: Security posture needs attention, but the item is still open.",
            "follow_through_recovery_persistence_summary": "RepoC has started calming down, but the recovery path still looks fragile and needs another confirming run.",
            "follow_through_relapse_churn_summary": "RepoC had a mild wobble after recovery began, so keep it visible until another calmer run confirms the recovery is holding.",
            "follow_through_recovery_freshness_summary": "RepoC still has some calmer carry-forward, but the recovery memory is already mixed-age and needs fresh confirmation.",
            "follow_through_recovery_memory_reset_summary": "RepoC is calmer, but the older recovery confidence should step down if the next run does not refresh it.",
            "top_fragile_recovery_items": [
                {
                    "repo": "RepoC",
                    "title": "Security posture needs attention",
                    "follow_through_recovery_persistence_summary": "RepoC is recovering, but the calmer path still looks fragile rather than settled.",
                }
            ],
            "top_sustained_recovery_items": [
                {
                    "repo": "RepoB",
                    "title": "Dependency hygiene is holding steady",
                    "follow_through_recovery_persistence_summary": "RepoB now has a calmer follow-through path that is actively holding.",
                }
            ],
            "top_churn_follow_through_items": [
                {
                    "repo": "RepoC",
                    "title": "Security posture needs attention",
                    "follow_through_relapse_churn_summary": "RepoC had one mild wobble and the recovery still looks fragile.",
                }
            ],
            "top_stale_recovery_items": [
                {
                    "repo": "RepoC",
                    "title": "Security posture needs attention",
                    "follow_through_recovery_freshness_summary": "RepoC still has some calmer carry-forward, but the recovery memory is already mixed-age and needs fresh confirmation.",
                }
            ],
            "top_softening_recovery_items": [
                {
                    "repo": "RepoC",
                    "title": "Security posture needs attention",
                    "follow_through_recovery_freshness_summary": "RepoC still has some calmer carry-forward, but the recovery memory is already mixed-age and needs fresh confirmation.",
                }
            ],
            "top_reset_recovery_items": [
                {
                    "repo": "RepoC",
                    "title": "Security posture needs attention",
                    "follow_through_recovery_memory_reset_summary": "RepoC is calmer, but the older recovery confidence should step down if the next run does not refresh it.",
                }
            ],
            "top_rebuilding_recovery_items": [],
            "next_action_confidence_score": 0.75,
            "next_action_confidence_label": "high",
            "next_action_confidence_reasons": ["The next step is tied directly to the current top target."],
            "primary_target_trust_policy": "act-with-review",
            "primary_target_trust_policy_reason": "Urgent work has enough tuned confidence to act, with a quick operator review.",
            "next_action_trust_policy": "act-with-review",
            "next_action_trust_policy_reason": "Healthy calibration supports a confident next step, with light operator judgment.",
            "primary_target_exception_status": "none",
            "primary_target_exception_reason": "",
            "primary_target_exception_pattern_status": "candidate",
            "primary_target_exception_pattern_reason": "This target is stabilizing under healthy calibration, but it has not held steady long enough to earn stronger trust yet.",
            "primary_target_trust_recovery_status": "candidate",
            "primary_target_trust_recovery_reason": "This target is stabilizing under healthy calibration, but it has not held steady long enough to earn stronger trust yet.",
            "primary_target_recovery_confidence_score": 0.72,
            "primary_target_recovery_confidence_label": "medium",
            "primary_target_recovery_confidence_reasons": [
                "Healthy calibration supports relaxing the earlier soft caution.",
                "Recent runs are stabilizing, but the retirement window is still short.",
            ],
            "recovery_confidence_summary": "RepoC: Security posture needs attention is building recovery confidence (medium, 0.72), but the earlier caution has not retired yet.",
            "primary_target_exception_retirement_status": "candidate",
            "primary_target_exception_retirement_reason": "This target is trending toward retirement, but it has not earned it yet.",
            "exception_retirement_summary": "RepoC: Security posture needs attention is trending toward exception retirement, but the evidence is not strong enough to retire it yet.",
            "retired_exception_hotspots": [],
            "sticky_exception_hotspots": [],
            "exception_retirement_window_runs": 4,
            "primary_target_policy_debt_status": "watch",
            "primary_target_policy_debt_reason": "This class has enough recent exception activity to watch for lingering caution, but it is not yet clearly sticky or clearly normalization-friendly.",
            "primary_target_class_normalization_status": "candidate",
            "primary_target_class_normalization_reason": "This class is trending healthier, but the current target has not earned class-level normalization yet.",
            "policy_debt_summary": "RepoC: Security posture needs attention sits in a class with mixed recent caution behavior, so watch for policy debt before normalizing further.",
            "trust_normalization_summary": "RepoC: Security posture needs attention belongs to a healthier class trend, but it has not earned class-level normalization yet.",
            "policy_debt_hotspots": [],
            "normalized_class_hotspots": [],
            "class_normalization_window_runs": 4,
            "primary_target_class_memory_freshness_status": "mixed-age",
            "primary_target_class_memory_freshness_reason": "Class memory is still useful, but it is partly aging: 50% of the weighted signal is recent and the rest is older carry-forward.",
            "primary_target_class_decay_status": "none",
            "primary_target_class_decay_reason": "",
            "class_memory_summary": "RepoC: Security posture needs attention still has useful class memory, but part of that signal is aging and should be treated more cautiously.",
            "class_decay_summary": "Fresh class signals are still strong enough that no class-level trust posture needs to decay yet.",
            "primary_target_weighted_class_support_score": 0.48,
            "primary_target_weighted_class_caution_score": 0.24,
            "primary_target_class_trust_reweight_score": 0.24,
            "primary_target_class_trust_reweight_direction": "supporting-normalization",
            "primary_target_class_trust_reweight_reasons": [
                "Class memory is still useful, but it is partly aging: 50% of the weighted signal is recent and the rest is older carry-forward.",
                "Existing class normalization support is still contributing to a stronger posture.",
                "Fresh sticky class evidence is still carrying meaningful caution.",
            ],
            "class_reweighting_summary": "RepoC: Security posture needs attention inherited a stronger posture because fresh class support crossed the reweight threshold (0.24).",
            "supporting_class_hotspots": [],
            "caution_class_hotspots": [],
            "class_reweighting_window_runs": 4,
            "primary_target_class_trust_momentum_score": 0.26,
            "primary_target_class_trust_momentum_status": "building",
            "primary_target_class_reweight_stability_status": "watch",
            "primary_target_class_reweight_transition_status": "pending-support",
            "primary_target_class_reweight_transition_reason": "The class signal is visible, but it has not stayed strong long enough to confirm broader normalization yet.",
            "class_momentum_summary": "RepoC: Security posture needs attention shows healthier class support, but it has not stayed persistent enough to confirm broader normalization yet (0.26).",
            "class_reweight_stability_summary": "Class guidance for RepoC: Security posture needs attention is still settling and should be watched for one more stable stretch: supporting-normalization -> neutral.",
            "class_transition_window_runs": 4,
            "primary_target_class_transition_health_status": "building",
            "primary_target_class_transition_health_reason": "The pending class signal is still accumulating in the same direction and may confirm soon.",
            "primary_target_class_transition_resolution_status": "none",
            "primary_target_class_transition_resolution_reason": "",
            "class_transition_health_summary": "RepoC: Security posture needs attention still has a pending class signal that is accumulating and may confirm soon (1 run(s)).",
            "class_transition_resolution_summary": "No pending class transition has just confirmed, cleared, or expired in the recent window.",
            "class_transition_age_window_runs": 4,
            "primary_target_transition_closure_confidence_score": 0.75,
            "primary_target_transition_closure_confidence_label": "high",
            "primary_target_transition_closure_likely_outcome": "confirm-soon",
            "primary_target_transition_closure_confidence_reasons": [
                "The pending class signal is still accumulating in the same direction and may confirm soon.",
                "Class momentum is improving in the same direction and stability remains good enough to trust.",
                "The reweight score is still strengthening in the same direction.",
            ],
            "transition_closure_confidence_summary": "RepoC: Security posture needs attention has a pending class signal that looks strong enough to confirm soon if the next run stays aligned.",
            "transition_closure_window_runs": 4,
            "primary_target_class_pending_debt_status": "watch",
            "primary_target_class_pending_debt_reason": "This class has recent pending transitions, but they are not yet resolving cleanly enough or accumulating enough debt to call it a class-level problem.",
            "class_pending_debt_summary": "RepoC: Security posture needs attention belongs to a class with some recent pending-transition drag, but not enough to call it active pending debt yet.",
            "class_pending_resolution_summary": "Recent pending transitions for RepoC: Security posture needs attention are mixed, so keep watching whether they resolve or continue to stall.",
            "class_pending_debt_window_runs": 10,
            "pending_debt_hotspots": [],
            "healthy_pending_resolution_hotspots": [],
            "primary_target_pending_debt_freshness_status": "mixed-age",
            "primary_target_pending_debt_freshness_reason": "Pending-transition memory is still useful, but it is partly aging: 50% of the weighted signal is recent and the rest is older carry-forward.",
            "pending_debt_freshness_summary": "RepoC: Security posture needs attention still has useful pending-transition memory, but some of that signal is aging and should be weighted more cautiously.",
            "pending_debt_decay_summary": "No strong pending-debt freshness trend is dominating the closure forecast yet.",
            "stale_pending_debt_hotspots": [],
            "fresh_pending_resolution_hotspots": [],
            "pending_debt_decay_window_runs": 4,
            "primary_target_weighted_pending_resolution_support_score": 0.58,
            "primary_target_weighted_pending_debt_caution_score": 0.31,
            "primary_target_closure_forecast_reweight_score": 0.27,
            "primary_target_closure_forecast_reweight_direction": "supporting-confirmation",
            "primary_target_closure_forecast_reweight_reasons": [
                "Pending-transition memory is still useful, but it is partly aging: 50% of the weighted signal is recent and the rest is older carry-forward.",
                "Recent class resolution behavior is still strong enough that this pending signal could confirm soon.",
                "The live pending signal is still building in the same direction.",
            ],
            "closure_forecast_reweighting_summary": "RepoC: Security posture needs attention still needs persistence before confirmation, but fresh class resolution behavior is strengthening the pending forecast (0.27).",
            "closure_forecast_reweighting_window_runs": 4,
            "primary_target_closure_forecast_momentum_score": 0.18,
            "primary_target_closure_forecast_momentum_status": "building",
            "primary_target_closure_forecast_stability_status": "watch",
            "primary_target_closure_forecast_hysteresis_status": "pending-confirmation",
            "primary_target_closure_forecast_hysteresis_reason": "The confirmation-leaning forecast is visible, but it has not stayed persistent enough to trust fully yet.",
            "closure_forecast_momentum_summary": "The closure forecast for RepoC: Security posture needs attention is trending in one direction, but it has not held long enough to lock in (0.18).",
            "closure_forecast_stability_summary": "Closure forecasting for RepoC: Security posture needs attention is still settling and should be watched for one more stable stretch: supporting-confirmation -> neutral.",
            "closure_forecast_hysteresis_summary": "The confirmation-leaning forecast for RepoC: Security posture needs attention is visible but not yet persistent enough to trust fully.",
            "primary_target_closure_forecast_freshness_status": "mixed-age",
            "primary_target_closure_forecast_freshness_reason": "Closure-forecast memory is still useful, but it is partly aging: 50% of the weighted forecast signal is recent and the rest is older carry-forward.",
            "primary_target_closure_forecast_decay_status": "none",
            "primary_target_closure_forecast_decay_reason": "",
            "closure_forecast_freshness_summary": "RepoC: Security posture needs attention still has useful closure-forecast memory, but some of that signal is aging and should be weighted more cautiously.",
            "closure_forecast_decay_summary": "Recent closure-forecast evidence is still fresh enough that no forecast carry-forward needs to decay yet.",
            "primary_target_closure_forecast_refresh_recovery_score": 0.16,
            "primary_target_closure_forecast_refresh_recovery_status": "recovering-confirmation",
            "primary_target_closure_forecast_reacquisition_status": "pending-confirmation-reacquisition",
            "primary_target_closure_forecast_reacquisition_reason": "Fresh confirmation-side forecast evidence is returning, but it has not fully re-earned stronger carry-forward yet.",
            "closure_forecast_refresh_recovery_summary": "Fresh confirmation-side forecast evidence is returning for RepoC: Security posture needs attention, but it has not fully re-earned stronger carry-forward yet (0.16).",
            "closure_forecast_reacquisition_summary": "Fresh confirmation-side forecast evidence is returning, but it has not fully re-earned stronger carry-forward yet.",
            "primary_target_closure_forecast_reacquisition_age_runs": 1,
            "primary_target_closure_forecast_reacquisition_persistence_score": 0.19,
            "primary_target_closure_forecast_reacquisition_persistence_status": "just-reacquired",
            "primary_target_closure_forecast_reacquisition_persistence_reason": "Stronger closure-forecast posture has returned, but it has not yet proved it can hold.",
            "closure_forecast_reacquisition_persistence_summary": "RepoC: Security posture needs attention has only just re-earned stronger closure-forecast posture, so it is still fragile (0.19; 1 run).",
            "primary_target_closure_forecast_recovery_churn_score": 0.22,
            "primary_target_closure_forecast_recovery_churn_status": "watch",
            "primary_target_closure_forecast_recovery_churn_reason": "Recovery is wobbling and may lose its restored strength soon.",
            "closure_forecast_recovery_churn_summary": "Recovery for RepoC: Security posture needs attention is wobbling enough that restored forecast strength may soften soon (0.22).",
            "primary_target_closure_forecast_reacquisition_freshness_status": "mixed-age",
            "primary_target_closure_forecast_reacquisition_freshness_reason": "Reacquired closure-forecast memory is still useful, but it is partly aging: 50% of the weighted signal is recent and the rest is older carry-forward.",
            "closure_forecast_reacquisition_freshness_summary": "RepoC: Security posture needs attention still has useful reacquired closure-forecast memory, but the restored posture is no longer getting fully fresh reinforcement.",
            "primary_target_closure_forecast_persistence_reset_status": "none",
            "primary_target_closure_forecast_persistence_reset_reason": "",
            "closure_forecast_persistence_reset_summary": "Reacquired posture for RepoC: Security posture needs attention is aging enough that it can keep holding, but it should no longer stay indefinitely at sustained strength.",
            "primary_target_closure_forecast_reset_refresh_recovery_score": 0.18,
            "primary_target_closure_forecast_reset_refresh_recovery_status": "recovering-confirmation-reset",
            "primary_target_closure_forecast_reset_reentry_status": "pending-confirmation-reentry",
            "primary_target_closure_forecast_reset_reentry_reason": "Fresh confirmation-side evidence is returning after a reset, but it has not yet re-earned re-entry.",
            "closure_forecast_reset_refresh_recovery_summary": "Fresh confirmation-side evidence is returning for RepoC: Security posture needs attention after a reset, but it has not yet re-earned re-entry (0.18).",
            "closure_forecast_reset_reentry_summary": "Fresh confirmation-side evidence is returning after a reset, but it has not yet re-earned re-entry.",
            "primary_target_closure_forecast_reset_reentry_age_runs": 1,
            "primary_target_closure_forecast_reset_reentry_persistence_score": 0.24,
            "primary_target_closure_forecast_reset_reentry_persistence_status": "just-reentered",
            "primary_target_closure_forecast_reset_reentry_persistence_reason": "Stronger closure-forecast posture has re-entered after reset, but it has not yet proved it can hold.",
            "closure_forecast_reset_reentry_persistence_summary": "RepoC: Security posture needs attention has only just re-entered stronger closure-forecast posture after reset, so it is still fragile (0.24; 1 run).",
            "primary_target_closure_forecast_reset_reentry_churn_score": 0.18,
            "primary_target_closure_forecast_reset_reentry_churn_status": "none",
            "primary_target_closure_forecast_reset_reentry_churn_reason": "",
            "closure_forecast_reset_reentry_churn_summary": "No meaningful reset re-entry churn is active right now.",
            "primary_target_closure_forecast_reset_reentry_freshness_status": "mixed-age",
            "primary_target_closure_forecast_reset_reentry_freshness_reason": "Reset re-entry memory is still useful, but it is partly aging: 50% of the weighted signal is recent and the rest is older carry-forward.",
            "closure_forecast_reset_reentry_freshness_summary": "RepoC: Security posture needs attention still has useful reset re-entry memory, but the restored posture is no longer getting fully fresh reinforcement.",
            "primary_target_closure_forecast_reset_reentry_reset_status": "none",
            "primary_target_closure_forecast_reset_reentry_reset_reason": "",
            "closure_forecast_reset_reentry_reset_summary": "Reset re-entry posture for RepoC: Security posture needs attention is aging enough that it can keep holding, but it should no longer stay indefinitely at sustained strength.",
            "primary_target_closure_forecast_reset_reentry_refresh_recovery_score": 0.31,
            "primary_target_closure_forecast_reset_reentry_refresh_recovery_status": "rebuilding-confirmation-reentry",
            "primary_target_closure_forecast_reset_reentry_rebuild_status": "pending-confirmation-rebuild",
            "primary_target_closure_forecast_reset_reentry_rebuild_reason": "Fresh confirmation-side evidence is rebuilding after the reset re-entry aged out, but it has not yet fully re-earned stronger posture.",
            "closure_forecast_reset_reentry_refresh_recovery_summary": "Fresh confirmation-side evidence is rebuilding for RepoC: Security posture needs attention after reset re-entry aged out (0.31).",
            "closure_forecast_reset_reentry_rebuild_summary": "RepoC: Security posture needs attention is rebuilding confirmation-side reset re-entry, but it has not fully re-earned stronger posture yet.",
            "primary_target_closure_forecast_reset_reentry_rebuild_freshness_status": "mixed-age",
            "primary_target_closure_forecast_reset_reentry_rebuild_freshness_reason": "Rebuilt reset re-entry memory is still useful, but it is partly aging: 50% of the weighted signal is recent and the rest is older carry-forward.",
            "closure_forecast_reset_reentry_rebuild_freshness_summary": "RepoC: Security posture needs attention still has useful rebuilt reset re-entry memory, but the restored posture is no longer getting fully fresh reinforcement.",
            "primary_target_closure_forecast_reset_reentry_rebuild_reset_status": "none",
            "primary_target_closure_forecast_reset_reentry_rebuild_reset_reason": "",
            "closure_forecast_reset_reentry_rebuild_reset_summary": "Rebuilt posture for RepoC: Security posture needs attention is aging enough that it can keep holding, but it should no longer stay indefinitely at sustained strength.",
            "primary_target_closure_forecast_reset_reentry_rebuild_refresh_recovery_score": 0.27,
            "primary_target_closure_forecast_reset_reentry_rebuild_refresh_recovery_status": "recovering-confirmation-rebuild-reset",
            "primary_target_closure_forecast_reset_reentry_rebuild_reentry_status": "pending-confirmation-rebuild-reentry",
            "primary_target_closure_forecast_reset_reentry_rebuild_reentry_reason": "Fresh confirmation-side evidence is returning after rebuilt posture was softened or reset, but it has not yet re-earned stronger rebuilt posture.",
            "closure_forecast_reset_reentry_rebuild_refresh_recovery_summary": "Fresh confirmation-side evidence is returning for RepoC: Security posture needs attention after rebuilt posture softened, but it has not yet re-earned stronger rebuilt posture (0.27).",
            "closure_forecast_reset_reentry_rebuild_reentry_summary": "RepoC: Security posture needs attention is recovering after rebuilt posture softened, but stronger rebuilt confirmation posture still needs more fresh follow-through before it is re-earned.",
            "primary_target_closure_forecast_reset_reentry_rebuild_reentry_age_runs": 1,
            "primary_target_closure_forecast_reset_reentry_rebuild_reentry_persistence_score": 0.26,
            "primary_target_closure_forecast_reset_reentry_rebuild_reentry_persistence_status": "just-reentered",
            "primary_target_closure_forecast_reset_reentry_rebuild_reentry_persistence_reason": "Stronger rebuilt posture has been re-earned, but it has not yet proved it can hold.",
            "closure_forecast_reset_reentry_rebuild_reentry_persistence_summary": "RepoC: Security posture needs attention has only just re-earned stronger rebuilt posture, so it is still fragile (0.26; 1 run).",
            "primary_target_closure_forecast_reset_reentry_rebuild_reentry_churn_score": 0.0,
            "primary_target_closure_forecast_reset_reentry_rebuild_reentry_churn_status": "none",
            "primary_target_closure_forecast_reset_reentry_rebuild_reentry_churn_reason": "",
            "closure_forecast_reset_reentry_rebuild_reentry_churn_summary": "No meaningful rebuilt re-entry churn is active right now.",
            "primary_target_closure_forecast_reset_reentry_rebuild_age_runs": 1,
            "primary_target_closure_forecast_reset_reentry_rebuild_persistence_score": 0.29,
            "primary_target_closure_forecast_reset_reentry_rebuild_persistence_status": "just-rebuilt",
            "primary_target_closure_forecast_reset_reentry_rebuild_persistence_reason": "Stronger reset re-entry posture has been rebuilt, but it has not yet proved it can hold.",
            "closure_forecast_reset_reentry_rebuild_persistence_summary": "RepoC: Security posture needs attention has only just rebuilt stronger reset re-entry posture, so it is still fragile (0.29; 1 run).",
            "primary_target_closure_forecast_reset_reentry_rebuild_churn_score": 0.14,
            "primary_target_closure_forecast_reset_reentry_rebuild_churn_status": "none",
            "primary_target_closure_forecast_reset_reentry_rebuild_churn_reason": "",
            "closure_forecast_reset_reentry_rebuild_churn_summary": "No meaningful reset re-entry rebuild churn is active right now.",
            "just_rebuilt_hotspots": [],
            "just_reentered_rebuild_hotspots": [],
            "holding_reset_reentry_rebuild_hotspots": [],
            "holding_reset_reentry_rebuild_reentry_hotspots": [],
            "reset_reentry_rebuild_churn_hotspots": [],
            "reset_reentry_rebuild_reentry_churn_hotspots": [],
            "stale_reset_reentry_rebuild_hotspots": [],
            "fresh_reset_reentry_rebuild_signal_hotspots": [],
            "stale_reset_reentry_hotspots": [],
            "fresh_reset_reentry_signal_hotspots": [],
            "closure_forecast_reset_reentry_decay_window_runs": 4,
            "closure_forecast_reset_reentry_refresh_window_runs": 4,
            "closure_forecast_reset_reentry_rebuild_window_runs": 4,
            "closure_forecast_reset_reentry_rebuild_decay_window_runs": 4,
            "closure_forecast_reset_reentry_rebuild_refresh_window_runs": 4,
            "closure_forecast_reset_reentry_rebuild_reentry_window_runs": 4,
            "stale_closure_forecast_hotspots": [],
            "fresh_closure_forecast_signal_hotspots": [],
            "closure_forecast_decay_window_runs": 4,
            "closure_forecast_refresh_window_runs": 4,
            "closure_forecast_reacquisition_window_runs": 4,
            "closure_forecast_reacquisition_decay_window_runs": 4,
            "closure_forecast_transition_window_runs": 4,
            "sustained_confirmation_hotspots": [],
            "sustained_clearance_hotspots": [],
            "oscillating_closure_forecast_hotspots": [],
            "recovering_confirmation_hotspots": [],
            "recovering_clearance_hotspots": [],
            "just_reacquired_hotspots": [],
            "holding_reacquisition_hotspots": [],
            "recovery_churn_hotspots": [],
            "stale_reacquisition_hotspots": [],
            "fresh_reacquisition_signal_hotspots": [],
            "recovering_from_confirmation_reentry_reset_hotspots": [],
            "recovering_from_clearance_reentry_reset_hotspots": [],
            "recovering_from_confirmation_rebuild_reset_hotspots": [],
            "recovering_from_clearance_rebuild_reset_hotspots": [],
            "supporting_pending_resolution_hotspots": [],
            "caution_pending_debt_hotspots": [],
            "stalled_transition_hotspots": [],
            "resolving_transition_hotspots": [],
            "sustained_class_hotspots": [],
            "oscillating_class_hotspots": [],
            "stale_class_memory_hotspots": [],
            "fresh_class_signal_hotspots": [],
            "class_decay_window_runs": 4,
            "recommendation_drift_status": "stable",
            "recommendation_drift_summary": "Recent trust-policy behavior is stable enough that no meaningful recommendation drift is recorded.",
            "policy_flip_hotspots": [],
            "exception_pattern_summary": "RepoC: Security posture needs attention is stabilizing, but it has not yet earned stronger trust.",
            "false_positive_exception_hotspots": [],
            "trust_recovery_window_runs": 3,
            "adaptive_confidence_summary": "Calibration is validating well, so the recommendation can be acted on with light operator review.",
            "recommendation_quality_summary": "Strong recommendation because the next step is tied directly to the current top target.",
            "confidence_validation_status": "healthy",
            "confidence_window_runs": 8,
            "validated_recommendation_count": 4,
            "partially_validated_recommendation_count": 1,
            "unresolved_recommendation_count": 1,
            "reopened_recommendation_count": 0,
            "insufficient_future_runs_count": 2,
            "high_confidence_hit_rate": 0.75,
            "medium_confidence_hit_rate": 0.5,
            "low_confidence_caution_rate": 1.0,
            "recent_validation_outcomes": [
                {
                    "run_id": "user:2026-03-27T10:00:00+00:00",
                    "target_label": "RepoC: Security posture needs attention",
                    "confidence_label": "high",
                    "outcome": "validated",
                    "validated_in_runs": 2,
                }
            ],
            "confidence_calibration_summary": "Recent high-confidence recommendations are validating well: 75% high-confidence hit rate across 6 judged runs with no reopen noise.",
            "primary_target_reason": "This remains the top target because urgent review work should be closed before lower-pressure ready items.",
            "primary_target_done_criteria": "Complete the recommended review action and confirm the item exits urgent state on the next run.",
            "closure_guidance": "Preview the governance controls and confirm this urgent review target clears on the next run.",
            "attention_age_bands": {"0-1 days": 1, "2-7 days": 0, "8-21 days": 0, "22+ days": 0},
            "chronic_item_count": 0,
            "newly_stale_count": 0,
            "longest_persisting_item": {
                "repo": "RepoC",
                "title": "Security posture needs attention",
                "lane": "urgent",
                "age_days": 0,
                "aging_status": "watch",
            },
            "accountability_summary": "The urgent queue is still active, but no items have crossed into chronic aging pressure yet.",
            "primary_target": {
                "item_id": "review-target:RepoC",
                "repo": "RepoC",
                "title": "Security posture needs attention",
                "trust_recovery_status": "candidate",
                "trust_recovery_reason": "This target is stabilizing under healthy calibration, but it has not held steady long enough to earn stronger trust yet.",
                "exception_pattern_status": "recovering",
                "exception_pattern_reason": "This target is stabilizing under healthy calibration, but it has not held steady long enough to earn stronger trust yet.",
                "class_transition_age_runs": 1,
                "recent_transition_path": "pending-support",
            },
        },
        "operator_queue": [
            {
                "item_id": "review-target:RepoC",
                "kind": "review",
                "lane": "urgent",
                "priority": 83,
                "repo": "RepoC",
                "title": "Security posture needs attention",
                "summary": "Governance approval is ready.",
                "recommended_action": "Preview governance controls",
                "source_run_id": "user:2026-03-29T10:00:00+00:00",
                "age_days": 0,
                "aging_status": "watch",
                "newly_stale": False,
                "links": [],
                "follow_through_status": "waiting-on-evidence",
                "follow_through_age_runs": 1,
                "follow_through_checkpoint_status": "due-soon",
                "follow_through_summary": "RepoC has recent follow-up recorded and is now waiting for confirming evidence.",
                "follow_through_next_checkpoint": "Wait for the next run to confirm the pressure falls.",
                "follow_through_escalation_status": "watch",
                "follow_through_escalation_summary": "RepoC is not late yet, but it should stay visible until the next checkpoint lands.",
                "follow_through_recovery_age_runs": 1,
                "follow_through_recovery_status": "recovering",
                "follow_through_recovery_summary": "RepoC is recovering from recent escalation, but the lower-pressure state still needs to hold.",
                "follow_through_recovery_persistence_age_runs": 1,
                "follow_through_recovery_persistence_status": "fragile-recovery",
                "follow_through_recovery_persistence_summary": "RepoC is recovering, but the calmer path still looks fragile rather than settled.",
                "follow_through_relapse_churn_status": "fragile",
                "follow_through_relapse_churn_summary": "RepoC had one mild wobble and the recovery still looks fragile.",
                "follow_through_recovery_freshness_age_runs": 1,
                "follow_through_recovery_freshness_status": "mixed-age",
                "follow_through_recovery_freshness_summary": "RepoC still has some calmer carry-forward, but the recovery memory is already mixed-age and needs fresh confirmation.",
                "follow_through_recovery_decay_status": "softening",
                "follow_through_recovery_decay_summary": "RepoC still looks calmer, but the recovery memory is softening without a fully fresh confirming run yet.",
                "follow_through_recovery_memory_reset_status": "reset-watch",
                "follow_through_recovery_memory_reset_summary": "RepoC is calmer, but the older recovery confidence should step down if the next run does not refresh it.",
            }
        ],
        "governance_preview": {"applyable_count": 1},
        "governance_drift": [{"repo": "RepoC", "control": "secret_scanning"}],
    }


class TestRadarChart:
    def test_creates_sheet(self):
        wb = Workbook()
        _build_repo_profiles(wb, _make_report())
        assert "Repo Profiles" in wb.sheetnames

    def test_has_charts(self):
        wb = Workbook()
        _build_repo_profiles(wb, _make_report())
        ws = wb["Repo Profiles"]
        assert len(ws._charts) >= 1

    def test_dimension_labels(self):
        wb = Workbook()
        _build_repo_profiles(wb, _make_report())
        ws = wb["Repo Profiles"]
        labels = [ws.cell(row=i+2, column=1).value for i in range(len(RADAR_LABELS))]
        assert labels == RADAR_LABELS


class TestSecuritySheet:
    def test_creates_sheet(self):
        wb = Workbook()
        _build_security(wb, _make_report())
        assert "Security" in wb.sheetnames

    def test_sorts_by_score_ascending(self):
        wb = Workbook()
        _build_security(wb, _make_report())
        ws = wb["Security"]
        # Row 2 should be worst security score
        assert ws.cell(row=2, column=2).value <= ws.cell(row=3, column=2).value

    def test_shows_secrets_found(self):
        wb = Workbook()
        _build_security(wb, _make_report())
        ws = wb["Security"]
        # Find RepoC (has 2 secrets) — should be first (lowest score)
        assert ws.cell(row=2, column=3).value == 2


class TestChangesSheet:
    def test_no_sheet_without_diff(self):
        wb = Workbook()
        _build_changes(wb, _make_report(), None)
        assert "Changes" not in wb.sheetnames

    def test_creates_sheet_with_diff(self):
        wb = Workbook()
        diff = {
            "tier_changes": [{"name": "RepoA", "old_tier": "functional", "new_tier": "shipped", "direction": "promotion", "old_score": 0.72, "new_score": 0.80}],
            "score_changes": [{"name": "RepoA", "old_score": 0.72, "new_score": 0.80, "delta": 0.08}],
            "average_score_delta": 0.03,
        }
        _build_changes(wb, _make_report(), diff)
        assert "Changes" in wb.sheetnames

    def test_shows_promotions(self):
        wb = Workbook()
        diff = {
            "tier_changes": [{"name": "RepoA", "old_tier": "functional", "new_tier": "shipped", "direction": "promotion", "old_score": 0.72, "new_score": 0.80}],
            "score_changes": [],
            "average_score_delta": 0.03,
        }
        _build_changes(wb, _make_report(), diff)
        ws = wb["Changes"]
        assert ws.cell(row=3, column=2).value == 1  # 1 promotion


class TestAllReposSheet:
    def _build(self, score_history=None):
        wb = Workbook()
        _build_all_repos(wb, _make_report(), score_history)
        return wb["All Repos"]

    def test_creates_sheet(self):
        wb = Workbook()
        _build_all_repos(wb, _make_report())
        assert "All Repos" in wb.sheetnames

    def test_headers_include_interest_breakdown(self):
        ws = self._build()
        header_values = [ws.cell(row=1, column=c).value for c in range(1, 40)]
        for label in ("Tech Novelty", "Burst", "Ambition", "Storytelling"):
            assert label in header_values, f"Missing header: {label}"

    def test_trend_is_last_column(self):
        ws = self._build()
        # Trend is dynamically the last header column (currently 36)
        trend_col = None
        for c in range(1, 40):
            if ws.cell(row=1, column=c).value == "Trend":
                trend_col = c
                break
        assert trend_col is not None, "Trend header not found"

    def test_iconset_rule_on_score_column(self):
        wb = Workbook()
        _build_all_repos(wb, _make_report())
        ws = wb["All Repos"]
        rules = [
            rule
            for rule_list in ws.conditional_formatting._cf_rules.values()
            for rule in rule_list
            if rule.type == "iconSet"
        ]
        assert len(rules) >= 1, "Expected at least one iconSet rule on All Repos sheet"

    def test_sparkline_written_to_trend_column(self):
        score_history = {"RepoA": [0.5, 0.6, 0.7, 0.8, 0.85]}
        ws = self._build(score_history=score_history)
        # Find the Trend column dynamically
        trend_col = None
        for c in range(1, 40):
            if ws.cell(row=1, column=c).value == "Trend":
                trend_col = c
                break
        assert trend_col is not None, "Trend header not found"
        # Find RepoA row (sorted by score descending, RepoA has highest score)
        repoA_row = None
        for r in range(2, 10):
            if ws.cell(row=r, column=1).value == "RepoA":
                repoA_row = r
                break
        assert repoA_row is not None, "RepoA row not found"
        spark_val = ws.cell(row=repoA_row, column=trend_col).value
        assert spark_val is not None, "Expected sparkline in Trend column"


class TestSecurityIconSetRule:
    def test_iconset_rule_on_security_score_column(self):
        wb = Workbook()
        _build_security(wb, _make_report())
        ws = wb["Security"]
        rules = [
            rule
            for rule_list in ws.conditional_formatting._cf_rules.values()
            for rule in rule_list
            if rule.type == "iconSet"
        ]
        assert len(rules) >= 1, "Expected at least one iconSet rule on Security sheet"


class TestHotspotsAndDataSheets:
    def test_creates_hotspots_sheet(self):
        wb = Workbook()
        _build_hotspots(wb, _make_report())
        assert "Hotspots" in wb.sheetnames

    def test_creates_implementation_hotspots_sheet(self):
        wb = Workbook()
        _build_implementation_hotspots(wb, _make_report())
        assert "Implementation Hotspots" in wb.sheetnames
        ws = wb["Implementation Hotspots"]
        assert ws["A1"].value == "Implementation Hotspots"
        assert "Start with RepoC" in str(ws["A2"].value)

    def test_creates_operator_outcomes_sheet(self):
        wb = Workbook()
        _build_operator_outcomes(wb, _make_report())
        assert "Operator Outcomes" in wb.sheetnames
        ws = wb["Operator Outcomes"]
        assert ws["A1"].value == "Operator Outcomes"

    def test_hidden_data_sheets_are_created(self):
        wb = Workbook()
        _build_hidden_data_sheets(wb, _make_report(), trend_data=[{"average_score": 0.6}], score_history={"RepoA": [0.5, 0.8]})
        assert "Data_Repos" in wb.sheetnames
        assert "Data_Lenses" in wb.sheetnames
        assert "Data_TrendMatrix" in wb.sheetnames
        assert "Data_PortfolioHistory" in wb.sheetnames
        assert "Data_Rollups" in wb.sheetnames
        assert "Data_ReviewTargets" in wb.sheetnames
        assert "Data_OperatorQueue" in wb.sheetnames
        assert "Data_OperatorOutcomes" in wb.sheetnames
        assert "Data_ActionSyncOutcomes" in wb.sheetnames
        assert "Data_CampaignTuning" in wb.sheetnames
        assert "Data_InterventionLedger" in wb.sheetnames
        assert "Data_ActionSyncAutomation" in wb.sheetnames
        assert "Data_OperatorRepoRollups" in wb.sheetnames
        assert "Data_MaterialChangeRollups" in wb.sheetnames
        assert "Data_RepoDetail" in wb.sheetnames
        assert "Data_ImplementationHotspots" in wb.sheetnames
        assert "Data_RepoDimensionRollups" in wb.sheetnames
        assert "Data_RepoHistoryRollups" in wb.sheetnames
        assert "Data_RunChangeRollups" in wb.sheetnames
        assert "Data_RunChangeRepoData" in wb.sheetnames
        assert wb["Data_Repos"].sheet_state == "hidden"

    def test_hidden_repo_sheet_has_structured_table(self):
        wb = Workbook()
        _build_hidden_data_sheets(wb, _make_report())
        ws = wb["Data_Repos"]
        assert ws.tables
        assert "tblRepos" in ws.tables


class TestAnalystWorkbookSheets:
    def test_creates_portfolio_explorer_and_by_lens(self):
        wb = Workbook()
        _build_portfolio_explorer(wb, _make_report(), portfolio_profile="default", collection="showcase")
        _build_by_lens(wb, _make_report(), portfolio_profile="default", collection="showcase")
        assert "Portfolio Explorer" in wb.sheetnames
        assert "By Lens" in wb.sheetnames

    def test_creates_campaign_and_writeback_sheets(self):
        wb = Workbook()
        _build_campaigns(wb, _make_report())
        _build_writeback_audit(wb, _make_report())
        assert "Campaigns" in wb.sheetnames
        assert "Writeback Audit" in wb.sheetnames

    def test_creates_historical_intelligence_sheet(self):
        wb = Workbook()
        report = _make_report()
        report["intervention_ledger_summary"] = {
            "summary": "RepoC is improving after intervention while RepoB still looks relapsing.",
        }
        report["next_historical_focus"] = {
            "summary": "Read RepoC next: it is the clearest current example of improvement after intervention.",
        }
        report["top_improving_repos"] = [
            {
                "repo": "RepoC",
                "historical_intelligence_status": "improving-after-intervention",
                "pressure_trend": "improving",
                "hotspot_persistence": "changing",
                "scorecard_trend": "improving",
                "summary": "RepoC is improving after intervention.",
            }
        ]

        _build_historical_intelligence(wb, report)
        ws = wb["Historical Intelligence"]

        assert ws["A1"].value == "Historical Portfolio Intelligence"
        assert "RepoC is improving after intervention" in str(ws["A2"].value)
        assert "Read RepoC next" in str(ws["A3"].value)

    def test_creates_collection_trend_review_governance_and_print_sheets(self):
        wb = Workbook()
        report = _make_report()
        _build_by_collection(wb, report, portfolio_profile="default")
        _build_trend_summary(wb, report, trend_data=[{"date": "2026-03-28", "average_score": 0.6, "repos_audited": 3, "tier_distribution": {"shipped": 1, "functional": 1}}], score_history={"RepoA": [0.5, 0.8]})
        _build_repo_detail(wb, report)
        _build_run_changes(wb, report, {"score_changes": []})
        _build_review_queue(wb, report)
        _build_review_history_sheet(wb, report)
        _build_governance_controls(wb, report)
        _build_governance_audit(wb, report)
        _build_print_pack(wb, report, None, portfolio_profile="default", collection="showcase")
        assert "By Collection" in wb.sheetnames
        assert "Trend Summary" in wb.sheetnames
        assert "Repo Detail" in wb.sheetnames
        assert "Run Changes" in wb.sheetnames
        assert "Review Queue" in wb.sheetnames
        assert "Review History" in wb.sheetnames
        assert "Governance Controls" in wb.sheetnames
        assert "Governance Audit" in wb.sheetnames
        assert "Print Pack" in wb.sheetnames

    def test_repo_detail_uses_selector_and_data_validation(self):
        wb = Workbook()
        _build_repo_detail(wb, _make_report())
        ws = wb["Repo Detail"]
        assert "Choose one repo" in str(ws["A3"].value)
        assert ws["A13"].value == "Current State"
        assert ws["E13"].value == "Why This Repo Looks This Way"
        assert ws["F20"].value == "What Changed"
        assert ws["F25"].value == "What To Do Next"
        assert ws["A28"].value == "Where To Start"
        assert ws["A29"].value == "Summary"
        assert ws["F28"].value == "Follow-Through Status"
        assert ws["F29"].value == "Follow-Through Summary"
        assert ws["F30"].value == "Checkpoint Timing"
        assert ws["F31"].value == "Escalation"
        assert ws["F32"].value == "Escalation Summary"
        assert ws["F33"].value == "Recovery / Retirement"
        assert ws["F34"].value == "Recovery Summary"
        assert ws["F35"].value == "Recovery Persistence"
        assert ws["F36"].value == "Recovery Persistence Summary"
        assert ws["F37"].value == "Relapse Churn"
        assert ws["F38"].value == "Relapse Churn Summary"
        assert ws["A4"].value == "Select Repo"
        assert ws["B4"].value == "RepoA"
        assert ws.data_validations.dataValidation
        assert "VLOOKUP" in str(ws["B6"].value)
        assert "No description recorded yet." in str(ws["B11"].value)
        assert "VLOOKUP" in str(ws["B29"].value)

    def test_repo_detail_preserves_existing_selection_when_still_valid(self):
        wb = Workbook()
        _build_repo_detail(wb, _make_report())
        ws = wb["Repo Detail"]
        ws["B4"] = "RepoC"

        _build_repo_detail(wb, _make_report())

        assert wb["Repo Detail"]["B4"].value == "RepoC"

    def test_repo_detail_formulas_include_safe_placeholder_copy(self):
        audits = [
            _make_audit("RepoA", 0.85, "A", "shipped")
        ]
        audits[0]["metadata"]["description"] = ""
        audits[0]["metadata"]["language"] = None
        audits[0]["action_candidates"] = []
        audits[0]["hotspots"] = []
        audits[0]["score_explanation"] = {
            "top_positive_drivers": [],
            "top_negative_drivers": [],
            "next_tier_gap_summary": "",
            "next_best_action": "",
            "next_best_action_rationale": "",
        }

        wb = Workbook()
        _build_hidden_data_sheets(wb, _make_report(audits=audits))
        _build_repo_detail(wb, _make_report(audits=audits))
        ws = wb["Repo Detail"]

        assert "Unknown" in str(ws["E6"].value)
        assert "No description recorded yet." in str(ws["B11"].value)
        assert "No briefing detail recorded yet." in str(ws["F18"].value)
        assert "Unknown" in str(ws["G28"].value)
        assert "Unknown" in str(ws["G30"].value)
        assert "No follow-through recovery or escalation-retirement signal is currently surfaced." in str(ws["G34"].value)
        assert "No follow-through recovery persistence signal is currently surfaced." in str(ws["G36"].value)
        assert "No relapse churn is currently surfaced." in str(ws["G38"].value)
        assert "No clear next action is recorded yet." in str(ws["G26"].value)

    def test_run_changes_surfaces_summary(self):
        wb = Workbook()
        diff = {
            "score_changes": [{"name": "RepoA", "old_score": 0.7, "new_score": 0.8, "delta": 0.1}],
            "tier_changes": [],
            "repo_changes": [],
        }
        _build_run_changes(wb, _make_report(), diff)
        ws = wb["Run Changes"]
        assert ws["A1"].value == "Run Changes"
        assert ws["A3"].value is not None
        assert "Read this page top to bottom" in str(ws["A4"].value)
        assert ws["A5"].value == "Comparison Window"

    def test_run_changes_shows_safe_first_run_copy_without_diff(self):
        wb = Workbook()
        _build_run_changes(wb, _make_report(), None)
        ws = wb["Run Changes"]
        assert no_baseline_summary() == str(ws["B5"].value)

    def test_review_queue_has_summary_and_freeze_panes(self):
        wb = Workbook()
        _build_review_queue(wb, _make_report())
        ws = wb["Review Queue"]
        assert "Work this page in lane order" in str(ws["A3"].value)
        assert ws["A4"].value == "Summary"
        assert ws["E4"].value == "Top 10 To Act On"
        header_row = next(row for row in range(20, 80) if ws.cell(row=row, column=1).value == "Repo")
        assert header_row > 24
        assert ws.freeze_panes == f"A{header_row + 1}"
        assert ws.cell(row=header_row, column=8).value == "Catalog"
        assert ws.cell(row=header_row, column=9).value == "Intent Alignment"
        assert ws.cell(row=header_row, column=10).value == "Maturity"
        assert ws.cell(row=header_row, column=11).value == "Scorecard Gap"
        assert ws.cell(row=header_row, column=12).value == "Last Movement"
        assert ws.cell(row=header_row, column=13).value == "Follow-Through"
        assert ws.cell(row=header_row, column=14).value == "Next Checkpoint"
        assert ws.cell(row=header_row, column=15).value == "Checkpoint Timing"
        assert ws.cell(row=header_row, column=16).value == "Escalation"
        assert ws.cell(row=header_row, column=17).value == "Escalation Summary"
        assert ws.cell(row=header_row, column=18).value == "Recovery / Retirement"
        assert ws.cell(row=header_row, column=19).value == "Recovery Summary"
        assert ws.cell(row=header_row, column=20).value == "Recovery Persistence"
        assert ws.cell(row=header_row, column=21).value == "Persistence Summary"
        assert ws.cell(row=header_row, column=22).value == "Relapse Churn"
        assert ws.cell(row=header_row, column=23).value == "Churn Summary"
        assert ws.cell(row=header_row, column=24).value == "Recovery Freshness"
        assert ws.cell(row=header_row, column=25).value == "Freshness Summary"
        assert ws.cell(row=header_row, column=26).value == "Recovery Memory Reset"
        assert ws.cell(row=header_row, column=27).value == "Reset Summary"
        assert ws.cell(row=header_row, column=28).value == "Recovery Rebuild Strength"
        assert ws.cell(row=header_row, column=29).value == "Rebuild Summary"
        assert ws.cell(row=header_row, column=30).value == "Recovery Reacquisition"
        assert ws.cell(row=header_row, column=31).value == "Reacquisition Summary"
        assert ws.cell(row=header_row, column=32).value == "Reacquisition Durability"
        assert ws.cell(row=header_row, column=33).value == "Durability Summary"
        assert ws.cell(row=header_row, column=34).value == "Reacquisition Confidence"
        assert ws.cell(row=header_row, column=35).value == "Confidence Summary"
        assert ws.cell(row=header_row, column=36).value == "Operator Focus"
        assert ws.cell(row=header_row, column=37).value == "Focus Summary"
        assert ws.cell(row=header_row, column=38).value == "Focus Line"
        assert ws.cell(row=header_row, column=39).value == "Open Artifact"
        assert no_linked_artifact_summary() in {
            ws.cell(row=row, column=39).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "No portfolio catalog contract is recorded yet." in {
            ws.cell(row=row, column=8).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert any("missing-contract" in value for value in {
            str(ws.cell(row=row, column=9).value)
            for row in range(header_row + 1, header_row + 10)
        })
        assert "—" in {
            ws.cell(row=row, column=10).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "No maturity gap summary is recorded yet." in {
            ws.cell(row=row, column=11).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "RepoC has recent follow-up recorded and is now waiting for confirming evidence." in {
            ws.cell(row=row, column=13).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "Due Soon" in {
            ws.cell(row=row, column=15).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "Recovering" in {
            ws.cell(row=row, column=18).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "Fragile Recovery" in {
            ws.cell(row=row, column=20).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "Fragile" in {
            ws.cell(row=row, column=22).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "Mixed Age" in {
            ws.cell(row=row, column=24).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "Reset Watch" in {
            ws.cell(row=row, column=26).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "None" in {
            ws.cell(row=row, column=32).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "None" in {
            ws.cell(row=row, column=34).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "Act Now" in {
            ws.cell(row=row, column=36).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert any(
            "Act Now:" in str(ws.cell(row=row, column=38).value)
            for row in range(header_row + 1, header_row + 10)
        )
        assert "no" in {
            ws.cell(row=row, column=40).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert any(
            "stay visible" in str(ws.cell(row=row, column=37).value).lower()
            for row in range(header_row + 1, header_row + 10)
        )
        assert "None" in {
            ws.cell(row=row, column=28).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "None" in {
            ws.cell(row=row, column=30).value
            for row in range(header_row + 1, header_row + 10)
        }
        assert "None" in {
            ws.cell(row=row, column=34).value
            for row in range(header_row + 1, header_row + 10)
        }

    def test_core_sheets_expose_consistent_navigation_strip(self, tmp_path):
        report_path = tmp_path / "report.json"
        report_path.write_text(json.dumps(_make_report()))

        output = export_excel(report_path, tmp_path / "out.xlsx", excel_mode="standard")
        wb = load_workbook(output)

        for sheet_name in ["Dashboard", "Run Changes", "Review Queue", "Repo Detail", "Executive Summary"]:
            ws = wb[sheet_name]
            quick_link_cells = [
                cell
                for row in ws.iter_rows(min_row=1, max_row=7, values_only=False)
                for cell in row
                if cell.value == "Quick Links"
            ]
            assert quick_link_cells, f"missing quick links header on {sheet_name}"
            start_col = quick_link_cells[0].column
            labels = [
                ws.cell(row=row, column=start_col).value
                for row in range(2, 7)
            ]
            assert labels == [
                "Dashboard",
                "Review Queue",
                "Repo Detail",
                "Run Changes",
                "Executive Summary",
            ]

    def test_standard_operator_sheets_include_trend_callouts(self):
        wb = Workbook()
        report = _make_report()
        _build_review_queue(wb, report, excel_mode="standard")
        _build_dashboard(wb, report, excel_mode="standard")
        _build_executive_summary(wb, report, None, portfolio_profile="default", collection="showcase", excel_mode="standard")
        _build_print_pack(wb, report, None, portfolio_profile="default", collection="showcase", excel_mode="standard")

        review_ws = wb["Review Queue"]
        executive_ws = wb["Executive Summary"]

        review_labels = [
            review_ws.cell(row=row, column=1).value
            for row in range(10, 90)
            if review_ws.cell(row=row, column=1).value is not None
        ]
        assert "Trend" in review_labels
        assert "Why Top Target" in review_labels
        assert "Closure Guidance" in review_labels
        assert "Aging Pressure" in review_labels
        assert "What We Tried" in review_labels
        assert "Last Outcome" in review_labels
        assert "Resolution Evidence" in review_labels
        assert "Recommendation Confidence" in review_labels
        assert "Confidence Rationale" in review_labels
        assert "Next Action Confidence" in review_labels
        assert "Trust Policy" in review_labels
        assert "Trust Rationale" in review_labels
        assert "Trust Exception" in review_labels
        assert "Trust Recovery" in review_labels
        assert "Recovery Confidence" in review_labels
        assert "Exception Retirement" in review_labels
        assert "Retirement Summary" in review_labels
        assert "Policy Debt" in review_labels
        assert "Class Normalization" in review_labels
        assert "Class Memory" in review_labels
        assert "Trust Decay" in review_labels
        assert "Class Reweighting" in review_labels
        assert "Class Reweighting Why" in review_labels
        assert "Class Momentum" in review_labels
        assert "Reweight Stability" in review_labels
        assert "Transition Health" in review_labels
        assert "Transition Resolution" in review_labels
        assert "Transition Summary" in review_labels
        assert "Transition Closure" in review_labels
        assert "Transition Likely Outcome" in review_labels
        assert "Pending Debt Freshness" in review_labels
        assert "Closure Forecast" in review_labels
        assert "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Persistence" in review_labels
        assert "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Churn Controls" in review_labels
        assert "Closure Forecast Summary" in review_labels
        assert "Momentum Summary" in review_labels
        assert "Exception Learning" in review_labels
        assert "Recommendation Drift" in review_labels
        assert "Adaptive Confidence" in review_labels
        review_values = [
            review_ws.cell(row=row, column=2).value
            for row in range(10, 60)
            if review_ws.cell(row=row, column=2).value is not None
        ]
        assert any("stable" in str(value).lower() for value in review_values)
        assert any("urgent review work" in str(value).lower() for value in review_values)
        executive_labels = [
            executive_ws.cell(row=row, column=4).value
            for row in range(20, 75)
            if executive_ws.cell(row=row, column=4).value is not None
        ]
        assert "Trend" in executive_labels
        assert "Why Top Target" in executive_labels
        assert "Recovery Freshness" in executive_labels
        assert "Recovery Memory Reset" in executive_labels
        assert "Freshness Hotspot" in executive_labels
        assert "Rebuild Hotspot" in executive_labels

    def test_dashboard_and_executive_summary_share_operator_story_lines(self):
        wb = Workbook()
        report = _make_report()
        _build_dashboard(wb, report, excel_mode="standard")
        _build_executive_summary(wb, report, None, portfolio_profile="default", collection="showcase", excel_mode="standard")
        _build_print_pack(wb, report, None, portfolio_profile="default", collection="showcase", excel_mode="standard")

        dashboard_ws = wb["Dashboard"]
        executive_ws = wb["Executive Summary"]
        print_ws = wb["Print Pack"]

        dashboard_values = {
            cell
            for row in dashboard_ws.iter_rows(min_row=1, max_row=110, min_col=1, max_col=25, values_only=True)
            for cell in row
            if cell is not None
        }
        executive_values = {
            cell
            for row in executive_ws.iter_rows(min_row=1, max_row=140, min_col=1, max_col=20, values_only=True)
            for cell in row
            if cell is not None
        }

        queue_summary = build_queue_pressure_summary(report)
        top_recommendation = build_top_recommendation_summary(report)
        trust_summary = build_trust_actionability_summary(report)
        weekly_pack = build_weekly_review_pack(report)

        assert queue_summary in dashboard_values
        assert top_recommendation in dashboard_values
        assert trust_summary in dashboard_values
        assert queue_summary in executive_values
        assert top_recommendation in executive_values
        assert trust_summary in executive_values
        assert "No portfolio catalog contract is recorded yet." in executive_values
        assert "Intent alignment cannot be judged until a portfolio catalog contract exists." in executive_values
        assert print_ws["D4"].value == "Workflow Guidance"
        assert print_ws["D5"].value == "Product Mode"
        assert any(label in str(print_ws["E5"].value) for label in ("First Run:", "Weekly Review:", "Deep Dive:", "Action Sync:"))
        assert print_ws["D6"].value == "Artifact Role"
        assert print_ws["D7"].value == "Reading Order"
        assert print_ws["D8"].value == "Next Best Step"
        assert print_ws["B7"].value == weekly_pack["portfolio_headline"]
        assert print_ws["B8"].value == weekly_pack["queue_pressure_summary"]
        assert print_ws["B9"].value == weekly_pack["run_change_summary"]
        assert print_ws["B10"].value == weekly_pack["what_to_do_this_week"]
        assert print_ws["A70"].value == "Top Attention"
        assert print_ws["E70"].value == "Top Repo Drilldowns"
        executive_labels = {
            cell
            for row in executive_ws.iter_rows(min_row=20, max_row=110, min_col=1, max_col=20, values_only=True)
            for cell in row
            if isinstance(cell, str)
        }
        for expected_label in {
            "Closure Guidance",
            "Escalation",
            "Recovery Rebuild Strength",
            "Recovery Reacquisition",
            "Reacquisition Durability",
            "Reacquisition Confidence",
            "Operator Focus",
            "Focus Summary",
            "Focus Line",
            "Reacquisition Softening Hotspot",
            "Revalidation Needed Hotspot",
            "Retired Confidence Hotspot",
            "Under Revalidation Hotspot",
            "Rebuilding Restored Confidence Hotspot",
            "Re-Earning Confidence Hotspot",
            "Just Re-Earned Confidence Hotspot",
            "Holding Re-Earned Confidence Hotspot",
            "What We Tried",
            "Last Outcome",
            "Resolution Evidence",
            "Recommendation Confidence",
            "Confidence Rationale",
            "Next Action Confidence",
            "Trust Policy",
            "Trust Rationale",
            "Trust Recovery",
            "Recovery Confidence",
            "Exception Retirement",
            "Retirement Summary",
            "Policy Debt",
            "Class Normalization",
            "Class Memory",
            "Trust Decay",
            "Class Reweighting",
            "Class Reweighting Why",
            "Class Momentum",
            "Reweight Stability",
            "Transition Health",
            "Transition Resolution",
            "Transition Summary",
            "Transition Closure",
            "Transition Likely Outcome",
            "Pending Debt Freshness",
            "Closure Forecast",
            "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Persistence",
            "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Churn Controls",
            "Closure Forecast Summary",
            "Momentum Summary",
            "Exception Learning",
            "Recommendation Drift",
            "Adaptive Confidence",
            "Recommendation Quality",
            "Confidence Validation",
        }:
            assert expected_label in executive_labels
        assert print_ws["A17"].value == "Primary Target"
        assert print_ws["A18"].value == "Why Top Target"
        assert print_ws["A19"].value == "Recovery / Retirement"
        assert print_ws["A20"].value == "Recovery Persistence"
        assert print_ws["A21"].value == "Relapse Churn"
        assert print_ws["A22"].value == "Recovery Freshness"
        assert print_ws["A23"].value == "Recovery Memory Reset"
        assert print_ws["A24"].value == "Recovery Rebuild Strength"
        assert print_ws["A25"].value == "Recovery Reacquisition"
        assert print_ws["A26"].value == "Reacquisition Durability"
        assert print_ws["A27"].value == "Reacquisition Confidence"
        assert print_ws["A28"].value == "Operator Focus"
        assert print_ws["A29"].value == "Focus Summary"
        assert print_ws["A30"].value == "Focus Line"
        assert print_ws["A31"].value == "What We Tried"
        assert print_ws["A32"].value == "Last Outcome"
        assert print_ws["A33"].value == "Resolution Evidence"
        assert print_ws["A34"].value == "Recovery Counts"
        assert print_ws["A35"].value == "Recommendation Confidence"
        assert print_ws["A36"].value == "Confidence Rationale"
        assert print_ws["A37"].value == "Next Action Confidence"
        assert print_ws["A38"].value == "Trust Policy"
        assert print_ws["A39"].value == "Trust Rationale"
        assert print_ws["A40"].value == "Trust Exception"
        assert print_ws["A41"].value == "Trust Recovery"
        assert print_ws["A42"].value == "Recovery Confidence"
        assert print_ws["A43"].value == "Exception Retirement"
        assert print_ws["A44"].value == "Retirement Summary"
        assert print_ws["A45"].value == "Policy Debt"
        assert print_ws["A46"].value == "Class Normalization"
        assert print_ws["A47"].value == "Class Memory"
        assert print_ws["A48"].value == "Trust Decay"
        assert print_ws["A49"].value == "Class Reweighting"
        assert print_ws["A50"].value == "Class Reweighting Why"
        assert print_ws["A51"].value == "Class Momentum"
        assert print_ws["A52"].value == "Reweight Stability"
        assert print_ws["A53"].value == "Transition Health"
        assert print_ws["A54"].value == "Transition Resolution"
        assert print_ws["A55"].value == "Transition Summary"
        assert print_ws["A56"].value == "Transition Closure"
        assert print_ws["A57"].value == "Transition Likely Outcome"
        assert print_ws["A58"].value == "Pending Debt Freshness"
        assert print_ws["A59"].value == "Closure Forecast"
        assert print_ws["A60"].value == "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Persistence"
        assert print_ws["A61"].value == "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Churn Controls"
        assert print_ws["A62"].value == "Closure Forecast Summary"
        assert print_ws["A63"].value == "Momentum Summary"
        assert print_ws["A64"].value == "Exception Learning"
        assert print_ws["A65"].value == "Recommendation Drift"
        assert print_ws["A66"].value == "Adaptive Confidence"
        assert print_ws["A67"].value == "Recommendation Quality"
        assert print_ws["A68"].value == "Confidence Validation"
        assert print_ws["A69"].value == "Calibration Snapshot"
        assert "Why Top Target" in dashboard_values
        assert "Closure Guidance" in dashboard_values
        assert "What We Tried" in dashboard_values
        assert "Recommendation Confidence" in dashboard_values
        assert "Trust Policy" in dashboard_values
        assert "Trust Exception" in dashboard_values
        assert "Trust Recovery" in dashboard_values
        assert "Recovery Confidence" in dashboard_values
        assert "Exception Retirement" in dashboard_values
        assert "Policy Debt" in dashboard_values
        assert "Class Normalization" in dashboard_values
        assert "Class Memory" in dashboard_values
        assert "Trust Decay" in dashboard_values
        assert "Class Reweighting" in dashboard_values
        assert "Class Momentum" in dashboard_values
        assert "Reweight Stability" in dashboard_values
        assert "Transition Health" in dashboard_values
        assert "Transition Resolution" in dashboard_values
        assert "Transition Closure" in dashboard_values
        assert "Transition Likely Outcome" in dashboard_values
        assert "Pending Debt Freshness" in dashboard_values
        assert "Closure Forecast" in dashboard_values
        assert "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Persistence" in dashboard_values
        assert "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Churn Controls" in dashboard_values
        assert "Exception Learning" in dashboard_values
        assert "Recommendation Drift" in dashboard_values

    def test_campaigns_show_empty_state_when_no_preview_rows(self):
        wb = Workbook()
        report = _make_report()
        report["campaign_summary"] = {}
        report["writeback_preview"] = {"repos": []}
        _build_campaigns(wb, report)
        ws = wb["Campaigns"]
        values = [ws.cell(row=row, column=1).value for row in range(1, 30)]
        assert any(isinstance(value, str) and "No active campaign rows" in value for value in values)

    def test_campaigns_show_github_projects_status_and_counts(self):
        wb = Workbook()
        report = _make_report()
        report["action_sync_summary"] = {
            "summary": "Action Sync is preview-ready: Security Review is the strongest next campaign to preview from the current local facts.",
        }
        report["next_action_sync_step"] = "Preview Security Review next, then decide whether it is ready to sync to all."
        report["apply_readiness_summary"] = {
            "summary": "Apply handoff says preview Security Review next before deciding on apply to all."
        }
        report["next_apply_candidate"] = {
            "summary": "Preview Security Review next, then decide whether it is ready to apply to all.",
            "preview_command": "audit user --campaign security-review --writeback-target all",
        }
        report["campaign_outcomes_summary"] = {
            "summary": "Security Review was applied recently; monitor it now before treating it as stable.",
        }
        report["next_monitoring_step"] = {
            "summary": "Monitor Security Review for at least 2 post-apply runs before treating it as stable.",
        }
        report["campaign_tuning_summary"] = {
            "summary": "Security Review should win ties because recent outcomes are proven.",
        }
        report["next_tuned_campaign"] = {
            "summary": "Security Review should win ties inside the preview-ready group because recent outcome history is proven.",
        }
        report["intervention_ledger_summary"] = {
            "summary": "RepoC is improving after intervention while one repo still looks relapsing.",
        }
        report["next_historical_focus"] = {
            "summary": "Read RepoC next: it is the clearest current example of improvement after intervention.",
        }
        report["automation_guidance_summary"] = {
            "summary": "Preview Security Review next; that is the strongest safe automation step right now.",
        }
        report["next_safe_automation_step"] = {
            "summary": "Preview Security Review next; that is the strongest safe automation step right now.",
            "recommended_command": "audit user --campaign security-review --writeback-target all",
        }
        report["writeback_preview"] = {
            "sync_mode": "reconcile",
            "github_projects": {
                "enabled": True,
                "status": "configured",
                "project_owner": "octo-org",
                "project_number": 7,
            },
            "repos": [
                {
                    "repo": "RepoC",
                    "issue_title": "[Repo Auditor] Security Review",
                    "topics": ["ghra-call-security-review"],
                    "github_project_field_count": 3,
                    "notion_action_count": 1,
                    "action_ids": ["security-review-abc123"],
                }
            ],
        }
        _build_campaigns(wb, report)
        ws = wb["Campaigns"]

        assert ws["A13"].value == "GitHub Projects"
        assert ws["B13"].value == "configured (octo-org #7)"
        assert ws["J5"].value == "Action Sync Readiness"
        assert "preview-ready" in str(ws["K5"].value).lower()
        assert ws["J7"].value == "Apply Packet"
        assert "preview security review next" in str(ws["K7"].value).lower()
        assert ws["J9"].value == "Post-Apply Monitoring"
        assert "monitor it now" in str(ws["K9"].value).lower()
        assert ws["J11"].value == "Campaign Tuning"
        assert "win ties" in str(ws["K11"].value).lower()
        assert ws["J13"].value == "Historical Portfolio Intelligence"
        assert "improving after intervention" in str(ws["K13"].value).lower()
        assert ws["J15"].value == "Automation Guidance"
        assert "strongest safe automation step" in str(ws["K15"].value).lower()
        assert ws["J16"].value == "Next Safe Automation Step"
        assert ws["D15"].value == "Projects"
        assert ws["D16"].value == 3

    def test_print_pack_shows_action_sync_readiness_guidance(self):
        wb = Workbook()
        report = _make_report()
        report["operator_summary"] = {
            "headline": "Campaign work can move forward.",
            "counts": {"blocked": 0, "urgent": 0, "ready": 1, "deferred": 0},
            "action_sync_summary": {
                "summary": "Action Sync is preview-ready: Security Review is the strongest next campaign to preview from the current local facts.",
            },
            "next_action_sync_step": "Preview Security Review next, then decide whether it is ready to sync to all.",
            "apply_readiness_summary": {
                "summary": "Apply handoff says preview Security Review next before deciding on apply to all."
            },
            "next_apply_candidate": {
                "summary": "Preview Security Review next, then decide whether it is ready to apply to all.",
                "preview_command": "audit user --campaign security-review --writeback-target all",
            },
            "campaign_outcomes_summary": {
                "summary": "Security Review was applied recently; monitor it now before treating it as stable.",
            },
            "next_monitoring_step": {
                "summary": "Monitor Security Review for at least 2 post-apply runs before treating it as stable.",
            },
            "campaign_tuning_summary": {
                "summary": "Security Review should win ties because recent outcomes are proven.",
            },
            "next_tuned_campaign": {
                "summary": "Security Review should win ties inside the preview-ready group because recent outcome history is proven.",
            },
            "intervention_ledger_summary": {
                "summary": "RepoC is improving after intervention while one repo still looks relapsing.",
            },
            "next_historical_focus": {
                "summary": "Read RepoC next: it is the clearest current example of improvement after intervention.",
            },
            "automation_guidance_summary": {
                "summary": "Preview Security Review next; that is the strongest safe automation step right now.",
            },
            "next_safe_automation_step": {
                "summary": "Preview Security Review next; that is the strongest safe automation step right now.",
                "recommended_command": "audit user --campaign security-review --writeback-target all",
            },
        }
        _build_print_pack(wb, report, None)
        ws = wb["Print Pack"]

        assert ws["D9"].value == "Action Sync Readiness"
        assert "preview-ready" in str(ws["E9"].value).lower()
        assert ws["D10"].value == "Next Action Sync Step"
        assert ws["D11"].value == "Apply Packet"
        assert "preview security review next" in str(ws["E11"].value).lower()
        assert ws["D12"].value == "Command Hint"
        assert ws["D13"].value == "Post-Apply Monitoring"
        assert "monitor it now" in str(ws["E13"].value).lower()
        assert ws["D14"].value == "Next Monitoring Step"
        assert ws["D15"].value == "Campaign Tuning"
        assert "win ties" in str(ws["E15"].value).lower()
        assert ws["D16"].value == "Next Tie-Break Candidate"
        assert ws["D17"].value == "Historical Portfolio Intelligence"
        assert "improving after intervention" in str(ws["E17"].value).lower()
        assert ws["D18"].value == "Next Historical Focus"
        assert ws["D19"].value == "Automation Guidance"
        assert "strongest safe automation step" in str(ws["E19"].value).lower()
        assert ws["D20"].value == "Next Safe Automation Step"

    def test_writeback_audit_shows_empty_state_when_no_results(self):
        wb = Workbook()
        report = _make_report()
        report["writeback_results"] = {"results": []}
        _build_writeback_audit(wb, report)
        ws = wb["Writeback Audit"]
        assert "No writeback results are recorded" in ws["A11"].value

    def test_governance_controls_use_human_readable_state_labels(self):
        wb = Workbook()
        _build_governance_controls(wb, _make_report())
        ws = wb["Governance Controls"]
        assert ws["C12"].value == "Preview only"

    def test_dashboard_handles_repos_without_actions_or_hotspots(self):
        wb = Workbook()
        report = _make_report()
        report["audits"][0]["action_candidates"] = []
        report["audits"][0]["hotspots"] = []
        _build_dashboard(wb, report)
        assert "Dashboard" in wb.sheetnames

    def test_compare_sheet_only_when_diff_present(self):
        wb = Workbook()
        _build_compare_sheet(wb, None)
        assert "Compare" not in wb.sheetnames

        diff = {
            "previous_date": "2026-03-28T00:00:00+00:00",
            "current_date": "2026-03-29T00:00:00+00:00",
            "average_score_delta": 0.03,
            "lens_deltas": {"ship_readiness": 0.1},
            "repo_changes": [
                {
                    "name": "RepoA",
                    "delta": 0.1,
                    "old_tier": "functional",
                    "new_tier": "shipped",
                    "security_change": {"old_label": "watch", "new_label": "healthy"},
                    "hotspot_change": {"old_count": 1, "new_count": 0},
                    "collection_change": {"old": [], "new": ["showcase"]},
                }
            ],
        }
        _build_compare_sheet(wb, diff)
        assert "Compare" in wb.sheetnames

    def test_creates_scenario_and_executive_summary(self):
        wb = Workbook()
        _build_scenario_planner(wb, _make_report(), portfolio_profile="default", collection="showcase")
        _build_executive_summary(wb, _make_report(), None, portfolio_profile="default", collection="showcase")
        assert "Scenario Planner" in wb.sheetnames
        assert "Executive Summary" in wb.sheetnames

    def test_secondary_visible_sheets_gain_freeze_panes(self):
        wb = Workbook()
        report = _make_report()
        _build_action_items(wb, report)
        _build_hotspots(wb, report)
        _build_repo_profiles(wb, report)
        _build_scenario_planner(wb, report, portfolio_profile="default", collection="showcase")
        _build_security_debt(wb, report)
        _build_score_explainer(wb)
        _build_trends(wb, report, trend_data=[{"date": "2026-03-28", "average_score": 0.6, "repos_audited": 3, "tier_distribution": {"shipped": 1, "functional": 1}}])

        assert wb["Action Items"].freeze_panes == "A5"
        assert wb["Hotspots"].freeze_panes == "A2"
        assert wb["Repo Profiles"].freeze_panes == "B2"
        assert wb["Scenario Planner"].freeze_panes == "B6"
        assert wb["Security Debt"].freeze_panes == "A2"
        assert wb["Score Explainer"].freeze_panes == "A5"
        assert wb["Trends"].freeze_panes == "B4"

    def test_creates_security_phase_sheets(self):
        from src.excel_export import (
            _build_security_controls,
            _build_security_debt,
            _build_supply_chain,
        )

        wb = Workbook()
        _build_security_controls(wb, _make_report())
        _build_supply_chain(wb, _make_report())
        _build_security_debt(wb, _make_report())
        assert "Security Controls" in wb.sheetnames
        assert "Supply Chain" in wb.sheetnames
        assert "Security Debt" in wb.sheetnames


class TestWorkbookModes:
    def test_template_mode_requires_existing_template(self, tmp_path):
        report_path = tmp_path / "report.json"
        report_path.write_text(json.dumps(_make_report()))

        with pytest.raises(FileNotFoundError):
            export_excel(
                report_path,
                tmp_path / "out.xlsx",
                excel_mode="template",
                template_path=tmp_path / "missing.xlsx",
            )

    def test_standard_mode_generates_operator_workbook(self, tmp_path):
        report_path = tmp_path / "report.json"
        report_path.write_text(json.dumps(_make_report()))

        output = export_excel(
            report_path,
            tmp_path / "out.xlsx",
            excel_mode="standard",
        )

        wb = load_workbook(output)
        assert "Dashboard" in wb.sheetnames
        assert "By Collection" in wb.sheetnames
        assert "Trend Summary" in wb.sheetnames
        assert "Run Changes" in wb.sheetnames
        assert "Review Queue" in wb.sheetnames
        assert "Implementation Hotspots" in wb.sheetnames
        assert "Operator Outcomes" in wb.sheetnames
        assert "Repo Detail" in wb.sheetnames
        assert "Governance Controls" in wb.sheetnames
        assert "Print Pack" in wb.sheetnames

    def test_standard_mode_hides_advanced_tabs_by_default(self, tmp_path):
        report_path = tmp_path / "report.json"
        report_path.write_text(json.dumps(_make_report()))

        output = export_excel(
            report_path,
            tmp_path / "out.xlsx",
            excel_mode="standard",
        )

        wb = load_workbook(output)
        visible_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        assert set(visible_sheets) == {
            "Index",
            "Dashboard",
            "All Repos",
            "Review Queue",
            "Portfolio Explorer",
            "Portfolio Catalog",
            "Scorecards",
            "Implementation Hotspots",
            "Operator Outcomes",
            "Approval Ledger",
            "Historical Intelligence",
            "Repo Detail",
            "Executive Summary",
            "By Lens",
            "By Collection",
            "Trend Summary",
            "Run Changes",
            "Campaigns",
            "Governance Controls",
            "Print Pack",
        }
        assert visible_sheets[:2] == ["Index", "Dashboard"]
        assert wb["Hotspots"].sheet_state == "hidden"

    def test_template_mode_generates_native_sparkline_xml_and_named_ranges(self, tmp_path):
        report_path = tmp_path / "report.json"
        report_path.write_text(json.dumps(_make_report()))

        output = export_excel(
            report_path,
            tmp_path / "out.xlsx",
            trend_data=[{"date": "2026-03-28", "average_score": 0.58, "repos_audited": 3, "tier_distribution": {"shipped": 1, "functional": 1}}],
            score_history={"RepoA": [0.5, 0.7], "RepoB": [0.4, 0.6]},
            excel_mode="template",
            template_path=DEFAULT_TEMPLATE_PATH,
        )

        wb = load_workbook(output)
        assert TEMPLATE_INFO_SHEET in wb.sheetnames
        assert "nrReviewOpenCount" in wb.defined_names
        assert "nrSelectedProfileLabel" in wb.defined_names

        with zipfile.ZipFile(output) as archive:
            worksheet_names = [name for name in archive.namelist() if name.startswith("xl/worksheets/sheet")]
            assert any(
                "sparkline" in archive.read(name).decode("utf-8", "ignore").lower()
                for name in worksheet_names
            )

    def test_template_and_standard_modes_match_key_visible_facts(self, tmp_path):
        report_path = tmp_path / "report.json"
        report_path.write_text(json.dumps(_make_report()))

        standard_output = export_excel(report_path, tmp_path / "standard.xlsx", excel_mode="standard")
        template_output = export_excel(
            report_path,
            tmp_path / "template.xlsx",
            excel_mode="template",
            template_path=DEFAULT_TEMPLATE_PATH,
        )

        standard_wb = load_workbook(standard_output)
        template_wb = load_workbook(template_output)

        assert standard_wb["Dashboard"]["A1"].value == template_wb["Dashboard"]["A1"].value
        assert standard_wb["Review Queue"]["B6"].value == template_wb["Review Queue"]["B6"].value
        assert standard_wb["Governance Controls"]["B5"].value == template_wb["Governance Controls"]["B5"].value
        assert standard_wb["Print Pack"]["B7"].value == template_wb["Print Pack"]["B7"].value
        assert standard_wb["Print Pack"]["B8"].value == template_wb["Print Pack"]["B8"].value
        assert standard_wb["Print Pack"]["B9"].value == template_wb["Print Pack"]["B9"].value
        assert standard_wb["Print Pack"]["B10"].value == template_wb["Print Pack"]["B10"].value
        assert standard_wb["Print Pack"]["B16"].value == template_wb["Print Pack"]["B16"].value
        assert template_wb["Review Queue"]["A4"].value == "Summary"
        template_header_row = next(
            row for row in range(15, 45) if template_wb["Review Queue"].cell(row=row, column=1).value == "Repo"
        )
        assert template_header_row >= 17

    def test_internal_navigation_links_use_locations_not_external_relationships(self, tmp_path):
        report_path = tmp_path / "report.json"
        report_path.write_text(json.dumps(_make_report()))

        output = export_excel(report_path, tmp_path / "out.xlsx", excel_mode="standard")

        with zipfile.ZipFile(output) as archive:
            rel_name = "xl/worksheets/_rels/sheet1.xml.rels"
            if rel_name in archive.namelist():
                rel_xml = archive.read(rel_name).decode("utf-8", "ignore")
                assert "relationships/hyperlink" not in rel_xml

    def test_review_queue_uses_autofilter_not_structured_table(self, tmp_path):
        report_path = tmp_path / "report.json"
        report_path.write_text(json.dumps(_make_report()))

        output = export_excel(report_path, tmp_path / "out.xlsx", excel_mode="standard")

        wb = load_workbook(output)
        ws = wb["Review Queue"]
        header_row = next(row for row in range(20, 70) if ws.cell(row=row, column=1).value == "Repo")
        assert ws.auto_filter.ref == f"A{header_row}:AN{header_row + 1}"
        assert not ws.tables

    def test_portfolio_catalog_and_scorecards_sheets_are_present(self, tmp_path):
        report_path = tmp_path / "report.json"
        report_path.write_text(json.dumps(_make_report()))

        output = export_excel(report_path, tmp_path / "out.xlsx", excel_mode="standard")

        wb = load_workbook(output)
        catalog_ws = wb["Portfolio Catalog"]
        assert catalog_ws["A1"].value == "Portfolio Catalog"
        assert "No portfolio catalog contract is recorded yet." in str(catalog_ws["A3"].value)
        assert "Intent alignment cannot be judged" in str(catalog_ws["A4"].value)

        hidden_ws = wb["Data_PortfolioCatalog"]
        headers = [hidden_ws.cell(row=1, column=col).value for col in range(1, 16)]
        assert headers == [
            "Repo",
            "Full Name",
            "Owner",
            "Team",
            "Purpose",
            "Lifecycle",
            "Criticality",
            "Review Cadence",
            "Disposition",
            "Maturity Program",
            "Target Maturity",
            "Notes",
            "Intent Alignment",
            "Intent Alignment Reason",
            "Catalog Line",
        ]

        scorecards_ws = wb["Scorecards"]
        assert scorecards_ws["A1"].value == "Scorecards"
        assert "No maturity scorecard is recorded yet." in str(scorecards_ws["A3"].value)

        hidden_scorecards_ws = wb["Data_Scorecards"]
        scorecard_headers = [hidden_scorecards_ws.cell(row=1, column=col).value for col in range(1, 14)]
        assert scorecard_headers == [
            "Repo",
            "Full Name",
            "Program",
            "Program Label",
            "Score",
            "Maturity Level",
            "Target Maturity",
            "Status",
            "Passed Rules",
            "Applicable Rules",
            "Failed Rule Keys",
            "Top Gaps",
            "Summary",
        ]

    def test_visible_sheets_use_filters_while_hidden_data_sheets_keep_tables(self, tmp_path):
        report_path = tmp_path / "report.json"
        report_path.write_text(json.dumps(_make_report()))

        output = export_excel(report_path, tmp_path / "out.xlsx", excel_mode="standard")

        with zipfile.ZipFile(output) as archive:
            workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
            workbook_rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            rel_targets = {
                rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
                for rel in workbook_rels
                if rel.attrib.get("Type", "").endswith("/worksheet")
            }
            visible_targets = []
            hidden_targets = []
            for sheet in workbook_root.findall("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets/{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
                target = rel_targets[sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]]
                if sheet.attrib.get("state") == "hidden":
                    hidden_targets.append(target)
                else:
                    visible_targets.append(target)

            visible_table_rels = []
            hidden_table_rels = []
            for target in visible_targets:
                rel_name = target.replace("worksheets/", "worksheets/_rels/") + ".rels"
                if rel_name in archive.namelist():
                    rel_xml = archive.read(rel_name).decode("utf-8", "ignore")
                    if "relationships/table" in rel_xml:
                        visible_table_rels.append(rel_name)
            for target in hidden_targets:
                rel_name = target.replace("worksheets/", "worksheets/_rels/") + ".rels"
                if rel_name in archive.namelist():
                    rel_xml = archive.read(rel_name).decode("utf-8", "ignore")
                    if "relationships/table" in rel_xml:
                        hidden_table_rels.append(rel_name)

            assert not visible_table_rels
            assert hidden_table_rels
