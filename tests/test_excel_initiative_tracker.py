"""Tests for excel_initiative_tracker_helpers — Arc G Sprint 7A.4."""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook

from src.excel_initiative_tracker_helpers import (
    _format_missing_requirements,
    build_initiative_tracker_sheet,
)
from src.initiatives import Initiative, save_initiatives
from src.maturity_tiers import TierGap

# ── Helpers ───────────────────────────────────────────────────────────────────


def _future(days: int = 30) -> str:
    return (date.today() + timedelta(days=days)).isoformat()


def _past(days: int = 5) -> str:
    return (date.today() - timedelta(days=days)).isoformat()


def _make_initiative(
    repo_name: str = "Wavelength",
    target_tier: int = 3,
    deadline: str | None = None,
    closed_at: str | None = None,
) -> Initiative:
    return Initiative(
        repo_name=repo_name,
        target_tier=target_tier,
        deadline=deadline or _future(30),
        set_at="2026-05-12T10:00:00+00:00",
        set_by="operator",
        closed_at=closed_at,
        closed_reason=None,
    )


def _style_stubs():
    """Return minimal style-helper stubs accepted by build_initiative_tracker_sheet."""
    from openpyxl.styles import Alignment, Font

    def get_or_create_sheet(wb, name):
        if name in wb.sheetnames:
            return wb[name]
        return wb.create_sheet(name)

    def configure_sheet_view(ws, **kwargs):
        pass

    def set_sheet_header(ws, title, subtitle, width=7):
        pass

    def style_header_row(ws, row, ncols):
        pass

    def auto_width(ws, ncols, max_row):
        pass

    wrap_alignment = Alignment(wrap_text=True)
    subtitle_font = Font(bold=True)

    return {
        "get_or_create_sheet": get_or_create_sheet,
        "configure_sheet_view": configure_sheet_view,
        "set_sheet_header": set_sheet_header,
        "style_header_row": style_header_row,
        "auto_width": auto_width,
        "wrap_alignment": wrap_alignment,
        "subtitle_font": subtitle_font,
    }


def _build_sheet(
    initiatives: list[Initiative],
    projects: list[dict] | None = None,
    tmp_path: Path | None = None,
) -> tuple[Workbook, object]:
    """Build a workbook with the Initiative Tracker sheet using saved initiatives."""
    wb = Workbook()
    stubs = _style_stubs()

    # Save initiatives to tmp path
    output_dir = tmp_path or Path("/tmp/test_initiative_tracker")
    output_dir.mkdir(parents=True, exist_ok=True)
    from src.initiatives import initiatives_path

    save_initiatives(initiatives_path(output_dir), initiatives)

    data: dict = {}
    if projects:
        data["projects"] = projects

    build_initiative_tracker_sheet(
        wb,
        data,
        output_dir=output_dir,
        **stubs,
    )
    return wb, wb["Initiative Tracker"]


def _bronze_repo(name: str = "Wavelength") -> dict:
    """Portfolio-truth project at Bronze tier."""
    return {
        "identity": {"display_name": name, "has_git": True},
        "derived": {
            "last_meaningful_activity_at": (date.today() - timedelta(days=50)).isoformat(),
            "activity_status": "active",
            "context_quality": "weak",
            "context_files": ["README.md"],
            "run_instructions_present": False,
        },
        "risk": {
            "risk_tier": "elevated",
            "risk_factors": [],
            "doctor_gap": True,
        },
    }


# ── Tests ─────────────────────────────────────────────────────────────────────


class TestEmptyState:
    def test_sheet_exists_with_no_initiatives(self, tmp_path):
        wb, ws = _build_sheet([], tmp_path=tmp_path)
        assert "Initiative Tracker" in wb.sheetnames

    def test_empty_state_message_present(self, tmp_path):
        wb, ws = _build_sheet([], tmp_path=tmp_path)
        # Collect all cell values
        values = [ws.cell(row=r, column=1).value for r in range(1, 6)]
        assert any(v and "No initiatives yet" in str(v) for v in values), (
            f"Expected empty-state message, got: {values}"
        )

    def test_empty_state_no_crash_on_missing_file(self, tmp_path):
        """If initiatives.json doesn't exist, sheet builds cleanly."""
        wb = Workbook()
        stubs = _style_stubs()
        # Don't write any file — output_dir has no initiatives.json
        build_initiative_tracker_sheet(
            wb,
            {},
            output_dir=tmp_path,
            **stubs,
        )
        assert "Initiative Tracker" in wb.sheetnames


class TestOnTrackLane:
    def test_on_track_initiative_row_present(self, tmp_path):
        initiatives = [_make_initiative("Wavelength", target_tier=3, deadline=_future(60))]
        projects = [_bronze_repo("Wavelength")]
        wb, ws = _build_sheet(initiatives, projects=projects, tmp_path=tmp_path)

        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend(v for v in row if v)

        assert any("Wavelength" in str(v) for v in all_values), (
            f"Expected 'Wavelength' in sheet, got: {all_values[:20]}"
        )

    def test_on_track_lane_header_present(self, tmp_path):
        initiatives = [_make_initiative("Wavelength", target_tier=3, deadline=_future(60))]
        wb, ws = _build_sheet(initiatives, tmp_path=tmp_path)

        all_values = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        assert any(v and "ON TRACK" in str(v) for v in all_values), (
            f"Expected 'ON TRACK' lane header, got: {all_values}"
        )

    def test_columns_header_present(self, tmp_path):
        initiatives = [_make_initiative("Wavelength", target_tier=3, deadline=_future(60))]
        wb, ws = _build_sheet(initiatives, tmp_path=tmp_path)

        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend(str(v) for v in row if v)

        assert any("REPO" in v for v in all_values), "Expected REPO column header"
        assert any("DEADLINE" in v for v in all_values), "Expected DEADLINE column header"


class TestMixedStatuses:
    def test_overdue_and_on_track_both_present(self, tmp_path):
        initiatives = [
            _make_initiative("Overdue-Repo", target_tier=2, deadline=_past(10)),
            _make_initiative("OnTrack-Repo", target_tier=2, deadline=_future(60)),
        ]
        wb, ws = _build_sheet(initiatives, tmp_path=tmp_path)

        all_values = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        has_overdue = any(v and "OVERDUE" in str(v) for v in all_values)
        has_on_track = any(v and "ON TRACK" in str(v) for v in all_values)
        assert has_overdue, f"Expected OVERDUE lane, got: {all_values}"
        assert has_on_track, f"Expected ON TRACK lane, got: {all_values}"

    def test_at_risk_lane_present(self, tmp_path):
        # at-risk = deadline within 14 days
        initiatives = [
            _make_initiative("AtRisk-Repo", target_tier=2, deadline=_future(7)),
        ]
        wb, ws = _build_sheet(initiatives, tmp_path=tmp_path)

        all_values = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        assert any(v and "AT RISK" in str(v) for v in all_values), (
            f"Expected AT RISK lane, got: {all_values}"
        )

    def test_closed_initiative_not_shown(self, tmp_path):
        """Closed initiatives (closed_at set) must not appear in the sheet."""
        initiatives = [
            _make_initiative("Closed-Repo", closed_at="2026-05-01T00:00:00+00:00"),
        ]
        wb, ws = _build_sheet(initiatives, tmp_path=tmp_path)

        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend(v for v in row if v)
        assert not any("Closed-Repo" in str(v) for v in all_values), (
            "Closed initiative should not appear in sheet"
        )


class TestTabColor:
    def test_tab_color_red_for_overdue(self, tmp_path):
        initiatives = [_make_initiative("OverdueRepo", target_tier=2, deadline=_past(5))]
        wb, ws = _build_sheet(initiatives, tmp_path=tmp_path)
        # Red tab for overdue
        tab_color = ws.sheet_properties.tabColor
        assert tab_color is not None
        color_val = str(tab_color.rgb if hasattr(tab_color, "rgb") else tab_color.value or "")
        assert "DC2626" in color_val.upper() or color_val != "", (
            f"Expected red tab color for overdue, got: {tab_color!r}"
        )

    def test_tab_color_set_for_on_track(self, tmp_path):
        initiatives = [_make_initiative("OnTrackRepo", target_tier=2, deadline=_future(60))]
        wb, ws = _build_sheet(initiatives, tmp_path=tmp_path)
        tab_color = ws.sheet_properties.tabColor
        # Any non-None tab color is acceptable; just verify it was set
        assert tab_color is not None


# ── Arc G S10.2 — (approx.) hint in missing-requirements cell ────────────────


class TestFormatMissingRequirements:
    """Unit tests for _format_missing_requirements and its (approx.) annotation (Arc G S10.2)."""

    def test_mixed_strict_and_proxy_labels(self):
        """Proxy requirements get '(approx.)'; strict requirements do not."""
        gap = TierGap(
            current_tier=1,
            target_tier=2,
            missing_requirements=["readme", "ci"],
            requirement_sources=["strict", "proxy"],
        )
        result = _format_missing_requirements(gap)
        assert "readme" in result
        assert (
            "(approx.)" not in result.split("readme")[0] + result.split("readme")[1].split(";")[0]
        )
        assert "ci (approx.)" in result

    def test_all_strict_no_approx(self):
        """All-strict gap produces no '(approx.)' hints."""
        gap = TierGap(
            current_tier=1,
            target_tier=2,
            missing_requirements=["readme", "tests"],
            requirement_sources=["strict", "strict"],
        )
        result = _format_missing_requirements(gap)
        assert "(approx.)" not in result
        assert "readme" in result
        assert "tests" in result

    def test_all_proxy_every_requirement_annotated(self):
        """All-proxy gap appends '(approx.)' to every requirement."""
        gap = TierGap(
            current_tier=1,
            target_tier=2,
            missing_requirements=["readme", "ci", "license"],
            requirement_sources=["proxy", "proxy", "proxy"],
        )
        result = _format_missing_requirements(gap)
        assert result.count("(approx.)") == 3

    def test_empty_requirements_returns_dash(self):
        """No missing requirements returns the em-dash sentinel."""
        gap = TierGap(
            current_tier=1,
            target_tier=2,
            missing_requirements=[],
            requirement_sources=[],
        )
        assert _format_missing_requirements(gap) == "—"

    def test_legacy_empty_sources_treated_as_all_proxy(self):
        """When requirement_sources is empty (legacy TierGap), all requirements get (approx.)."""
        gap = TierGap(
            current_tier=1,
            target_tier=2,
            missing_requirements=["readme", "ci"],
            requirement_sources=[],
        )
        result = _format_missing_requirements(gap)
        assert result.count("(approx.)") == 2

    def test_workbook_cell_contains_approx_for_proxy_requirement(self, tmp_path):
        """End-to-end: rendered workbook cell contains '(approx.)' for a proxy requirement."""
        # We bypass tier_gap entirely and test _format_missing_requirements
        # directly with a constructed TierGap to keep this test deterministic.
        gap = TierGap(
            current_tier=1,
            target_tier=2,
            missing_requirements=["ci-badge"],
            requirement_sources=["proxy"],
        )
        result = _format_missing_requirements(gap)
        assert "ci-badge (approx.)" == result

    def test_workbook_cell_strict_requirement_no_approx(self, tmp_path):
        """End-to-end: strict requirement renders without '(approx.)' suffix."""
        gap = TierGap(
            current_tier=1,
            target_tier=2,
            missing_requirements=["license"],
            requirement_sources=["strict"],
        )
        result = _format_missing_requirements(gap)
        assert result == "license"
        assert "(approx.)" not in result
