"""Helpers for dashboard workbook content."""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference, ScatterChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.series import Series as BubbleSeries
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font

X_MID = 0.55
Y_MID = 0.45

QUADRANT_NAMES = [
    ("Flagships", "high completeness + high interest"),
    ("Hidden Gems", "high interest, incomplete — worth finishing"),
    ("Workhorses", "solid but routine — bread and butter"),
    ("Archive Candidates", "low both — consider archiving"),
]


def write_dashboard_header(
    *,
    ws,
    data: dict[str, Any],
    diff_data: dict[str, Any] | None,
    title_font,
    subtitle_font,
    narrative_font,
    center_alignment,
    wrap_alignment,
    section_font,
    generate_narrative,
) -> None:
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"GitHub Portfolio Dashboard: {data['username']}"
    title_cell.font = title_font
    title_cell.alignment = center_alignment

    ws.merge_cells("A2:L2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = (
        f"Generated: {data['generated_at'][:10]} | {data['repos_audited']} repos audited"
    )
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = center_alignment

    ws.merge_cells("A3:L3")
    narrative_cell = ws["A3"]
    narrative_cell.value = generate_narrative(data, diff_data)
    narrative_cell.font = narrative_font
    narrative_cell.alignment = wrap_alignment
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "A5"
    ws["A4"] = "Portfolio Health Snapshot"
    ws["A4"].font = section_font
    ws["O4"] = "Operator Attention Snapshot"
    ws["O4"].font = section_font


def build_dashboard_ranked_content(
    *,
    data: dict[str, Any],
    operator_block_end: int,
    build_workbook_rollups,
    build_dashboard_top_attention_rows,
    build_dashboard_top_opportunity_rows,
    build_dashboard_laggard_rows,
) -> dict[str, Any]:
    repo_rollups = build_workbook_rollups(data)[1]
    top_attention_start = max(19, operator_block_end + 2)
    return {
        "top_attention_start": top_attention_start,
        "top_attention_rows": build_dashboard_top_attention_rows(repo_rollups),
        "top_opportunities": build_dashboard_top_opportunity_rows(data.get("audits", [])),
        "laggard_rows": build_dashboard_laggard_rows(data.get("audits", [])),
    }


def write_dashboard_story_sections(
    *,
    ws,
    dashboard_story: dict[str, Any],
    subheader_font,
    wrap_alignment,
    write_key_value_block,
) -> int:
    for offset, (label, value) in enumerate(dashboard_story["sidebar_rows"], start=8):
        ws.cell(row=offset, column=13, value=label).font = subheader_font
        ws.cell(row=offset, column=14, value=value).alignment = wrap_alignment

    return write_key_value_block(
        ws,
        5,
        15,
        dashboard_story["operator_rows"],
        title="Operator Snapshot",
    )


def write_dashboard_ranked_sections(
    *,
    ws,
    ranked_content: dict[str, Any],
    write_ranked_list,
) -> None:
    write_ranked_list(
        ws,
        ranked_content["top_attention_start"],
        15,
        "Top Attention Items",
        ["Repo", "Counts", "Why Now", "Next Step"],
        ranked_content["top_attention_rows"],
    )
    write_ranked_list(
        ws,
        ranked_content["top_attention_start"] + 8,
        15,
        "Top Opportunities",
        ["Repo", "Score", "Tier", "Best Next Move"],
        ranked_content["top_opportunities"],
    )
    write_ranked_list(
        ws,
        ranked_content["top_attention_start"] + 16,
        15,
        "Top Laggards",
        ["Repo", "Score", "Tier", "What Is Dragging It Down"],
        ranked_content["laggard_rows"],
    )


def build_dashboard_operator_rows(
    *,
    data: dict[str, Any],
    excel_mode: str,
    operator_summary: dict[str, Any],
    governance_summary: dict[str, Any],
    campaign_summary: dict[str, Any],
    setup_health: dict[str, Any],
    lane_counts: dict[str, Any],
    run_change_summary: str,
    queue_pressure_summary: str,
    top_recommendation_summary: str,
    operator_context: dict[str, Any],
    format_lane_counts,
    display_operator_state,
    build_portfolio_catalog_summary,
    build_operating_paths_summary,
    build_portfolio_intent_alignment_summary,
    build_scorecards_summary,
) -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = [
        ("Setup Health", display_operator_state(setup_health.get("status", "ok"))),
        ("Operator Headline", operator_summary.get("headline", "Portfolio health is stable.")),
        ("Queue Counts", format_lane_counts(lane_counts)),
        (
            "Governance",
            governance_summary.get(
                "headline", "Governance preview is aligned with the latest report."
            ),
        ),
        (
            "Campaign State",
            campaign_summary.get("label")
            or campaign_summary.get("campaign_type")
            or "No active managed campaign in this run.",
        ),
        ("Run Changes", run_change_summary),
        ("Queue Pressure", queue_pressure_summary),
        ("What Changed", operator_context["what_changed"]),
        ("Why It Matters", operator_context["why_it_matters"]),
        ("Follow-Through", operator_context["follow_through"]),
        ("Next Checkpoint", operator_context["follow_through_checkpoint"]),
        ("Escalation", operator_context["follow_through_escalation"]),
        ("Recovery / Retirement", operator_context["follow_through_recovery"]),
        (
            "Recovery Persistence",
            operator_context["follow_through_recovery_persistence"],
        ),
        ("Relapse Churn", operator_context["follow_through_relapse_churn"]),
        (
            "Recovery Freshness",
            operator_context["follow_through_recovery_freshness"],
        ),
        (
            "Recovery Memory Reset",
            operator_context["follow_through_recovery_memory_reset"],
        ),
        (
            "Recovery Rebuild Strength",
            operator_context["follow_through_rebuild_strength"],
        ),
        ("Recovery Reacquisition", operator_context["follow_through_reacquisition"]),
        (
            "Reacquisition Durability",
            operator_context["follow_through_reacquisition_durability"],
        ),
        (
            "Reacquisition Confidence",
            operator_context["follow_through_reacquisition_confidence"],
        ),
        ("Operator Focus", operator_context["operator_focus"]),
        ("Focus Summary", operator_context["operator_focus_summary"]),
        ("Focus Line", operator_context["operator_focus_line"]),
        ("Portfolio Catalog", build_portfolio_catalog_summary(data)),
        ("Operating Paths", build_operating_paths_summary(data)),
        ("Intent Alignment", build_portfolio_intent_alignment_summary(data)),
        ("Scorecards", build_scorecards_summary(data)),
        ("Next Action", operator_context["next_action"]),
        ("Top Recommendation", top_recommendation_summary),
    ]
    if excel_mode == "standard":
        rows.extend(
            [
                (
                    "Trend",
                    f"{operator_context['trend_status']} — {operator_context['trend_summary']}",
                ),
                ("Primary Target", operator_context["primary_target"]),
                ("Resolution Counts", operator_context["resolution_counts"]),
                (
                    "Recovery Hotspot",
                    operator_context["follow_through_relapsing_hotspot"],
                ),
                (
                    "Retiring Watch Hotspot",
                    operator_context["follow_through_retiring_hotspot"],
                ),
                ("Churn Hotspot", operator_context["follow_through_churn_hotspot"]),
                (
                    "Freshness Hotspot",
                    operator_context["follow_through_recovery_freshness_hotspot"],
                ),
                (
                    "Freshness Detail",
                    operator_context[
                        "follow_through_recovery_freshness_hotspot_summary"
                    ],
                ),
                (
                    "Rebuild Hotspot",
                    operator_context["follow_through_recovery_rebuild_hotspot"],
                ),
                (
                    "Rebuild Strength Hotspot",
                    operator_context["follow_through_rebuild_strength_hotspot"],
                ),
                (
                    "Reacquiring Hotspot",
                    operator_context["follow_through_reacquiring_hotspot"],
                ),
                (
                    "Reacquired Hotspot",
                    operator_context["follow_through_reacquired_hotspot"],
                ),
                (
                    "Fragile Reacquisition Hotspot",
                    operator_context["follow_through_fragile_reacquisition_hotspot"],
                ),
                (
                    "Just Reacquired Hotspot",
                    operator_context["follow_through_just_reacquired_hotspot"],
                ),
                (
                    "Holding Reacquired Hotspot",
                    operator_context["follow_through_holding_reacquired_hotspot"],
                ),
                (
                    "Durable Reacquired Hotspot",
                    operator_context["follow_through_durable_reacquired_hotspot"],
                ),
                (
                    "Softening Reacquired Hotspot",
                    operator_context["follow_through_softening_reacquired_hotspot"],
                ),
                (
                    "Fragile Reacquisition Confidence Hotspot",
                    operator_context[
                        "follow_through_fragile_reacquisition_confidence_hotspot"
                    ],
                ),
                (
                    "Reacquisition Softening Hotspot",
                    operator_context["follow_through_reacquisition_softening_hotspot"],
                ),
                (
                    "Revalidation Needed Hotspot",
                    operator_context[
                        "follow_through_reacquisition_revalidation_hotspot"
                    ],
                ),
                (
                    "Retired Confidence Hotspot",
                    operator_context[
                        "follow_through_reacquisition_retired_confidence_hotspot"
                    ],
                ),
                (
                    "Under Revalidation Hotspot",
                    operator_context["follow_through_under_revalidation_hotspot"],
                ),
                (
                    "Rebuilding Restored Confidence Hotspot",
                    operator_context[
                        "follow_through_rebuilding_restored_confidence_hotspot"
                    ],
                ),
                (
                    "Re-Earning Confidence Hotspot",
                    operator_context["follow_through_reearning_confidence_hotspot"],
                ),
                (
                    "Just Re-Earned Confidence Hotspot",
                    operator_context[
                        "follow_through_just_reearned_confidence_hotspot"
                    ],
                ),
                (
                    "Holding Re-Earned Confidence Hotspot",
                    operator_context[
                        "follow_through_holding_reearned_confidence_hotspot"
                    ],
                ),
            ]
        )
    rows.extend(
        [
            ("Next Run", operator_context["next_mode"]),
            ("Watch Strategy", operator_context["watch_strategy"]),
            ("Watch Decision", operator_context["watch_decision"]),
            ("Source Run", operator_summary.get("source_run_id", "")),
        ]
    )
    if excel_mode == "standard":
        rows.extend(
            [
                ("Why Top Target", operator_context["primary_target_reason"]),
                ("Follow-Through Hotspot", operator_context["follow_through_hotspot"]),
                (
                    "Escalation Hotspot",
                    operator_context["follow_through_escalation_hotspot"],
                ),
                (
                    "Recovery Hotspot",
                    operator_context["follow_through_relapsing_hotspot"],
                ),
                (
                    "Retiring Watch Hotspot",
                    operator_context["follow_through_retiring_hotspot"],
                ),
                ("Closure Guidance", operator_context["closure_guidance"]),
                ("Aging Pressure", operator_context["aging_pressure"]),
                ("What We Tried", operator_context["last_intervention"]),
                ("Last Outcome", operator_context["last_outcome"]),
                ("Resolution Evidence", operator_context["resolution_evidence"]),
                ("Recovery Counts", operator_context["recovery_counts"]),
                (
                    "Recommendation Confidence",
                    operator_context["primary_confidence"],
                ),
                ("Confidence Rationale", operator_context["confidence_reason"]),
                (
                    "Next Action Confidence",
                    operator_context["next_action_confidence"],
                ),
                ("Trust Policy", operator_context["trust_policy"]),
                ("Trust Rationale", operator_context["trust_policy_reason"]),
                (
                    "Trust Exception",
                    f"{operator_context['exception_status']} — {operator_context['exception_reason']}",
                ),
                (
                    "Trust Recovery",
                    f"{operator_context['trust_recovery_status']} — {operator_context['trust_recovery_reason']}",
                ),
                ("Recovery Confidence", operator_context["recovery_confidence"]),
                (
                    "Exception Retirement",
                    f"{operator_context['retirement_status']} — {operator_context['retirement_reason']}",
                ),
                ("Retirement Summary", operator_context["retirement_summary"]),
                (
                    "Policy Debt",
                    f"{operator_context['policy_debt_status']} — {operator_context['policy_debt_reason']}",
                ),
                (
                    "Class Normalization",
                    f"{operator_context['class_normalization_status']} — {operator_context['trust_normalization_summary']}",
                ),
                (
                    "Class Memory",
                    f"{operator_context['class_memory_status']} — {operator_context['class_memory_reason']}",
                ),
                (
                    "Trust Decay",
                    f"{operator_context['class_decay_status']} — {operator_context['class_decay_summary']}",
                ),
                (
                    "Class Reweighting",
                    f"{operator_context['class_reweight_direction']} ({operator_context['class_reweight_score']}) — {operator_context['class_reweight_summary']}",
                ),
                ("Class Reweighting Why", operator_context["class_reweight_reason"]),
                ("Class Momentum", operator_context["class_momentum_status"]),
                ("Reweight Stability", operator_context["class_reweight_stability"]),
                ("Transition Health", operator_context["class_transition_health"]),
                (
                    "Transition Resolution",
                    operator_context["class_transition_resolution"],
                ),
                ("Transition Summary", operator_context["class_transition_summary"]),
                (
                    "Transition Closure",
                    operator_context["transition_closure_confidence"],
                ),
                (
                    "Transition Likely Outcome",
                    operator_context["transition_likely_outcome"],
                ),
                (
                    "Pending Debt Freshness",
                    operator_context["pending_debt_freshness"],
                ),
                ("Closure Forecast", operator_context["closure_forecast_direction"]),
                (
                    "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Persistence",
                    operator_context[
                        "reset_reentry_rebuild_reentry_restore_rerererestore_persistence"
                    ],
                ),
                (
                    "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Churn Controls",
                    operator_context[
                        "reset_reentry_rebuild_reentry_restore_rerererestore_churn"
                    ],
                ),
                (
                    "Closure Forecast Summary",
                    operator_context["transition_closure_summary"],
                ),
                ("Momentum Summary", operator_context["class_momentum_summary"]),
                (
                    "Exception Learning",
                    f"{operator_context['exception_pattern_status']} — {operator_context['exception_pattern_summary']}",
                ),
                (
                    "Recommendation Drift",
                    f"{operator_context['drift_status']} — {operator_context['drift_summary']}",
                ),
                (
                    "Adaptive Confidence",
                    operator_context["adaptive_confidence_summary"],
                ),
                (
                    "Recommendation Quality",
                    operator_context["recommendation_quality"],
                ),
                (
                    "Confidence Validation",
                    f"{operator_context['calibration_status']} — {operator_context['calibration_summary']}",
                ),
                ("High-Confidence Hit Rate", operator_context["high_hit_rate"]),
                (
                    "Reopened Recommendations",
                    operator_context["reopened_recommendations"],
                ),
            ]
        )
    return rows


def build_dashboard_top_attention_rows(
    repo_rollups: list[list[Any]],
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for (
        repo,
        _total,
        blocked,
        urgent,
        ready,
        _deferred,
        _kind,
        _priority,
        title,
        action,
    ) in repo_rollups[:5]:
        rows.append(
            [
                repo,
                f"B{blocked} / U{urgent} / R{ready}",
                title or "See detailed queue",
                action or "Review repo detail",
            ]
        )
    return rows or [
        ["Portfolio", "No open items", "Nothing is currently queued.", "Monitor future audits"]
    ]


def build_dashboard_top_opportunity_rows(audits: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    audits_sorted = sorted(
        audits, key=lambda audit: audit.get("overall_score", 0), reverse=True
    )
    for audit in audits_sorted:
        action = (audit.get("action_candidates") or [{}])[0]
        hotspots = audit.get("hotspots") or []
        best_next_move = action.get("title", "")
        if not best_next_move and hotspots:
            best_next_move = hotspots[0].get("recommended_action", "")
        rows.append(
            [
                audit.get("metadata", {}).get("name", ""),
                round(audit.get("overall_score", 0), 3),
                audit.get("completeness_tier", ""),
                best_next_move,
            ]
        )
        if len(rows) == 5:
            break
    return rows


def build_dashboard_laggard_rows(audits: list[dict[str, Any]]) -> list[list[Any]]:
    bottom_repos = sorted(audits, key=lambda audit: audit.get("overall_score", 0))[:5]
    return [
        [
            audit.get("metadata", {}).get("name", ""),
            round(audit.get("overall_score", 0), 3),
            audit.get("completeness_tier", ""),
            (audit.get("hotspots") or [{}])[0].get("title", "Needs follow-through"),
        ]
        for audit in bottom_repos
    ]


def build_dashboard_kpi_specs(
    *,
    grade: str,
    average_score: float,
    tiers: dict[str, Any],
) -> list[tuple[str, Any, str | None, str | None]]:
    return [
        ("Portfolio Grade", grade, None, None),
        ("Avg Score", f"{average_score:.2f}", None, None),
        ("Shipped", tiers.get("shipped", 0), "166534", "Tier Breakdown"),
        ("Functional", tiers.get("functional", 0), "1565C0", "Tier Breakdown"),
        ("WIP", tiers.get("wip", 0), "D97706", "Quick Wins"),
        (
            "Needs Work",
            tiers.get("skeleton", 0) + tiers.get("abandoned", 0),
            "C2410C",
            None,
        ),
    ]


def build_dashboard_sidebar_rows(
    *,
    run_change_summary: str,
    queue_pressure_summary: str,
    trust_actionability_summary: str,
    top_recommendation_summary: str,
) -> list[tuple[str, Any]]:
    return [
        ("Run Summary", run_change_summary),
        ("Queue Pressure", queue_pressure_summary),
        ("Trust Summary", trust_actionability_summary),
        ("Top Recommendation", top_recommendation_summary),
        (
            "Use This Page",
            "Read the left side for portfolio shape, the right side for operator pressure, then jump into Review Queue or Repo Detail.",
        ),
    ]


def build_dashboard_story_blocks(
    *,
    data: dict[str, Any],
    diff_data: dict[str, Any] | None,
    excel_mode: str,
    build_run_change_summary,
    build_queue_pressure_summary,
    build_trust_actionability_summary,
    build_top_recommendation_summary,
    resolve_weekly_story_value,
    build_weekly_review_pack,
    build_executive_operator_context,
    operator_counts,
    format_lane_counts,
    display_operator_state,
    build_portfolio_catalog_summary,
    build_operating_paths_summary,
    build_portfolio_intent_alignment_summary,
    build_scorecards_summary,
) -> dict[str, Any]:
    run_change_summary = data.get("run_change_summary") or build_run_change_summary(diff_data)
    weekly_pack = build_weekly_review_pack(data, diff_data)
    queue_pressure_summary = resolve_weekly_story_value(
        weekly_pack,
        "why_this_week",
        weekly_pack.get("queue_pressure_summary"),
        build_queue_pressure_summary(data, diff_data),
    )
    trust_actionability_summary = build_trust_actionability_summary(data)
    top_recommendation_summary = resolve_weekly_story_value(
        weekly_pack,
        "decision",
        weekly_pack.get("what_to_do_this_week"),
        build_top_recommendation_summary(data),
    )
    operator_context = build_executive_operator_context(data, weekly_pack)
    operator_summary = data.get("operator_summary") or {}
    governance_summary = data.get("governance_summary") or {}
    campaign_summary = data.get("campaign_summary") or {}
    setup_health = operator_summary.get("operator_setup_health") or {}
    lane_counts = operator_counts(data)
    return {
        "sidebar_rows": build_dashboard_sidebar_rows(
            run_change_summary=run_change_summary,
            queue_pressure_summary=queue_pressure_summary,
            trust_actionability_summary=trust_actionability_summary,
            top_recommendation_summary=top_recommendation_summary,
        ),
        "operator_rows": build_dashboard_operator_rows(
            data=data,
            excel_mode=excel_mode,
            operator_summary=operator_summary,
            governance_summary=governance_summary,
            campaign_summary=campaign_summary,
            setup_health=setup_health,
            lane_counts=lane_counts,
            run_change_summary=run_change_summary,
            queue_pressure_summary=queue_pressure_summary,
            top_recommendation_summary=top_recommendation_summary,
            operator_context=operator_context,
            format_lane_counts=format_lane_counts,
            display_operator_state=display_operator_state,
            build_portfolio_catalog_summary=build_portfolio_catalog_summary,
            build_operating_paths_summary=build_operating_paths_summary,
            build_portfolio_intent_alignment_summary=build_portfolio_intent_alignment_summary,
            build_scorecards_summary=build_scorecards_summary,
        ),
    }


def build_dashboard_portfolio_trend_sparkline(
    *,
    score_history: dict[str, list[float]] | None,
    load_trends,
    render_sparkline,
) -> str:
    if not score_history:
        return ""
    average_scores = [trend.get("average_score", 0) for trend in (load_trends() or [])]
    sparkline = render_sparkline(average_scores)
    if not sparkline:
        return ""
    return f"Trend: {sparkline}"


def build_dashboard_dna_tiers(audits: list[dict[str, Any]]) -> list[str]:
    audits_sorted = sorted(
        audits, key=lambda audit: audit.get("overall_score", 0), reverse=True
    )
    return [
        audit.get("completeness_tier", "abandoned") for audit in audits_sorted[:24]
    ]


def build_dashboard_highlight_rows(data: dict[str, Any]) -> dict[str, list[str]]:
    return {
        "best_work": list(
            data.get("best_work") or data.get("summary", {}).get("highest_scored", [])
        )[:5],
        "needs_attention": list(
            data.get("summary", {}).get("lowest_scored", [])
        )[:5],
    }


def dashboard_preferred_widths() -> dict[str, int]:
    return {
        "A": 24,
        "B": 12,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 11,
        "H": 11,
        "I": 11,
        "J": 11,
        "K": 12,
        "L": 14,
        "O": 18,
        "P": 22,
        "Q": 16,
        "R": 24,
    }


def build_dashboard_language_rows(
    language_distribution: dict[str, Any],
) -> list[tuple[str, Any]]:
    return list(language_distribution.items())[:8]


def write_dashboard_visual_sections(
    *,
    ws,
    data: dict[str, Any],
    tiers: dict[str, Any],
    pie_colors: list[str],
    tier_order: list[str],
    section_font,
    subheader_font,
    subtitle_font,
    wrap_alignment,
    tier_fills: dict[str, Any],
    get_column_letter,
) -> None:
    _write_dashboard_dna_row(ws, data.get("audits", []), subheader_font, tier_fills)
    _write_dashboard_usage_note(ws, subtitle_font, wrap_alignment)
    _write_dashboard_distribution_charts(ws, tiers, data.get("audits", []), pie_colors, tier_order)
    _write_dashboard_highlights(ws, data, section_font, subheader_font)
    for column_index in range(24, 42):
        ws.column_dimensions[get_column_letter(column_index)].hidden = True
    for column_letter, width in dashboard_preferred_widths().items():
        ws.column_dimensions[column_letter].width = width
    _write_dashboard_language_chart(ws, build_dashboard_language_rows(data.get("language_distribution", {})))
    _write_dashboard_scatter_section(ws, data.get("audits", []), section_font, subheader_font)


def _write_dashboard_dna_row(ws, audits: list[dict[str, Any]], subheader_font, tier_fills) -> None:
    dna_row = 8
    ws.cell(row=dna_row, column=1, value="Portfolio DNA").font = subheader_font
    for i, tier in enumerate(build_dashboard_dna_tiers(audits)):
        cell = ws.cell(row=dna_row, column=2 + i, value="")
        if tier in tier_fills:
            cell.fill = tier_fills[tier]


def _write_dashboard_usage_note(ws, subtitle_font, wrap_alignment) -> None:
    ws.merge_cells("A9:L10")
    ws["A9"] = (
        "Use the left side for portfolio shape and the right side for operator pressure. "
        "If you only have a minute, read the narrative, scan the queue snapshot, and then open Review Queue."
    )
    ws["A9"].font = subtitle_font
    ws["A9"].alignment = wrap_alignment


def _write_dashboard_distribution_charts(
    ws,
    tiers: dict[str, Any],
    audits: list[dict[str, Any]],
    pie_colors: list[str],
    tier_order: list[str],
) -> None:
    pie_label_col = 24
    pie_value_col = 25
    pie_start = 10
    for i, tier in enumerate(tier_order):
        ws.cell(row=pie_start + i, column=pie_label_col, value=tier.capitalize())
        ws.cell(row=pie_start + i, column=pie_value_col, value=tiers.get(tier, 0))

    pie = PieChart()
    pie.title = "Tier Distribution"
    pie.style = 10
    labels = Reference(ws, min_col=pie_label_col, min_row=pie_start, max_row=pie_start + 4)
    values = Reference(ws, min_col=pie_value_col, min_row=pie_start, max_row=pie_start + 4)
    pie.add_data(values, titles_from_data=False)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    for i, color in enumerate(pie_colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)
    pie.width = 8.4
    pie.height = 6.5
    ws.add_chart(pie, "A17")

    grade_dist = {}
    for audit in audits:
        grade = audit.get("grade", "F")
        grade_dist[grade] = grade_dist.get(grade, 0) + 1
    grade_label_col = 27
    grade_value_col = 28
    grade_row = 10
    for i, grade in enumerate(["A", "B", "C", "D", "F"]):
        ws.cell(row=grade_row + i, column=grade_label_col, value=grade)
        ws.cell(row=grade_row + i, column=grade_value_col, value=grade_dist.get(grade, 0))

    bar = BarChart()
    bar.type = "col"
    bar.title = "Grade Distribution"
    bar.style = 10
    bar_data = Reference(ws, min_col=grade_value_col, min_row=grade_row, max_row=grade_row + 4)
    bar_cats = Reference(ws, min_col=grade_label_col, min_row=grade_row, max_row=grade_row + 4)
    bar.add_data(bar_data, titles_from_data=False)
    bar.set_categories(bar_cats)
    bar.width = 8.4
    bar.height = 6.5
    ws.add_chart(bar, "J17")


def _write_dashboard_highlights(ws, data: dict[str, Any], section_font, subheader_font) -> None:
    highlight_row = 31
    ws.cell(row=highlight_row, column=1, value="Highlights").font = section_font
    highlight_rows = build_dashboard_highlight_rows(data)
    if highlight_rows["best_work"]:
        ws.cell(row=highlight_row + 1, column=1, value="Best Work:").font = subheader_font
        for i, name in enumerate(highlight_rows["best_work"]):
            ws.cell(row=highlight_row + 1, column=2 + i, value=name)
    if highlight_rows["needs_attention"]:
        ws.cell(row=highlight_row + 2, column=1, value="Needs Attention:").font = subheader_font
        for i, name in enumerate(highlight_rows["needs_attention"]):
            ws.cell(row=highlight_row + 2, column=2 + i, value=name)


def _write_dashboard_language_chart(ws, language_rows: list[tuple[str, Any]]) -> None:
    lang_label_col = 30
    lang_value_col = 31
    lang_row = 10
    for i, (lang, count) in enumerate(language_rows):
        ws.cell(row=lang_row + i, column=lang_label_col, value=lang)
        ws.cell(row=lang_row + i, column=lang_value_col, value=count)

    if not language_rows:
        return

    lang_bar = BarChart()
    lang_bar.type = "bar"
    lang_bar.title = "Top Languages"
    lang_bar.style = 10
    lang_data = Reference(
        ws,
        min_col=lang_value_col,
        min_row=lang_row,
        max_row=lang_row + len(language_rows) - 1,
    )
    lang_cats = Reference(
        ws,
        min_col=lang_label_col,
        min_row=lang_row,
        max_row=lang_row + len(language_rows) - 1,
    )
    lang_bar.add_data(lang_data, titles_from_data=False)
    lang_bar.set_categories(lang_cats)
    lang_bar.width = 8.0
    lang_bar.height = 6.5
    ws.add_chart(lang_bar, "A35")


def _write_dashboard_scatter_section(ws, audits: list[dict[str, Any]], section_font, subheader_font) -> None:
    if len(audits) < 2:
        return

    data_start_row = 10
    col_name = 33
    col_x = 34
    col_y = 35
    for i, audit in enumerate(audits):
        row = data_start_row + i
        ws.cell(row=row, column=col_name, value=audit["metadata"]["name"])
        ws.cell(row=row, column=col_x, value=round(audit.get("overall_score", 0), 3))
        ws.cell(row=row, column=col_y, value=round(audit.get("interest_score", 0), 3))

    data_end_row = data_start_row + len(audits) - 1
    chart = ScatterChart()
    chart.title = "Completeness vs Interest"
    chart.x_axis.title = "Completeness"
    chart.y_axis.title = "Interest"
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 1
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.style = 13

    xvalues = Reference(ws, min_col=col_x, min_row=data_start_row, max_row=data_end_row)
    yvalues = Reference(ws, min_col=col_y, min_row=data_start_row, max_row=data_end_row)
    chart.add_data(yvalues, titles_from_data=False)
    chart.set_categories(xvalues)
    if chart.series:
        chart.series[0].graphicalProperties.line.noFill = True

    line_col = 36
    ws.cell(row=data_start_row, column=line_col, value=0.55)
    ws.cell(row=data_start_row, column=line_col + 1, value=0.0)
    ws.cell(row=data_start_row + 1, column=line_col, value=0.55)
    ws.cell(row=data_start_row + 1, column=line_col + 1, value=1.0)

    vline_y = Reference(
        ws, min_col=line_col + 1, min_row=data_start_row, max_row=data_start_row + 1
    )
    vline_x = Reference(ws, min_col=line_col, min_row=data_start_row, max_row=data_start_row + 1)
    chart.add_data(vline_y, titles_from_data=False)
    chart.set_categories(vline_x)
    if len(chart.series) > 1:
        chart.series[-1].graphicalProperties.line = LineProperties(
            w=12700, prstDash="dash", solidFill="808080"
        )

    ws.cell(row=data_start_row, column=line_col + 2, value=0.0)
    ws.cell(row=data_start_row, column=line_col + 3, value=0.45)
    ws.cell(row=data_start_row + 1, column=line_col + 2, value=1.0)
    ws.cell(row=data_start_row + 1, column=line_col + 3, value=0.45)

    hline_y = Reference(
        ws, min_col=line_col + 3, min_row=data_start_row, max_row=data_start_row + 1
    )
    hline_x = Reference(
        ws, min_col=line_col + 2, min_row=data_start_row, max_row=data_start_row + 1
    )
    chart.add_data(hline_y, titles_from_data=False)
    chart.set_categories(hline_x)
    if len(chart.series) > 2:
        chart.series[-1].graphicalProperties.line = LineProperties(
            w=12700, prstDash="dash", solidFill="808080"
        )

    chart.width = 10.5
    chart.height = 8
    ws.add_chart(chart, "J34")
    _write_quadrant_table(ws, audits, legend_row=52, section_font=section_font, subheader_font=subheader_font)


def _write_quadrant_table(ws, audits: list[dict[str, Any]], *, legend_row: int, section_font, subheader_font) -> None:
    buckets: list[list[str]] = [[], [], [], []]
    for audit in audits:
        x = audit.get("overall_score", 0)
        y = audit.get("interest_score", 0)
        if x >= X_MID and y >= Y_MID:
            buckets[0].append(audit["metadata"]["name"])
        elif x < X_MID and y >= Y_MID:
            buckets[1].append(audit["metadata"]["name"])
        elif x >= X_MID and y < Y_MID:
            buckets[2].append(audit["metadata"]["name"])
        else:
            buckets[3].append(audit["metadata"]["name"])

    ws.cell(row=legend_row, column=10, value="Scatter Quadrants").font = section_font
    for j, header in enumerate(["Quadrant", "Count", "Repos"]):
        ws.cell(row=legend_row + 1, column=10 + j, value=header).font = subheader_font

    for i, ((name, _desc), repos) in enumerate(zip(QUADRANT_NAMES, buckets)):
        row = legend_row + 2 + i
        ws.cell(row=row, column=10, value=name).font = Font("Calibri", 11, bold=True)
        ws.cell(row=row, column=11, value=len(repos))
        ws.cell(
            row=row,
            column=12,
            value=", ".join(repos[:8]) + ("..." if len(repos) > 8 else ""),
        )


def build_dashboard_sheet(
    wb: Workbook,
    data: dict[str, Any],
    diff_data: dict[str, Any] | None = None,
    score_history: dict[str, list[float]] | None = None,
    *,
    excel_mode: str,
    get_or_create_sheet,
    clear_worksheet,
    configure_sheet_view,
    write_dashboard_header,
    build_dashboard_kpi_specs,
    write_kpi_card,
    sheet_location,
    build_dashboard_portfolio_trend_sparkline,
    load_trend_data,
    render_sparkline,
    build_dashboard_story_blocks,
    write_dashboard_story_sections,
    build_dashboard_ranked_content,
    write_dashboard_ranked_sections,
    write_dashboard_visual_sections,
    build_run_change_summary,
    build_queue_pressure_summary,
    build_trust_actionability_summary,
    build_top_recommendation_summary,
    resolve_weekly_story_value,
    build_weekly_review_pack,
    build_executive_operator_context,
    operator_counts,
    format_lane_counts,
    display_operator_state,
    build_portfolio_catalog_summary,
    build_operating_paths_summary,
    build_portfolio_intent_alignment_summary,
    build_scorecards_summary,
    build_workbook_rollups,
    build_dashboard_top_attention_rows,
    build_dashboard_top_opportunity_rows,
    build_dashboard_laggard_rows,
    write_key_value_block,
    write_ranked_list,
    title_font,
    subtitle_font,
    narrative_font,
    center_alignment,
    wrap_alignment,
    section_font,
    subheader_font,
    navy: str,
    grade_colors: dict[str, str],
    pie_colors,
    tier_order,
    tier_fills,
    sparkline_font,
    get_column_letter,
    generate_narrative,
) -> None:
    if "Dashboard" in wb.sheetnames:
        ws = get_or_create_sheet(wb, "Dashboard")
    elif len(wb.sheetnames) == 1 and wb.active.title == "Sheet":
        ws = wb.active
        ws.title = "Dashboard"
        clear_worksheet(ws)
    else:
        ws = get_or_create_sheet(wb, "Dashboard")
    ws.sheet_properties.tabColor = navy
    configure_sheet_view(ws, zoom=120, show_grid_lines=False)
    write_dashboard_header(
        ws=ws,
        data=data,
        diff_data=diff_data,
        title_font=title_font,
        subtitle_font=subtitle_font,
        narrative_font=narrative_font,
        center_alignment=center_alignment,
        wrap_alignment=wrap_alignment,
        section_font=section_font,
        generate_narrative=generate_narrative,
    )

    grade = data.get("portfolio_grade", "?")
    grade_color = grade_colors.get(grade, navy)
    tiers = data.get("tier_distribution", {})
    for offset, (label, value, accent, target_sheet) in enumerate(
        build_dashboard_kpi_specs(
            grade=grade,
            average_score=data["average_score"],
            tiers=tiers,
        )
    ):
        link = f"#{sheet_location(target_sheet)}" if target_sheet else None
        write_kpi_card(
            ws,
            5,
            1 + offset * 2,
            label,
            value,
            grade_color if offset == 0 else accent,
            link,
        )

    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 40

    trend_summary = build_dashboard_portfolio_trend_sparkline(
        score_history=score_history,
        load_trends=load_trend_data,
        render_sparkline=render_sparkline,
    )
    if trend_summary:
        cell = ws.cell(row=7, column=3, value=trend_summary)
        cell.font = sparkline_font

    dashboard_story = build_dashboard_story_blocks(
        data=data,
        diff_data=diff_data,
        excel_mode=excel_mode,
        build_run_change_summary=build_run_change_summary,
        build_queue_pressure_summary=build_queue_pressure_summary,
        build_trust_actionability_summary=build_trust_actionability_summary,
        build_top_recommendation_summary=build_top_recommendation_summary,
        resolve_weekly_story_value=resolve_weekly_story_value,
        build_weekly_review_pack=build_weekly_review_pack,
        build_executive_operator_context=build_executive_operator_context,
        operator_counts=operator_counts,
        format_lane_counts=format_lane_counts,
        display_operator_state=display_operator_state,
        build_portfolio_catalog_summary=build_portfolio_catalog_summary,
        build_operating_paths_summary=build_operating_paths_summary,
        build_portfolio_intent_alignment_summary=build_portfolio_intent_alignment_summary,
        build_scorecards_summary=build_scorecards_summary,
    )
    operator_block_end = write_dashboard_story_sections(
        ws=ws,
        dashboard_story=dashboard_story,
        subheader_font=subheader_font,
        wrap_alignment=wrap_alignment,
        write_key_value_block=write_key_value_block,
    )
    ranked_content = build_dashboard_ranked_content(
        data=data,
        operator_block_end=operator_block_end,
        build_workbook_rollups=build_workbook_rollups,
        build_dashboard_top_attention_rows=build_dashboard_top_attention_rows,
        build_dashboard_top_opportunity_rows=build_dashboard_top_opportunity_rows,
        build_dashboard_laggard_rows=build_dashboard_laggard_rows,
    )
    write_dashboard_ranked_sections(
        ws=ws,
        ranked_content=ranked_content,
        write_ranked_list=write_ranked_list,
    )
    write_dashboard_visual_sections(
        ws=ws,
        data=data,
        tiers=tiers,
        pie_colors=pie_colors,
        tier_order=tier_order,
        section_font=section_font,
        subheader_font=subheader_font,
        subtitle_font=subtitle_font,
        wrap_alignment=wrap_alignment,
        tier_fills=tier_fills,
        get_column_letter=get_column_letter,
    )


def build_dashboard_bubble_chart(
    ws,
    data: dict[str, Any],
    *,
    bubble_chart_cls,
    reference_cls,
) -> None:
    """Add a bubble chart: x=completeness, y=interest, size=LOC."""
    audits = data.get("audits", [])
    if len(audits) < 2:
        return

    bcol_x, bcol_y, bcol_z = 20, 21, 22
    data_start = 10

    for index, audit in enumerate(audits):
        row = data_start + index
        code_quality = next(
            (
                result.get("details", {})
                for result in audit.get("analyzer_results", [])
                if result["dimension"] == "code_quality"
            ),
            {},
        )
        loc = code_quality.get("total_loc", 100)
        ws.cell(row=row, column=bcol_x, value=round(audit.get("overall_score", 0), 3))
        ws.cell(row=row, column=bcol_y, value=round(audit.get("interest_score", 0), 3))
        ws.cell(row=row, column=bcol_z, value=max(loc, 10))

    data_end = data_start + len(audits) - 1

    chart = bubble_chart_cls()
    chart.title = "Portfolio Map (size = LOC)"
    chart.x_axis.title = "Completeness"
    chart.y_axis.title = "Interest"
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 1
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.style = 18

    xvalues = reference_cls(ws, min_col=bcol_x, min_row=data_start, max_row=data_end)
    yvalues = reference_cls(ws, min_col=bcol_y, min_row=data_start, max_row=data_end)
    sizes = reference_cls(ws, min_col=bcol_z, min_row=data_start, max_row=data_end)
    series = BubbleSeries(values=yvalues, xvalues=xvalues, zvalues=sizes, title="Repos")
    chart.series.append(series)

    chart.width = 16
    chart.height = 12
    ws.add_chart(chart, "A50")


def build_dashboard_workbook_sheet(
    wb: Workbook,
    data: dict[str, Any],
    diff_data: dict[str, Any] | None = None,
    score_history: dict[str, list[float]] | None = None,
    *,
    excel_mode: str,
    get_or_create_sheet,
    clear_worksheet,
    configure_sheet_view,
    write_dashboard_header,
    build_dashboard_kpi_specs,
    write_kpi_card,
    sheet_location,
    build_dashboard_portfolio_trend_sparkline,
    load_trend_data,
    render_sparkline,
    build_dashboard_story_blocks,
    write_dashboard_story_sections,
    build_dashboard_ranked_content,
    write_dashboard_ranked_sections,
    write_dashboard_visual_sections,
    build_run_change_summary,
    build_queue_pressure_summary,
    build_trust_actionability_summary,
    build_top_recommendation_summary,
    resolve_weekly_story_value,
    build_weekly_review_pack,
    build_executive_operator_context,
    operator_counts,
    format_lane_counts,
    display_operator_state,
    build_portfolio_catalog_summary,
    build_operating_paths_summary,
    build_portfolio_intent_alignment_summary,
    build_scorecards_summary,
    build_workbook_rollups,
    build_dashboard_top_attention_rows,
    build_dashboard_top_opportunity_rows,
    build_dashboard_laggard_rows,
    write_key_value_block,
    write_ranked_list,
    title_font,
    subtitle_font,
    narrative_font,
    center_alignment,
    wrap_alignment,
    section_font,
    subheader_font,
    navy: str,
    grade_colors: dict[str, str],
    pie_colors,
    tier_order,
    tier_fills,
    sparkline_font,
    get_column_letter,
    generate_narrative,
) -> None:
    build_dashboard_sheet(
        wb,
        data,
        diff_data,
        score_history,
        excel_mode=excel_mode,
        get_or_create_sheet=get_or_create_sheet,
        clear_worksheet=clear_worksheet,
        configure_sheet_view=configure_sheet_view,
        write_dashboard_header=write_dashboard_header,
        build_dashboard_kpi_specs=build_dashboard_kpi_specs,
        write_kpi_card=write_kpi_card,
        sheet_location=sheet_location,
        build_dashboard_portfolio_trend_sparkline=build_dashboard_portfolio_trend_sparkline,
        load_trend_data=load_trend_data,
        render_sparkline=render_sparkline,
        build_dashboard_story_blocks=build_dashboard_story_blocks,
        write_dashboard_story_sections=write_dashboard_story_sections,
        build_dashboard_ranked_content=build_dashboard_ranked_content,
        write_dashboard_ranked_sections=write_dashboard_ranked_sections,
        write_dashboard_visual_sections=write_dashboard_visual_sections,
        build_run_change_summary=build_run_change_summary,
        build_queue_pressure_summary=build_queue_pressure_summary,
        build_trust_actionability_summary=build_trust_actionability_summary,
        build_top_recommendation_summary=build_top_recommendation_summary,
        resolve_weekly_story_value=resolve_weekly_story_value,
        build_weekly_review_pack=build_weekly_review_pack,
        build_executive_operator_context=build_executive_operator_context,
        operator_counts=operator_counts,
        format_lane_counts=format_lane_counts,
        display_operator_state=display_operator_state,
        build_portfolio_catalog_summary=build_portfolio_catalog_summary,
        build_operating_paths_summary=build_operating_paths_summary,
        build_portfolio_intent_alignment_summary=build_portfolio_intent_alignment_summary,
        build_scorecards_summary=build_scorecards_summary,
        build_workbook_rollups=build_workbook_rollups,
        build_dashboard_top_attention_rows=build_dashboard_top_attention_rows,
        build_dashboard_top_opportunity_rows=build_dashboard_top_opportunity_rows,
        build_dashboard_laggard_rows=build_dashboard_laggard_rows,
        write_key_value_block=write_key_value_block,
        write_ranked_list=write_ranked_list,
        title_font=title_font,
        subtitle_font=subtitle_font,
        narrative_font=narrative_font,
        center_alignment=center_alignment,
        wrap_alignment=wrap_alignment,
        section_font=section_font,
        subheader_font=subheader_font,
        navy=navy,
        grade_colors=grade_colors,
        pie_colors=pie_colors,
        tier_order=tier_order,
        tier_fills=tier_fills,
        sparkline_font=sparkline_font,
        get_column_letter=get_column_letter,
        generate_narrative=generate_narrative,
    )
