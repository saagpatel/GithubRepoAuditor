"""Helpers for review-queue workbook content."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font


def build_review_queue_summary_content(
    *,
    data: dict,
    excel_mode: str,
    counts: dict,
    queue: list[dict],
    ordered_queue: list[dict],
    repo_rollups: list[list[object]],
    top_issue_families: list[tuple[str, int]],
    top_recommendation_summary: str,
    queue_pressure_summary: str,
    operator_context: dict,
    format_lane_counts,
    format_repo_rollup_counts,
    summarize_top_actions,
) -> tuple[list[tuple[str, object]], list[list[object]], list[list[object]], list[list[object]]]:
    summary_rows = [
        (
            "Headline",
            (data.get("operator_summary") or {}).get(
                "headline", "Review activity is available below."
            ),
        ),
        ("Queue Counts", format_lane_counts(counts)),
        ("Queue Pressure", queue_pressure_summary),
        ("Total Queue Items", len(queue)),
        (
            "Immediate Focus",
            (ordered_queue[0].get("recommended_action") or ordered_queue[0].get("title", ""))
            if ordered_queue
            else "No immediate queue item is open.",
        ),
        ("Top Recommendation", top_recommendation_summary),
        (
            "Top Issue Family",
            f"{top_issue_families[0][0]} ({top_issue_families[0][1]})"
            if top_issue_families
            else "No material change families",
        ),
    ]
    if excel_mode == "standard":
        summary_rows.extend(
            [
                ("Trend", f"{operator_context['trend_status']} — {operator_context['trend_summary']}"),
                ("Primary Target", operator_context["primary_target"]),
                ("Resolution Counts", operator_context["resolution_counts"]),
                ("Why Top Target", operator_context["primary_target_reason"]),
                ("Closure Guidance", operator_context["closure_guidance"]),
                ("Aging Pressure", operator_context["aging_pressure"]),
                ("What We Tried", operator_context["last_intervention"]),
                ("Last Outcome", operator_context["last_outcome"]),
                ("Resolution Evidence", operator_context["resolution_evidence"]),
                ("Recovery Counts", operator_context["recovery_counts"]),
                ("Recommendation Confidence", operator_context["primary_confidence"]),
                ("Confidence Rationale", operator_context["confidence_reason"]),
                ("Next Action Confidence", operator_context["next_action_confidence"]),
                ("Trust Policy", operator_context["trust_policy"]),
                ("Trust Rationale", operator_context["trust_policy_reason"]),
                (
                    "Trust Exception",
                    f"{operator_context['exception_status']} — {operator_context['exception_reason']}",
                ),
                (
                    "Trust Recovery",
                    f"{operator_context['trust_recovery_status']} — {operator_context['trust_recovery_reason']}",
                ),
                ("Recovery Confidence", operator_context["recovery_confidence"]),
                (
                    "Exception Retirement",
                    f"{operator_context['retirement_status']} — {operator_context['retirement_reason']}",
                ),
                ("Retirement Summary", operator_context["retirement_summary"]),
                (
                    "Policy Debt",
                    f"{operator_context['policy_debt_status']} — {operator_context['policy_debt_reason']}",
                ),
                (
                    "Class Normalization",
                    f"{operator_context['class_normalization_status']} — {operator_context['trust_normalization_summary']}",
                ),
                (
                    "Class Memory",
                    f"{operator_context['class_memory_status']} — {operator_context['class_memory_reason']}",
                ),
                (
                    "Trust Decay",
                    f"{operator_context['class_decay_status']} — {operator_context['class_decay_summary']}",
                ),
                (
                    "Class Reweighting",
                    f"{operator_context['class_reweight_direction']} ({operator_context['class_reweight_score']}) — {operator_context['class_reweight_summary']}",
                ),
                ("Class Reweighting Why", operator_context["class_reweight_reason"]),
                ("Class Momentum", operator_context["class_momentum_status"]),
                ("Reweight Stability", operator_context["class_reweight_stability"]),
                ("Transition Health", operator_context["class_transition_health"]),
                ("Transition Resolution", operator_context["class_transition_resolution"]),
                ("Transition Summary", operator_context["class_transition_summary"]),
                ("Transition Closure", operator_context["transition_closure_confidence"]),
                ("Transition Likely Outcome", operator_context["transition_likely_outcome"]),
                ("Pending Debt Freshness", operator_context["pending_debt_freshness"]),
                ("Closure Forecast", operator_context["closure_forecast_direction"]),
                (
                    "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Persistence",
                    operator_context[
                        "reset_reentry_rebuild_reentry_restore_rerererestore_persistence"
                    ],
                ),
                (
                    "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Churn Controls",
                    operator_context[
                        "reset_reentry_rebuild_reentry_restore_rerererestore_churn"
                    ],
                ),
                ("Closure Forecast Summary", operator_context["transition_closure_summary"]),
                ("Momentum Summary", operator_context["class_momentum_summary"]),
                (
                    "Exception Learning",
                    f"{operator_context['exception_pattern_status']} — {operator_context['exception_pattern_summary']}",
                ),
                (
                    "Recommendation Drift",
                    f"{operator_context['drift_status']} — {operator_context['drift_summary']}",
                ),
                ("Adaptive Confidence", operator_context["adaptive_confidence_summary"]),
                ("Recommendation Quality", operator_context["recommendation_quality"]),
                (
                    "Confidence Validation",
                    f"{operator_context['calibration_status']} — {operator_context['calibration_summary']}",
                ),
                ("High-Confidence Hit Rate", operator_context["high_hit_rate"]),
                ("Reopened Recommendations", operator_context["reopened_recommendations"]),
            ]
        )
    summary_rows.append(("Source Run", (data.get("operator_summary") or {}).get("source_run_id", "")))

    top_repo_rows = [
        [
            repo,
            _primary_lane_label(blocked, urgent, ready, deferred),
            format_repo_rollup_counts(blocked, urgent, ready, deferred),
            title or "See detailed queue rows below.",
            action or "Open the repo queue details.",
        ]
        for repo, _total, blocked, urgent, ready, deferred, _kind, _priority, title, action in repo_rollups[:10]
    ] or [
        ["Portfolio", "Clear", "0 blocked, 0 urgent, 0 ready, 0 deferred", "No open review items.", "Monitor future audits."]
    ]

    issue_family_rows = [[label, count] for label, count in top_issue_families] or [["No material change families", 0]]
    action_rows = [[action, count] for action, count in summarize_top_actions(queue)] or [["No recommended actions", 0]]
    return summary_rows, top_repo_rows, issue_family_rows, action_rows


def build_review_queue_sheet_content(
    *,
    data: dict,
    excel_mode: str,
    build_workbook_rollups,
    operator_counts,
    ordered_queue_items,
    summarize_top_issue_families,
    summarize_top_actions,
    build_queue_pressure_summary,
    build_top_recommendation_summary,
    build_weekly_review_pack,
    build_executive_operator_context,
    format_lane_counts,
    format_repo_rollup_counts,
    build_review_queue_table_rows,
) -> dict[str, object]:
    counts = operator_counts(data)
    queue = data.get("operator_queue", []) or []
    queue_pressure_summary = build_queue_pressure_summary(data)
    top_recommendation_summary = build_top_recommendation_summary(data)
    material_changes = data.get("material_changes", []) or []
    ordered_queue = ordered_queue_items(queue)
    repo_rollups = build_workbook_rollups(data)[1]
    top_issue_families = summarize_top_issue_families(material_changes)
    weekly_pack = build_weekly_review_pack(data, None)
    operator_context = build_executive_operator_context(data, weekly_pack)
    summary_rows, top_repo_rows, issue_family_rows, action_rows = build_review_queue_summary_content(
        data=data,
        excel_mode=excel_mode,
        counts=counts,
        queue=queue,
        ordered_queue=ordered_queue,
        repo_rollups=repo_rollups,
        top_issue_families=top_issue_families,
        top_recommendation_summary=top_recommendation_summary,
        queue_pressure_summary=queue_pressure_summary,
        operator_context=operator_context,
        format_lane_counts=format_lane_counts,
        format_repo_rollup_counts=format_repo_rollup_counts,
        summarize_top_actions=summarize_top_actions,
    )
    targets = ordered_queue or data.get("review_targets", [])
    table_rows = build_review_queue_table_rows(targets, data.get("review_summary") or {})
    return {
        "summary_rows": summary_rows,
        "top_repo_rows": top_repo_rows,
        "issue_family_rows": issue_family_rows,
        "action_rows": action_rows,
        "targets": targets,
        "table_rows": table_rows,
    }


def write_review_queue_overview(
    ws,
    content: dict[str, object],
    *,
    write_key_value_block,
    write_ranked_list,
    guidance_text: str,
    subtitle_font,
    wrap_alignment,
) -> int:
    summary_end = write_key_value_block(ws, 4, 1, content["summary_rows"], title="Summary")
    top_repo_end = write_ranked_list(
        ws,
        4,
        5,
        "Top 10 To Act On",
        ["Repo", "Primary Lane", "Counts", "Why Now", "Next Step"],
        content["top_repo_rows"],
    )
    secondary_start = max(summary_end, top_repo_end) + 2
    issue_family_end = write_ranked_list(
        ws,
        secondary_start,
        1,
        "Top Issue Families",
        ["Issue Family", "Count"],
        content["issue_family_rows"],
    )
    action_end = write_ranked_list(
        ws,
        secondary_start,
        5,
        "Top Recommended Actions",
        ["Action", "Count"],
        content["action_rows"],
    )
    guidance_row = max(issue_family_end, action_end) + 2
    ws.merge_cells(start_row=guidance_row, start_column=1, end_row=guidance_row, end_column=8)
    ws.cell(row=guidance_row, column=1, value=guidance_text).font = subtitle_font
    ws.cell(row=guidance_row, column=1).alignment = wrap_alignment
    return guidance_row + 1


def write_review_queue_table(
    ws,
    content: dict[str, object],
    *,
    start_row: int,
    headers: list[str],
    center_aligned_columns: set[int],
    style_header_row,
    style_data_cell,
    apply_lane_row_fill,
    apply_zebra_stripes,
    set_autofilter,
    auto_width,
    link_color: str,
) -> int:
    for col, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=header)
    style_header_row(ws, start_row, len(headers))
    ws.freeze_panes = f"A{start_row + 1}"

    for row, item in enumerate(content["table_rows"], start_row + 1):
        for col, value in enumerate(item["values"], 1):
            align = "center" if col in center_aligned_columns else "left"
            style_data_cell(ws.cell(row=row, column=col, value=value), align)
        repo_cell = ws.cell(row=row, column=1)
        if item["repo_url"]:
            repo_cell.hyperlink = item["repo_url"]
            repo_cell.font = Font("Calibri", 10, color=link_color, underline="single")
        artifact_cell = ws.cell(row=row, column=40)
        if item["artifact_url"]:
            artifact_cell.hyperlink = item["artifact_url"]
            artifact_cell.font = Font("Calibri", 10, color=link_color, underline="single")
        apply_lane_row_fill(ws, row, len(headers), item["lane"])

    max_row = len(content["targets"]) + start_row
    if content["targets"]:
        apply_zebra_stripes(ws, start_row + 1, max_row, len(headers))
        for row, item in enumerate(content["table_rows"], start_row + 1):
            apply_lane_row_fill(ws, row, len(headers), item["lane"])
        set_autofilter(ws, len(headers), max_row, start_row=start_row)
    auto_width(ws, len(headers), max_row)
    return max_row


def _primary_lane_label(blocked: object, urgent: object, ready: object, deferred: object) -> str:
    if blocked:
        return "Blocked"
    if urgent:
        return "Needs Attention Now"
    if ready:
        return "Ready for Manual Action"
    if deferred:
        return "Safe to Defer"
    return "Clear"


def build_review_queue_sheet(
    wb: Workbook,
    data: dict,
    *,
    excel_mode: str,
    get_or_create_sheet,
    configure_sheet_view,
    set_sheet_header,
    write_instruction_banner,
    build_review_queue_sheet_content,
    build_workbook_rollups,
    operator_counts,
    ordered_queue_items,
    summarize_top_issue_families,
    summarize_top_actions,
    build_queue_pressure_summary,
    build_top_recommendation_summary,
    build_weekly_review_pack,
    build_executive_operator_context,
    format_lane_counts,
    format_repo_rollup_counts,
    build_review_queue_table_rows,
    write_review_queue_overview,
    write_review_queue_table,
    write_key_value_block,
    write_ranked_list,
    guidance_text: str,
    review_queue_headers: list[str],
    review_queue_center_aligned_columns: set[int],
    style_header_row,
    style_data_cell,
    apply_lane_row_fill,
    apply_zebra_stripes,
    set_autofilter,
    auto_width,
    subtitle_font,
    wrap_alignment,
    section_font,
    teal: str,
) -> None:
    ws = get_or_create_sheet(wb, "Review Queue")
    ws.sheet_properties.tabColor = "2563EB"
    configure_sheet_view(ws, zoom=115, show_grid_lines=False)
    set_sheet_header(
        ws,
        "Review Queue",
        "Start with the summary strip, then use the full queue below when you need row-level facts and decision context.",
        width=12,
    )
    write_instruction_banner(
        ws,
        3,
        10,
        "Work this page in lane order: clear Blocked first, handle Needs Attention Now next, and only then move into Ready for Manual Action or Safe to Defer.",
    )
    content = build_review_queue_sheet_content(
        data=data,
        excel_mode=excel_mode,
        build_workbook_rollups=build_workbook_rollups,
        operator_counts=operator_counts,
        ordered_queue_items=ordered_queue_items,
        summarize_top_issue_families=summarize_top_issue_families,
        summarize_top_actions=summarize_top_actions,
        build_queue_pressure_summary=build_queue_pressure_summary,
        build_top_recommendation_summary=build_top_recommendation_summary,
        build_weekly_review_pack=build_weekly_review_pack,
        build_executive_operator_context=build_executive_operator_context,
        format_lane_counts=format_lane_counts,
        format_repo_rollup_counts=format_repo_rollup_counts,
        build_review_queue_table_rows=build_review_queue_table_rows,
    )
    start_row = write_review_queue_overview(
        ws,
        content,
        write_key_value_block=write_key_value_block,
        write_ranked_list=write_ranked_list,
        guidance_text=guidance_text,
        subtitle_font=subtitle_font,
        wrap_alignment=wrap_alignment,
    )
    write_review_queue_table(
        ws,
        content,
        start_row=start_row,
        headers=list(review_queue_headers),
        center_aligned_columns=set(review_queue_center_aligned_columns),
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_lane_row_fill=apply_lane_row_fill,
        apply_zebra_stripes=apply_zebra_stripes,
        set_autofilter=set_autofilter,
        auto_width=auto_width,
        link_color=teal,
    )


def build_review_queue_workbook_sheet(
    wb: Workbook,
    data: dict,
    *,
    excel_mode: str,
    get_or_create_sheet,
    configure_sheet_view,
    set_sheet_header,
    write_instruction_banner,
    build_review_queue_sheet_content,
    build_workbook_rollups,
    operator_counts,
    ordered_queue_items,
    summarize_top_issue_families,
    summarize_top_actions,
    build_queue_pressure_summary,
    build_top_recommendation_summary,
    build_weekly_review_pack,
    build_executive_operator_context,
    format_lane_counts,
    format_repo_rollup_counts,
    build_review_queue_table_rows,
    write_review_queue_overview,
    write_review_queue_table,
    write_key_value_block,
    write_ranked_list,
    guidance_text: str,
    review_queue_headers: list[str],
    review_queue_center_aligned_columns: set[int],
    style_header_row,
    style_data_cell,
    apply_lane_row_fill,
    apply_zebra_stripes,
    set_autofilter,
    auto_width,
    subtitle_font,
    wrap_alignment,
    section_font,
    teal: str,
) -> None:
    build_review_queue_sheet(
        wb,
        data,
        excel_mode=excel_mode,
        get_or_create_sheet=get_or_create_sheet,
        configure_sheet_view=configure_sheet_view,
        set_sheet_header=set_sheet_header,
        write_instruction_banner=write_instruction_banner,
        build_review_queue_sheet_content=build_review_queue_sheet_content,
        build_workbook_rollups=build_workbook_rollups,
        operator_counts=operator_counts,
        ordered_queue_items=ordered_queue_items,
        summarize_top_issue_families=summarize_top_issue_families,
        summarize_top_actions=summarize_top_actions,
        build_queue_pressure_summary=build_queue_pressure_summary,
        build_top_recommendation_summary=build_top_recommendation_summary,
        build_weekly_review_pack=build_weekly_review_pack,
        build_executive_operator_context=build_executive_operator_context,
        format_lane_counts=format_lane_counts,
        format_repo_rollup_counts=format_repo_rollup_counts,
        build_review_queue_table_rows=build_review_queue_table_rows,
        write_review_queue_overview=write_review_queue_overview,
        write_review_queue_table=write_review_queue_table,
        write_key_value_block=write_key_value_block,
        write_ranked_list=write_ranked_list,
        guidance_text=guidance_text,
        review_queue_headers=review_queue_headers,
        review_queue_center_aligned_columns=review_queue_center_aligned_columns,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_lane_row_fill=apply_lane_row_fill,
        apply_zebra_stripes=apply_zebra_stripes,
        set_autofilter=set_autofilter,
        auto_width=auto_width,
        subtitle_font=subtitle_font,
        wrap_alignment=wrap_alignment,
        section_font=section_font,
        teal=teal,
    )
