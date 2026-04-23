"""Helpers for one-off visible portfolio workbook sheets."""

from __future__ import annotations

from typing import Any

from openpyxl.chart import BarChart, RadarChart, Reference

TECH_STACK_HEADERS = ["Language", "Repos", "Bytes", "Avg Score", "Proficiency"]
REGISTRY_MATCHED_HEADERS = ["GitHub", "Registry", "Status", "Tier", "Score"]
TIER_BREAKDOWN_HEADERS = ["Repo", "Grade", "Score", "Language", "Badges", "Description"]
RISK_SUMMARY_HEADERS = ["Risk Tier", "Count"]
REPO_PROFILE_RADAR_DIMS = [
    "readme",
    "structure",
    "code_quality",
    "testing",
    "cicd",
    "dependencies",
    "activity",
    "documentation",
    "build_readiness",
    "community_profile",
]
REPO_PROFILE_RADAR_LABELS = [
    "README",
    "Structure",
    "Code Quality",
    "Testing",
    "CI/CD",
    "Deps",
    "Activity",
    "Docs",
    "Build Ready",
    "Community",
]


def build_tech_stack_rows(tech_stack: dict[str, dict[str, Any]]) -> list[list[Any]]:
    return [
        [
            language,
            info.get("repos", 0),
            info.get("bytes", 0),
            info.get("avg_score", 0),
            info.get("proficiency", 0),
        ]
        for language, info in tech_stack.items()
    ]


def build_tech_stack_chart(ws, *, start_row: int, end_row: int):
    chart = BarChart()
    chart.type = "col"
    chart.title = "Language Proficiency (bytes x quality)"
    chart.style = 10
    chart_data = Reference(ws, min_col=5, min_row=start_row - 1, max_row=end_row)
    chart_cats = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(chart_cats)
    chart.width = 20
    chart.height = 12
    return chart


def write_tech_stack_table(
    ws,
    rows: list[list[Any]],
    *,
    start_row: int,
    style_header_row,
    style_data_cell,
) -> int:
    for col, header in enumerate(TECH_STACK_HEADERS, 1):
        ws.cell(row=start_row, column=col, value=header)
    style_header_row(ws, start_row, len(TECH_STACK_HEADERS))
    ws.freeze_panes = "B2"

    for row_number, values in enumerate(rows, start_row + 1):
        for col, value in enumerate(values, 1):
            style_data_cell(ws.cell(row=row_number, column=col, value=value))
    return start_row + len(rows)


def write_tech_stack_best_work(
    ws,
    best_work: list[str],
    *,
    start_row: int,
    section_font,
    thin_border,
) -> int:
    if not best_work:
        return start_row

    ws.cell(row=start_row, column=1, value="Best Work (Top 5)").font = section_font
    for row_number, name in enumerate(best_work, start_row + 1):
        ws.cell(row=row_number, column=1, value=f"{row_number - start_row}. {name}").border = (
            thin_border
        )
    return start_row + len(best_work)


def build_tech_stack_sheet(
    wb,
    data: dict[str, Any],
    *,
    get_or_create_sheet,
    configure_sheet_view,
    auto_width,
    section_font,
    thin_border,
    style_header_row,
    style_data_cell,
) -> None:
    tech_stack = data.get("tech_stack", {})
    if not tech_stack:
        return

    ws = get_or_create_sheet(wb, "Tech Stack")
    ws.sheet_properties.tabColor = "4A148C"
    configure_sheet_view(ws, zoom=110, show_grid_lines=True)

    ws.merge_cells("A1:E1")
    ws["A1"].value = "Technology Stack"
    ws["A1"].font = section_font

    tech_stack_rows = build_tech_stack_rows(tech_stack)
    max_row = write_tech_stack_table(
        ws,
        tech_stack_rows,
        start_row=3,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
    )
    if len(tech_stack) > 1:
        ws.add_chart(build_tech_stack_chart(ws, start_row=4, end_row=max_row), "G3")

    final_row = write_tech_stack_best_work(
        ws,
        data.get("best_work", []),
        start_row=max_row + 3,
        section_font=section_font,
        thin_border=thin_border,
    )

    auto_width(ws, len(TECH_STACK_HEADERS), max(final_row, max_row) + 10)


def build_registry_content(reconciliation: dict[str, Any]) -> dict[str, Any]:
    matched_rows = [
        [
            item.get("github_name", ""),
            item.get("registry_name", ""),
            item.get("registry_status", ""),
            item.get("audit_tier", ""),
            item.get("score", 0),
        ]
        for item in reconciliation.get("matched", []) or []
    ]
    return {
        "title": f"Registry Reconciliation ({reconciliation.get('registry_total', 0)} projects)",
        "matched_rows": matched_rows,
        "unmatched_github": list(reconciliation.get("on_github_not_registry", []) or []),
        "unmatched_registry": list(reconciliation.get("in_registry_not_github", []) or []),
    }


def write_registry_sections(
    ws,
    content: dict[str, Any],
    *,
    section_font,
    subheader_font,
    thin_border,
    style_header_row,
    style_data_cell,
    color_tier_cell,
) -> int:
    row = 1
    ws.merge_cells("A1:E1")
    ws["A1"].value = content["title"]
    ws["A1"].font = section_font

    if content["matched_rows"]:
        row = 3
        ws.cell(row=row, column=1, value=f"Matched ({len(content['matched_rows'])})").font = (
            subheader_font
        )
        row += 1
        for col, header in enumerate(REGISTRY_MATCHED_HEADERS, 1):
            ws.cell(row=row, column=col, value=header)
        style_header_row(ws, row, len(REGISTRY_MATCHED_HEADERS))
        row += 1
        for values in content["matched_rows"]:
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 4:
                    color_tier_cell(cell, str(value))
                style_data_cell(cell)
            row += 1

    if content["unmatched_github"]:
        row += 2
        ws.cell(
            row=row,
            column=1,
            value=f"On GitHub, NOT in Registry ({len(content['unmatched_github'])})",
        ).font = subheader_font
        row += 1
        for name in content["unmatched_github"]:
            ws.cell(row=row, column=1, value=name).border = thin_border
            row += 1

    if content["unmatched_registry"]:
        row += 1
        ws.cell(
            row=row,
            column=1,
            value=f"In Registry, NOT on GitHub ({len(content['unmatched_registry'])})",
        ).font = subheader_font
        row += 1
        for name in content["unmatched_registry"]:
            ws.cell(row=row, column=1, value=name).border = thin_border
            row += 1

    return row


def build_registry_sheet(
    wb,
    data: dict[str, Any],
    *,
    get_or_create_sheet,
    build_registry_content_fn,
    write_registry_sections_fn,
    section_font,
    subheader_font,
    thin_border,
    style_header_row,
    style_data_cell,
    color_tier_cell,
    auto_width,
) -> None:
    recon = data.get("reconciliation")
    if not recon:
        return

    ws = get_or_create_sheet(wb, "Registry")
    ws.sheet_properties.tabColor = "00695C"
    row = write_registry_sections_fn(
        ws,
        build_registry_content_fn(recon),
        section_font=section_font,
        subheader_font=subheader_font,
        thin_border=thin_border,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        color_tier_cell=color_tier_cell,
    )

    auto_width(ws, 5, row)


def write_trends_sheet(
    ws,
    trend_data: list[dict[str, Any]],
    *,
    tier_order: list[str],
    section_font,
    subtitle_font,
    subheader_font,
    thin_border,
) -> int:
    ws.freeze_panes = "B4"
    ws.merge_cells("A1:F1")
    ws["A1"].value = "Portfolio Trends"
    ws["A1"].font = section_font

    if len(trend_data) < 2:
        ws.cell(row=3, column=1, value="Run more audits to see trends (need 2+ historical runs)")
        ws.cell(row=3, column=1).font = subtitle_font
        return 3

    ws.cell(row=3, column=1, value="Date").font = subheader_font
    ws.cell(row=4, column=1, value="Avg Score").font = subheader_font
    ws.cell(row=5, column=1, value="Repos").font = subheader_font
    for col, run in enumerate(trend_data, 2):
        ws.cell(row=3, column=col, value=run["date"]).border = thin_border
        ws.cell(row=4, column=col, value=run["average_score"]).border = thin_border
        ws.cell(row=5, column=col, value=run["repos_audited"]).border = thin_border

    row = 7
    ws.cell(row=row, column=1, value="Tier").font = subheader_font
    for offset, tier in enumerate(tier_order, start=1):
        ws.cell(row=row + offset, column=1, value=tier.capitalize())
        for col, run in enumerate(trend_data, 2):
            ws.cell(
                row=row + offset,
                column=col,
                value=run.get("tier_distribution", {}).get(tier, 0),
            )
    return row + len(tier_order)


def build_tier_breakdown_sections(
    audits: list[dict[str, Any]],
    *,
    tier_order: list[str],
) -> list[dict[str, Any]]:
    audits_by_tier: dict[str, list[dict[str, Any]]] = {}
    for audit in audits:
        audits_by_tier.setdefault(audit.get("completeness_tier", ""), []).append(audit)

    sections: list[dict[str, Any]] = []
    for tier in tier_order:
        tier_audits = audits_by_tier.get(tier, [])
        if not tier_audits:
            continue
        tier_audits.sort(key=lambda audit: audit.get("overall_score", 0), reverse=True)
        rows = []
        for audit in tier_audits:
            metadata = audit["metadata"]
            rows.append(
                [
                    metadata["name"],
                    audit.get("grade", "F"),
                    round(audit.get("overall_score", 0), 2),
                    metadata.get("language") or "—",
                    len(audit.get("badges", [])),
                    (metadata.get("description") or "")[:50],
                ]
            )
        sections.append({"tier": tier, "rows": rows, "count": len(rows)})
    return sections


def write_tier_breakdown_sections(
    ws,
    sections: list[dict[str, Any]],
    *,
    header_font_factory,
    white_color,
    center_alignment,
    tier_fills: dict[str, Any],
    style_header_row,
    style_data_cell,
    color_grade_cell,
) -> int:
    current_row = 1
    for section in sections:
        tier = section["tier"]
        rows = section["rows"]
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=len(TIER_BREAKDOWN_HEADERS),
        )
        header_cell = ws.cell(
            row=current_row,
            column=1,
            value=f"{tier.upper()} ({section['count']} repos)",
        )
        header_cell.font = header_font_factory(bold=True, size=13, color=white_color)
        header_cell.fill = tier_fills.get(tier)
        header_cell.alignment = center_alignment
        current_row += 1

        for col, header in enumerate(TIER_BREAKDOWN_HEADERS, 1):
            ws.cell(row=current_row, column=col, value=header)
        style_header_row(ws, current_row, len(TIER_BREAKDOWN_HEADERS))
        current_row += 1

        for values in rows:
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                style_data_cell(cell)
                if col == 2:
                    color_grade_cell(cell, str(value))
            current_row += 1
        current_row += 2
    return current_row


def build_tier_breakdown_sheet(
    wb,
    data: dict[str, Any],
    *,
    tier_order: list[str],
    header_font_factory,
    white_color,
    center_alignment,
    tier_fills: dict[str, Any],
    get_or_create_sheet,
    configure_sheet_view,
    auto_width,
    build_tier_breakdown_sections_fn,
    write_tier_breakdown_sections_fn,
    style_header_row,
    style_data_cell,
    color_grade_cell,
) -> None:
    ws = get_or_create_sheet(wb, "Tier Breakdown")
    ws.sheet_properties.tabColor = "166534"
    configure_sheet_view(ws, zoom=105, show_grid_lines=True)
    ws.freeze_panes = "B2"

    current_row = write_tier_breakdown_sections_fn(
        ws,
        build_tier_breakdown_sections_fn(data.get("audits", []), tier_order=tier_order),
        header_font_factory=header_font_factory,
        white_color=white_color,
        center_alignment=center_alignment,
        tier_fills=tier_fills,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        color_grade_cell=color_grade_cell,
    )

    auto_width(ws, len(TIER_BREAKDOWN_HEADERS), current_row)


def build_repo_profile_matrix(audits: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[list[Any]]]:
    ranked_audits = sorted(
        audits,
        key=lambda audit: audit.get("overall_score", 0),
        reverse=True,
    )[:20]
    rows: list[list[Any]] = []
    for label, dimension in zip(REPO_PROFILE_RADAR_LABELS, REPO_PROFILE_RADAR_DIMS, strict=True):
        row = [label]
        for audit in ranked_audits:
            scores = {result["dimension"]: result["score"] for result in audit.get("analyzer_results", [])}
            row.append(round(scores.get(dimension, 0), 2))
        rows.append(row)
    return ranked_audits, rows


def write_repo_profile_matrix(ws, audits: list[dict[str, Any]], matrix_rows: list[list[Any]]) -> None:
    for col_idx, audit in enumerate(audits, start=2):
        ws.cell(row=1, column=col_idx, value=audit["metadata"]["name"])
    for row_idx, row_values in enumerate(matrix_rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def add_repo_profile_charts(ws, audits: list[dict[str, Any]]) -> None:
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(REPO_PROFILE_RADAR_DIMS) + 1)
    chart_row = len(REPO_PROFILE_RADAR_DIMS) + 4

    for batch_start in range(0, min(len(audits), 20), 4):
        batch_end = min(batch_start + 4, len(audits))
        chart = RadarChart()
        chart.type = "filled"
        chart.style = 26
        chart.title = f"Repos {batch_start + 1}-{batch_end}"
        chart.y_axis.delete = True

        chart_data = Reference(
            ws,
            min_col=batch_start + 2,
            max_col=batch_end + 1,
            min_row=1,
            max_row=len(REPO_PROFILE_RADAR_DIMS) + 1,
        )
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(labels)
        chart.width = 18
        chart.height = 14

        col_letter = "A" if (batch_start // 4) % 2 == 0 else "K"
        ws.add_chart(chart, f"{col_letter}{chart_row}")
        if (batch_start // 4) % 2 == 1:
            chart_row += 18


def build_risk_summary_rows(risk_posture: dict[str, int]) -> list[tuple[str, int]]:
    return [
        ("elevated", risk_posture.get("elevated", 0)),
        ("moderate", risk_posture.get("moderate", 0)),
        ("baseline", risk_posture.get("baseline", 0)),
        ("deferred", risk_posture.get("deferred", 0)),
    ]


def write_risk_summary_table(
    ws,
    rows: list[tuple[str, int]],
    *,
    start_row: int,
    style_header_row,
    add_table,
) -> int:
    for col, header in enumerate(RISK_SUMMARY_HEADERS, 1):
        ws.cell(row=start_row, column=col, value=header)
    style_header_row(ws, start_row, len(RISK_SUMMARY_HEADERS))

    for offset, (tier, count) in enumerate(rows, start=1):
        ws.cell(row=start_row + offset, column=1, value=tier)
        ws.cell(row=start_row + offset, column=2, value=count)

    max_row = start_row + len(rows)
    add_table(ws, "tblRiskSummary", len(RISK_SUMMARY_HEADERS), max_row, start_row=start_row)
    return max_row
