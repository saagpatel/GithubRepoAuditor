from openpyxl.styles import PatternFill

from src.excel_styles import (
    CENTER,
    SECTION_FONT,
    SUBHEADER_FILL,
    SUBHEADER_FONT,
    SUBTITLE_FONT,
    THIN_BORDER,
    TITLE_FONT,
    WRAP,
    apply_zebra_stripes,
    style_data_cell,
)


def set_sheet_header(ws, title: str, subtitle: str, *, width: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = WRAP


def write_instruction_banner(ws, row: int, end_col: int, message: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=1, value=message)
    cell.font = SUBTITLE_FONT
    cell.alignment = WRAP
    cell.fill = PatternFill(fill_type="solid", fgColor="E0F2FE")


def write_key_value_block(
    ws, start_row: int, start_col: int, rows: list[tuple[str, object]], *, title: str | None = None
) -> int:
    row = start_row
    if title:
        ws.cell(row=row, column=start_col, value=title).font = SECTION_FONT
        row += 1
    for label, value in rows:
        ws.cell(row=row, column=start_col, value=label).font = SUBHEADER_FONT
        style_data_cell(ws.cell(row=row, column=start_col + 1, value=value), "left")
        row += 1
    return row - 1


def write_ranked_list(
    ws,
    start_row: int,
    start_col: int,
    title: str,
    headers: list[str],
    rows: list[list[object]],
) -> int:
    ws.cell(row=start_row, column=start_col, value=title).font = SECTION_FONT
    header_row = start_row + 1
    for offset, header in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + offset, value=header)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 24
    for row_index, values in enumerate(rows, header_row + 1):
        for col_offset, value in enumerate(values):
            align = "center" if isinstance(value, (int, float)) else "left"
            style_data_cell(
                ws.cell(row=row_index, column=start_col + col_offset, value=value), align
            )
    if rows:
        apply_zebra_stripes(
            ws, header_row + 1, header_row + len(rows), start_col + len(headers) - 1
        )
    return header_row + len(rows)
