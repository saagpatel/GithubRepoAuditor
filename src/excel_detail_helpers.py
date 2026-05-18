from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink

from src.excel_styles import (
    TEAL,
    apply_zebra_stripes,
    auto_width,
    style_data_cell,
    style_header_row,
)
from src.excel_workbook_helpers import add_table, sheet_location

REVIEW_QUEUE_LANE_FILLS = {
    "blocked": "FEE2E2",
    "urgent": "FEF3C7",
    "ready": "DBEAFE",
    "deferred": "DCFCE7",
}


def write_hidden_table_sheet(
    wb,
    title: str,
    table_name: str,
    headers: list[str],
    rows: list[list[object]],
) -> None:
    ws = wb.create_sheet(title)
    ws.sheet_state = "hidden"
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    for row_index, row_values in enumerate(rows, 2):
        for col_index, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            style_data_cell(cell, "center" if isinstance(value, (int, float)) else "left")

    max_row = len(rows) + 1
    if max_row > 1:
        add_table(ws, table_name, len(headers), max_row)
        apply_zebra_stripes(ws, 2, max_row, len(headers))
    auto_width(ws, len(headers), max_row)


def set_internal_hyperlink(
    cell, sheet_name: str, *, target_cell: str = "A1", display: str | None = None
) -> None:
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        location=sheet_location(sheet_name, target_cell),
        display=display or str(cell.value or ""),
    )
    cell.font = Font("Calibri", 10, bold=True, color=TEAL, underline="single")


def repo_detail_lookup_formula(
    column_index: int, fallback: str, *, allow_blank: bool = False
) -> str:
    escaped_fallback = fallback.replace('"', '""')
    lookup = f"VLOOKUP($B$4,Data_RepoDetail!$A:$BZ,{column_index},FALSE)"
    if allow_blank:
        return f'=IFERROR({lookup},"{escaped_fallback}")'
    return f'=IFERROR(IF({lookup}="","{escaped_fallback}",{lookup}),"{escaped_fallback}")'


def repo_detail_last_movement(scores: list[float]) -> str:
    if len(scores) < 2:
        return "No prior comparison yet."
    delta = round(scores[-1] - scores[-2], 3)
    if abs(delta) < 0.005:
        return "Holding flat versus the last run."
    if delta > 0:
        return f"Up {delta:.3f} versus the last run."
    return f"Down {abs(delta):.3f} versus the last run."


def lane_fill_hex(lane: str | None) -> str | None:
    if not lane:
        return None
    return REVIEW_QUEUE_LANE_FILLS.get(str(lane).strip().lower())


def apply_lane_row_fill(ws, row: int, max_col: int, lane: str | None) -> None:
    fill_hex = lane_fill_hex(lane)
    if not fill_hex:
        return
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor=fill_hex)
