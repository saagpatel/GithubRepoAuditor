"""Helpers for Executive Summary workbook layout."""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook

from src.excel_operator_closure_helpers import (
    operator_calibration_values,
    operator_class_memory_values,
    operator_class_momentum_values,
    operator_class_normalization_values,
    operator_class_reweight_values,
    operator_class_transition_values,
    operator_transition_closure_values,
)
from src.excel_operator_follow_through_helpers import (
    operator_follow_through_details,
    operator_follow_through_freshness_details,
    operator_follow_through_reacquisition_retirement_details,
    operator_follow_through_rebuild_details,
    operator_follow_through_recovery_details,
    operator_follow_through_revalidation_recovery_details,
)
from src.excel_operator_queue_helpers import operator_focus_snapshot
from src.excel_operator_summary_helpers import (
    operator_accountability_values,
    operator_confidence_values,
    operator_decision_memory_values,
    operator_exception_values,
    operator_handoff_values,
    operator_learning_values,
    operator_retirement_values,
    operator_trend_values,
    operator_trust_values,
    operator_watch_values,
)

EXECUTIVE_SCENARIO_START_ROW = 22
EXECUTIVE_OPERATOR_START_ROW = 22


def build_executive_operator_context(
    data: dict[str, Any], weekly_pack: dict[str, Any]
) -> dict[str, Any]:
    operator_focus, operator_focus_summary, operator_focus_line = operator_focus_snapshot(
        weekly_pack
    )
    next_mode, watch_strategy, watch_decision = operator_watch_values(data)
    what_changed, why_it_matters, next_action = operator_handoff_values(data)
    (
        follow_through,
        follow_through_checkpoint,
        follow_through_escalation,
        follow_through_hotspot,
        follow_through_escalation_hotspot,
    ) = operator_follow_through_details(data)
    (
        follow_through_recovery,
        follow_through_recovery_persistence,
        follow_through_relapse_churn,
        follow_through_relapsing_hotspot,
        follow_through_retiring_hotspot,
        follow_through_churn_hotspot,
    ) = operator_follow_through_recovery_details(data)
    (
        follow_through_recovery_freshness,
        follow_through_recovery_memory_reset,
        follow_through_recovery_freshness_hotspot,
        follow_through_recovery_freshness_hotspot_summary,
        follow_through_recovery_rebuild_hotspot,
    ) = operator_follow_through_freshness_details(data)
    (
        follow_through_rebuild_strength,
        follow_through_reacquisition,
        follow_through_reacquisition_durability,
        follow_through_reacquisition_confidence,
        follow_through_rebuild_strength_hotspot,
        follow_through_reacquiring_hotspot,
        follow_through_reacquired_hotspot,
        follow_through_fragile_reacquisition_hotspot,
        follow_through_just_reacquired_hotspot,
        follow_through_holding_reacquired_hotspot,
        follow_through_durable_reacquired_hotspot,
        follow_through_softening_reacquired_hotspot,
        follow_through_fragile_reacquisition_confidence_hotspot,
    ) = operator_follow_through_rebuild_details(data)
    (
        follow_through_reacquisition_softening_decay,
        follow_through_reacquisition_confidence_retirement,
        follow_through_reacquisition_softening_hotspot,
        follow_through_reacquisition_revalidation_hotspot,
        follow_through_reacquisition_retired_confidence_hotspot,
    ) = operator_follow_through_reacquisition_retirement_details(data)
    (
        follow_through_revalidation_recovery,
        follow_through_under_revalidation_hotspot,
        follow_through_rebuilding_restored_confidence_hotspot,
        follow_through_reearning_confidence_hotspot,
        follow_through_just_reearned_confidence_hotspot,
        follow_through_holding_reearned_confidence_hotspot,
    ) = operator_follow_through_revalidation_recovery_details(data)
    trend_status, trend_summary, primary_target, resolution_counts = operator_trend_values(data)
    primary_target_reason, closure_guidance, aging_pressure = operator_accountability_values(data)
    last_intervention, last_outcome, resolution_evidence, recovery_counts = (
        operator_decision_memory_values(data)
    )
    primary_confidence, confidence_reason, next_action_confidence, recommendation_quality = (
        operator_confidence_values(data)
    )
    trust_policy, trust_policy_reason, adaptive_confidence_summary = operator_trust_values(data)
    exception_status, exception_reason, drift_status, drift_summary = operator_exception_values(
        data
    )
    (
        trust_recovery_status,
        trust_recovery_reason,
        exception_pattern_status,
        exception_pattern_summary,
    ) = operator_learning_values(data)
    recovery_confidence, retirement_status, retirement_reason, retirement_summary = (
        operator_retirement_values(data)
    )
    (
        policy_debt_status,
        policy_debt_reason,
        class_normalization_status,
        trust_normalization_summary,
    ) = operator_class_normalization_values(data)
    class_memory_status, class_memory_reason, class_decay_status, class_decay_summary = (
        operator_class_memory_values(data)
    )
    (
        class_reweight_direction,
        class_reweight_score,
        class_reweight_reason,
        class_reweight_summary,
    ) = operator_class_reweight_values(data)
    class_momentum_status, class_reweight_stability, class_momentum_summary = (
        operator_class_momentum_values(data)
    )
    class_transition_health, class_transition_resolution, class_transition_summary = (
        operator_class_transition_values(data)
    )
    (
        transition_closure_confidence,
        transition_likely_outcome,
        pending_debt_freshness,
        closure_forecast_direction,
        reset_reentry_rebuild_reentry_restore_rererestore_refresh_recovery,
        reset_reentry_rebuild_reentry_restore_rerererestore,
        reset_reentry_rebuild_reentry_restore_rerererestore_persistence,
        reset_reentry_rebuild_reentry_restore_rerererestore_churn,
        transition_closure_summary,
    ) = operator_transition_closure_values(data)
    calibration_status, calibration_summary, high_hit_rate, reopened_recommendations = (
        operator_calibration_values(data)
    )
    return locals()


def build_executive_scenario_rows(preview: dict[str, Any]) -> list[tuple[str, Any]]:
    return [
        ("Projected Avg Score Delta", preview.get("projected_average_score_delta", 0.0)),
        ("Projected Promotions", preview.get("projected_tier_promotions", 0)),
    ]


def build_executive_preflight_rows(preflight: dict[str, Any]) -> list[tuple[str, Any]]:
    return [
        ("Status", preflight.get("status", "unknown")),
        ("Errors", preflight.get("blocking_errors", 0)),
        ("Warnings", preflight.get("warnings", 0)),
    ]


def build_executive_kpi_cards(
    *,
    data: dict[str, Any],
    critical_repos: list[Any],
    run_change_counts: dict[str, Any],
) -> list[tuple[str, Any, str | None]]:
    return [
        ("Portfolio Grade", data.get("portfolio_grade", "F"), None),
        ("Avg Score", f"{data.get('average_score', 0):.2f}", None),
        ("Critical Repos", len(critical_repos), None),
        ("Improving", run_change_counts.get("score_improvements", 0), None),
        ("Regressing", run_change_counts.get("score_regressions", 0), "C2410C"),
    ]


def build_executive_story_inputs(
    *,
    data: dict[str, Any],
    diff_data: dict[str, Any] | None,
    weekly_pack: dict[str, Any],
    leaders: list[dict[str, Any]],
    build_run_change_counts,
    build_run_change_summary,
    build_queue_pressure_summary,
    build_trust_actionability_summary,
    build_top_recommendation_summary,
    resolve_weekly_story_value,
    build_workbook_rollups,
) -> dict[str, Any]:
    recommended_focus = ""
    if data.get("operator_queue"):
        recommended_focus = data["operator_queue"][0].get("recommended_action", "")
    elif leaders:
        recommended_focus = f"Protect momentum around {leaders[0]['name']}"

    if diff_data:
        change_summary = (
            f"Average score moved {diff_data.get('average_score_delta', 0.0):+.3f} across "
            f"{len(diff_data.get('repo_changes', []) or [])} repos with notable changes."
        )
    else:
        change_summary = (
            f"{len(data.get('material_changes', []) or [])} material changes and "
            f"{len(data.get('governance_drift', []) or [])} governance drift signals were captured in this run."
        )

    run_change_counts = data.get("run_change_counts") or build_run_change_counts(diff_data)
    run_change_summary = data.get("run_change_summary") or build_run_change_summary(diff_data)
    queue_pressure_summary = resolve_weekly_story_value(
        weekly_pack,
        "why_this_week",
        weekly_pack.get("queue_pressure_summary"),
        build_queue_pressure_summary(data, diff_data),
    )
    trust_actionability_summary = build_trust_actionability_summary(data)
    top_attention = build_workbook_rollups(data)[1][:5]
    top_recommendation = resolve_weekly_story_value(
        weekly_pack,
        "decision",
        weekly_pack.get("what_to_do_this_week"),
        build_top_recommendation_summary(data),
        recommended_focus,
        "Start with the highest-pressure queue item, then protect the current leaders.",
    )
    biggest_opportunity = leaders[0]["name"] if leaders else "Portfolio-wide follow-through"

    return {
        "recommended_focus": recommended_focus,
        "change_summary": change_summary,
        "run_change_counts": run_change_counts,
        "run_change_summary": run_change_summary,
        "queue_pressure_summary": queue_pressure_summary,
        "trust_actionability_summary": trust_actionability_summary,
        "top_attention": top_attention,
        "top_recommendation": top_recommendation,
        "biggest_opportunity": biggest_opportunity,
    }


def build_executive_ranked_lists(
    *,
    leaders: list[dict[str, Any]],
    top_attention: list[tuple[Any, ...]],
    diff_data: dict[str, Any] | None,
) -> tuple[list[list[Any]], list[list[Any]], list[list[Any]] | None]:
    leader_rows = [[entry["name"], entry["profile_score"], entry["tier"]] for entry in leaders] or [
        ["No leader", 0, "—"]
    ]
    attention_rows = [
        [
            repo,
            f"B{blocked}/U{urgent}/R{ready}",
            title or "Queue pressure",
            action or "Review the repo detail page.",
        ]
        for repo, _total, blocked, urgent, ready, _deferred, _kind, _priority, title, action in top_attention
    ] or [["Portfolio", "No open items", "Queue is clear", "Monitor future runs."]]
    mover_rows = None
    if diff_data:
        mover_rows = [
            [
                change.get("name", ""),
                round(change.get("delta", 0.0), 3),
                f"{change.get('old_tier', '—')} -> {change.get('new_tier', '—')}",
            ]
            for change in (diff_data.get("repo_changes", []) or [])[:5]
        ] or [["No movers", 0.0, "—"]]
    return leader_rows, attention_rows, mover_rows


def build_executive_narrative_rows(
    *,
    data: dict[str, Any],
    leaders: list[dict[str, Any]],
    critical_repos: list[Any],
    operator_summary: dict[str, Any],
    operator_context: dict[str, Any],
    story_inputs: dict[str, Any],
    excel_mode: str,
    build_portfolio_catalog_summary,
    build_operating_paths_summary,
    build_portfolio_intent_alignment_summary,
    build_scorecards_summary,
) -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = [
        (
            "Portfolio Health",
            f"{data.get('tier_distribution', {}).get('shipped', 0)} repos are shipped, the portfolio average is {data.get('average_score', 0):.2f}, and {leaders[0]['name'] if leaders else 'the current leaders'} set the current high-water mark.",
        ),
        (
            "Top Attention",
            operator_summary.get("headline", "No urgent operator headline is present.")
            + (
                f" Critical security pressure is concentrated in {len(critical_repos)} repos."
                if critical_repos
                else ""
            ),
        ),
        ("Run Changes", story_inputs["run_change_summary"] or story_inputs["change_summary"]),
        ("Queue Pressure", story_inputs["queue_pressure_summary"]),
        ("What Changed", operator_context["what_changed"] or story_inputs["change_summary"]),
        ("Why It Matters", operator_context["why_it_matters"]),
        ("Top Recommendation", story_inputs["top_recommendation"]),
        ("Portfolio Catalog", build_portfolio_catalog_summary(data)),
        ("Operating Paths", build_operating_paths_summary(data)),
        ("Intent Alignment", build_portfolio_intent_alignment_summary(data)),
        ("Scorecards", build_scorecards_summary(data)),
        ("Follow-Through", operator_context["follow_through"]),
        ("Next Checkpoint", operator_context["follow_through_checkpoint"]),
        ("Escalation", operator_context["follow_through_escalation"]),
        ("Recovery / Retirement", operator_context["follow_through_recovery"]),
        ("Recovery Persistence", operator_context["follow_through_recovery_persistence"]),
        ("Relapse Churn", operator_context["follow_through_relapse_churn"]),
        ("Recovery Freshness", operator_context["follow_through_recovery_freshness"]),
        ("Recovery Memory Reset", operator_context["follow_through_recovery_memory_reset"]),
        ("Recovery Rebuild Strength", operator_context["follow_through_rebuild_strength"]),
        ("Recovery Reacquisition", operator_context["follow_through_reacquisition"]),
        ("Reacquisition Durability", operator_context["follow_through_reacquisition_durability"]),
        ("Reacquisition Confidence", operator_context["follow_through_reacquisition_confidence"]),
        ("Operator Focus", operator_context["operator_focus"]),
        ("Focus Summary", operator_context["operator_focus_summary"]),
        ("Focus Line", operator_context["operator_focus_line"]),
        ("Trust Summary", story_inputs["trust_actionability_summary"]),
        ("Biggest Opportunity", story_inputs["biggest_opportunity"]),
        (
            "Focus This Week",
            operator_context["next_action"]
            or story_inputs["recommended_focus"]
            or "Review the top queue items first, then protect the highest-value repos from drift.",
        ),
    ]
    if excel_mode == "standard":
        rows[5:5] = [
            ("Trend", f"{operator_context['trend_status']} — {operator_context['trend_summary']}"),
            ("Why Top Target", operator_context["primary_target_reason"]),
            ("Follow-Through Hotspot", operator_context["follow_through_hotspot"]),
            ("Escalation Hotspot", operator_context["follow_through_escalation_hotspot"]),
            ("Recovery Hotspot", operator_context["follow_through_relapsing_hotspot"]),
            ("Retiring Watch Hotspot", operator_context["follow_through_retiring_hotspot"]),
            ("Churn Hotspot", operator_context["follow_through_churn_hotspot"]),
            ("Freshness Hotspot", operator_context["follow_through_recovery_freshness_hotspot"]),
            ("Freshness Detail", operator_context["follow_through_recovery_freshness_hotspot_summary"]),
            ("Rebuild Hotspot", operator_context["follow_through_recovery_rebuild_hotspot"]),
            ("Rebuild Strength Hotspot", operator_context["follow_through_rebuild_strength_hotspot"]),
            ("Reacquiring Hotspot", operator_context["follow_through_reacquiring_hotspot"]),
            ("Reacquired Hotspot", operator_context["follow_through_reacquired_hotspot"]),
            ("Fragile Reacquisition Hotspot", operator_context["follow_through_fragile_reacquisition_hotspot"]),
            ("Just Reacquired Hotspot", operator_context["follow_through_just_reacquired_hotspot"]),
            ("Holding Reacquired Hotspot", operator_context["follow_through_holding_reacquired_hotspot"]),
            ("Durable Reacquired Hotspot", operator_context["follow_through_durable_reacquired_hotspot"]),
            ("Softening Reacquired Hotspot", operator_context["follow_through_softening_reacquired_hotspot"]),
            (
                "Fragile Reacquisition Confidence Hotspot",
                operator_context["follow_through_fragile_reacquisition_confidence_hotspot"],
            ),
            ("Reacquisition Softening Hotspot", operator_context["follow_through_reacquisition_softening_hotspot"]),
            ("Revalidation Needed Hotspot", operator_context["follow_through_reacquisition_revalidation_hotspot"]),
            (
                "Retired Confidence Hotspot",
                operator_context["follow_through_reacquisition_retired_confidence_hotspot"],
            ),
            ("Under Revalidation Hotspot", operator_context["follow_through_under_revalidation_hotspot"]),
            (
                "Rebuilding Restored Confidence Hotspot",
                operator_context["follow_through_rebuilding_restored_confidence_hotspot"],
            ),
            ("Re-Earning Confidence Hotspot", operator_context["follow_through_reearning_confidence_hotspot"]),
            ("Just Re-Earned Confidence Hotspot", operator_context["follow_through_just_reearned_confidence_hotspot"]),
            (
                "Holding Re-Earned Confidence Hotspot",
                operator_context["follow_through_holding_reearned_confidence_hotspot"],
            ),
            ("Closure Guidance", operator_context["closure_guidance"]),
            ("What We Tried", operator_context["last_intervention"]),
            ("Recommendation Confidence", operator_context["primary_confidence"]),
            ("Resolution Evidence", operator_context["resolution_evidence"]),
            ("Confidence Rationale", operator_context["confidence_reason"]),
            ("Trust Policy", operator_context["trust_policy"]),
            ("Trust Rationale", operator_context["trust_policy_reason"]),
            ("Trust Exception", f"{operator_context['exception_status']} — {operator_context['exception_reason']}"),
            ("Trust Recovery", f"{operator_context['trust_recovery_status']} — {operator_context['trust_recovery_reason']}"),
            ("Recovery Confidence", operator_context["recovery_confidence"]),
            ("Exception Retirement", f"{operator_context['retirement_status']} — {operator_context['retirement_reason']}"),
            ("Retirement Summary", operator_context["retirement_summary"]),
            ("Policy Debt", f"{operator_context['policy_debt_status']} — {operator_context['policy_debt_reason']}"),
            ("Class Normalization", f"{operator_context['class_normalization_status']} — {operator_context['trust_normalization_summary']}"),
            ("Class Memory", f"{operator_context['class_memory_status']} — {operator_context['class_memory_reason']}"),
            ("Trust Decay", f"{operator_context['class_decay_status']} — {operator_context['class_decay_summary']}"),
            (
                "Class Reweighting",
                f"{operator_context['class_reweight_direction']} ({operator_context['class_reweight_score']}) — {operator_context['class_reweight_summary']}",
            ),
            ("Class Reweighting Why", operator_context["class_reweight_reason"]),
            ("Class Momentum", operator_context["class_momentum_status"]),
            ("Reweight Stability", operator_context["class_reweight_stability"]),
            ("Transition Health", operator_context["class_transition_health"]),
            ("Transition Summary", operator_context["class_transition_summary"]),
            ("Transition Closure", operator_context["transition_closure_confidence"]),
            ("Transition Resolution", operator_context["class_transition_resolution"]),
            ("Transition Likely Outcome", operator_context["transition_likely_outcome"]),
            ("Pending Debt Freshness", operator_context["pending_debt_freshness"]),
            ("Closure Forecast", operator_context["closure_forecast_direction"]),
            (
                "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Persistence",
                operator_context["reset_reentry_rebuild_reentry_restore_rerererestore_persistence"],
            ),
            (
                "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Churn Controls",
                operator_context["reset_reentry_rebuild_reentry_restore_rerererestore_churn"],
            ),
            ("Closure Forecast Summary", operator_context["transition_closure_summary"]),
            ("Momentum Summary", operator_context["class_momentum_summary"]),
            ("Exception Learning", f"{operator_context['exception_pattern_status']} — {operator_context['exception_pattern_summary']}"),
            ("Recommendation Drift", f"{operator_context['drift_status']} — {operator_context['drift_summary']}"),
            ("Adaptive Confidence", operator_context["adaptive_confidence_summary"]),
            ("Confidence Validation", f"{operator_context['calibration_status']} — {operator_context['calibration_summary']}"),
        ]
    return rows


def build_executive_operator_rows(
    *,
    operator_summary: dict[str, Any],
    operator_context: dict[str, Any],
    queue_counts: str,
    fallback_action: str,
    excel_mode: str,
) -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = [
        ("Headline", operator_summary.get("headline", "")),
        ("Queue Counts", queue_counts),
        ("Source Run", operator_summary.get("source_run_id", "")),
        ("Next Run", f"{operator_context['next_mode']} via {operator_context['watch_strategy']}"),
        ("Watch Decision", operator_context["watch_decision"]),
        ("What To Do Next", operator_context["next_action"] or fallback_action),
    ]
    if excel_mode == "standard":
        rows.extend(
            [
                ("Trend", f"{operator_context['trend_status']} — {operator_context['trend_summary']}"),
                ("Primary Target", operator_context["primary_target"]),
                ("Resolution Counts", operator_context["resolution_counts"]),
                ("Why Top Target", operator_context["primary_target_reason"]),
                ("Closure Guidance", operator_context["closure_guidance"]),
                ("Aging Pressure", operator_context["aging_pressure"]),
                ("What We Tried", operator_context["last_intervention"]),
                ("Last Outcome", operator_context["last_outcome"]),
                ("Resolution Evidence", operator_context["resolution_evidence"]),
                ("Recovery Counts", operator_context["recovery_counts"]),
                ("Recommendation Confidence", operator_context["primary_confidence"]),
                ("Confidence Rationale", operator_context["confidence_reason"]),
                ("Next Action Confidence", operator_context["next_action_confidence"]),
                ("Trust Policy", operator_context["trust_policy"]),
                ("Trust Rationale", operator_context["trust_policy_reason"]),
                ("Trust Recovery", f"{operator_context['trust_recovery_status']} — {operator_context['trust_recovery_reason']}"),
                ("Recovery Confidence", operator_context["recovery_confidence"]),
                ("Recovery Persistence", operator_context["follow_through_recovery_persistence"]),
                ("Relapse Churn", operator_context["follow_through_relapse_churn"]),
                ("Recovery Freshness", operator_context["follow_through_recovery_freshness"]),
                ("Recovery Memory Reset", operator_context["follow_through_recovery_memory_reset"]),
                ("Recovery Rebuild Strength", operator_context["follow_through_rebuild_strength"]),
                ("Recovery Reacquisition", operator_context["follow_through_reacquisition"]),
                ("Recovery Hotspot", operator_context["follow_through_relapsing_hotspot"]),
                ("Retiring Watch Hotspot", operator_context["follow_through_retiring_hotspot"]),
                ("Churn Hotspot", operator_context["follow_through_churn_hotspot"]),
                ("Freshness Hotspot", operator_context["follow_through_recovery_freshness_hotspot"]),
                ("Freshness Detail", operator_context["follow_through_recovery_freshness_hotspot_summary"]),
                ("Rebuild Hotspot", operator_context["follow_through_recovery_rebuild_hotspot"]),
                ("Rebuild Strength Hotspot", operator_context["follow_through_rebuild_strength_hotspot"]),
                ("Reacquiring Hotspot", operator_context["follow_through_reacquiring_hotspot"]),
                ("Reacquired Hotspot", operator_context["follow_through_reacquired_hotspot"]),
                ("Fragile Reacquisition Hotspot", operator_context["follow_through_fragile_reacquisition_hotspot"]),
                ("Exception Retirement", f"{operator_context['retirement_status']} — {operator_context['retirement_reason']}"),
                ("Retirement Summary", operator_context["retirement_summary"]),
                ("Policy Debt", f"{operator_context['policy_debt_status']} — {operator_context['policy_debt_reason']}"),
                ("Class Normalization", f"{operator_context['class_normalization_status']} — {operator_context['trust_normalization_summary']}"),
                ("Class Memory", f"{operator_context['class_memory_status']} — {operator_context['class_memory_reason']}"),
                ("Trust Decay", f"{operator_context['class_decay_status']} — {operator_context['class_decay_summary']}"),
                (
                    "Class Reweighting",
                    f"{operator_context['class_reweight_direction']} ({operator_context['class_reweight_score']}) — {operator_context['class_reweight_summary']}",
                ),
                ("Class Reweighting Why", operator_context["class_reweight_reason"]),
                ("Class Momentum", operator_context["class_momentum_status"]),
                ("Reweight Stability", operator_context["class_reweight_stability"]),
                ("Transition Health", operator_context["class_transition_health"]),
                ("Transition Resolution", operator_context["class_transition_resolution"]),
                ("Transition Summary", operator_context["class_transition_summary"]),
                ("Transition Closure", operator_context["transition_closure_confidence"]),
                ("Transition Likely Outcome", operator_context["transition_likely_outcome"]),
                ("Pending Debt Freshness", operator_context["pending_debt_freshness"]),
                ("Closure Forecast", operator_context["closure_forecast_direction"]),
                (
                    "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Persistence",
                    operator_context["reset_reentry_rebuild_reentry_restore_rerererestore_persistence"],
                ),
                (
                    "Reset Re-entry Rebuild Re-Entry Restore Re-Re-Re-Restore Churn Controls",
                    operator_context["reset_reentry_rebuild_reentry_restore_rerererestore_churn"],
                ),
                ("Closure Forecast Summary", operator_context["transition_closure_summary"]),
                ("Momentum Summary", operator_context["class_momentum_summary"]),
                ("Exception Learning", f"{operator_context['exception_pattern_status']} — {operator_context['exception_pattern_summary']}"),
                ("Recommendation Drift", f"{operator_context['drift_status']} — {operator_context['drift_summary']}"),
                ("Adaptive Confidence", operator_context["adaptive_confidence_summary"]),
                ("Recommendation Quality", operator_context["recommendation_quality"]),
                ("Confidence Validation", f"{operator_context['calibration_status']} — {operator_context['calibration_summary']}"),
                (
                    "Calibration Snapshot",
                    f"High-confidence hit rate {operator_context['high_hit_rate']} | {operator_context['reopened_recommendations']}",
                ),
            ]
        )
    return rows


def build_executive_sheet_content(
    *,
    data: dict[str, Any],
    diff_data: dict[str, Any] | None,
    context: dict[str, Any],
    excel_mode: str,
    build_weekly_review_pack,
    build_run_change_counts,
    build_run_change_summary,
    build_queue_pressure_summary,
    build_trust_actionability_summary,
    build_top_recommendation_summary,
    resolve_weekly_story_value,
    build_workbook_rollups,
    format_lane_counts,
    operator_counts,
    build_portfolio_catalog_summary,
    build_operating_paths_summary,
    build_portfolio_intent_alignment_summary,
    build_scorecards_summary,
) -> dict[str, Any]:
    leaders = context["profile_leaderboard"].get("leaders", [])[:5]
    critical_repos = data.get("security_posture", {}).get("critical_repos", []) or []
    operator_summary = data.get("operator_summary") or {}
    weekly_pack = build_weekly_review_pack(data, diff_data)
    operator_context = build_executive_operator_context(data, weekly_pack)
    preview = context["scenario_preview"].get("portfolio_projection", {})
    story_inputs = build_executive_story_inputs(
        data=data,
        diff_data=diff_data,
        weekly_pack=weekly_pack,
        leaders=leaders,
        build_run_change_counts=build_run_change_counts,
        build_run_change_summary=build_run_change_summary,
        build_queue_pressure_summary=build_queue_pressure_summary,
        build_trust_actionability_summary=build_trust_actionability_summary,
        build_top_recommendation_summary=build_top_recommendation_summary,
        resolve_weekly_story_value=resolve_weekly_story_value,
        build_workbook_rollups=build_workbook_rollups,
    )
    fallback_action = (
        data.get("operator_queue", [{}])[0].get("recommended_action")
        if data.get("operator_queue")
        else ""
    ) or "Start with the top review queue item, then protect the current profile leaders."
    preflight = data.get("preflight_summary") or {}
    return {
        "leaders": leaders,
        "critical_repos": critical_repos,
        "operator_summary": operator_summary,
        "operator_context": operator_context,
        "story_inputs": story_inputs,
        "kpi_cards": build_executive_kpi_cards(
            data=data,
            critical_repos=critical_repos,
            run_change_counts=story_inputs["run_change_counts"],
        ),
        "narrative_rows": build_executive_narrative_rows(
            data=data,
            leaders=leaders,
            critical_repos=critical_repos,
            operator_summary=operator_summary,
            operator_context=operator_context,
            story_inputs=story_inputs,
            excel_mode=excel_mode,
            build_portfolio_catalog_summary=build_portfolio_catalog_summary,
            build_operating_paths_summary=build_operating_paths_summary,
            build_portfolio_intent_alignment_summary=build_portfolio_intent_alignment_summary,
            build_scorecards_summary=build_scorecards_summary,
        ),
        "ranked_lists": build_executive_ranked_lists(
            leaders=leaders,
            top_attention=story_inputs["top_attention"],
            diff_data=diff_data,
        ),
        "scenario_rows": build_executive_scenario_rows(preview),
        "operator_rows": build_executive_operator_rows(
            operator_summary=operator_summary,
            operator_context=operator_context,
            queue_counts=format_lane_counts(operator_counts(data)),
            fallback_action=fallback_action,
            excel_mode=excel_mode,
        ),
        "preflight_rows": build_executive_preflight_rows(preflight),
        "show_preflight": bool(
            preflight and (preflight.get("blocking_errors", 0) or preflight.get("warnings", 0))
        ),
    }


def write_executive_support_sections(
    ws,
    content: dict[str, Any],
    *,
    excel_mode: str,
    section_font,
    subheader_font,
) -> None:
    ws.cell(row=EXECUTIVE_SCENARIO_START_ROW, column=1, value="Scenario Preview").font = (
        section_font
    )
    for row_index, (label, value) in enumerate(
        content["scenario_rows"], start=EXECUTIVE_SCENARIO_START_ROW + 1
    ):
        ws.cell(row=row_index, column=1, value=label).font = subheader_font
        ws.cell(row=row_index, column=2, value=value)

    if content["operator_summary"]:
        ws.cell(row=EXECUTIVE_OPERATOR_START_ROW, column=4, value="Operator Control Center").font = (
            section_font
        )
        for row_index, (label, value) in enumerate(
            content["operator_rows"], start=EXECUTIVE_OPERATOR_START_ROW + 1
        ):
            ws.cell(row=row_index, column=4, value=label).font = subheader_font
            ws.cell(row=row_index, column=5, value=value)

    if content["show_preflight"]:
        preflight_row = 59 if excel_mode == "standard" else 33
        ws.cell(row=preflight_row, column=1, value="Preflight Diagnostics").font = section_font
        for offset, (label, value) in enumerate(content["preflight_rows"], start=1):
            ws.cell(row=preflight_row + offset, column=1, value=label).font = subheader_font
            ws.cell(row=preflight_row + offset, column=2, value=value)


def build_executive_summary_sheet(
    wb: Workbook,
    data: dict[str, Any],
    diff_data: dict[str, Any] | None,
    *,
    portfolio_profile: str,
    collection: str | None,
    excel_mode: str,
    build_analyst_context,
    build_weekly_review_pack,
    build_run_change_counts,
    build_run_change_summary,
    build_queue_pressure_summary,
    build_trust_actionability_summary,
    build_top_recommendation_summary,
    resolve_weekly_story_value,
    build_workbook_rollups,
    format_lane_counts,
    operator_counts,
    build_portfolio_catalog_summary,
    build_operating_paths_summary,
    build_portfolio_intent_alignment_summary,
    build_scorecards_summary,
    get_or_create_sheet,
    configure_sheet_view,
    set_sheet_header,
    write_instruction_banner,
    write_key_value_block,
    write_ranked_list,
    write_kpi_card,
    auto_width,
    navy: str,
    section_font,
    subheader_font,
) -> None:
    context = build_analyst_context(
        data, profile_name=portfolio_profile, collection_name=collection
    )
    content = build_executive_sheet_content(
        data=data,
        diff_data=diff_data,
        context=context,
        excel_mode=excel_mode,
        build_weekly_review_pack=build_weekly_review_pack,
        build_run_change_counts=build_run_change_counts,
        build_run_change_summary=build_run_change_summary,
        build_queue_pressure_summary=build_queue_pressure_summary,
        build_trust_actionability_summary=build_trust_actionability_summary,
        build_top_recommendation_summary=build_top_recommendation_summary,
        resolve_weekly_story_value=resolve_weekly_story_value,
        build_workbook_rollups=build_workbook_rollups,
        format_lane_counts=format_lane_counts,
        operator_counts=operator_counts,
        build_portfolio_catalog_summary=build_portfolio_catalog_summary,
        build_operating_paths_summary=build_operating_paths_summary,
        build_portfolio_intent_alignment_summary=build_portfolio_intent_alignment_summary,
        build_scorecards_summary=build_scorecards_summary,
    )
    ws = get_or_create_sheet(wb, "Executive Summary")
    ws.sheet_properties.tabColor = navy
    configure_sheet_view(ws, zoom=120, show_grid_lines=False)
    set_sheet_header(
        ws,
        "Executive Summary",
        f"Profile: {context['profile_name']} | Collection: {context['collection_name'] or 'all'}",
        width=6,
    )
    write_instruction_banner(
        ws,
        3,
        12,
        "Use the left side for the short leadership story and the right side for operator evidence you may need to defend the next move.",
    )
    ws.freeze_panes = "A4"

    write_key_value_block(ws, 4, 1, content["narrative_rows"], title="Leadership Brief")

    for offset, (label, value, accent) in enumerate(content["kpi_cards"]):
        write_kpi_card(ws, 10, 1 + offset * 2, label, value, accent)

    leader_rows, attention_rows, mover_rows = content["ranked_lists"]
    write_ranked_list(ws, 14, 1, "Top Profile Leaders", ["Repo", "Profile Score", "Tier"], leader_rows)
    write_ranked_list(
        ws,
        14,
        9,
        "Top 5 Attention Items",
        ["Repo", "Counts", "Why Now", "Next Step"],
        attention_rows,
    )

    if mover_rows:
        write_ranked_list(ws, 14, 5, "Top Movers", ["Repo", "Delta", "Tier Change"], mover_rows)

    write_executive_support_sections(
        ws,
        content,
        excel_mode=excel_mode,
        section_font=section_font,
        subheader_font=subheader_font,
    )
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:4"
    ws.print_area = "A1:L40"
    auto_width(ws, 6, 88 if excel_mode == "standard" else 35)


def build_executive_summary_workbook_sheet(
    wb: Workbook,
    data: dict[str, Any],
    diff_data: dict[str, Any] | None,
    *,
    portfolio_profile: str = "default",
    collection: str | None = None,
    excel_mode: str = "standard",
    build_weekly_review_pack,
    build_run_change_counts,
    build_run_change_summary,
    build_queue_pressure_summary,
    build_trust_actionability_summary,
    build_top_recommendation_summary,
    resolve_weekly_story_value,
    build_workbook_rollups,
    format_lane_counts,
    operator_counts,
    build_portfolio_catalog_summary,
    build_operating_paths_summary,
    build_portfolio_intent_alignment_summary,
    build_scorecards_summary,
    get_or_create_sheet,
    configure_sheet_view,
    set_sheet_header,
    write_instruction_banner,
    write_key_value_block,
    write_ranked_list,
    write_kpi_card,
    auto_width,
    navy: str,
    section_font,
    subheader_font,
) -> None:
    from src.analyst_views import build_analyst_context

    build_executive_summary_sheet(
        wb,
        data,
        diff_data,
        portfolio_profile=portfolio_profile,
        collection=collection,
        excel_mode=excel_mode,
        build_analyst_context=build_analyst_context,
        build_weekly_review_pack=build_weekly_review_pack,
        build_run_change_counts=build_run_change_counts,
        build_run_change_summary=build_run_change_summary,
        build_queue_pressure_summary=build_queue_pressure_summary,
        build_trust_actionability_summary=build_trust_actionability_summary,
        build_top_recommendation_summary=build_top_recommendation_summary,
        resolve_weekly_story_value=resolve_weekly_story_value,
        build_workbook_rollups=build_workbook_rollups,
        format_lane_counts=format_lane_counts,
        operator_counts=operator_counts,
        build_portfolio_catalog_summary=build_portfolio_catalog_summary,
        build_operating_paths_summary=build_operating_paths_summary,
        build_portfolio_intent_alignment_summary=build_portfolio_intent_alignment_summary,
        build_scorecards_summary=build_scorecards_summary,
        get_or_create_sheet=get_or_create_sheet,
        configure_sheet_view=configure_sheet_view,
        set_sheet_header=set_sheet_header,
        write_instruction_banner=write_instruction_banner,
        write_key_value_block=write_key_value_block,
        write_ranked_list=write_ranked_list,
        write_kpi_card=write_kpi_card,
        auto_width=auto_width,
        navy=navy,
        section_font=section_font,
        subheader_font=subheader_font,
    )
