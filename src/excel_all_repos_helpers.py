"""Helpers for the All Repos workbook sheet."""

from __future__ import annotations

from collections.abc import Callable
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ALL_REPOS_HEADERS = [
    "Repo",
    "Grade",
    "Score",
    "Interest",
    "Interest Grade",
    "Interest Tier",
    "Tier",
    "Badges",
    "Next Badge",
    "Language",
    "Topics",
    "Commit Pattern",
    "Bus Factor",
    "Days Since Push",
    "Commits",
    "Releases",
    "Test Files",
    "Test Framework",
    "LOC",
    "TODO Density",
    "PR Merge %",
    "Comment Ratio",
    "Dep Count",
    "Libyears",
    "Stars",
    "Private",
    "Flags",
    "Description",
    "Biggest Drag",
    "Why This Grade",
    "Tech Novelty",
    "Burst",
    "Ambition",
    "Storytelling",
    "Created",
    "Size (KB)",
    "Trend",
    "Risk Tier",
]

ALL_REPOS_LONG_TEXT_HEADERS = ("Description", "Topics", "Badges")
ALL_REPOS_GRADE_COLUMN = 2
ALL_REPOS_SCORE_COLUMN = 3
ALL_REPOS_INTEREST_COLUMN = 4
ALL_REPOS_TIER_COLUMN = 7
ALL_REPOS_PATTERN_COLUMN = 12
ALL_REPOS_TREND_COLUMN_OFFSET = 1
ALL_REPOS_LONG_TEXT_WIDTH = 60


def build_all_repo_rows(
    audits: list[dict[str, Any]],
    *,
    score_history: dict[str, list[float]] | None,
    risk_lookup: dict[str, str] | None,
    render_sparkline: Callable[[list[float]], str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    ranked_audits = sorted(audits, key=lambda audit: audit.get("overall_score", 0), reverse=True)
    for audit in ranked_audits:
        metadata = audit.get("metadata", {})
        details = {
            result["dimension"]: result.get("details", {})
            for result in audit.get("analyzer_results", [])
        }
        activity = details.get("activity", {})
        code_quality = details.get("code_quality", {})
        testing = details.get("testing", {})
        dependencies = details.get("dependencies", {})
        documentation = details.get("documentation", {})
        interest = details.get("interest", {})

        badges = audit.get("badges", [])
        next_badges = audit.get("next_badges", [])
        repo_name = metadata.get("name", "")
        trend_scores = (score_history or {}).get(repo_name, [])

        rows.append(
            {
                "repo_name": repo_name,
                "html_url": metadata.get("html_url", ""),
                "grade": audit.get("grade", "F"),
                "tier": audit.get("completeness_tier", ""),
                "commit_pattern": activity.get("commit_pattern", ""),
                "values": [
                    repo_name,
                    audit.get("grade", "F"),
                    round(audit.get("overall_score", 0), 3),
                    round(audit.get("interest_score", 0), 3),
                    audit.get("interest_grade", "—"),
                    audit.get("interest_tier", "—"),
                    audit.get("completeness_tier", ""),
                    ", ".join(badges[:4]),
                    next_badges[0]["action"][:50] if next_badges else "",
                    metadata.get("language") or "—",
                    ", ".join(metadata.get("topics", [])[:8]) or "—",
                    activity.get("commit_pattern", "—"),
                    activity.get("bus_factor", "—"),
                    activity.get("days_since_push", "—"),
                    activity.get("total_commits", "—"),
                    activity.get("release_count", "—"),
                    testing.get("test_file_count", 0),
                    testing.get("framework", "—"),
                    code_quality.get("total_loc", 0),
                    round(code_quality.get("todo_density_per_1k", 0) or 0, 1),
                    round((code_quality.get("pr_merge_ratio", 0) or 0) * 100, 0),
                    round(documentation.get("comment_ratio", 0) or 0, 2),
                    (
                        dependencies.get("dep_count", "—")
                        if dependencies.get("dep_count") is not None
                        else "—"
                    ),
                    dependencies.get("total_libyears", "—"),
                    metadata.get("stars", 0),
                    "Yes" if metadata.get("private") else "No",
                    ", ".join(audit.get("flags", [])),
                    metadata.get("description") or "",
                    _build_biggest_drag(audit),
                    _build_grade_reason(audit),
                    round(interest.get("tech_novelty", 0), 2),
                    round(interest.get("burst_coefficient", 0), 2),
                    round(interest.get("ambition_score") or 0, 2),
                    round(interest.get("readme_storytelling", 0), 2),
                    _build_created_date(metadata.get("created_at", "")),
                    metadata.get("size_kb", 0),
                    render_sparkline(trend_scores),
                    (risk_lookup or {}).get(str(repo_name), ""),
                ],
            }
        )
    return rows


def _build_biggest_drag(audit: dict[str, Any]) -> str:
    dimension_scores = {
        result["dimension"]: result["score"]
        for result in audit.get("analyzer_results", [])
        if result["dimension"] != "interest"
    }
    if not dimension_scores:
        return "—"
    worst_dimension = min(dimension_scores, key=dimension_scores.get)
    return f"{worst_dimension} ({dimension_scores[worst_dimension]:.1f})"


def _build_grade_reason(audit: dict[str, Any]) -> str:
    dimension_scores = {
        result["dimension"]: result["score"]
        for result in audit.get("analyzer_results", [])
        if result["dimension"] != "interest"
    }
    grade = audit.get("grade", "F")
    if not dimension_scores:
        return grade
    weakest_dimensions = sorted(dimension_scores.items(), key=lambda item: item[1])[:2]
    if len(weakest_dimensions) < 2:
        return grade
    return (
        f"{grade}: {weakest_dimensions[0][0]}={weakest_dimensions[0][1]:.1f}, "
        f"{weakest_dimensions[1][0]}={weakest_dimensions[1][1]:.1f}"
    )


def _build_created_date(created_at: str) -> str:
    if created_at and len(created_at) >= 10:
        return created_at[:10]
    return created_at


def write_all_repo_rows(
    ws,
    repo_rows: list[dict[str, Any]],
    *,
    headers: list[str],
    style_data_cell,
    color_grade_cell,
    color_tier_cell,
    color_pattern_cell,
    sparkline_font,
    link_color: str,
) -> int:
    for row_index, repo_row in enumerate(repo_rows, 2):
        for col_index, value in enumerate(repo_row["values"], 1):
            style_data_cell(ws.cell(row=row_index, column=col_index, value=value))

        name_cell = ws.cell(row=row_index, column=1)
        if repo_row["html_url"]:
            name_cell.hyperlink = repo_row["html_url"]
            name_cell.font = Font("Calibri", 10, color=link_color, underline="single")

        color_grade_cell(ws.cell(row=row_index, column=ALL_REPOS_GRADE_COLUMN), repo_row["grade"])
        color_tier_cell(ws.cell(row=row_index, column=ALL_REPOS_TIER_COLUMN), repo_row["tier"])
        pattern = repo_row["commit_pattern"]
        if pattern and pattern != "—":
            color_pattern_cell(ws.cell(row=row_index, column=ALL_REPOS_PATTERN_COLUMN), pattern)

        trend_cell = ws.cell(row=row_index, column=len(headers) - ALL_REPOS_TREND_COLUMN_OFFSET)
        if trend_cell.value:
            trend_cell.font = sparkline_font

    return len(repo_rows) + 1


def finalize_all_repos_layout(
    ws,
    *,
    headers: list[str],
    max_row: int,
    apply_zebra_stripes,
    add_table,
    auto_width,
) -> None:
    apply_zebra_stripes(
        ws,
        2,
        max_row,
        len(headers),
        skip_cols={ALL_REPOS_GRADE_COLUMN, ALL_REPOS_TIER_COLUMN},
    )
    add_table(ws, "tblAllRepos", len(headers), max_row)
    auto_width(ws, len(headers), max_row + 2)
    ws.column_dimensions[get_column_letter(ALL_REPOS_SCORE_COLUMN)].width = 10
    ws.column_dimensions[get_column_letter(ALL_REPOS_INTEREST_COLUMN)].width = 10

    for column_name in ALL_REPOS_LONG_TEXT_HEADERS:
        col_index = headers.index(column_name) + 1
        ws.column_dimensions[get_column_letter(col_index)].width = ALL_REPOS_LONG_TEXT_WIDTH
        for row_index in range(2, max_row + 1):
            ws.cell(row=row_index, column=col_index).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )


def configure_all_repos_sheet(
    ws,
    *,
    headers: list[str],
    audit_count: int,
    style_header_row,
) -> None:
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{audit_count + 1}"


def build_all_repos_sheet(
    wb: Workbook,
    data: dict[str, Any],
    score_history: dict[str, list[float]] | None = None,
    *,
    risk_lookup: dict[str, str] | None,
    get_or_create_sheet,
    configure_sheet_view,
    headers: list[str],
    configure_all_repos_sheet,
    style_header_row,
    build_all_repo_rows,
    render_sparkline,
    write_all_repo_rows,
    style_data_cell,
    color_grade_cell,
    color_tier_cell,
    color_pattern_cell,
    sparkline_font,
    link_color: str,
    apply_all_repos_postprocessing,
    data_bar_rule_factory,
    icon_set_rule_factory,
    subheader_font,
    finalize_all_repos_layout,
    apply_zebra_stripes,
    add_table,
    auto_width,
) -> None:
    ws = get_or_create_sheet(wb, "All Repos")
    ws.sheet_properties.tabColor = "1565C0"
    configure_sheet_view(ws, zoom=105, show_grid_lines=True)

    configure_all_repos_sheet(
        ws,
        headers=headers,
        audit_count=len(data.get("audits", [])),
        style_header_row=style_header_row,
    )

    repo_rows = build_all_repo_rows(
        data.get("audits", []),
        score_history=score_history,
        risk_lookup=risk_lookup,
        render_sparkline=render_sparkline,
    )
    max_row = write_all_repo_rows(
        ws,
        repo_rows,
        headers=headers,
        style_data_cell=style_data_cell,
        color_grade_cell=color_grade_cell,
        color_tier_cell=color_tier_cell,
        color_pattern_cell=color_pattern_cell,
        sparkline_font=sparkline_font,
        link_color=link_color,
    )

    apply_all_repos_postprocessing(
        ws,
        max_row=max_row,
        data_bar_rule_factory=data_bar_rule_factory,
        icon_set_rule_factory=icon_set_rule_factory,
        subheader_font=subheader_font,
    )

    finalize_all_repos_layout(
        ws,
        headers=headers,
        max_row=max_row,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=add_table,
        auto_width=auto_width,
    )


def apply_all_repos_postprocessing(
    ws,
    *,
    max_row: int,
    data_bar_rule_factory,
    icon_set_rule_factory,
    subheader_font,
) -> None:
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=ALL_REPOS_SCORE_COLUMN).number_format = "0.000"
        ws.cell(row=row, column=ALL_REPOS_INTEREST_COLUMN).number_format = "0.000"

    if max_row > 1:
        ws.conditional_formatting.add(
            f"C2:C{max_row}",
            data_bar_rule_factory(
                start_type="num", start_value=0, end_type="num", end_value=1, color="166534"
            ),
        )
        ws.conditional_formatting.add(
            f"D2:D{max_row}",
            data_bar_rule_factory(
                start_type="num", start_value=0, end_type="num", end_value=1, color="0EA5E9"
            ),
        )
        ws.conditional_formatting.add(
            f"C2:C{max_row}",
            icon_set_rule_factory("3TrafficLights1", "num", [0, 0.55, 0.7]),
        )

    score_dv = DataValidation(
        allow_blank=True,
        prompt="Weighted average of 10 dimensions. See Score Explainer sheet.",
        promptTitle="Overall Score",
    )
    score_dv.sqref = f"C2:C{max_row}"
    ws.add_data_validation(score_dv)

    interest_dv = DataValidation(
        allow_blank=True,
        prompt="How interesting/ambitious (separate from completeness). Based on tech novelty, commit patterns, scope.",
        promptTitle="Interest Score",
    )
    interest_dv.sqref = f"D2:D{max_row}"
    ws.add_data_validation(interest_dv)

    grade_dv = DataValidation(
        allow_blank=True,
        prompt="A (>=85%) B (>=70%) C (>=55%) D (>=35%) F (<35%)",
        promptTitle="Letter Grade",
    )
    grade_dv.sqref = f"B2:B{max_row}"
    ws.add_data_validation(grade_dv)

    summary_row = max_row + 1
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = subheader_font
    ws.cell(
        row=summary_row,
        column=ALL_REPOS_SCORE_COLUMN,
        value=f"=AVERAGE(C2:C{max_row})",
    ).font = subheader_font
    ws.cell(
        row=summary_row,
        column=ALL_REPOS_INTEREST_COLUMN,
        value=f"=AVERAGE(D2:D{max_row})",
    ).font = subheader_font


def build_all_repos_workbook_sheet(
    wb: Workbook,
    data: dict[str, Any],
    score_history: dict[str, list[float]] | None = None,
    *,
    risk_lookup: dict[str, str] | None,
    get_or_create_sheet,
    configure_sheet_view,
    headers: list[str],
    configure_all_repos_sheet,
    style_header_row,
    build_all_repo_rows,
    render_sparkline,
    write_all_repo_rows,
    style_data_cell,
    color_grade_cell,
    color_tier_cell,
    color_pattern_cell,
    sparkline_font,
    link_color: str,
    apply_all_repos_postprocessing,
    data_bar_rule_factory,
    icon_set_rule_factory,
    subheader_font,
    finalize_all_repos_layout,
    apply_zebra_stripes,
    add_table,
    auto_width,
) -> None:
    build_all_repos_sheet(
        wb,
        data,
        score_history,
        risk_lookup=risk_lookup,
        get_or_create_sheet=get_or_create_sheet,
        configure_sheet_view=configure_sheet_view,
        headers=headers,
        configure_all_repos_sheet=configure_all_repos_sheet,
        style_header_row=style_header_row,
        build_all_repo_rows=build_all_repo_rows,
        render_sparkline=render_sparkline,
        write_all_repo_rows=write_all_repo_rows,
        style_data_cell=style_data_cell,
        color_grade_cell=color_grade_cell,
        color_tier_cell=color_tier_cell,
        color_pattern_cell=color_pattern_cell,
        sparkline_font=sparkline_font,
        link_color=link_color,
        apply_all_repos_postprocessing=apply_all_repos_postprocessing,
        data_bar_rule_factory=data_bar_rule_factory,
        icon_set_rule_factory=icon_set_rule_factory,
        subheader_font=subheader_font,
        finalize_all_repos_layout=finalize_all_repos_layout,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=add_table,
        auto_width=auto_width,
    )
