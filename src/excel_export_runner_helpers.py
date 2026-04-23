from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def build_excel_workbook(
    wb: Workbook,
    data: dict[str, Any],
    *,
    trend_data: list[dict[str, Any]] | None,
    diff_data: dict[str, Any] | None,
    score_history: dict[str, list[float]] | None,
    portfolio_profile: str,
    collection: str | None,
    excel_mode: str,
    risk_lookup: dict[str, str] | None,
    run_workbook_build_steps,
    build_dashboard,
    build_all_repos,
    build_portfolio_explorer,
    build_portfolio_catalog_sheet,
    build_scorecards_sheet,
    build_repo_detail,
    build_by_lens,
    build_by_collection,
    build_trend_summary,
    build_run_changes,
    build_heatmap,
    build_quick_wins,
    build_badges,
    build_tech_stack,
    build_trends,
    build_tier_breakdown,
    build_activity,
    build_repo_profiles,
    build_security,
    build_security_controls,
    build_supply_chain,
    build_security_debt,
    build_campaigns,
    build_writeback_audit,
    build_governance_controls,
    build_governance_audit,
    build_review_queue,
    build_review_history_sheet,
    build_hotspots,
    build_implementation_hotspots,
    build_operator_outcomes,
    build_approval_ledger,
    build_historical_intelligence,
    build_compare_sheet,
    build_scenario_planner,
    build_executive_summary,
    build_print_pack,
    build_changes,
    build_reconciliation,
    build_dependency_graph,
    build_score_explainer,
    build_action_items,
    build_hidden_data_sheets,
    build_navigation,
    inject_sheet_navigation,
    apply_workbook_named_ranges,
    finalize_workbook_structure,
    core_visible_sheets: set[str],
    template_info_sheet: str,
    preferred_order: list[str],
) -> None:
    run_workbook_build_steps(
        [
            (build_dashboard, (wb, data, diff_data, score_history), {"excel_mode": excel_mode}),
            (build_all_repos, (wb, data, score_history), {"risk_lookup": risk_lookup}),
            (
                build_portfolio_explorer,
                (wb, data),
                {"portfolio_profile": portfolio_profile, "collection": collection},
            ),
            (build_portfolio_catalog_sheet, (wb, data), {}),
            (build_scorecards_sheet, (wb, data), {}),
            (build_repo_detail, (wb, data), {}),
            (
                build_by_lens,
                (wb, data),
                {"portfolio_profile": portfolio_profile, "collection": collection},
            ),
            (build_by_collection, (wb, data), {"portfolio_profile": portfolio_profile}),
            (build_trend_summary, (wb, data, trend_data, score_history), {}),
            (build_run_changes, (wb, data, diff_data), {}),
            (build_heatmap, (wb, data), {}),
            (build_quick_wins, (wb, data), {}),
            (build_badges, (wb, data), {}),
            (build_tech_stack, (wb, data), {}),
            (build_trends, (wb, data, trend_data), {}),
            (build_tier_breakdown, (wb, data), {}),
            (build_activity, (wb, data), {}),
            (build_repo_profiles, (wb, data), {}),
            (build_security, (wb, data), {}),
            (build_security_controls, (wb, data), {}),
            (build_supply_chain, (wb, data), {}),
            (build_security_debt, (wb, data), {}),
            (build_campaigns, (wb, data), {}),
            (build_writeback_audit, (wb, data), {}),
            (build_governance_controls, (wb, data), {}),
            (build_governance_audit, (wb, data), {}),
            (build_review_queue, (wb, data), {"excel_mode": excel_mode}),
            (build_review_history_sheet, (wb, data), {}),
            (build_hotspots, (wb, data), {}),
            (build_implementation_hotspots, (wb, data), {}),
            (build_operator_outcomes, (wb, data), {}),
            (build_approval_ledger, (wb, data), {}),
            (build_historical_intelligence, (wb, data), {}),
            (build_compare_sheet, (wb, diff_data), {}),
            (
                build_scenario_planner,
                (wb, data),
                {"portfolio_profile": portfolio_profile, "collection": collection},
            ),
            (
                build_executive_summary,
                (wb, data, diff_data),
                {
                    "portfolio_profile": portfolio_profile,
                    "collection": collection,
                    "excel_mode": excel_mode,
                },
            ),
            (
                build_print_pack,
                (wb, data, diff_data),
                {
                    "portfolio_profile": portfolio_profile,
                    "collection": collection,
                    "excel_mode": excel_mode,
                },
            ),
            (build_changes, (wb, data, diff_data), {}),
            (build_reconciliation, (wb, data), {}),
            (build_dependency_graph, (wb, data), {}),
            (build_score_explainer, (wb,), {}),
            (build_action_items, (wb, data), {}),
            (build_hidden_data_sheets, (wb, data, trend_data, score_history, diff_data), {}),
            (
                build_navigation,
                (wb, data),
                {
                    "excel_mode": excel_mode,
                    "portfolio_profile": portfolio_profile,
                    "collection": collection,
                },
            ),
        ]
    )
    inject_sheet_navigation(wb)
    apply_workbook_named_ranges(
        wb,
        data,
        portfolio_profile=portfolio_profile,
        collection=collection,
        excel_mode=excel_mode,
    )
    finalize_workbook_structure(
        wb,
        core_visible_sheets=core_visible_sheets,
        template_info_sheet=template_info_sheet,
        preferred_order=preferred_order,
    )


def export_excel_workbook(
    report_path: Path,
    output_path: Path,
    *,
    trend_data: list[dict[str, Any]] | None,
    diff_data: dict[str, Any] | None,
    score_history: dict[str, list[float]] | None,
    portfolio_profile: str,
    collection: str | None,
    excel_mode: str,
    template_path: Path | None,
    truth_dir: Path | None,
    load_risk_truth,
    default_template_path: Path,
    resolve_template_path,
    copy_template_to_output,
    load_workbook_allowing_native_sparklines,
    build_excel_workbook,
    build_risk_summary_sheet,
    inject_native_sparklines,
    build_template_sparkline_specs,
) -> Path:
    data = json.loads(report_path.read_text())

    risk_lookup, risk_posture = load_risk_truth(truth_dir)

    if excel_mode not in {"template", "standard"}:
        raise ValueError(f"Unsupported excel mode: {excel_mode}")

    if excel_mode == "template":
        template = resolve_template_path(template_path or default_template_path)
        copy_template_to_output(output_path, template)
        wb = load_workbook_allowing_native_sparklines(output_path)
    else:
        wb = Workbook()

    build_excel_workbook(
        wb,
        data,
        trend_data=trend_data,
        diff_data=diff_data,
        score_history=score_history,
        portfolio_profile=portfolio_profile,
        collection=collection,
        excel_mode=excel_mode,
        risk_lookup=risk_lookup,
    )

    if risk_posture:
        build_risk_summary_sheet(wb, risk_posture)

    wb.save(str(output_path))
    if excel_mode == "template":
        inject_native_sparklines(
            output_path,
            build_template_sparkline_specs(
                data,
                trend_data=trend_data,
                score_history=score_history,
            ),
        )
    return output_path


def build_template_sparkline_specs(
    data: dict[str, Any],
    *,
    trend_data: list[dict[str, Any]] | None,
    score_history: dict[str, list[float]] | None,
    extend_score_history_with_current,
    extend_portfolio_trend_with_current,
    sparkline_spec_cls,
    trend_history_window: int,
) -> list[Any]:
    specs: list[Any] = []
    extended_score_history = extend_score_history_with_current(data, score_history)
    row_map = {
        repo_name: index + 2 for index, repo_name in enumerate(extended_score_history.keys())
    }
    audits_sorted = sorted(
        data.get("audits", []), key=lambda audit: audit.get("overall_score", 0), reverse=True
    )

    for offset, audit in enumerate(audits_sorted, 2):
        repo_name = audit.get("metadata", {}).get("name", "")
        scores = extended_score_history.get(repo_name)
        matrix_row = row_map.get(repo_name)
        if scores and matrix_row:
            start = max(1, trend_history_window - len(scores) + 1)
            end = trend_history_window
            specs.append(
                sparkline_spec_cls(
                    sheet_name="All Repos",
                    location=f"AA{offset}",
                    data_range=(
                        f"Data_TrendMatrix!"
                        f"{get_column_letter(start + 1)}{matrix_row}:"
                        f"{get_column_letter(end + 1)}{matrix_row}"
                    ),
                )
            )

    extended_trends = extend_portfolio_trend_with_current(data, trend_data)
    if extended_trends:
        specs.append(
            sparkline_spec_cls(
                sheet_name="Dashboard",
                location="L7",
                data_range=f"Data_PortfolioHistory!C2:C{len(extended_trends) + 1}",
            )
        )
    return specs
