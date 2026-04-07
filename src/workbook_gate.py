from __future__ import annotations

import argparse
import json
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from src.excel_export import CORE_VISIBLE_SHEETS, export_excel
from src.excel_template import DEFAULT_TEMPLATE_PATH

DEFAULT_GATE_DIR = Path("output") / "workbook-gate"
RESULT_FILENAME = "workbook-gate-result.json"
MANUAL_SIGNOFF_STATES = {"pending", "passed", "failed"}
MANUAL_SIGNOFF_OUTCOMES = {"passed", "failed"}
MANUAL_CHECK_STATES = {"passed", "failed", "n/a"}


def _sample_report_data() -> dict:
    return {
        "username": "sample-user",
        "generated_at": "2026-04-07T12:00:00+00:00",
        "audits": [
            {
                "metadata": {
                    "name": "RepoA",
                    "html_url": "https://github.com/sample-user/RepoA",
                    "description": "Strong shipped project",
                    "language": "Python",
                },
                "overall_score": 0.87,
                "interest_score": 0.62,
                "grade": "A",
                "completeness_tier": "shipped",
                "badges": ["fresh", "tested"],
                "flags": [],
                "lenses": {
                    "ship_readiness": {"score": 0.88, "summary": "Ready"},
                    "showcase_value": {"score": 0.9, "summary": "Strong story"},
                    "security_posture": {"score": 0.82, "summary": "Healthy"},
                },
                "security_posture": {"label": "healthy", "score": 0.82},
                "hotspots": [{"title": "Keep momentum", "severity": 0.2, "category": "finish-line"}],
                "action_candidates": [{"title": "Protect current momentum"}],
                "analyzer_results": [],
            },
            {
                "metadata": {
                    "name": "RepoB",
                    "html_url": "https://github.com/sample-user/RepoB",
                    "description": "Functional but under-finished",
                    "language": "TypeScript",
                },
                "overall_score": 0.61,
                "interest_score": 0.35,
                "grade": "C",
                "completeness_tier": "functional",
                "badges": [],
                "flags": [],
                "lenses": {
                    "ship_readiness": {"score": 0.63, "summary": "Needs work"},
                    "showcase_value": {"score": 0.55, "summary": "Average"},
                    "security_posture": {"score": 0.66, "summary": "Watch"},
                },
                "security_posture": {"label": "watch", "score": 0.66},
                "hotspots": [{"title": "Finish testing", "severity": 0.54, "category": "quality"}],
                "action_candidates": [{"title": "Finish the remaining delivery work"}],
                "analyzer_results": [],
            },
            {
                "metadata": {
                    "name": "RepoC",
                    "html_url": "https://github.com/sample-user/RepoC",
                    "description": "Risky work in progress",
                    "language": "Python",
                },
                "overall_score": 0.41,
                "interest_score": 0.14,
                "grade": "D",
                "completeness_tier": "wip",
                "badges": [],
                "flags": ["no-tests"],
                "lenses": {
                    "ship_readiness": {"score": 0.41, "summary": "Thin"},
                    "showcase_value": {"score": 0.3, "summary": "Weak"},
                    "security_posture": {"score": 0.29, "summary": "Risky"},
                },
                "security_posture": {"label": "critical", "score": 0.29},
                "hotspots": [{"title": "Security posture needs attention", "severity": 0.83, "category": "security-debt"}],
                "action_candidates": [{"title": "Preview governance controls"}],
                "analyzer_results": [],
            },
        ],
        "repos_audited": 3,
        "total_repos": 3,
        "average_score": 0.63,
        "portfolio_grade": "B",
        "portfolio_health_score": 0.67,
        "tier_distribution": {"shipped": 1, "functional": 1, "wip": 1, "skeleton": 0, "abandoned": 0},
        "language_distribution": {"Python": 2, "TypeScript": 1},
        "collections": {
            "showcase": {"description": "Best examples", "repos": [{"name": "RepoA", "reason": "Strong showcase"}]},
        },
        "profiles": {"default": {"description": "Balanced"}},
        "lenses": {"ship_readiness": {"description": "Delivery readiness", "average_score": 0.71}},
        "scenario_summary": {
            "top_levers": [
                {
                    "key": "testing",
                    "title": "Strengthen tests",
                    "lens": "ship_readiness",
                    "repo_count": 2,
                    "average_expected_lens_delta": 0.1,
                    "projected_tier_promotions": 1,
                }
            ],
            "portfolio_projection": {
                "selected_repo_count": 3,
                "projected_average_score_delta": 0.04,
                "projected_tier_promotions": 1,
            },
        },
        "security_posture": {
            "average_score": 0.59,
            "critical_repos": ["RepoC"],
            "repos_with_secrets": ["RepoC"],
            "provider_coverage": {
                "github": {"available_repos": 2, "total_repos": 3},
                "scorecard": {"available_repos": 1, "total_repos": 3},
            },
            "open_alerts": {"code_scanning": 2, "secret_scanning": 1},
        },
        "security_governance_preview": [
            {
                "repo": "RepoC",
                "priority": "high",
                "title": "Enable CodeQL default setup",
                "expected_posture_lift": 0.12,
                "effort": "medium",
                "source": "github",
                "why": "Code scanning is not configured",
            }
        ],
        "campaign_summary": {
            "campaign_type": "security-review",
            "label": "Security Review",
            "action_count": 1,
            "repo_count": 1,
        },
        "writeback_preview": {
            "sync_mode": "reconcile",
            "repos": [
                {
                    "repo": "RepoC",
                    "topics": ["ghra-call-security-review"],
                    "issue_title": "[Repo Auditor] Security Review",
                    "notion_action_count": 1,
                }
            ],
        },
        "writeback_results": {
            "mode": "apply",
            "target": "github",
            "results": [
                {
                    "repo_full_name": "sample-user/RepoC",
                    "target": "github-issue",
                    "status": "created",
                    "url": "https://github.com/sample-user/RepoC/issues/1",
                }
            ],
        },
        "managed_state_drift": [
            {
                "repo_full_name": "sample-user/RepoC",
                "target": "github-issue",
                "drift_state": "managed-issue-edited",
            }
        ],
        "rollback_preview": {"available": True, "item_count": 1, "fully_reversible_count": 1},
        "review_summary": {
            "review_id": "sample-review-1",
            "status": "open",
            "source_run_id": "sample-user:2026-04-07T12:00:00+00:00",
        },
        "review_targets": [
            {
                "repo": "RepoC",
                "title": "Security posture needs attention",
                "severity": 0.83,
                "next_step": "Preview governance controls",
                "recommended_next_step": "Preview governance controls",
                "decision_hint": "ready-for-governance-approval",
                "safe_to_defer": False,
            }
        ],
        "review_history": [
            {
                "review_id": "sample-review-1",
                "generated_at": "2026-04-07T12:00:00+00:00",
                "material_change_count": 2,
                "status": "open",
                "decision_state": "needs-review",
                "sync_state": "local-only",
                "emitted": True,
            }
        ],
        "material_changes": [
            {
                "change_type": "security-change",
                "repo": "RepoC",
                "severity": 0.83,
                "title": "Security posture needs attention",
            }
        ],
        "operator_summary": {
            "headline": "There is live drift or high-severity change that needs attention now.",
            "counts": {"blocked": 0, "urgent": 2, "ready": 1, "deferred": 0},
            "watch_strategy": "adaptive",
            "watch_enabled": True,
            "watch_chosen_mode": "incremental",
            "watch_decision_reason": "adaptive-incremental",
            "watch_decision_summary": "The current baseline is still compatible, so incremental watch remains safe for the next run.",
            "next_recommended_run_mode": "incremental",
            "full_refresh_due": False,
            "source_run_id": "sample-user:2026-04-07T12:00:00+00:00",
            "trend_status": "stable",
            "new_attention_count": 1,
            "resolved_attention_count": 0,
            "persisting_attention_count": 1,
            "reopened_attention_count": 0,
            "history_window_runs": 3,
            "aging_status": "watch",
            "decision_memory_status": "attempted",
            "primary_target_last_seen_at": "2026-04-07T12:00:00+00:00",
            "primary_target_last_intervention": {
                "item_id": "review-target:RepoC",
                "repo": "RepoC",
                "title": "Security posture needs attention",
                "event_type": "drifted",
                "recorded_at": "2026-04-07T12:00:00+00:00",
                "outcome": "drifted",
            },
            "primary_target_last_outcome": "no-change",
            "primary_target_resolution_evidence": "The last intervention was drifted for RepoC: Security posture needs attention, but the item is still open.",
            "primary_target_confidence_score": 0.7,
            "primary_target_confidence_label": "medium",
            "primary_target_confidence_reasons": [
                "Urgent drift or regression needs attention before ready work.",
                "A prior intervention happened, but the item is still open.",
                "This item has repeated recently and is no longer brand new.",
            ],
            "recent_interventions": [
                {
                    "item_id": "review-target:RepoC",
                    "repo": "RepoC",
                    "title": "Security posture needs attention",
                    "event_type": "drifted",
                    "recorded_at": "2026-04-07T12:00:00+00:00",
                    "outcome": "drifted",
                }
            ],
            "recently_quieted_count": 0,
            "confirmed_resolved_count": 0,
            "reopened_after_resolution_count": 0,
            "decision_memory_window_runs": 3,
            "resolution_evidence_summary": "The last intervention was drifted for RepoC: Security posture needs attention, but the item is still open.",
            "next_action_confidence_score": 0.75,
            "next_action_confidence_label": "high",
            "next_action_confidence_reasons": ["The next step is tied directly to the current top target."],
            "primary_target_trust_policy": "act-with-review",
            "primary_target_trust_policy_reason": "Urgent work has enough tuned confidence to act, with a quick operator review.",
            "next_action_trust_policy": "act-with-review",
            "next_action_trust_policy_reason": "Healthy calibration supports a confident next step, with light operator judgment.",
            "primary_target_exception_status": "none",
            "primary_target_exception_reason": "",
            "primary_target_exception_pattern_status": "none",
            "primary_target_exception_pattern_reason": "",
            "primary_target_trust_recovery_status": "none",
            "primary_target_trust_recovery_reason": "",
            "primary_target_recovery_confidence_score": 0.72,
            "primary_target_recovery_confidence_label": "medium",
            "primary_target_recovery_confidence_reasons": [
                "Healthy calibration supports relaxing the earlier soft caution.",
                "Recent runs are stabilizing, but the retirement window is still short.",
                "Recent exception history is still too light to prove the softer posture can retire.",
            ],
            "recovery_confidence_summary": "RepoC: Security posture needs attention is building recovery confidence (medium, 0.72), but the earlier caution has not retired yet.",
            "primary_target_exception_retirement_status": "candidate",
            "primary_target_exception_retirement_reason": "This target is trending toward retirement, but it has not earned it yet.",
            "exception_retirement_summary": "RepoC: Security posture needs attention is trending toward exception retirement, but the evidence is not strong enough to retire it yet.",
            "retired_exception_hotspots": [],
            "sticky_exception_hotspots": [],
            "exception_retirement_window_runs": 4,
            "primary_target_policy_debt_status": "watch",
            "primary_target_policy_debt_reason": "This class has enough recent exception activity to watch for lingering caution, but it is not yet clearly sticky or clearly normalization-friendly.",
            "primary_target_class_normalization_status": "candidate",
            "primary_target_class_normalization_reason": "This class is trending healthier, but the current target has not earned class-level normalization yet.",
            "policy_debt_summary": "RepoC: Security posture needs attention sits in a class with mixed recent caution behavior, so watch for policy debt before normalizing further.",
            "trust_normalization_summary": "RepoC: Security posture needs attention belongs to a healthier class trend, but it has not earned class-level normalization yet.",
            "policy_debt_hotspots": [],
            "normalized_class_hotspots": [],
            "class_normalization_window_runs": 4,
            "primary_target_class_memory_freshness_status": "mixed-age",
            "primary_target_class_memory_freshness_reason": "Class memory is still useful, but it is partly aging: 50% of the weighted signal is recent and the rest is older carry-forward.",
            "primary_target_class_decay_status": "none",
            "primary_target_class_decay_reason": "",
            "class_memory_summary": "RepoC: Security posture needs attention still has useful class memory, but part of that signal is aging and should be treated more cautiously.",
            "class_decay_summary": "Fresh class signals are still strong enough that no class-level trust posture needs to decay yet.",
            "primary_target_weighted_class_support_score": 0.48,
            "primary_target_weighted_class_caution_score": 0.24,
            "primary_target_class_trust_reweight_score": 0.24,
            "primary_target_class_trust_reweight_direction": "supporting-normalization",
            "primary_target_class_trust_reweight_reasons": [
                "Class memory is still useful, but it is partly aging: 50% of the weighted signal is recent and the rest is older carry-forward.",
                "Existing class normalization support is still contributing to a stronger posture.",
                "Fresh sticky class evidence is still carrying meaningful caution.",
            ],
            "class_reweighting_summary": "RepoC: Security posture needs attention inherited a stronger posture because fresh class support crossed the reweight threshold (0.24).",
            "supporting_class_hotspots": [],
            "caution_class_hotspots": [],
            "class_reweighting_window_runs": 4,
            "primary_target_class_trust_momentum_score": 0.26,
            "primary_target_class_trust_momentum_status": "building",
            "primary_target_class_reweight_stability_status": "watch",
            "primary_target_class_reweight_transition_status": "pending-support",
            "primary_target_class_reweight_transition_reason": "The class signal is visible, but it has not stayed strong long enough to confirm broader normalization yet.",
            "class_momentum_summary": "RepoC: Security posture needs attention shows healthier class support, but it has not stayed persistent enough to confirm broader normalization yet (0.26).",
            "class_reweight_stability_summary": "Class guidance for RepoC: Security posture needs attention is still settling and should be watched for one more stable stretch: supporting-normalization -> neutral.",
            "class_transition_window_runs": 4,
            "primary_target_class_transition_health_status": "building",
            "primary_target_class_transition_health_reason": "The pending class signal is still accumulating in the same direction and may confirm soon.",
            "primary_target_class_transition_resolution_status": "none",
            "primary_target_class_transition_resolution_reason": "",
            "class_transition_health_summary": "RepoC: Security posture needs attention still has a pending class signal that is accumulating and may confirm soon (1 run(s)).",
            "class_transition_resolution_summary": "No pending class transition has just confirmed, cleared, or expired in the recent window.",
            "class_transition_age_window_runs": 4,
            "primary_target_transition_closure_confidence_score": 0.75,
            "primary_target_transition_closure_confidence_label": "high",
            "primary_target_transition_closure_likely_outcome": "confirm-soon",
            "primary_target_transition_closure_confidence_reasons": [
                "The pending class signal is still accumulating in the same direction and may confirm soon.",
                "Recent class momentum is still aligned with the pending direction.",
                "The reweight score improved by 0.06 in the pending direction.",
            ],
            "transition_closure_confidence_summary": "RepoC: Security posture needs attention still has a pending class signal that looks strong enough to confirm soon if the next run stays aligned (0.75).",
            "transition_closure_window_runs": 4,
            "primary_target_class_pending_debt_status": "watch",
            "primary_target_class_pending_debt_reason": "This class has mixed recent pending-transition outcomes, so watch whether new pending signals resolve cleanly or start to accumulate debt.",
            "class_pending_debt_summary": "RepoC: Security posture needs attention belongs to a class with mixed pending-transition outcomes, so watch whether new pending signals confirm or start to linger.",
            "class_pending_resolution_summary": "No class-level pending-resolution pattern is strong enough to call out yet.",
            "class_pending_debt_window_runs": 10,
            "pending_debt_hotspots": [],
            "healthy_pending_resolution_hotspots": [],
            "primary_target_pending_debt_freshness_status": "mixed-age",
            "primary_target_pending_debt_freshness_reason": "Pending-transition memory is still useful, but it is partly aging: 50% of the weighted signal is recent and the rest is older carry-forward.",
            "pending_debt_freshness_summary": "RepoC: Security posture needs attention still has useful pending-transition memory, but some of that signal is aging and should be weighted more cautiously.",
            "pending_debt_decay_summary": "No strong pending-debt freshness trend is dominating the closure forecast yet.",
            "stale_pending_debt_hotspots": [],
            "fresh_pending_resolution_hotspots": [],
            "pending_debt_decay_window_runs": 4,
            "primary_target_weighted_pending_resolution_support_score": 0.58,
            "primary_target_weighted_pending_debt_caution_score": 0.31,
            "primary_target_closure_forecast_reweight_score": 0.27,
            "primary_target_closure_forecast_reweight_direction": "supporting-confirmation",
            "primary_target_closure_forecast_reweight_reasons": [
                "Pending-transition memory is still useful, but it is partly aging: 50% of the weighted signal is recent and the rest is older carry-forward.",
                "Recent class resolution behavior is still strong enough that this pending signal could confirm soon.",
                "The live pending signal is still building in the same direction.",
            ],
            "closure_forecast_reweighting_summary": "RepoC: Security posture needs attention still needs persistence before confirmation, but fresh class resolution behavior is strengthening the pending forecast (0.27).",
            "closure_forecast_reweighting_window_runs": 4,
            "supporting_pending_resolution_hotspots": [],
            "caution_pending_debt_hotspots": [],
            "stalled_transition_hotspots": [],
            "resolving_transition_hotspots": [],
            "sustained_class_hotspots": [],
            "oscillating_class_hotspots": [],
            "stale_class_memory_hotspots": [],
            "fresh_class_signal_hotspots": [],
            "class_decay_window_runs": 4,
            "recommendation_drift_status": "stable",
            "recommendation_drift_summary": "Recent trust-policy behavior is stable enough that no meaningful recommendation drift is recorded.",
            "policy_flip_hotspots": [],
            "exception_pattern_summary": "Recent exception behavior does not yet show a strong overcautious or recovery pattern.",
            "false_positive_exception_hotspots": [],
            "trust_recovery_window_runs": 3,
            "adaptive_confidence_summary": "Calibration is validating well, so the recommendation can be acted on with light operator review.",
            "recommendation_quality_summary": "Strong recommendation because the next step is tied directly to the current top target.",
            "confidence_validation_status": "healthy",
            "confidence_window_runs": 8,
            "validated_recommendation_count": 4,
            "partially_validated_recommendation_count": 1,
            "unresolved_recommendation_count": 1,
            "reopened_recommendation_count": 0,
            "insufficient_future_runs_count": 2,
            "high_confidence_hit_rate": 0.75,
            "medium_confidence_hit_rate": 0.5,
            "low_confidence_caution_rate": 1.0,
            "recent_validation_outcomes": [
                {
                    "run_id": "sample-user:2026-04-05T12:00:00+00:00",
                    "target_label": "RepoC: Security posture needs attention",
                    "confidence_label": "high",
                    "outcome": "validated",
                    "validated_in_runs": 2,
                }
            ],
            "confidence_calibration_summary": "Recent high-confidence recommendations are validating well: 75% high-confidence hit rate across 6 judged runs with no reopen noise.",
            "primary_target_reason": "This urgent item is already being watched across recent runs, so it stays ahead of ready work until it clears.",
            "primary_target_done_criteria": "Complete the recommended action and confirm the item exits the blocked or urgent queue on the next run.",
            "closure_guidance": "Preview governance controls. Treat this as done only when complete the recommended action and confirm the item exits the blocked or urgent queue on the next run.",
            "attention_age_bands": {"0-1 days": 0, "2-7 days": 1, "8-21 days": 0, "22+ days": 0},
            "chronic_item_count": 0,
            "newly_stale_count": 0,
            "longest_persisting_item": {
                "item_id": "review-target:RepoC",
                "repo": "RepoC",
                "title": "Security posture needs attention",
                "lane": "urgent",
                "age_days": 2,
                "aging_status": "watch",
            },
            "accountability_summary": "This urgent item is already being watched across recent runs, so it stays ahead of ready work until it clears. Preview governance controls. Treat this as done only when complete the recommended action and confirm the item exits the blocked or urgent queue on the next run. Aging pressure: 0 chronic item(s) and 0 newly stale item(s).",
            "primary_target": {
                "item_id": "review-target:RepoC",
                "repo": "RepoC",
                "title": "Security posture needs attention",
                "lane": "urgent",
                "kind": "review",
                "priority": 83,
                "recommended_action": "Preview governance controls",
                "age_days": 2,
                "aging_status": "watch",
                "stale": False,
                "reopened": False,
                "repeat_urgent": True,
                "newly_stale": False,
                "reason": "This urgent item is already being watched across recent runs, so it stays ahead of ready work until it clears.",
                "done_criteria": "Complete the recommended action and confirm the item exits the blocked or urgent queue on the next run.",
                "closure_guidance": "Preview governance controls. Treat this as done only when complete the recommended action and confirm the item exits the blocked or urgent queue on the next run.",
                "decision_memory_status": "attempted",
                "last_intervention": {
                    "item_id": "review-target:RepoC",
                    "repo": "RepoC",
                    "title": "Security posture needs attention",
                    "event_type": "drifted",
                    "recorded_at": "2026-04-07T12:00:00+00:00",
                    "outcome": "drifted",
                },
                "last_outcome": "no-change",
                "resolution_evidence": "The last intervention was drifted for RepoC: Security posture needs attention, but the item is still open.",
                "exception_pattern_status": "none",
                "exception_pattern_reason": "",
                "trust_recovery_status": "none",
                "trust_recovery_reason": "",
                "recovery_confidence_score": 0.72,
                "recovery_confidence_label": "medium",
                "exception_retirement_status": "candidate",
                "exception_retirement_reason": "This target is trending toward retirement, but it has not earned it yet.",
                "stable_policy_run_count": 1,
                "recent_exception_path": "",
            },
            "resolution_targets": [
                {
                    "item_id": "review-target:RepoC",
                    "repo": "RepoC",
                    "title": "Security posture needs attention",
                    "lane": "urgent",
                    "kind": "review",
                    "priority": 83,
                    "recommended_action": "Preview governance controls",
                    "age_days": 2,
                    "aging_status": "watch",
                    "stale": False,
                    "reopened": False,
                    "repeat_urgent": True,
                    "newly_stale": False,
                    "decision_memory_status": "attempted",
                    "last_intervention": {
                        "item_id": "review-target:RepoC",
                        "repo": "RepoC",
                        "title": "Security posture needs attention",
                        "event_type": "drifted",
                        "recorded_at": "2026-04-07T12:00:00+00:00",
                        "outcome": "drifted",
                    },
                    "last_outcome": "no-change",
                    "resolution_evidence": "The last intervention was drifted for RepoC: Security posture needs attention, but the item is still open.",
                    "confidence_score": 0.7,
                    "confidence_label": "medium",
                    "confidence_reasons": [
                        "Urgent drift or regression needs attention before ready work.",
                        "A prior intervention happened, but the item is still open.",
                        "This item has repeated recently and is no longer brand new.",
                    ],
                    "calibration_adjustment": 0.05,
                    "calibration_adjustment_reason": "Healthy calibration slightly strengthens blocked and urgent recommendations.",
                    "trust_policy": "act-with-review",
                    "trust_policy_reason": "Urgent work has enough tuned confidence to act, with a quick operator review.",
                    "trust_exception_status": "none",
                    "trust_exception_reason": "",
                    "exception_pattern_status": "none",
                    "exception_pattern_reason": "",
                    "trust_recovery_status": "none",
                    "trust_recovery_reason": "",
                    "recovery_confidence_score": 0.72,
                    "recovery_confidence_label": "medium",
                    "exception_retirement_status": "candidate",
                    "exception_retirement_reason": "This target is trending toward retirement, but it has not earned it yet.",
                    "stable_policy_run_count": 1,
                    "recent_exception_path": "",
                    "policy_flip_count": 0,
                    "recent_policy_path": "",
                }
            ],
            "trend_summary": "The queue is stable but still sticky: 1 attention item is persisting from the last run. Close RepoC: Security posture needs attention next.",
            "follow_through_summary": "1 urgent item repeated in the recent window, 0 open items now look stale, and the oldest open item has been visible for about 2 day(s).",
            "quiet_streak_runs": 0,
        },
        "operator_queue": [
            {
                "item_id": "review-target:RepoC",
                "kind": "review",
                "lane": "urgent",
                "lane_label": "Needs Attention Now",
                "lane_reason": "This item shows live drift, high-severity change, or rollback exposure.",
                "priority": 83,
                "repo": "RepoC",
                "title": "Security posture needs attention",
                "summary": "Governance approval is ready.",
                "recommended_action": "Preview governance controls",
                "source_run_id": "sample-user:2026-04-07T12:00:00+00:00",
                "age_days": 0,
                "links": [],
            }
        ],
        "governance_preview": {"applyable_count": 1},
        "governance_drift": [{"repo": "RepoC", "control": "secret_scanning"}],
        "governance_summary": {
            "headline": "Governed control drift needs operator review.",
            "status": "ready",
            "needs_reapproval": False,
            "drift_count": 1,
            "applyable_count": 1,
            "applied_count": 0,
            "rollback_available_count": 1,
            "top_actions": [
                {
                    "repo": "RepoC",
                    "title": "Enable CodeQL default setup",
                    "operator_state": "ready",
                    "expected_posture_lift": 0.12,
                    "source": "github",
                    "why": "Code scanning is not configured",
                }
            ],
        },
        "watch_state": {
            "watch_enabled": True,
            "requested_strategy": "adaptive",
            "chosen_mode": "incremental",
            "next_recommended_run_mode": "incremental",
            "reason": "adaptive-incremental",
            "reason_summary": "The current baseline is still compatible, so incremental watch remains safe for the next run.",
            "full_refresh_due": False,
        },
    }


def _sheet_targets(workbook_path: Path) -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    with zipfile.ZipFile(workbook_path) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        workbook_rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {
            rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
            for rel in workbook_rels
            if rel.attrib.get("Type", "").endswith("/worksheet")
        }
        visible: list[tuple[str, str]] = []
        hidden: list[tuple[str, str]] = []
        for sheet in workbook_root.findall("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets/{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
            title = sheet.attrib["name"]
            target = rel_targets[sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]]
            if sheet.attrib.get("state") == "hidden":
                hidden.append((title, target))
            else:
                visible.append((title, target))
        return visible, hidden


def _check_result(name: str, status: str, details: str) -> dict:
    return {
        "name": name,
        "status": status,
        "details": details,
    }


def _section_result(name: str, checks: list[dict]) -> dict:
    return {
        "name": name,
        "status": "passed" if all(check["status"] == "passed" for check in checks) else "failed",
        "checks": checks,
    }


def _validate_workbook_artifact(
    workbook_path: Path,
    *,
    workbook_mode: str,
    expected_visible: set[str],
) -> dict:
    checks: list[dict] = []
    wb = load_workbook(workbook_path)
    visible_sheets = {ws.title for ws in wb.worksheets if ws.sheet_state == "visible"}
    if visible_sheets != expected_visible:
        checks.append(
            _check_result(
                "visible-sheet-set",
                "failed",
                f"expected={sorted(expected_visible)} actual={sorted(visible_sheets)}",
            )
        )
    else:
        checks.append(
            _check_result(
                "visible-sheet-set",
                "passed",
                f"Visible sheets match the expected core set: {sorted(visible_sheets)}",
            )
        )

    hidden_data_sheets = [ws.title for ws in wb.worksheets if ws.title.startswith("Data_")]
    if not hidden_data_sheets:
        checks.append(_check_result("hidden-data-sheets", "failed", "No hidden Data_* sheets were found."))
    elif any(wb[name].sheet_state != "hidden" for name in hidden_data_sheets):
        checks.append(_check_result("hidden-data-sheets", "failed", "One or more Data_* sheets are visible."))
    else:
        checks.append(
            _check_result(
                "hidden-data-sheets",
                "passed",
                f"Hidden Data_* sheets stayed hidden: {sorted(hidden_data_sheets)}",
            )
        )

    visible_targets, hidden_targets = _sheet_targets(workbook_path)
    hidden_data_targets = [target for title, target in hidden_targets if title.startswith("Data_")]
    visible_table_parts: list[str] = []
    hidden_table_parts_ok = False
    with zipfile.ZipFile(workbook_path) as archive:
        for title, target in visible_targets:
            xml = archive.read(target).decode("utf-8", "ignore")
            if "<tableParts" in xml:
                visible_table_parts.append(title)
        if hidden_data_targets:
            hidden_table_parts_ok = any(
                "<tableParts" in archive.read(target).decode("utf-8", "ignore")
                for target in hidden_data_targets
            )
    if visible_table_parts:
        checks.append(
            _check_result(
                "visible-sheet-table-parts",
                "failed",
                f"Visible sheets contain table parts: {sorted(visible_table_parts)}",
            )
        )
    else:
        checks.append(
            _check_result(
                "visible-sheet-table-parts",
                "passed",
                "Visible sheets stayed filter-based with no structured table parts.",
            )
        )
    if hidden_data_targets and hidden_table_parts_ok:
        checks.append(
            _check_result(
                "hidden-data-table-parts",
                "passed",
                "Hidden Data_* sheets still contain workbook table parts for downstream bindings.",
            )
        )
    else:
        checks.append(
            _check_result(
                "hidden-data-table-parts",
                "failed",
                "Hidden Data_* sheets no longer contain workbook table parts.",
            )
        )
    return _section_result(f"{workbook_mode}-workbook-invariants", checks)


def _validate_parity(standard_path: Path, template_path: Path) -> dict:
    checks: list[dict] = []
    standard_wb = load_workbook(standard_path)
    template_wb = load_workbook(template_path)
    parity_checks = [
        ("Dashboard", "A1"),
        ("Review Queue", "B6"),
        ("Governance Controls", "B5"),
        ("Print Pack", "B9"),
    ]
    for sheet_name, cell in parity_checks:
        if standard_wb[sheet_name][cell].value != template_wb[sheet_name][cell].value:
            checks.append(
                _check_result(
                    f"{sheet_name}!{cell}",
                    "failed",
                    "Standard and template workbooks diverged on a shared top-line fact.",
                )
            )
        else:
            checks.append(
                _check_result(
                    f"{sheet_name}!{cell}",
                    "passed",
                    f"Matched value `{standard_wb[sheet_name][cell].value}` across both workbook modes.",
                )
            )
    return _section_result("cross-mode-parity", checks)


def _manual_signoff_template(standard_path: Path) -> dict:
    return {
        "status": "pending",
        "authority": "desktop-excel",
        "workbook": str(standard_path),
        "note": "Complete this manual signoff before releasing workbook-facing changes.",
        "checks": [
            {
                "id": "excel-open-no-repair",
                "label": "Open the standard workbook in desktop Excel with no repair prompt.",
                "status": "pending",
            },
            {
                "id": "visible-tabs-present",
                "label": "Confirm Index, Dashboard, Review Queue, Executive Summary, and Print Pack are present and readable.",
                "status": "pending",
            },
            {
                "id": "normal-zoom-readable",
                "label": "Confirm the standard workbook is readable at normal zoom on the core visible sheets.",
                "status": "pending",
            },
            {
                "id": "chart-placement-clean",
                "label": "Confirm charts and layout blocks are placed cleanly with no overlap.",
                "status": "pending",
            },
            {
                "id": "filters-work",
                "label": "Confirm visible filters work on the core operator sheets.",
                "status": "pending",
            },
        ],
    }


def _manual_check_lookup(manual_signoff: dict) -> dict[str, dict]:
    return {item.get("id", ""): item for item in manual_signoff.get("checks", []) if item.get("id")}


def _release_status(result: dict) -> str:
    if result.get("automated_checks", {}).get("status") != "passed":
        return "blocked"
    manual_status = (result.get("manual_signoff") or {}).get("status", "pending")
    if manual_status == "passed":
        return "ready"
    if manual_status == "failed":
        return "blocked"
    return "pending_manual_signoff"


def _format_manual_check_marker(status: str) -> str:
    if status == "passed":
        return "[x]"
    if status == "failed":
        return "[!]"
    if status == "n/a":
        return "[-]"
    return "[ ]"


def _write_manual_checklist(output_dir: Path, manual_signoff: dict) -> Path:
    checklist_path = output_dir / "workbook-gate-checklist.md"
    reviewer = manual_signoff.get("reviewer", "")
    reviewed_at = manual_signoff.get("reviewed_at", "")
    notes = manual_signoff.get("notes", "")
    checklist_path.write_text(
        "\n".join(
            [
                "# Workbook Gate Manual Checklist",
                "",
                "Open the standard workbook in desktop Excel before releasing workbook-facing changes.",
                "",
                f"- Workbook: `{manual_signoff['workbook']}`",
                f"- Status: `{manual_signoff['status']}`",
                *( [f"- Reviewer: `{reviewer}`"] if reviewer else [] ),
                *( [f"- Reviewed At: `{reviewed_at}`"] if reviewed_at else [] ),
                *( [f"- Notes: {notes}"] if notes else [] ),
                "",
                "## Required Checks",
                "",
                *[
                    f"- {_format_manual_check_marker(item.get('status', 'pending'))} {item['label']}"
                    for item in manual_signoff.get("checks", [])
                ],
                "",
            ]
        )
    )
    return checklist_path


def _write_gate_summary(output_dir: Path, result: dict) -> Path:
    summary_path = output_dir / "workbook-gate-summary.md"
    automated = result.get("automated_checks", {})
    manual = result.get("manual_signoff", {})
    artifacts = result.get("artifacts", {})
    lines = [
        "# Workbook Gate Summary",
        "",
        f"- Status: `{result.get('status', 'error')}`",
        f"- Generated At: `{result.get('release_metadata', {}).get('generated_at', '')}`",
        f"- Workbook Modes Checked: `{', '.join(result.get('release_metadata', {}).get('workbook_modes', []))}`",
        f"- Standard Workbook: `{artifacts.get('standard_workbook', '')}`",
        f"- Template Workbook: `{artifacts.get('template_workbook', '')}`",
        f"- Manual Checklist: `{artifacts.get('manual_checklist', '')}`",
        "",
        "## Automated Checks",
        "",
        f"- Status: `{automated.get('status', 'failed')}`",
        "",
    ]
    for section in automated.get("sections", []):
        lines.append(f"### {section.get('name', 'section')}")
        lines.append("")
        for check in section.get("checks", []):
            marker = "PASS" if check.get("status") == "passed" else "FAIL"
            lines.append(f"- [{marker}] {check.get('name')}: {check.get('details')}")
        lines.append("")
    lines.extend(
        [
            "## Manual Excel Signoff",
            "",
            f"- Status: `{manual.get('status', 'pending')}`",
            f"- Authority: `{manual.get('authority', 'desktop-excel')}`",
            *( [f"- Reviewer: `{manual.get('reviewer', '')}`"] if manual.get("reviewer") else [] ),
            *( [f"- Reviewed At: `{manual.get('reviewed_at', '')}`"] if manual.get("reviewed_at") else [] ),
            *( [f"- Notes: {manual.get('notes', '')}"] if manual.get("notes") else [] ),
            "",
        ]
    )
    for check in manual.get("checks", []):
        lines.append(f"- {_format_manual_check_marker(check.get('status', 'pending'))} {check.get('label', '')}")
    history = result.get("manual_signoff_history", []) or []
    if history:
        lines.extend(
            [
                "",
                "## Signoff History",
                "",
            ]
        )
        for entry in history[-5:]:
            lines.append(
                f"- `{entry.get('reviewed_at', '')}` {entry.get('reviewer', 'unknown')} -> "
                f"`{entry.get('outcome', 'pending')}`"
            )
    lines.append("")
    summary_path.write_text("\n".join(lines))
    return summary_path


def _flatten_failed_checks(sections: list[dict]) -> list[str]:
    errors: list[str] = []
    for section in sections:
        for check in section.get("checks", []):
            if check.get("status") == "failed":
                errors.append(f"{section.get('name', 'section')}::{check.get('name', 'check')} - {check.get('details', '')}")
    return errors


def run_workbook_gate(output_dir: Path = DEFAULT_GATE_DIR) -> dict:
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / "workbook-gate-report.json"
    report_path.write_text(json.dumps(_sample_report_data(), indent=2))

    standard_path = export_excel(report_path, output_dir / "workbook-gate-standard.xlsx", excel_mode="standard")
    template_path = export_excel(
        report_path,
        output_dir / "workbook-gate-template.xlsx",
        excel_mode="template",
        template_path=DEFAULT_TEMPLATE_PATH,
    )

    automated_sections = [
        _validate_workbook_artifact(standard_path, workbook_mode="standard", expected_visible=CORE_VISIBLE_SHEETS),
        _validate_workbook_artifact(template_path, workbook_mode="template", expected_visible=CORE_VISIBLE_SHEETS),
        _validate_parity(standard_path, template_path),
    ]
    validation_errors = _flatten_failed_checks(automated_sections)
    manual_signoff = _manual_signoff_template(standard_path)
    checklist_path = _write_manual_checklist(output_dir, manual_signoff)
    result = {
        "status": "ok" if not validation_errors else "error",
        "report_path": str(report_path),
        "standard_workbook": str(standard_path),
        "template_workbook": str(template_path),
        "manual_checklist": str(checklist_path),
        "artifacts": {
            "report_path": str(report_path),
            "standard_workbook": str(standard_path),
            "template_workbook": str(template_path),
            "manual_checklist": str(checklist_path),
        },
        "release_metadata": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "workbook_modes": ["standard", "template"],
            "parity_status": automated_sections[2]["status"],
            "invariant_status": "passed" if all(section["status"] == "passed" for section in automated_sections[:2]) else "failed",
        },
        "automated_checks": {
            "status": "passed" if not validation_errors else "failed",
            "sections": automated_sections,
        },
        "manual_signoff": manual_signoff,
        "manual_signoff_history": [],
        "errors": validation_errors,
    }
    result["release_status"] = _release_status(result)
    summary_path = _write_gate_summary(output_dir, result)
    result["gate_summary"] = str(summary_path)
    result["artifacts"]["gate_summary"] = str(summary_path)
    (output_dir / RESULT_FILENAME).write_text(json.dumps(result, indent=2))
    return result


def _load_gate_result(output_dir: Path) -> dict:
    result_path = output_dir / RESULT_FILENAME
    if not result_path.exists():
        raise FileNotFoundError("No workbook gate result exists yet. Run workbook-gate first.")
    return json.loads(result_path.read_text())


def _validate_signoff_args(result: dict, *, reviewer: str, outcome: str, checks: list[str]) -> dict[str, str]:
    if result.get("automated_checks", {}).get("status") != "passed":
        raise ValueError("Workbook signoff cannot be recorded while automated checks are failing.")
    if not reviewer.strip():
        raise ValueError("--reviewer is required when recording signoff.")
    if outcome not in MANUAL_SIGNOFF_OUTCOMES:
        raise ValueError(f"--outcome must be one of {sorted(MANUAL_SIGNOFF_OUTCOMES)}.")
    if not checks:
        raise ValueError("At least one --check entry is required when recording signoff.")

    parsed: dict[str, str] = {}
    for raw in checks:
        if "=" not in raw:
            raise ValueError(f"Invalid --check value `{raw}`. Use <id>=passed|failed|n/a.")
        check_id, status = raw.split("=", 1)
        check_id = check_id.strip()
        status = status.strip()
        if not check_id:
            raise ValueError(f"Invalid --check value `{raw}`. Check id cannot be empty.")
        if status not in MANUAL_CHECK_STATES:
            raise ValueError(f"Invalid check status `{status}` for `{check_id}`.")
        if check_id in parsed:
            raise ValueError(f"Duplicate check status provided for `{check_id}`.")
        parsed[check_id] = status

    manual_lookup = _manual_check_lookup(result.get("manual_signoff") or {})
    expected_ids = set(manual_lookup)
    provided_ids = set(parsed)
    missing = sorted(expected_ids - provided_ids)
    extra = sorted(provided_ids - expected_ids)
    if missing:
        raise ValueError(f"Missing manual signoff checks: {', '.join(missing)}.")
    if extra:
        raise ValueError(f"Unknown manual signoff checks: {', '.join(extra)}.")

    statuses = list(parsed.values())
    if outcome == "passed" and any(status != "passed" for status in statuses):
        raise ValueError("A passing signoff requires every manual check to be marked passed.")
    if outcome == "failed" and "failed" not in statuses:
        raise ValueError("A failing signoff requires at least one manual check to be marked failed.")
    return parsed


def record_manual_signoff(
    output_dir: Path = DEFAULT_GATE_DIR,
    *,
    reviewer: str,
    outcome: str,
    checks: list[str],
    notes: str = "",
) -> dict:
    result = _load_gate_result(output_dir)
    parsed_checks = _validate_signoff_args(result, reviewer=reviewer, outcome=outcome, checks=checks)
    reviewed_at = datetime.now(timezone.utc).isoformat()

    manual_signoff = dict(result.get("manual_signoff") or {})
    updated_checks: list[dict] = []
    for item in manual_signoff.get("checks", []) or []:
        updated_item = dict(item)
        updated_item["status"] = parsed_checks[item["id"]]
        updated_checks.append(updated_item)
    manual_signoff.update(
        {
            "status": outcome,
            "reviewer": reviewer.strip(),
            "reviewed_at": reviewed_at,
            "outcome": outcome,
            "notes": notes.strip(),
            "checks": updated_checks,
        }
    )
    history_entry = {
        "reviewer": reviewer.strip(),
        "reviewed_at": reviewed_at,
        "outcome": outcome,
        "notes": notes.strip(),
        "checks": {item["id"]: item["status"] for item in updated_checks},
    }
    result["manual_signoff"] = manual_signoff
    result["manual_signoff_history"] = [*(result.get("manual_signoff_history") or []), history_entry]
    result["release_status"] = _release_status(result)

    checklist_path = _write_manual_checklist(output_dir, manual_signoff)
    result.setdefault("artifacts", {})["manual_checklist"] = str(checklist_path)
    result["manual_checklist"] = str(checklist_path)
    summary_path = _write_gate_summary(output_dir, result)
    result["artifacts"]["gate_summary"] = str(summary_path)
    result["gate_summary"] = str(summary_path)
    (output_dir / RESULT_FILENAME).write_text(json.dumps(result, indent=2))
    return result


def format_gate_result(result: dict) -> str:
    artifacts = result.get("artifacts", {})
    lines = [
        f"Workbook gate status: {result.get('status', 'error')}",
        f"Release status: {result.get('release_status', 'pending_manual_signoff')}",
        f"Sample report: {artifacts.get('report_path', '')}",
        f"Standard workbook: {artifacts.get('standard_workbook', '')}",
        f"Template workbook: {artifacts.get('template_workbook', '')}",
        f"Manual checklist: {artifacts.get('manual_checklist', '')}",
        f"Gate summary: {artifacts.get('gate_summary', '')}",
        f"Automated checks: {result.get('automated_checks', {}).get('status', 'failed')}",
        f"Manual signoff: {result.get('manual_signoff', {}).get('status', 'pending')}",
    ]
    errors = result.get("errors") or []
    if errors:
        lines.append("Validation errors:")
        lines.extend(f"- {error}" for error in errors)
    else:
        lines.append("Validation checks passed. Final release step: complete the manual desktop Excel checklist.")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="workbook-gate",
        description="Generate a canonical sample workbook pair and validate workbook release invariants.",
    )
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_GATE_DIR),
        help="Directory for workbook gate artifacts (default: output/workbook-gate)",
    )
    parser.add_argument(
        "--record-signoff",
        action="store_true",
        help="Record the outcome of the manual desktop Excel signoff against the existing workbook gate artifact.",
    )
    parser.add_argument(
        "--reviewer",
        default="",
        help="Reviewer name for a recorded manual signoff.",
    )
    parser.add_argument(
        "--outcome",
        choices=sorted(MANUAL_SIGNOFF_OUTCOMES),
        help="Outcome for a recorded manual signoff.",
    )
    parser.add_argument(
        "--check",
        action="append",
        default=[],
        metavar="ID=STATUS",
        help="Manual check result in the form <id>=passed|failed|n/a. Repeat for every manual signoff check.",
    )
    parser.add_argument(
        "--notes",
        default="",
        help="Optional notes to attach to a recorded manual signoff.",
    )
    args = parser.parse_args()
    if args.record_signoff:
        result = record_manual_signoff(
            Path(args.output_dir),
            reviewer=args.reviewer,
            outcome=args.outcome or "",
            checks=args.check,
            notes=args.notes,
        )
    else:
        result = run_workbook_gate(Path(args.output_dir))
    print(format_gate_result(result))
    if result["status"] != "ok":
        raise SystemExit(1)


if __name__ == "__main__":
    main()
