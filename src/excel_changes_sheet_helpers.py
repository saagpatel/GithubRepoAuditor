"""Helpers for the Changes workbook sheet."""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Font

CHANGES_TIER_HEADERS = ["Repo", "Old Tier", "New Tier", "Old Score", "New Score", "Direction"]
CHANGES_SCORE_HEADERS = ["Repo", "Old Score", "New Score", "Delta"]
CHANGES_MATERIAL_HEADERS = ["Repo", "Type", "Severity", "Title"]


def build_changes_sheet_content(
    data: dict[str, Any],
    diff_data: dict[str, Any],
) -> dict[str, Any]:
    tier_changes = diff_data.get("tier_changes", [])
    promotions = [change for change in tier_changes if change.get("direction") == "promotion"]
    demotions = [change for change in tier_changes if change.get("direction") == "demotion"]
    score_changes = [
        change
        for change in diff_data.get("score_changes", [])
        if abs(change.get("delta", 0)) > 0.05
    ]
    material_rows: list[dict[str, Any]] = []
    material_by_type: dict[str, list[dict[str, Any]]] = {}
    for material_change in data.get("material_changes", []):
        material_by_type.setdefault(material_change.get("change_type", "other"), []).append(
            material_change
        )
    for change_type in sorted(material_by_type):
        for item in material_by_type[change_type][:30]:
            material_rows.append(
                {
                    "repo": item.get("repo", ""),
                    "change_type": change_type,
                    "severity": item.get("severity", ""),
                    "title": item.get("title", ""),
                }
            )
    return {
        "summary_rows": [
            ("Promotions", len(promotions)),
            ("Demotions", len(demotions)),
            ("Avg Score Delta", round(diff_data.get("average_score_delta", 0), 4)),
        ],
        "tier_change_rows": [
            {
                "name": change.get("name", ""),
                "old_tier": change.get("old_tier", ""),
                "new_tier": change.get("new_tier", ""),
                "old_score": round(change.get("old_score", 0), 3),
                "new_score": round(change.get("new_score", 0), 3),
                "direction": change.get("direction", ""),
            }
            for change in tier_changes
        ],
        "score_change_rows": [
            {
                "name": change.get("name", ""),
                "old_score": round(change.get("old_score", 0), 3),
                "new_score": round(change.get("new_score", 0), 3),
                "delta": round(change.get("delta", 0), 4),
            }
            for change in sorted(score_changes, key=lambda item: item.get("delta", 0), reverse=True)
        ],
        "material_rows": material_rows,
    }


def build_changes_sheet_sections(content: dict[str, Any]) -> list[dict[str, Any]]:
    sections: list[dict[str, Any]] = []
    if content["tier_change_rows"]:
        sections.append(
            {
                "title": "Tier Changes",
                "headers": CHANGES_TIER_HEADERS,
                "rows": [
                    [
                        change["name"],
                        change["old_tier"],
                        change["new_tier"],
                        change["old_score"],
                        change["new_score"],
                        change["direction"],
                    ]
                    for change in content["tier_change_rows"]
                ],
                "highlight_kind": "direction",
            }
        )
    if content["score_change_rows"]:
        sections.append(
            {
                "title": "Significant Score Changes",
                "headers": CHANGES_SCORE_HEADERS,
                "rows": [
                    [
                        change["name"],
                        change["old_score"],
                        change["new_score"],
                        change["delta"],
                    ]
                    for change in content["score_change_rows"]
                ],
                "highlight_kind": "delta",
            }
        )
    if content["material_rows"]:
        sections.append(
            {
                "title": "Material Changes",
                "headers": CHANGES_MATERIAL_HEADERS,
                "rows": [
                    [
                        item["repo"],
                        item["change_type"],
                        item["severity"],
                        item["title"],
                    ]
                    for item in content["material_rows"]
                ],
                "highlight_kind": "severity",
            }
        )
    return sections


def write_changes_sheet_sections(
    ws: Any,
    sections: list[dict[str, Any]],
    *,
    start_row: int,
    section_font: Any,
    style_header_row,
) -> int:
    row = start_row
    for index, section in enumerate(sections):
        if index:
            row += 2
        ws.cell(row=row, column=1, value=section["title"]).font = section_font
        header_row = row + 1
        for col, header in enumerate(section["headers"], 1):
            ws.cell(row=header_row, column=col, value=header)
        style_header_row(ws, header_row, len(section["headers"]))
        for row_values in section["rows"]:
            row = header_row + 1
            for col, value in enumerate(row_values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                _style_changes_highlight(
                    cell,
                    highlight_kind=section["highlight_kind"],
                    column=col,
                    value=value,
                )
            header_row = row
        row = header_row
    return row


def build_changes_sheet(
    wb: Any,
    data: dict[str, Any],
    diff_data: dict[str, Any] | None,
    *,
    get_or_create_sheet,
    auto_width,
    title_font: Any,
    subheader_font: Any,
    section_font: Any,
    build_changes_sheet_content_fn,
    build_changes_sheet_sections_fn,
    write_changes_sheet_sections_fn,
    style_header_row,
) -> None:
    if not diff_data:
        return

    ws = get_or_create_sheet(wb, "Changes")
    ws.sheet_properties.tabColor = "0891B2"
    content = build_changes_sheet_content_fn(data, diff_data)
    sections = build_changes_sheet_sections_fn(content)

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Changes Since Last Audit"
    ws["A1"].font = title_font

    for offset, (label, value) in enumerate(content["summary_rows"]):
        label_column = 1 + offset * 2
        ws.cell(row=3, column=label_column, value=label).font = subheader_font
        ws.cell(row=3, column=label_column + 1, value=value)

    row = write_changes_sheet_sections_fn(
        ws,
        sections,
        start_row=5,
        section_font=section_font,
        style_header_row=style_header_row,
    )

    auto_width(ws, 8, row)


def _style_changes_highlight(cell: Any, *, highlight_kind: str, column: int, value: Any) -> None:
    if highlight_kind == "direction" and column == 6:
        if value == "promotion":
            cell.font = Font("Calibri", 10, bold=True, color="166534")
        elif value == "demotion":
            cell.font = Font("Calibri", 10, bold=True, color="991B1B")
        return
    if highlight_kind == "delta" and column == 4:
        numeric_value = value if isinstance(value, (int, float)) else 0
        cell.font = Font(
            "Calibri",
            10,
            bold=True,
            color="166534" if numeric_value > 0 else "991B1B",
        )
        return
    if highlight_kind == "severity" and column == 3:
        if value == "high":
            cell.font = Font("Calibri", 10, bold=True, color="991B1B")
        elif value == "medium":
            cell.font = Font("Calibri", 10, color="92400E")
