from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

DEFAULT_PREFERRED_SHEET_ORDER = [
    "Index",
    "Dashboard",
    "Review Queue",
    "Portfolio Explorer",
    "Implementation Hotspots",
    "Operator Outcomes",
    "Approval Ledger",
    "Historical Intelligence",
    "Repo Detail",
    "Executive Summary",
    "By Lens",
    "By Collection",
    "Trend Summary",
    "Run Changes",
    "Campaigns",
    "Governance Controls",
    "Print Pack",
]


def set_autofilter(ws, max_col: int, max_row: int, start_row: int = 1) -> None:
    """Attach a plain AutoFilter to an already-populated range."""
    if max_row <= start_row:
        return
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(max_col)}{max_row}"


def add_table(ws, table_name: str, max_col: int, max_row: int, start_row: int = 1) -> None:
    """Attach a structured table to hidden sheets and a plain filter to visible ones."""
    if max_row <= start_row:
        return
    if ws.sheet_state == "visible":
        set_autofilter(ws, max_col, max_row, start_row=start_row)
        return
    ref = f"A{start_row}:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def clear_worksheet(ws) -> None:
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column:
        ws.delete_cols(1, ws.max_column)
    ws._charts = []
    ws.conditional_formatting._cf_rules.clear()
    ws.data_validations.dataValidation = []
    ws.merged_cells.ranges = set()
    ws.freeze_panes = None
    ws.auto_filter.ref = None
    if hasattr(ws, "_tables"):
        ws._tables.clear()


def get_or_create_sheet(wb: Workbook, title: str):
    if title in wb.sheetnames:
        ws = wb[title]
        clear_worksheet(ws)
        return ws
    return wb.create_sheet(title)


def set_defined_name(wb: Workbook, name: str, attr_text: str) -> None:
    try:
        del wb.defined_names[name]
    except KeyError:
        pass
    wb.defined_names.add(DefinedName(name, attr_text=attr_text))


def refresh_pivot_caches_on_load(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for pivot in getattr(ws, "_pivots", []):
            cache = getattr(pivot, "cache", None)
            if cache and getattr(cache, "refreshOnLoad", None) is not None:
                cache.refreshOnLoad = True


def configure_sheet_view(ws, *, zoom: int = 110, show_grid_lines: bool = True) -> None:
    ws.sheet_view.zoomScale = zoom
    ws.sheet_view.zoomScaleNormal = 100
    ws.sheet_view.showGridLines = show_grid_lines


def apply_visible_sheet_profile(
    wb: Workbook, *, core_visible_sheets: set[str], template_info_sheet: str
) -> None:
    for ws in wb.worksheets:
        if ws.title.startswith("Data_") or ws.title in {template_info_sheet, "Lookups"}:
            ws.sheet_state = "hidden"
            continue
        ws.sheet_state = "visible" if ws.title in core_visible_sheets else "hidden"


def reorder_workbook_sheets(wb: Workbook, *, preferred_order: list[str]) -> None:
    order_index = {name: index for index, name in enumerate(preferred_order)}
    wb._sheets.sort(key=lambda ws: (order_index.get(ws.title, len(preferred_order)), ws.title))
    if "Index" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Index")


def finalize_workbook_structure(
    wb: Workbook,
    *,
    core_visible_sheets: set[str],
    template_info_sheet: str,
    preferred_order: list[str] | None = None,
) -> None:
    apply_visible_sheet_profile(
        wb,
        core_visible_sheets=core_visible_sheets,
        template_info_sheet=template_info_sheet,
    )
    reorder_workbook_sheets(
        wb,
        preferred_order=preferred_order or list(DEFAULT_PREFERRED_SHEET_ORDER),
    )
    refresh_pivot_caches_on_load(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True


def run_workbook_build_steps(steps: list[tuple[object, tuple, dict]]) -> None:
    for builder, args, kwargs in steps:
        builder(*args, **kwargs)


def sheet_location(sheet_name: str, cell: str = "A1") -> str:
    escaped = sheet_name.replace("'", "''")
    if any(ch in sheet_name for ch in {" ", "'", "!", "-"}):
        return f"'{escaped}'!{cell}"
    return f"{escaped}!{cell}"
