from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

from src.excel_detail_helpers import set_internal_hyperlink
from src.excel_operator_follow_through_helpers import operator_follow_through_details
from src.excel_operator_summary_helpers import operator_handoff_values, operator_watch_values
from src.excel_sheet_layout_helpers import write_instruction_banner
from src.excel_styles import (
    CENTER,
    LEFT,
    NAVY,
    SECTION_FONT,
    SUBHEADER_FILL,
    SUBHEADER_FONT,
    SUBTITLE_FONT,
    TEAL,
    THIN_BORDER,
    TITLE_FONT,
    WRAP,
    auto_width,
    style_data_cell,
)
from src.excel_workbook_helpers import clear_worksheet, configure_sheet_view, sheet_location


def inject_sheet_navigation(wb: Workbook) -> None:
    strip_sheets = {
        "Dashboard",
        "All Repos",
        "Portfolio Explorer",
        "Repo Detail",
        "By Lens",
        "By Collection",
        "Trend Summary",
        "Run Changes",
        "Review Queue",
        "Campaigns",
        "Governance Controls",
        "Executive Summary",
        "Print Pack",
        "Index",
    }
    ordered_targets = [
        "Dashboard",
        "Review Queue",
        "Repo Detail",
        "Run Changes",
        "Executive Summary",
    ]
    for sheet_name in strip_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        start_col = max(ws.max_column + 2, 10)
        title_cell = ws.cell(row=1, column=start_col, value="Quick Links")
        title_cell.font = SUBHEADER_FONT
        for offset, target_sheet in enumerate(ordered_targets, 1):
            if target_sheet not in wb.sheetnames:
                continue
            cell = ws.cell(row=offset + 1, column=start_col, value=target_sheet)
            cell.alignment = LEFT
            if target_sheet == sheet_name:
                from openpyxl.styles import Font

                cell.font = Font("Calibri", 10, bold=True, color=NAVY)
                continue
            set_internal_hyperlink(cell, target_sheet, display=target_sheet)


def build_navigation(
    wb: Workbook,
    data: dict,
    *,
    excel_mode: str = "standard",
    portfolio_profile: str = "default",
    collection: str | None = None,
) -> None:
    """Navigation index as the first sheet."""
    ws = wb["Index"] if "Index" in wb.sheetnames else wb.create_sheet("Index", 0)
    clear_worksheet(ws)
    ws.sheet_properties.tabColor = "263238"
    configure_sheet_view(ws, zoom=125, show_grid_lines=False)
    ws.freeze_panes = "A12"

    ws.merge_cells("A1:G1")
    ws["A1"].value = f"GitHub Portfolio Audit: {data['username']}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:G2")
    ws[
        "A2"
    ].value = f"Last updated: {data['generated_at'][:10]} | {data['repos_audited']} repos | Grade: {data.get('portfolio_grade', '?')}"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = CENTER
    write_instruction_banner(
        ws,
        3,
        7,
        "Use this workbook in order: Dashboard for the brief, Run Changes for movement, Review Queue for action, Repo Detail for one-repo context, and Executive Summary for a shareable recap.",
    )
    next_mode, watch_strategy, watch_decision = operator_watch_values(data)
    what_changed, why_it_matters, next_action = operator_handoff_values(data)
    (
        follow_through,
        _follow_through_checkpoint,
        follow_through_escalation,
        _follow_through_hotspot,
        _follow_through_escalation_hotspot,
    ) = operator_follow_through_details(data)

    ws.cell(row=4, column=1, value="Start Here").font = SECTION_FONT
    ws.cell(row=5, column=1, value="Workbook Mode").font = SUBHEADER_FONT
    ws.cell(row=5, column=2, value=excel_mode)
    ws.cell(row=5, column=3, value="Profile").font = SUBHEADER_FONT
    ws.cell(row=5, column=4, value=portfolio_profile)
    ws.cell(row=6, column=1, value="Collection").font = SUBHEADER_FONT
    ws.cell(row=6, column=2, value=collection or "all")
    ws.cell(row=6, column=3, value="Source Run").font = SUBHEADER_FONT
    ws.cell(row=6, column=4, value=(data.get("operator_summary") or {}).get("source_run_id", ""))
    ws.cell(row=7, column=1, value="Report Reference").font = SUBHEADER_FONT
    ws.cell(row=7, column=2, value=(data.get("operator_summary") or {}).get("report_reference", ""))
    ws.cell(row=7, column=3, value="Operator Headline").font = SUBHEADER_FONT
    ws.cell(row=7, column=4, value=(data.get("operator_summary") or {}).get("headline", ""))
    ws.cell(row=8, column=1, value="Next Run").font = SUBHEADER_FONT
    ws.cell(row=8, column=2, value=next_mode)
    ws.cell(row=8, column=3, value="Watch Strategy").font = SUBHEADER_FONT
    ws.cell(row=8, column=4, value=watch_strategy)
    ws.merge_cells("A9:G9")
    ws["A9"] = watch_decision
    ws["A9"].font = SUBTITLE_FONT
    ws["A9"].alignment = WRAP
    ws.merge_cells("A10:G10")
    ws["A10"] = f"What changed: {what_changed}"
    ws["A10"].font = SUBTITLE_FONT
    ws["A10"].alignment = WRAP
    ws.merge_cells("A11:G11")
    ws["A11"] = f"Why it matters: {why_it_matters}"
    ws["A11"].font = SUBTITLE_FONT
    ws["A11"].alignment = WRAP
    ws.merge_cells("A12:G12")
    ws["A12"] = f"What to do next: {next_action}"
    ws["A12"].font = SUBTITLE_FONT
    ws["A12"].alignment = WRAP
    ws.merge_cells("A13:G13")
    ws["A13"] = f"Follow-through: {follow_through}"
    ws["A13"].font = SUBTITLE_FONT
    ws["A13"].alignment = WRAP
    ws.merge_cells("A14:G14")
    ws["A14"] = f"Escalation: {follow_through_escalation}"
    ws["A14"].font = SUBTITLE_FONT
    ws["A14"].alignment = WRAP
    ws.merge_cells("A15:G15")
    ws["A15"] = (
        "Start with Dashboard for the portfolio brief, move to Review Queue for action, then drill into Portfolio Explorer and Executive Summary for detail."
    )
    ws["A15"].font = SUBTITLE_FONT
    ws["A15"].alignment = WRAP
    ws.merge_cells("A16:G16")
    ws["A16"] = (
        "Operating rules: standard mode is the default path, visible sheets stay filter-based, and hidden Data_* sheets remain the workbook contract. Advanced sheets are hidden by default; use Excel Unhide when you need deeper diagnostics."
    )
    ws["A16"].font = SUBTITLE_FONT
    ws["A16"].alignment = WRAP

    groups = [
        (
            "Daily Triage",
            18,
            1,
            [
                ("Dashboard", "Start here for the big-picture health view and top attention items."),
                ("Review Queue", "Use this for blocked, urgent, ready, and safe-to-defer review work."),
                ("Run Changes", "See what changed since the last run before you decide where to spend attention."),
                ("Campaigns", "Check managed campaign state, reopen/closure context, and drift."),
                ("Governance Controls", "Review governed controls, approval posture, and rollback visibility."),
                ("Writeback Audit", "See what writeback changed and what is reversible."),
            ],
        ),
        (
            "Portfolio Analysis",
            17,
            5,
            [
                ("Portfolio Explorer", "Rank repos, compare score quality, and drill from summary into raw facts."),
                ("Repo Detail", "Select one repo and get a single-page briefing on score, risks, trend, and next action."),
                ("By Lens", "Compare the portfolio by ship readiness, momentum, security, and fit."),
                ("By Collection", "Understand collection-level leaders and concentration."),
                ("Trend Summary", "See portfolio-wide movement and repo trendlines over time."),
                ("All Repos", "Scan the full inventory with grades, tiers, and supporting evidence."),
            ],
        ),
        (
            "Executive Readout",
            27,
            1,
            [
                ("Executive Summary", "Readable leadership summary with what changed and what matters this week."),
                ("Print Pack", "Print-friendly handoff with risks, opportunities, and operator counts in plain language."),
            ],
        ),
        (
            "Deep Diagnostics",
            27,
            5,
            [
                ("Security", "Raw security posture, secrets, and dangerous-file findings."),
                ("Security Controls", "Control coverage and rollout detail by repo."),
                ("Supply Chain", "Dependency health, scorecard signals, and provider coverage."),
                ("Scoring Heatmap", "Dimension-by-dimension matrix for detailed score inspection."),
                ("Hotspots", "Highest-severity risks and opportunities across the portfolio."),
                ("Review History", "Recurring review history and decision-state ledger."),
                ("Governance Audit", "Approval, drift, and rollback evidence summary."),
                ("Score Explainer", "How scores, grades, and tiers are computed."),
                ("Action Items", "Prioritized improvement ideas with effort context."),
            ],
        ),
    ]

    from openpyxl.styles import Font

    for section, start_row, start_col, sheets in groups:
        ws.cell(row=start_row, column=start_col, value=section).font = SECTION_FONT
        header_row = start_row + 1
        for offset, header in enumerate(["Sheet", "Use This When", "Go"], 0):
            cell = ws.cell(row=header_row, column=start_col + offset, value=header)
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        row = header_row + 1
        for name, desc in sheets:
            if name not in wb.sheetnames:
                continue
            sheet_cell = ws.cell(row=row, column=start_col, value=name)
            sheet_cell.hyperlink = Hyperlink(
                ref=sheet_cell.coordinate,
                location=sheet_location(name),
                display=str(name),
            )
            sheet_cell.font = Font("Calibri", 11, bold=True, color=TEAL, underline="single")
            sheet_cell.border = THIN_BORDER
            sheet_cell.alignment = LEFT
            style_data_cell(ws.cell(row=row, column=start_col + 1, value=desc), "left")
            go_cell = ws.cell(row=row, column=start_col + 2, value="Open")
            go_cell.hyperlink = Hyperlink(
                ref=go_cell.coordinate,
                location=sheet_location(name),
                display="Open",
            )
            go_cell.font = Font("Calibri", 10, bold=True, color=TEAL, underline="single")
            go_cell.border = THIN_BORDER
            go_cell.alignment = CENTER
            row += 1

    auto_width(ws, 7, 35)
