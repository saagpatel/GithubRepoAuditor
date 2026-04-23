"""Flagship Excel dashboard generator.

Produces a 10-sheet workbook that serves as the primary way to understand
the portfolio: KPI dashboard, master table, heatmap, quick wins, badges,
tech stack, trends, tier breakdown, activity, and registry reconciliation.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BubbleChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.utils import get_column_letter

from src.excel_action_items_helpers import ACTION_ITEMS_HEADERS
from src.excel_action_items_helpers import (
    build_action_items_content as _build_action_items_content,
)
from src.excel_action_items_helpers import (
    collect_action_items as _collect_action_items,
)
from src.excel_action_items_helpers import (
    write_action_items_sections as _write_action_items_sections,
)
from src.excel_action_sync_readiness_helpers import (
    action_sync_readiness_rows as _action_sync_readiness_rows,
)
from src.excel_activity_sheet_helpers import (
    build_activity_rows as _build_activity_rows,
)
from src.excel_activity_sheet_helpers import (
    build_activity_sheet as _build_activity_sheet,
)
from src.excel_activity_sheet_helpers import (
    write_activity_table as _write_activity_table,
)
from src.excel_all_repos_helpers import ALL_REPOS_HEADERS
from src.excel_all_repos_helpers import (
    apply_all_repos_postprocessing as _apply_all_repos_postprocessing,
)
from src.excel_all_repos_helpers import build_all_repo_rows as _build_all_repo_rows
from src.excel_all_repos_helpers import (
    build_all_repos_workbook_sheet as _build_all_repos_workbook_sheet,
)
from src.excel_all_repos_helpers import (
    configure_all_repos_sheet as _configure_all_repos_sheet,
)
from src.excel_all_repos_helpers import (
    finalize_all_repos_layout as _finalize_all_repos_layout,
)
from src.excel_all_repos_helpers import (
    write_all_repo_rows as _write_all_repo_rows,
)
from src.excel_badges_helpers import (
    build_badge_distribution_chart as _build_badge_distribution_chart,
)
from src.excel_badges_helpers import (
    build_badges_sheet as _build_badges_sheet,
)
from src.excel_badges_helpers import (
    build_badges_sheet_content as _build_badges_sheet_content,
)
from src.excel_badges_helpers import (
    write_badges_sheet as _write_badges_sheet,
)
from src.excel_campaign_governance_helpers import (
    build_campaign_sheet_content as _build_campaign_sheet_content,
)
from src.excel_campaign_governance_helpers import (
    build_campaigns_sheet as _build_campaigns_sheet,
)
from src.excel_campaign_governance_helpers import (
    build_governance_controls_content as _build_governance_controls_content,
)
from src.excel_campaign_governance_helpers import (
    build_governance_controls_sheet as _build_governance_controls_sheet,
)
from src.excel_campaign_governance_helpers import (
    build_writeback_audit_content as _build_writeback_audit_content,
)
from src.excel_campaign_governance_helpers import (
    build_writeback_audit_sheet as _build_writeback_audit_sheet,
)
from src.excel_campaign_governance_helpers import (
    write_campaign_table as _write_campaign_table,
)
from src.excel_campaign_governance_helpers import (
    write_governance_controls_table as _write_governance_controls_table,
)
from src.excel_campaign_governance_helpers import (
    write_writeback_audit_table as _write_writeback_audit_table,
)
from src.excel_changes_sheet_helpers import (
    build_changes_sheet as _build_changes_sheet,
)
from src.excel_changes_sheet_helpers import (
    build_changes_sheet_content as _build_changes_sheet_content,
)
from src.excel_changes_sheet_helpers import (
    build_changes_sheet_sections as _build_changes_sheet_sections,
)
from src.excel_changes_sheet_helpers import (
    write_changes_sheet_sections as _write_changes_sheet_sections,
)
from src.excel_compare_scenario_helpers import (
    build_compare_lens_rows as _build_compare_lens_rows,
)
from src.excel_compare_scenario_helpers import (
    build_compare_repo_rows as _build_compare_repo_rows,
)
from src.excel_compare_scenario_helpers import (
    build_compare_sheet as _build_compare_sheet_helper,
)
from src.excel_compare_scenario_helpers import (
    build_scenario_planner_rows as _build_scenario_planner_rows,
)
from src.excel_compare_scenario_helpers import (
    build_scenario_planner_sheet as _build_scenario_planner_sheet,
)
from src.excel_compare_scenario_helpers import (
    write_compare_sections as _write_compare_sections,
)
from src.excel_compare_scenario_helpers import (
    write_scenario_planner_sections as _write_scenario_planner_sections,
)
from src.excel_dashboard_helpers import (
    build_dashboard_bubble_chart as _build_dashboard_bubble_chart,
)
from src.excel_dashboard_helpers import (
    build_dashboard_kpi_specs as _build_dashboard_kpi_specs,
)
from src.excel_dashboard_helpers import (
    build_dashboard_laggard_rows as _build_dashboard_laggard_rows,
)
from src.excel_dashboard_helpers import (
    build_dashboard_portfolio_trend_sparkline as _build_dashboard_portfolio_trend_sparkline,
)
from src.excel_dashboard_helpers import (
    build_dashboard_ranked_content as _build_dashboard_ranked_content,
)
from src.excel_dashboard_helpers import (
    build_dashboard_story_blocks as _build_dashboard_story_blocks,
)
from src.excel_dashboard_helpers import (
    build_dashboard_top_attention_rows as _build_dashboard_top_attention_rows,
)
from src.excel_dashboard_helpers import (
    build_dashboard_top_opportunity_rows as _build_dashboard_top_opportunity_rows,
)
from src.excel_dashboard_helpers import (
    build_dashboard_workbook_sheet as _build_dashboard_workbook_sheet,
)
from src.excel_dashboard_helpers import (
    write_dashboard_header as _write_dashboard_header,
)
from src.excel_dashboard_helpers import (
    write_dashboard_ranked_sections as _write_dashboard_ranked_sections,
)
from src.excel_dashboard_helpers import (
    write_dashboard_story_sections as _write_dashboard_story_sections,
)
from src.excel_dashboard_helpers import (
    write_dashboard_visual_sections as _write_dashboard_visual_sections,
)
from src.excel_detail_helpers import apply_lane_row_fill as _apply_lane_row_fill
from src.excel_executive_summary_helpers import (
    build_executive_operator_context as _build_executive_operator_context,
)
from src.excel_executive_summary_helpers import (
    build_executive_summary_workbook_sheet as _build_executive_summary_workbook_sheet,
)
from src.excel_export_registry_helpers import (
    build_excel_workbook_runtime as _build_excel_workbook_runtime,
)
from src.excel_export_runner_helpers import (
    build_excel_workbook as _build_excel_workbook_runner,
)
from src.excel_export_runner_helpers import (
    build_template_sparkline_specs as _build_template_sparkline_specs_helper,
)
from src.excel_export_runner_helpers import (
    export_excel_workbook as _export_excel_workbook,
)
from src.excel_export_truth_helpers import (
    load_risk_truth as _load_risk_truth,
)
from src.excel_heatmap_sheet_helpers import HEATMAP_HEADERS
from src.excel_heatmap_sheet_helpers import (
    build_heatmap_rows as _build_heatmap_rows,
)
from src.excel_heatmap_sheet_helpers import (
    build_heatmap_sheet as _build_heatmap_sheet,
)
from src.excel_heatmap_sheet_helpers import (
    write_heatmap_table as _write_heatmap_table,
)
from src.excel_hidden_data_content_helpers import (
    build_hidden_data_payload as _build_hidden_data_payload,
)
from src.excel_hidden_data_content_helpers import (
    build_hidden_data_sheets as _build_hidden_data_sheets_helper,
)
from src.excel_hidden_data_rows import build_core_hidden_rows as _build_core_hidden_rows
from src.excel_hidden_sheet_writer import write_hidden_data_tables as _write_hidden_data_tables
from src.excel_hotspot_sheet_helpers import (
    IMPLEMENTATION_HOTSPOTS_HEADERS,
)
from src.excel_hotspot_sheet_helpers import (
    build_dependency_graph_sheet as _build_dependency_graph_sheet,
)
from src.excel_hotspot_sheet_helpers import (
    build_hotspot_rows as _build_hotspot_rows,
)
from src.excel_hotspot_sheet_helpers import (
    build_hotspots_sheet as _build_hotspots_sheet,
)
from src.excel_hotspot_sheet_helpers import (
    build_implementation_hotspot_rows as _build_implementation_hotspot_rows,
)
from src.excel_hotspot_sheet_helpers import (
    build_implementation_hotspots_sheet as _build_implementation_hotspots_sheet,
)
from src.excel_hotspot_sheet_helpers import (
    write_hotspot_table as _write_hotspot_table,
)
from src.excel_ledger_sheet_helpers import (
    HISTORICAL_INTELLIGENCE_HEADERS,
)
from src.excel_ledger_sheet_helpers import (
    build_approval_ledger_content as _build_approval_ledger_content,
)
from src.excel_ledger_sheet_helpers import (
    build_approval_ledger_sheet as _build_approval_ledger_sheet,
)
from src.excel_ledger_sheet_helpers import (
    build_governance_audit_rows as _build_governance_audit_rows,
)
from src.excel_ledger_sheet_helpers import (
    build_governance_audit_sheet as _build_governance_audit_sheet,
)
from src.excel_ledger_sheet_helpers import (
    build_historical_intelligence_content as _build_historical_intelligence_content,
)
from src.excel_ledger_sheet_helpers import (
    build_review_history_rows as _build_review_history_rows,
)
from src.excel_ledger_sheet_helpers import (
    build_review_history_sheet as _build_review_history_sheet_helper,
)
from src.excel_ledger_sheet_helpers import (
    write_approval_ledger_sections as _write_approval_ledger_sections,
)
from src.excel_ledger_sheet_helpers import (
    write_governance_audit_table as _write_governance_audit_table,
)
from src.excel_ledger_sheet_helpers import (
    write_historical_intelligence_sheet as _write_historical_intelligence_sheet,
)
from src.excel_ledger_sheet_helpers import (
    write_review_history_table as _write_review_history_table,
)
from src.excel_navigation_helpers import build_navigation as _build_navigation
from src.excel_navigation_helpers import inject_sheet_navigation as _inject_sheet_navigation
from src.excel_operator_hidden_rows import build_operator_hidden_rows as _build_operator_hidden_rows
from src.excel_operator_outcomes_helpers import (
    build_operator_outcomes_content as _build_operator_outcomes_content,
)
from src.excel_operator_outcomes_helpers import (
    write_operator_outcomes_sections as _write_operator_outcomes_sections,
)
from src.excel_operator_queue_helpers import (
    format_lane_counts as _format_lane_counts,
)
from src.excel_operator_queue_helpers import (
    format_repo_rollup_counts as _format_repo_rollup_counts,
)
from src.excel_operator_queue_helpers import (
    operator_counts as _operator_counts,
)
from src.excel_operator_queue_helpers import (
    ordered_queue_items as _ordered_queue_items,
)
from src.excel_operator_queue_helpers import (
    summarize_top_actions as _summarize_top_actions,
)
from src.excel_operator_queue_helpers import (
    summarize_top_issue_families as _summarize_top_issue_families,
)
from src.excel_portfolio_misc_helpers import (
    REPO_PROFILE_RADAR_DIMS,
    REPO_PROFILE_RADAR_LABELS,
    RISK_SUMMARY_HEADERS,
)
from src.excel_portfolio_misc_helpers import (
    add_repo_profile_charts as _add_repo_profile_charts,
)
from src.excel_portfolio_misc_helpers import (
    build_registry_content as _build_registry_content,
)
from src.excel_portfolio_misc_helpers import (
    build_registry_sheet as _build_registry_sheet,
)
from src.excel_portfolio_misc_helpers import (
    build_repo_profile_matrix as _build_repo_profile_matrix,
)
from src.excel_portfolio_misc_helpers import (
    build_risk_summary_rows as _build_risk_summary_rows,
)
from src.excel_portfolio_misc_helpers import (
    build_tech_stack_sheet as _build_tech_stack_sheet,
)
from src.excel_portfolio_misc_helpers import (
    build_tier_breakdown_sections as _build_tier_breakdown_sections,
)
from src.excel_portfolio_misc_helpers import (
    build_tier_breakdown_sheet as _build_tier_breakdown_sheet,
)
from src.excel_portfolio_misc_helpers import (
    write_registry_sections as _write_registry_sections,
)
from src.excel_portfolio_misc_helpers import (
    write_repo_profile_matrix as _write_repo_profile_matrix,
)
from src.excel_portfolio_misc_helpers import (
    write_risk_summary_table as _write_risk_summary_table,
)
from src.excel_portfolio_misc_helpers import (
    write_tier_breakdown_sections as _write_tier_breakdown_sections,
)
from src.excel_portfolio_misc_helpers import (
    write_trends_sheet as _write_trends_sheet,
)
from src.excel_portfolio_sheet_helpers import (
    build_portfolio_catalog_sheet as _build_portfolio_catalog_sheet_helper,
)
from src.excel_portfolio_sheet_helpers import (
    build_portfolio_explorer_sheet as _build_portfolio_explorer_sheet,
)
from src.excel_portfolio_sheet_helpers import (
    build_scorecard_table_rows as _build_scorecard_table_rows,
)
from src.excel_portfolio_sheet_helpers import (
    build_scorecards_sheet as _build_scorecards_sheet_helper,
)
from src.excel_portfolio_sheet_helpers import (
    write_portfolio_table as _write_portfolio_table,
)
from src.excel_print_pack_helpers import (
    print_pack_workflow_guidance_rows as _print_pack_workflow_guidance_rows,
)
from src.excel_print_pack_layout_helpers import (
    build_print_pack_sheet_content as _build_print_pack_sheet_content,
)
from src.excel_print_pack_layout_helpers import (
    build_print_pack_workbook_sheet as _build_print_pack_workbook_sheet,
)
from src.excel_print_pack_layout_helpers import (
    write_print_pack_sheet as _write_print_pack_sheet,
)
from src.excel_profile_trend_helpers import (
    build_by_collection_rows as _build_by_collection_rows,
)
from src.excel_profile_trend_helpers import (
    build_by_collection_sheet as _build_by_collection_sheet,
)
from src.excel_profile_trend_helpers import (
    build_by_lens_rows as _build_by_lens_rows,
)
from src.excel_profile_trend_helpers import (
    build_by_lens_sheet as _build_by_lens_sheet,
)
from src.excel_profile_trend_helpers import (
    build_trend_summary_content as _build_trend_summary_content,
)
from src.excel_profile_trend_helpers import (
    build_trend_summary_sheet as _build_trend_summary_sheet,
)
from src.excel_profile_trend_helpers import (
    write_trend_summary_sections as _write_trend_summary_sections,
)
from src.excel_quick_wins_helpers import (
    build_quick_wins_rows as _build_quick_wins_rows,
)
from src.excel_quick_wins_helpers import (
    build_quick_wins_sheet as _build_quick_wins_sheet,
)
from src.excel_quick_wins_helpers import (
    write_quick_wins_sheet as _write_quick_wins_sheet,
)
from src.excel_repo_data_helpers import repo_detail_rows as _repo_detail_rows
from src.excel_repo_data_helpers import run_change_rows as _run_change_rows
from src.excel_repo_detail_helpers import (
    build_repo_detail_sheet as _build_repo_detail_sheet,
)
from src.excel_report_helpers import display_operator_state as _display_operator_state
from src.excel_report_helpers import generate_narrative as _generate_narrative
from src.excel_review_queue_helpers import (
    build_review_queue_sheet_content as _build_review_queue_sheet_content,
)
from src.excel_review_queue_helpers import (
    build_review_queue_workbook_sheet as _build_review_queue_workbook_sheet,
)
from src.excel_review_queue_helpers import (
    write_review_queue_overview as _write_review_queue_overview,
)
from src.excel_review_queue_helpers import (
    write_review_queue_table as _write_review_queue_table,
)
from src.excel_review_queue_table_helpers import (
    REVIEW_QUEUE_CENTER_ALIGNED_COLUMNS,
    REVIEW_QUEUE_GUIDANCE,
    REVIEW_QUEUE_HEADERS,
)
from src.excel_review_queue_table_helpers import (
    build_review_queue_table_rows as _build_review_queue_table_rows,
)
from src.excel_run_changes_helpers import (
    build_run_changes_content as _build_run_changes_content,
)
from src.excel_run_changes_helpers import (
    build_run_changes_sheet as _build_run_changes_sheet,
)
from src.excel_run_changes_helpers import (
    normalize_run_changes_sections as _normalize_run_changes_sections,
)
from src.excel_run_changes_helpers import (
    write_run_changes_sections as _write_run_changes_sections,
)
from src.excel_run_changes_helpers import (
    write_run_changes_summary as _write_run_changes_summary,
)
from src.excel_score_explainer_helpers import (
    build_score_explainer_content as _build_score_explainer_content,
)
from src.excel_score_explainer_helpers import (
    build_score_explainer_sheet as _build_score_explainer_sheet,
)
from src.excel_score_explainer_helpers import (
    write_score_explainer_sections as _write_score_explainer_sections,
)
from src.excel_security_sheet_helpers import (
    SECURITY_CONTROLS_HEADERS,
    SECURITY_DEBT_HEADERS,
    SECURITY_HEADERS,
    SUPPLY_CHAIN_HEADERS,
)
from src.excel_security_sheet_helpers import (
    build_security_controls_rows as _build_security_controls_rows,
)
from src.excel_security_sheet_helpers import (
    build_security_controls_sheet as _build_security_controls_sheet,
)
from src.excel_security_sheet_helpers import (
    build_security_debt_rows as _build_security_debt_rows,
)
from src.excel_security_sheet_helpers import (
    build_security_debt_sheet as _build_security_debt_sheet,
)
from src.excel_security_sheet_helpers import build_security_sheet as _build_security_sheet
from src.excel_security_sheet_helpers import (
    build_security_sheet_rows as _build_security_sheet_rows,
)
from src.excel_security_sheet_helpers import (
    build_supply_chain_rows as _build_supply_chain_rows,
)
from src.excel_security_sheet_helpers import build_supply_chain_sheet as _build_supply_chain_sheet
from src.excel_security_sheet_helpers import (
    write_security_table as _write_security_table,
)
from src.excel_sheet_layout_helpers import set_sheet_header as _set_sheet_header
from src.excel_sheet_layout_helpers import write_instruction_banner as _write_instruction_banner
from src.excel_sheet_layout_helpers import write_key_value_block as _write_key_value_block
from src.excel_sheet_layout_helpers import write_ranked_list as _write_ranked_list
from src.excel_styles import (
    CENTER,
    GRADE_COLORS,
    HEATMAP_AMBER,
    HEATMAP_GREEN,
    HEATMAP_RED,
    NARRATIVE_FONT,
    NAVY,
    SECTION_FONT,
    SPARKLINE_FONT,
    SUBHEADER_FILL,
    SUBHEADER_FONT,
    SUBTITLE_FONT,
    TEAL,
    THIN_BORDER,
    TIER_FILLS,
    TITLE_FONT,
    WHITE,
    WRAP,
    apply_zebra_stripes,
    auto_width,
    color_grade_cell,
    color_pattern_cell,
    color_tier_cell,
    style_data_cell,
    style_header_row,
    write_kpi_card,
)
from src.excel_template import (
    DEFAULT_TEMPLATE_PATH,
    TEMPLATE_INFO_SHEET,
    TREND_HISTORY_WINDOW,
    SparklineSpec,
    copy_template_to_output,
    inject_native_sparklines,
    load_workbook_allowing_native_sparklines,
    resolve_template_path,
)
from src.excel_template_info_helpers import (
    build_dashboard_generated_label as _build_dashboard_generated_label,
)
from src.excel_template_info_helpers import (
    build_named_range_targets as _build_named_range_targets,
)
from src.excel_template_info_helpers import (
    build_template_info_rows as _build_template_info_rows,
)
from src.excel_timeline_helpers import (
    extend_portfolio_trend_with_current as _extend_portfolio_trend_with_current,
)
from src.excel_timeline_helpers import (
    extend_score_history_with_current as _extend_score_history_with_current,
)
from src.excel_timeline_helpers import review_status_counts as _review_status_counts
from src.excel_workbook_helpers import (
    DEFAULT_PREFERRED_SHEET_ORDER,
)
from src.excel_workbook_helpers import add_table as _add_table
from src.excel_workbook_helpers import clear_worksheet as _clear_worksheet
from src.excel_workbook_helpers import configure_sheet_view as _configure_sheet_view
from src.excel_workbook_helpers import (
    finalize_workbook_structure as _finalize_workbook_structure,
)
from src.excel_workbook_helpers import get_or_create_sheet as _get_or_create_sheet
from src.excel_workbook_helpers import run_workbook_build_steps as _run_workbook_build_steps
from src.excel_workbook_helpers import set_autofilter as _set_autofilter
from src.excel_workbook_helpers import set_defined_name as _set_defined_name
from src.excel_workbook_helpers import sheet_location as _sheet_location
from src.excel_workbook_rollups import build_workbook_rollups as _build_workbook_rollups
from src.history import load_trend_data
from src.report_enrichment import (
    build_maturity_gap_summary,
    build_operating_paths_summary,
    build_portfolio_catalog_summary,
    build_portfolio_intent_alignment_summary,
    build_queue_pressure_summary,
    build_run_change_counts,
    build_run_change_summary,
    build_scorecards_summary,
    build_top_recommendation_summary,
    build_trust_actionability_summary,
    build_weekly_review_pack,
)
from src.sparkline import sparkline as render_sparkline
from src.terminology import ACTION_SYNC_CANONICAL_LABELS
from src.weekly_scheduling_overlay import resolve_weekly_story_value

# Tier display order
TIER_ORDER = ["shipped", "functional", "wip", "skeleton", "abandoned"]
PIE_COLORS = ["166534", "1565C0", "D97706", "C2410C", "6B7280"]
CORE_VISIBLE_SHEETS = {
    "Index",
    "Dashboard",
    "All Repos",
    "Portfolio Explorer",
    "Portfolio Catalog",
    "Scorecards",
    "Implementation Hotspots",
    "Operator Outcomes",
    "Approval Ledger",
    "Historical Intelligence",
    "Repo Detail",
    "By Lens",
    "By Collection",
    "Trend Summary",
    "Run Changes",
    "Review Queue",
    "Campaigns",
    "Governance Controls",
    "Executive Summary",
    "Print Pack",
}

RADAR_DIMS = REPO_PROFILE_RADAR_DIMS
RADAR_LABELS = REPO_PROFILE_RADAR_LABELS

def _apply_workbook_named_ranges(
    wb: Workbook,
    data: dict,
    *,
    portfolio_profile: str = "default",
    collection: str | None = None,
    excel_mode: str = "standard",
) -> None:
    for name, target in _build_named_range_targets(TEMPLATE_INFO_SHEET):
        _set_defined_name(wb, name, target)

    dashboard = wb["Dashboard"]
    dashboard["A2"] = _build_dashboard_generated_label(data)

    counts = _review_status_counts(data)
    info = (
        wb[TEMPLATE_INFO_SHEET]
        if TEMPLATE_INFO_SHEET in wb.sheetnames
        else wb.create_sheet(TEMPLATE_INFO_SHEET)
    )
    info.sheet_state = "hidden"
    for row_index, (label, value) in enumerate(
        _build_template_info_rows(
            data=data,
            counts=counts,
            portfolio_profile=portfolio_profile,
            collection=collection,
            excel_mode=excel_mode,
        ),
        1,
    ):
        info.cell(row=row_index, column=1, value=label)
        info.cell(row=row_index, column=2, value=value)


# ═══════════════════════════════════════════════════════════════════════
# Sheet 1: Dashboard (Executive Overview)
# ═══════════════════════════════════════════════════════════════════════


def _build_dashboard(
    wb: Workbook,
    data: dict,
    diff_data: dict | None = None,
    score_history: dict[str, list[float]] | None = None,
    *,
    excel_mode: str = "standard",
) -> None:
    _build_dashboard_workbook_sheet(
        wb,
        data,
        diff_data,
        score_history,
        excel_mode=excel_mode,
        get_or_create_sheet=_get_or_create_sheet,
        clear_worksheet=_clear_worksheet,
        configure_sheet_view=_configure_sheet_view,
        write_dashboard_header=_write_dashboard_header,
        build_dashboard_kpi_specs=_build_dashboard_kpi_specs,
        write_kpi_card=write_kpi_card,
        sheet_location=_sheet_location,
        build_dashboard_portfolio_trend_sparkline=_build_dashboard_portfolio_trend_sparkline,
        load_trend_data=load_trend_data,
        render_sparkline=render_sparkline,
        build_dashboard_story_blocks=_build_dashboard_story_blocks,
        write_dashboard_story_sections=_write_dashboard_story_sections,
        build_dashboard_ranked_content=_build_dashboard_ranked_content,
        write_dashboard_ranked_sections=_write_dashboard_ranked_sections,
        write_dashboard_visual_sections=_write_dashboard_visual_sections,
        build_run_change_summary=build_run_change_summary,
        build_queue_pressure_summary=build_queue_pressure_summary,
        build_trust_actionability_summary=build_trust_actionability_summary,
        build_top_recommendation_summary=build_top_recommendation_summary,
        resolve_weekly_story_value=resolve_weekly_story_value,
        build_weekly_review_pack=build_weekly_review_pack,
        build_executive_operator_context=_build_executive_operator_context,
        operator_counts=_operator_counts,
        format_lane_counts=_format_lane_counts,
        display_operator_state=_display_operator_state,
        build_portfolio_catalog_summary=build_portfolio_catalog_summary,
        build_operating_paths_summary=build_operating_paths_summary,
        build_portfolio_intent_alignment_summary=build_portfolio_intent_alignment_summary,
        build_scorecards_summary=build_scorecards_summary,
        build_workbook_rollups=_build_workbook_rollups,
        build_dashboard_top_attention_rows=_build_dashboard_top_attention_rows,
        build_dashboard_top_opportunity_rows=_build_dashboard_top_opportunity_rows,
        build_dashboard_laggard_rows=_build_dashboard_laggard_rows,
        write_key_value_block=_write_key_value_block,
        write_ranked_list=_write_ranked_list,
        title_font=TITLE_FONT,
        subtitle_font=SUBTITLE_FONT,
        narrative_font=NARRATIVE_FONT,
        center_alignment=CENTER,
        wrap_alignment=WRAP,
        section_font=SECTION_FONT,
        subheader_font=SUBHEADER_FONT,
        navy=NAVY,
        grade_colors=GRADE_COLORS,
        pie_colors=PIE_COLORS,
        tier_order=TIER_ORDER,
        tier_fills=TIER_FILLS,
        sparkline_font=SPARKLINE_FONT,
        get_column_letter=get_column_letter,
        generate_narrative=_generate_narrative,
    )


# ═══════════════════════════════════════════════════════════════════════
# Sheet 2: All Repos (Master Table)
# ═══════════════════════════════════════════════════════════════════════


def _build_all_repos(
    wb: Workbook,
    data: dict,
    score_history: dict[str, list[float]] | None = None,
    *,
    risk_lookup: dict[str, str] | None = None,
) -> None:
    _build_all_repos_workbook_sheet(
        wb,
        data,
        score_history,
        risk_lookup=risk_lookup,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        headers=ALL_REPOS_HEADERS,
        configure_all_repos_sheet=_configure_all_repos_sheet,
        style_header_row=style_header_row,
        build_all_repo_rows=_build_all_repo_rows,
        render_sparkline=render_sparkline,
        write_all_repo_rows=_write_all_repo_rows,
        style_data_cell=style_data_cell,
        color_grade_cell=color_grade_cell,
        color_tier_cell=color_tier_cell,
        color_pattern_cell=color_pattern_cell,
        sparkline_font=SPARKLINE_FONT,
        link_color=TEAL,
        apply_all_repos_postprocessing=_apply_all_repos_postprocessing,
        data_bar_rule_factory=DataBarRule,
        icon_set_rule_factory=IconSetRule,
        subheader_font=SUBHEADER_FONT,
        finalize_all_repos_layout=_finalize_all_repos_layout,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        auto_width=auto_width,
    )


# ═══════════════════════════════════════════════════════════════════════
# Sheet 3: Scoring Heatmap
# ═══════════════════════════════════════════════════════════════════════


def _build_heatmap(wb: Workbook, data: dict) -> None:
    _build_heatmap_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        build_heatmap_rows_fn=_build_heatmap_rows,
        write_heatmap_table_fn=_write_heatmap_table,
        get_column_letter=get_column_letter,
        color_scale_rule=ColorScaleRule,
        heatmap_headers=list(HEATMAP_HEADERS),
        heatmap_red=HEATMAP_RED,
        heatmap_amber=HEATMAP_AMBER,
        heatmap_green=HEATMAP_GREEN,
        center_alignment=CENTER,
        thin_border=THIN_BORDER,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        color_grade_cell=color_grade_cell,
        auto_width=auto_width,
    )


# ═══════════════════════════════════════════════════════════════════════
# Sheet 4: Quick Wins
# ═══════════════════════════════════════════════════════════════════════


def _build_quick_wins(wb: Workbook, data: dict) -> None:
    _build_quick_wins_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        build_quick_wins_rows_fn=_build_quick_wins_rows,
        write_quick_wins_sheet_fn=_write_quick_wins_sheet,
        data_bar_rule=DataBarRule,
        section_font=SECTION_FONT,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        color_grade_cell=color_grade_cell,
        color_tier_cell=color_tier_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        auto_width=auto_width,
    )


# ═══════════════════════════════════════════════════════════════════════
# Sheet 5: Badges Dashboard
# ═══════════════════════════════════════════════════════════════════════


def _build_badges(wb: Workbook, data: dict) -> None:
    _build_badges_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        build_badges_sheet_content_fn=_build_badges_sheet_content,
        write_badges_sheet_fn=_write_badges_sheet,
        section_font=SECTION_FONT,
        subtitle_font=SUBTITLE_FONT,
        subheader_font=SUBHEADER_FONT,
        subheader_fill=SUBHEADER_FILL,
        border=THIN_BORDER,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        build_badge_distribution_chart=_build_badge_distribution_chart,
        auto_width=auto_width,
    )


# ═══════════════════════════════════════════════════════════════════════
# Sheet 6: Tech Stack
# ═══════════════════════════════════════════════════════════════════════


def _build_tech_stack(wb: Workbook, data: dict) -> None:
    _build_tech_stack_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        auto_width=auto_width,
        section_font=SECTION_FONT,
        thin_border=THIN_BORDER,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
    )


# ═══════════════════════════════════════════════════════════════════════
# Sheet 7: Trends
# ═══════════════════════════════════════════════════════════════════════


def _build_trends(wb: Workbook, data: dict, trend_data: list[dict] | None = None) -> None:
    ws = _get_or_create_sheet(wb, "Trends")
    ws.sheet_properties.tabColor = "311B92"
    _configure_sheet_view(ws, zoom=110, show_grid_lines=True)
    final_row = _write_trends_sheet(
        ws,
        trend_data or [],
        tier_order=TIER_ORDER,
        section_font=SECTION_FONT,
        subtitle_font=SUBTITLE_FONT,
        subheader_font=SUBHEADER_FONT,
        thin_border=THIN_BORDER,
    )
    auto_width(ws, len((trend_data or [])) + 1, max(final_row, 15))


# ═══════════════════════════════════════════════════════════════════════
# Sheet 8: Tier Breakdown
# ═══════════════════════════════════════════════════════════════════════


def _build_tier_breakdown(wb: Workbook, data: dict) -> None:
    from openpyxl.styles import Font as XFont

    from src.excel_styles import TIER_FILLS

    _build_tier_breakdown_sheet(
        wb,
        data,
        tier_order=TIER_ORDER,
        header_font_factory=XFont,
        white_color=WHITE,
        center_alignment=CENTER,
        tier_fills=TIER_FILLS,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        auto_width=auto_width,
        build_tier_breakdown_sections_fn=_build_tier_breakdown_sections,
        write_tier_breakdown_sections_fn=_write_tier_breakdown_sections,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        color_grade_cell=color_grade_cell,
    )


# ═══════════════════════════════════════════════════════════════════════
# Sheet 9: Activity
# ═══════════════════════════════════════════════════════════════════════


def _build_activity(wb: Workbook, data: dict) -> None:
    _build_activity_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        build_activity_rows_fn=_build_activity_rows,
        write_activity_table_fn=_write_activity_table,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        color_pattern_cell=color_pattern_cell,
        auto_width=auto_width,
    )


# ═══════════════════════════════════════════════════════════════════════
# Sheet 10: Registry Reconciliation
# ═══════════════════════════════════════════════════════════════════════


def _build_reconciliation(wb: Workbook, data: dict) -> None:
    _build_registry_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        build_registry_content_fn=_build_registry_content,
        write_registry_sections_fn=_write_registry_sections,
        section_font=SECTION_FONT,
        subheader_font=SUBHEADER_FONT,
        thin_border=THIN_BORDER,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        color_tier_cell=color_tier_cell,
        auto_width=auto_width,
    )


# ═══════════════════════════════════════════════════════════════════════
# Main export
# ═══════════════════════════════════════════════════════════════════════


def _build_score_explainer(wb: Workbook) -> None:
    """Static reference sheet explaining the scoring system."""
    from src.scorer import COMPLETENESS_TIERS, GRADE_THRESHOLDS, WEIGHTS

    _build_score_explainer_sheet(
        wb,
        weights=WEIGHTS,
        grade_thresholds=GRADE_THRESHOLDS,
        completeness_tiers=COMPLETENESS_TIERS,
        get_or_create_sheet=_get_or_create_sheet,
        build_score_explainer_content_fn=_build_score_explainer_content,
        write_score_explainer_sections_fn=_write_score_explainer_sections,
        title_font=TITLE_FONT,
        section_font=SECTION_FONT,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        color_grade_cell=color_grade_cell,
        color_tier_cell=color_tier_cell,
        auto_width=auto_width,
    )


def _build_action_items(wb: Workbook, data: dict) -> None:
    """Prioritized action item list with weekly sprint."""
    ws = _get_or_create_sheet(wb, "Action Items")
    ws.sheet_properties.tabColor = "E65100"
    _configure_sheet_view(ws, zoom=110, show_grid_lines=True)

    actions = _collect_action_items(data)
    content = _build_action_items_content(actions)

    ws.merge_cells("A1:F1")
    ws["A1"].value = f"Action Items — {len(actions)} prioritized improvements"
    ws["A1"].font = SECTION_FONT
    final_row = _write_action_items_sections(
        ws,
        content,
        section_font=SECTION_FONT,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        set_autofilter=_set_autofilter,
    )
    auto_width(ws, len(ACTION_ITEMS_HEADERS), final_row + 1)


# ═══════════════════════════════════════════════════════════════════════
# Repo Profiles (Radar Charts)
# ═══════════════════════════════════════════════════════════════════════


def _build_repo_profiles(wb: Workbook, data: dict) -> None:
    ws = _get_or_create_sheet(wb, "Repo Profiles")
    ws.sheet_properties.tabColor = "7C3AED"
    _configure_sheet_view(ws, zoom=105, show_grid_lines=True)
    ws.freeze_panes = "B2"

    audits, matrix_rows = _build_repo_profile_matrix(data.get("audits", []))
    if len(audits) < 2:
        return

    _write_repo_profile_matrix(ws, audits, matrix_rows)
    _add_repo_profile_charts(ws, audits)


# ═══════════════════════════════════════════════════════════════════════
# Security Sheet
# ═══════════════════════════════════════════════════════════════════════


def _build_security(wb: Workbook, data: dict) -> None:
    _build_security_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        build_security_sheet_rows_fn=_build_security_sheet_rows,
        write_security_table_fn=_write_security_table,
        security_headers=list(SECURITY_HEADERS),
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        color_scale_rule=ColorScaleRule,
        icon_set_rule=IconSetRule,
        heatmap_red=HEATMAP_RED,
        heatmap_amber=HEATMAP_AMBER,
        heatmap_green=HEATMAP_GREEN,
        auto_width=auto_width,
    )


# ═══════════════════════════════════════════════════════════════════════
# Changes Since Last Audit (Diff Sheet)
# ═══════════════════════════════════════════════════════════════════════


def _build_changes(wb: Workbook, data: dict, diff_data: dict | None) -> None:
    _build_changes_sheet(
        wb,
        data,
        diff_data,
        get_or_create_sheet=_get_or_create_sheet,
        auto_width=auto_width,
        title_font=TITLE_FONT,
        subheader_font=SUBHEADER_FONT,
        section_font=SECTION_FONT,
        build_changes_sheet_content_fn=_build_changes_sheet_content,
        build_changes_sheet_sections_fn=_build_changes_sheet_sections,
        write_changes_sheet_sections_fn=_write_changes_sheet_sections,
        style_header_row=style_header_row,
    )


# ═══════════════════════════════════════════════════════════════════════
# Bubble Chart helper for Dashboard
# ═══════════════════════════════════════════════════════════════════════


def _build_bubble_on_dashboard(ws, data: dict) -> None:
    _build_dashboard_bubble_chart(
        ws,
        data,
        bubble_chart_cls=BubbleChart,
        reference_cls=Reference,
    )


# ═══════════════════════════════════════════════════════════════════════
# Dependency Graph Sheet
# ═══════════════════════════════════════════════════════════════════════


def _build_dependency_graph(wb: Workbook, data: dict) -> None:
    from src.dep_graph import build_dependency_graph

    _build_dependency_graph_sheet(
        wb,
        data,
        build_dependency_graph_fn=build_dependency_graph,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        color_scale_rule=ColorScaleRule,
        heatmap_amber=HEATMAP_AMBER,
        heatmap_green=HEATMAP_GREEN,
        auto_width=auto_width,
        section_font=SECTION_FONT,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
    )


def _build_hotspots(wb: Workbook, data: dict) -> None:
    _build_hotspots_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        build_hotspot_rows_fn=_build_hotspot_rows,
        write_hotspot_table_fn=_write_hotspot_table,
        data_bar_rule=DataBarRule,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_implementation_hotspots(wb: Workbook, data: dict) -> None:
    _build_implementation_hotspots_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        section_font=SECTION_FONT,
        wrap_alignment=WRAP,
        build_implementation_hotspot_rows_fn=_build_implementation_hotspot_rows,
        write_hotspot_table_fn=_write_hotspot_table,
        implementation_hotspots_headers=list(IMPLEMENTATION_HOTSPOTS_HEADERS),
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        data_bar_rule=DataBarRule,
        auto_width=auto_width,
    )


def _build_operator_outcomes(wb: Workbook, data: dict) -> None:
    ws = _get_or_create_sheet(wb, "Operator Outcomes")
    ws.sheet_properties.tabColor = "0369A1"
    _configure_sheet_view(ws, zoom=105, show_grid_lines=True)
    content = _build_operator_outcomes_content(
        data,
        join_outcome_examples=_join_outcome_examples,
    )
    max_row = _write_operator_outcomes_sections(
        ws,
        content,
        section_font=SECTION_FONT,
        wrap_alignment=WRAP,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
    )
    auto_width(ws, 6, max_row)


def _build_approval_ledger(wb: Workbook, data: dict) -> None:
    _build_approval_ledger_sheet(
        wb,
        data,
        approval_ledger_label=ACTION_SYNC_CANONICAL_LABELS["approval_ledger"],
        approved_but_manual_label=ACTION_SYNC_CANONICAL_LABELS["approved_but_manual"],
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        auto_width=auto_width,
        section_font=SECTION_FONT,
        subtitle_font=SUBTITLE_FONT,
        wrap_alignment=WRAP,
        build_approval_ledger_content_fn=_build_approval_ledger_content,
        write_approval_ledger_sections_fn=_write_approval_ledger_sections,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
    )


def _build_historical_intelligence(wb: Workbook, data: dict) -> None:
    ws = _get_or_create_sheet(wb, "Historical Intelligence")
    ws.sheet_properties.tabColor = "0F766E"
    _configure_sheet_view(ws, zoom=105, show_grid_lines=True)

    content = _build_historical_intelligence_content(data)
    row = _write_historical_intelligence_sheet(
        ws,
        content,
        title=ACTION_SYNC_CANONICAL_LABELS["historical_portfolio_intelligence"],
        section_font=SECTION_FONT,
        subtitle_font=SUBTITLE_FONT,
        wrap_alignment=WRAP,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
    )

    auto_width(ws, len(HISTORICAL_INTELLIGENCE_HEADERS), max(row - 1, 5))


def _join_outcome_examples(items: list[dict]) -> str:
    if not items:
        return "No recent examples are recorded yet."
    labels = []
    for item in items[:3]:
        repo = str(item.get("repo") or "").strip()
        title = str(item.get("title") or item.get("action_id") or "Operator outcome").strip()
        labels.append(f"{repo}: {title}" if repo else title)
    return "; ".join(labels)


def _build_hidden_data_sheets(
    wb: Workbook,
    data: dict,
    trend_data: list[dict] | None = None,
    score_history: dict[str, list[float]] | None = None,
    diff_data: dict | None = None,
) -> None:
    _build_hidden_data_sheets_helper(
        wb,
        data=data,
        trend_data=trend_data,
        score_history=score_history,
        diff_data=diff_data,
        build_hidden_data_payload_fn=_build_hidden_data_payload,
        write_hidden_data_tables=_write_hidden_data_tables,
        build_repo_detail_rows=_repo_detail_rows,
        build_run_change_rows=_run_change_rows,
        build_workbook_rollups=_build_workbook_rollups,
        build_operator_hidden_rows=_build_operator_hidden_rows,
        build_core_hidden_rows=_build_core_hidden_rows,
        trend_history_window=TREND_HISTORY_WINDOW,
        tier_order=TIER_ORDER,
    )


def _build_security_controls(wb: Workbook, data: dict) -> None:
    _build_security_controls_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        build_security_controls_rows_fn=_build_security_controls_rows,
        write_security_table_fn=_write_security_table,
        security_controls_headers=list(SECURITY_CONTROLS_HEADERS),
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_supply_chain(wb: Workbook, data: dict) -> None:
    _build_supply_chain_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        build_supply_chain_rows_fn=_build_supply_chain_rows,
        write_security_table_fn=_write_security_table,
        supply_chain_headers=list(SUPPLY_CHAIN_HEADERS),
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_security_debt(wb: Workbook, data: dict) -> None:
    _build_security_debt_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        build_security_debt_rows_fn=_build_security_debt_rows,
        write_security_table_fn=_write_security_table,
        security_debt_headers=list(SECURITY_DEBT_HEADERS),
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_campaigns(wb: Workbook, data: dict) -> None:
    _build_campaigns_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        build_campaign_sheet_content_fn=_build_campaign_sheet_content,
        write_key_value_block=_write_key_value_block,
        action_sync_readiness_rows=_action_sync_readiness_rows,
        action_sync_readiness_title=ACTION_SYNC_CANONICAL_LABELS["readiness"],
        write_campaign_table_fn=_write_campaign_table,
        style_header_row=style_header_row,
        add_table=_add_table,
        apply_zebra_stripes=apply_zebra_stripes,
        wrap_alignment=WRAP,
        subtitle_font=SUBTITLE_FONT,
        auto_width=auto_width,
    )


def _build_writeback_audit(wb: Workbook, data: dict) -> None:
    _build_writeback_audit_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        build_writeback_audit_content_fn=_build_writeback_audit_content,
        write_key_value_block=_write_key_value_block,
        action_sync_readiness_rows=_action_sync_readiness_rows,
        action_sync_readiness_title=ACTION_SYNC_CANONICAL_LABELS["readiness"],
        write_writeback_audit_table_fn=_write_writeback_audit_table,
        style_header_row=style_header_row,
        add_table=_add_table,
        apply_zebra_stripes=apply_zebra_stripes,
        wrap_alignment=WRAP,
        subtitle_font=SUBTITLE_FONT,
        auto_width=auto_width,
        serialize_details=json.dumps,
    )


def _build_portfolio_explorer(
    wb: Workbook,
    data: dict,
    *,
    portfolio_profile: str = "default",
    collection: str | None = None,
) -> None:
    from src.analyst_views import build_analyst_context

    _build_portfolio_explorer_sheet(
        wb,
        data,
        portfolio_profile=portfolio_profile,
        collection=collection,
        build_analyst_context=build_analyst_context,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        subtitle_font=SUBTITLE_FONT,
        wrap_alignment=WRAP,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        auto_width=auto_width,
        build_maturity_gap_summary=build_maturity_gap_summary,
    )


def _build_portfolio_catalog_sheet(wb: Workbook, data: dict) -> None:
    _build_portfolio_catalog_sheet_helper(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        build_portfolio_catalog_summary=build_portfolio_catalog_summary,
        build_portfolio_intent_alignment_summary=build_portfolio_intent_alignment_summary,
        build_operating_paths_summary=build_operating_paths_summary,
        subtitle_font=SUBTITLE_FONT,
        wrap_alignment=WRAP,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_scorecards_sheet(wb: Workbook, data: dict) -> None:
    _build_scorecards_sheet_helper(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        build_scorecards_summary=build_scorecards_summary,
        subtitle_font=SUBTITLE_FONT,
        wrap_alignment=WRAP,
        build_scorecard_table_rows_fn=_build_scorecard_table_rows,
        write_portfolio_table=_write_portfolio_table,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_by_lens(
    wb: Workbook,
    data: dict,
    *,
    portfolio_profile: str = "default",
    collection: str | None = None,
) -> None:
    from src.analyst_views import build_analyst_context

    _build_by_lens_sheet(
        wb,
        data,
        portfolio_profile=portfolio_profile,
        collection=collection,
        build_analyst_context=build_analyst_context,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        subtitle_font=SUBTITLE_FONT,
        wrap_alignment=WRAP,
        build_by_lens_rows_fn=_build_by_lens_rows,
        write_portfolio_table=_write_portfolio_table,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_by_collection(
    wb: Workbook,
    data: dict,
    *,
    portfolio_profile: str = "default",
) -> None:
    from src.analyst_views import build_analyst_context

    _build_by_collection_sheet(
        wb,
        data,
        portfolio_profile=portfolio_profile,
        build_analyst_context=build_analyst_context,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        subtitle_font=SUBTITLE_FONT,
        wrap_alignment=WRAP,
        build_by_collection_rows_fn=_build_by_collection_rows,
        write_portfolio_table=_write_portfolio_table,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_trend_summary(
    wb: Workbook,
    data: dict,
    trend_data: list[dict] | None = None,
    score_history: dict[str, list[float]] | None = None,
) -> None:
    _build_trend_summary_sheet(
        wb,
        data,
        trend_data=trend_data,
        score_history=score_history,
        extend_portfolio_trend_with_current=_extend_portfolio_trend_with_current,
        extend_score_history_with_current=_extend_score_history_with_current,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        subtitle_font=SUBTITLE_FONT,
        wrap_alignment=WRAP,
        build_trend_summary_content_fn=_build_trend_summary_content,
        write_trend_summary_sections_fn=_write_trend_summary_sections,
        render_sparkline=render_sparkline,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        section_font=SECTION_FONT,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_repo_detail(wb: Workbook, data: dict) -> None:
    _build_repo_detail_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        write_instruction_banner=_write_instruction_banner,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        auto_width=auto_width,
        subheader_font=SUBHEADER_FONT,
        section_font=SECTION_FONT,
        wrap_alignment=WRAP,
        teal=TEAL,
    )


def _build_run_changes(wb: Workbook, data: dict, diff_data: dict | None) -> None:
    _build_run_changes_sheet(
        wb,
        data,
        diff_data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        write_instruction_banner=_write_instruction_banner,
        build_run_changes_content_fn=_build_run_changes_content,
        normalize_run_changes_sections_fn=_normalize_run_changes_sections,
        write_run_changes_summary_fn=_write_run_changes_summary,
        write_run_changes_sections_fn=_write_run_changes_sections,
        subtitle_font=SUBTITLE_FONT,
        subheader_font=SUBHEADER_FONT,
        wrap_alignment=WRAP,
        write_kpi_card=write_kpi_card,
        section_font=SECTION_FONT,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        auto_width=auto_width,
    )


def _build_review_queue(wb: Workbook, data: dict, *, excel_mode: str = "standard") -> None:
    _build_review_queue_workbook_sheet(
        wb,
        data,
        excel_mode=excel_mode,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        write_instruction_banner=_write_instruction_banner,
        build_review_queue_sheet_content=_build_review_queue_sheet_content,
        build_workbook_rollups=_build_workbook_rollups,
        operator_counts=_operator_counts,
        ordered_queue_items=_ordered_queue_items,
        summarize_top_issue_families=_summarize_top_issue_families,
        summarize_top_actions=_summarize_top_actions,
        build_queue_pressure_summary=build_queue_pressure_summary,
        build_top_recommendation_summary=build_top_recommendation_summary,
        build_weekly_review_pack=build_weekly_review_pack,
        build_executive_operator_context=_build_executive_operator_context,
        format_lane_counts=_format_lane_counts,
        format_repo_rollup_counts=_format_repo_rollup_counts,
        build_review_queue_table_rows=_build_review_queue_table_rows,
        write_review_queue_overview=_write_review_queue_overview,
        write_review_queue_table=_write_review_queue_table,
        write_key_value_block=_write_key_value_block,
        write_ranked_list=_write_ranked_list,
        guidance_text=REVIEW_QUEUE_GUIDANCE,
        review_queue_headers=list(REVIEW_QUEUE_HEADERS),
        review_queue_center_aligned_columns=set(REVIEW_QUEUE_CENTER_ALIGNED_COLUMNS),
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_lane_row_fill=_apply_lane_row_fill,
        apply_zebra_stripes=apply_zebra_stripes,
        set_autofilter=_set_autofilter,
        auto_width=auto_width,
        subtitle_font=SUBTITLE_FONT,
        wrap_alignment=WRAP,
        section_font=SECTION_FONT,
        teal=TEAL,
    )


def _build_review_history_sheet(wb: Workbook, data: dict) -> None:
    _build_review_history_sheet_helper(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        write_key_value_block=_write_key_value_block,
        build_review_history_rows_fn=_build_review_history_rows,
        write_review_history_table_fn=_write_review_history_table,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_governance_controls(wb: Workbook, data: dict) -> None:
    _build_governance_controls_sheet(
        wb,
        data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        build_governance_controls_content_fn=_build_governance_controls_content,
        display_operator_state=_display_operator_state,
        approval_workflow_label=ACTION_SYNC_CANONICAL_LABELS["approval_workflow"],
        next_approval_review_label=ACTION_SYNC_CANONICAL_LABELS["next_approval_review"],
        write_key_value_block=_write_key_value_block,
        write_governance_controls_table_fn=_write_governance_controls_table,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        add_table=_add_table,
        apply_zebra_stripes=apply_zebra_stripes,
        wrap_alignment=WRAP,
        subtitle_font=SUBTITLE_FONT,
        auto_width=auto_width,
    )


def _build_governance_audit(wb: Workbook, data: dict) -> None:
    _build_governance_audit_sheet(
        wb,
        data,
        display_operator_state=_display_operator_state,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        build_governance_audit_rows_fn=_build_governance_audit_rows,
        write_governance_audit_table_fn=_write_governance_audit_table,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        apply_zebra_stripes=apply_zebra_stripes,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_compare_sheet(
    wb: Workbook,
    diff_data: dict | None,
) -> None:
    _build_compare_sheet_helper(
        wb,
        diff_data,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        title_font=TITLE_FONT,
        build_compare_lens_rows_fn=_build_compare_lens_rows,
        build_compare_repo_rows_fn=_build_compare_repo_rows,
        write_compare_sections_fn=_write_compare_sections,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_scenario_planner(
    wb: Workbook,
    data: dict,
    *,
    portfolio_profile: str = "default",
    collection: str | None = None,
) -> None:
    from src.analyst_views import build_analyst_context

    _build_scenario_planner_sheet(
        wb,
        data,
        portfolio_profile=portfolio_profile,
        collection=collection,
        build_analyst_context=build_analyst_context,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        title_font=TITLE_FONT,
        subheader_font=SUBHEADER_FONT,
        build_scenario_planner_rows_fn=_build_scenario_planner_rows,
        write_scenario_planner_sections_fn=_write_scenario_planner_sections,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        add_table=_add_table,
        auto_width=auto_width,
    )


def _build_executive_summary(
    wb: Workbook,
    data: dict,
    diff_data: dict | None,
    *,
    portfolio_profile: str = "default",
    collection: str | None = None,
    excel_mode: str = "standard",
) -> None:
    _build_executive_summary_workbook_sheet(
        wb,
        data,
        diff_data,
        portfolio_profile=portfolio_profile,
        collection=collection,
        excel_mode=excel_mode,
        build_weekly_review_pack=build_weekly_review_pack,
        build_run_change_counts=build_run_change_counts,
        build_run_change_summary=build_run_change_summary,
        build_queue_pressure_summary=build_queue_pressure_summary,
        build_trust_actionability_summary=build_trust_actionability_summary,
        build_top_recommendation_summary=build_top_recommendation_summary,
        resolve_weekly_story_value=resolve_weekly_story_value,
        build_workbook_rollups=_build_workbook_rollups,
        format_lane_counts=_format_lane_counts,
        operator_counts=_operator_counts,
        build_portfolio_catalog_summary=build_portfolio_catalog_summary,
        build_operating_paths_summary=build_operating_paths_summary,
        build_portfolio_intent_alignment_summary=build_portfolio_intent_alignment_summary,
        build_scorecards_summary=build_scorecards_summary,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        write_instruction_banner=_write_instruction_banner,
        write_key_value_block=_write_key_value_block,
        write_ranked_list=_write_ranked_list,
        write_kpi_card=write_kpi_card,
        auto_width=auto_width,
        navy=NAVY,
        section_font=SECTION_FONT,
        subheader_font=SUBHEADER_FONT,
    )


def _build_print_pack(
    wb: Workbook,
    data: dict,
    diff_data: dict | None,
    *,
    portfolio_profile: str = "default",
    collection: str | None = None,
    excel_mode: str = "standard",
) -> None:
    _build_print_pack_workbook_sheet(
        wb,
        data,
        diff_data,
        portfolio_profile=portfolio_profile,
        collection=collection,
        excel_mode=excel_mode,
        get_or_create_sheet=_get_or_create_sheet,
        configure_sheet_view=_configure_sheet_view,
        set_sheet_header=_set_sheet_header,
        build_print_pack_sheet_content_fn=_build_print_pack_sheet_content,
        write_print_pack_sheet_fn=_write_print_pack_sheet,
        build_weekly_review_pack=build_weekly_review_pack,
        build_executive_operator_context=_build_executive_operator_context,
        print_pack_workflow_guidance_rows=_print_pack_workflow_guidance_rows,
        build_run_change_summary=build_run_change_summary,
        build_queue_pressure_summary=build_queue_pressure_summary,
        build_top_recommendation_summary=build_top_recommendation_summary,
        build_trust_actionability_summary=build_trust_actionability_summary,
        display_operator_state=_display_operator_state,
        summarize_top_issue_families=_summarize_top_issue_families,
        section_font=SECTION_FONT,
        subheader_font=SUBHEADER_FONT,
        auto_width=auto_width,
    )


def _build_template_sparkline_specs(
    data: dict,
    *,
    trend_data: list[dict] | None = None,
    score_history: dict[str, list[float]] | None = None,
) -> list[SparklineSpec]:
    return _build_template_sparkline_specs_helper(
        data,
        trend_data=trend_data,
        score_history=score_history,
        extend_score_history_with_current=_extend_score_history_with_current,
        extend_portfolio_trend_with_current=_extend_portfolio_trend_with_current,
        sparkline_spec_cls=SparklineSpec,
        trend_history_window=TREND_HISTORY_WINDOW,
    )


def _build_excel_workbook(
    wb: Workbook,
    data: dict,
    *,
    trend_data: list[dict] | None = None,
    diff_data: dict | None = None,
    score_history: dict[str, list[float]] | None = None,
    portfolio_profile: str = "default",
    collection: str | None = None,
    excel_mode: str = "standard",
    risk_lookup: dict[str, str] | None = None,
) -> None:
    _build_excel_workbook_runner(
        wb,
        data,
        trend_data=trend_data,
        diff_data=diff_data,
        score_history=score_history,
        portfolio_profile=portfolio_profile,
        collection=collection,
        excel_mode=excel_mode,
        risk_lookup=risk_lookup,
        **_build_excel_workbook_runtime(
            run_workbook_build_steps=_run_workbook_build_steps,
            build_dashboard=_build_dashboard,
            build_all_repos=_build_all_repos,
            build_portfolio_explorer=_build_portfolio_explorer,
            build_portfolio_catalog_sheet=_build_portfolio_catalog_sheet,
            build_scorecards_sheet=_build_scorecards_sheet,
            build_repo_detail=_build_repo_detail,
            build_by_lens=_build_by_lens,
            build_by_collection=_build_by_collection,
            build_trend_summary=_build_trend_summary,
            build_run_changes=_build_run_changes,
            build_heatmap=_build_heatmap,
            build_quick_wins=_build_quick_wins,
            build_badges=_build_badges,
            build_tech_stack=_build_tech_stack,
            build_trends=_build_trends,
            build_tier_breakdown=_build_tier_breakdown,
            build_activity=_build_activity,
            build_repo_profiles=_build_repo_profiles,
            build_security=_build_security,
            build_security_controls=_build_security_controls,
            build_supply_chain=_build_supply_chain,
            build_security_debt=_build_security_debt,
            build_campaigns=_build_campaigns,
            build_writeback_audit=_build_writeback_audit,
            build_governance_controls=_build_governance_controls,
            build_governance_audit=_build_governance_audit,
            build_review_queue=_build_review_queue,
            build_review_history_sheet=_build_review_history_sheet,
            build_hotspots=_build_hotspots,
            build_implementation_hotspots=_build_implementation_hotspots,
            build_operator_outcomes=_build_operator_outcomes,
            build_approval_ledger=_build_approval_ledger,
            build_historical_intelligence=_build_historical_intelligence,
            build_compare_sheet=_build_compare_sheet,
            build_scenario_planner=_build_scenario_planner,
            build_executive_summary=_build_executive_summary,
            build_print_pack=_build_print_pack,
            build_changes=_build_changes,
            build_reconciliation=_build_reconciliation,
            build_dependency_graph=_build_dependency_graph,
            build_score_explainer=_build_score_explainer,
            build_action_items=_build_action_items,
            build_hidden_data_sheets=_build_hidden_data_sheets,
            build_navigation=_build_navigation,
            inject_sheet_navigation=_inject_sheet_navigation,
            apply_workbook_named_ranges=_apply_workbook_named_ranges,
            finalize_workbook_structure=_finalize_workbook_structure,
            core_visible_sheets=CORE_VISIBLE_SHEETS,
            template_info_sheet=TEMPLATE_INFO_SHEET,
            preferred_order=list(DEFAULT_PREFERRED_SHEET_ORDER),
        ),
    )


def _build_risk_summary_sheet(wb: Workbook, risk_posture: dict) -> None:
    ws = _get_or_create_sheet(wb, "Risk Summary")
    ws.sheet_properties.tabColor = "DC2626"
    _configure_sheet_view(ws, zoom=115, show_grid_lines=False)
    _set_sheet_header(
        ws,
        "Risk Summary",
        "Portfolio risk tier distribution sourced from portfolio-truth-latest.json.",
        width=4,
    )
    start_row = 3
    max_row = _write_risk_summary_table(
        ws,
        _build_risk_summary_rows(risk_posture),
        start_row=start_row,
        style_header_row=style_header_row,
        add_table=_add_table,
    )
    auto_width(ws, len(RISK_SUMMARY_HEADERS), max_row)


def export_excel(
    report_path: Path,
    output_path: Path,
    trend_data: list[dict] | None = None,
    diff_data: dict | None = None,
    score_history: dict[str, list[float]] | None = None,
    portfolio_profile: str = "default",
    collection: str | None = None,
    excel_mode: str = "standard",
    template_path: Path | None = None,
    truth_dir: Path | None = None,
) -> Path:
    """Generate the flagship Excel dashboard."""
    return _export_excel_workbook(
        report_path,
        output_path,
        trend_data=trend_data,
        diff_data=diff_data,
        score_history=score_history,
        portfolio_profile=portfolio_profile,
        collection=collection,
        excel_mode=excel_mode,
        template_path=template_path,
        truth_dir=truth_dir,
        load_risk_truth=_load_risk_truth,
        default_template_path=DEFAULT_TEMPLATE_PATH,
        resolve_template_path=resolve_template_path,
        copy_template_to_output=copy_template_to_output,
        load_workbook_allowing_native_sparklines=load_workbook_allowing_native_sparklines,
        build_excel_workbook=_build_excel_workbook,
        build_risk_summary_sheet=_build_risk_summary_sheet,
        inject_native_sparklines=inject_native_sparklines,
        build_template_sparkline_specs=_build_template_sparkline_specs,
    )
