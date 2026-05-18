"""Design system for the flagship Excel dashboard.

Centralized colors, fonts, named styles, and helper functions.
All sheets import from here for consistent visual language.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.worksheet.hyperlink import Hyperlink

# ── Named Style Registry ──────────────────────────────────────────────
# We need an idempotency marker per Workbook so we don't register the same
# NamedStyle twice (openpyxl raises on duplicate names). We previously used
# `dict[id(wb), ...]` but Python recycles id() values after GC, which caused
# a freshly-allocated Workbook to inherit a stale "already registered" entry
# from a GC'd predecessor and then fail at cell.style = "data_left".
#
# Storing the marker as a private attribute on the Workbook itself ties the
# lifetime of the marker to the lifetime of the workbook — no recycling.

_REGISTERED_ATTR = "_gha_named_styles_registered"

# ── Color Palette ────────────────────────────────────────────────────

NAVY = "1B2A4A"
TEAL = "0EA5E9"
SLATE = "64748B"
LIGHT_BG = "F4F7FB"
WHITE = "FFFFFF"
LIGHT_BORDER = "E2E8F0"
MED_BORDER = "CBD5E1"

# Tier colors
TIER_COLORS = {
    "shipped": "166534",
    "functional": "1565C0",
    "wip": "D97706",
    "skeleton": "C2410C",
    "abandoned": "6B7280",
}

# Grade colors
GRADE_COLORS = {
    "A": "166534",
    "B": "15803D",
    "C": "CA8A04",
    "D": "C2410C",
    "F": "991B1B",
}

# Interest tier colors
INTEREST_COLORS = {
    "flagship": "7C3AED",
    "notable": "2563EB",
    "standard": "6B7280",
    "mundane": "9CA3AF",
}

# Commit pattern colors
PATTERN_COLORS = {
    "steady": "166534",
    "burst": "1E40AF",
    "new": "0891B2",
    "seasonal": "7C3AED",
    "winding-down": "CA8A04",
    "dormant": "991B1B",
    "unknown": "9CA3AF",
}

# Heatmap gradient (for ColorScaleRule)
HEATMAP_RED = "FFCDD2"
HEATMAP_AMBER = "FFF9C4"
HEATMAP_GREEN = "C8E6C9"

# ── Fills ────────────────────────────────────────────────────────────

TIER_FILLS = {k: PatternFill("solid", fgColor=v) for k, v in TIER_COLORS.items()}
GRADE_FILLS = {k: PatternFill("solid", fgColor=v) for k, v in GRADE_COLORS.items()}
PATTERN_FILLS = {k: PatternFill("solid", fgColor=v) for k, v in PATTERN_COLORS.items()}

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SUBHEADER_FILL = PatternFill("solid", fgColor="EAF1F8")
KPI_CARD_FILL = PatternFill("solid", fgColor=LIGHT_BG)
ZEBRA_FILL = PatternFill("solid", fgColor="F7FAFD")

# ── Fonts ────────────────────────────────────────────────────────────

TITLE_FONT = Font("Calibri", 22, bold=True, color=NAVY)
SUBTITLE_FONT = Font("Calibri", 13, color=SLATE)
HEADER_FONT = Font("Calibri", 11, bold=True, color=WHITE)
SUBHEADER_FONT = Font("Calibri", 11, bold=True, color=NAVY)
KPI_NUMBER_FONT = Font("Calibri", 28, bold=True, color=NAVY)
KPI_LABEL_FONT = Font("Calibri", 11, color=SLATE)
DATA_FONT = Font("Calibri", 11)
DATA_BOLD_FONT = Font("Calibri", 11, bold=True)
TIER_FONT = Font("Calibri", 11, bold=True, color=WHITE)
SECTION_FONT = Font("Calibri", 15, bold=True, color=NAVY)
HIGHLIGHT_FONT = Font("Calibri", 11, bold=True, color=TEAL)
NARRATIVE_FONT = Font("Calibri", 12, italic=True, color=SLATE)
SPARKLINE_FONT = Font("Courier New", 11, color=TEAL)

# ── Borders ──────────────────────────────────────────────────────────

THIN_BORDER = Border(
    left=Side(style="thin", color=LIGHT_BORDER),
    right=Side(style="thin", color=LIGHT_BORDER),
    top=Side(style="thin", color=LIGHT_BORDER),
    bottom=Side(style="thin", color=LIGHT_BORDER),
)

CARD_BORDER = Border(
    left=Side(style="thin", color=MED_BORDER),
    right=Side(style="thin", color=MED_BORDER),
    top=Side(style="thin", color=MED_BORDER),
    bottom=Side(style="thin", color=MED_BORDER),
)

# ── Alignment ────────────────────────────────────────────────────────

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Helper Functions ─────────────────────────────────────────────────


def style_header_row(ws, row: int, max_col: int) -> None:
    """Apply dark navy header styling to a row."""
    ws.row_dimensions[row].height = 30
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _ensure_data_named_styles(wb) -> None:
    """Register data-cell NamedStyles on *wb* once per workbook instance.

    NamedStyle assignment (``cell.style = name``) bypasses openpyxl's
    per-attribute IndexedList hash chain, which is the dominant CPU hotspot
    when styling tens-of-thousands of data cells (see S2.0 profile findings).
    """
    if getattr(wb, _REGISTERED_ATTR, False):
        return

    _thin = Border(
        left=Side(style="thin", color=LIGHT_BORDER),
        right=Side(style="thin", color=LIGHT_BORDER),
        top=Side(style="thin", color=LIGHT_BORDER),
        bottom=Side(style="thin", color=LIGHT_BORDER),
    )
    _data_font = Font("Calibri", 11)

    existing = set(wb.named_styles)
    for name, alignment in (
        ("data_left", Alignment(horizontal="left", vertical="center")),
        ("data_center", Alignment(horizontal="center", vertical="center")),
        ("data_right", Alignment(horizontal="right", vertical="center")),
    ):
        if name in existing:
            continue
        ns = NamedStyle(name)
        ns.font = _data_font
        ns.border = _thin
        ns.alignment = alignment
        wb.add_named_style(ns)

    setattr(wb, _REGISTERED_ATTR, True)


def style_data_cell(cell, align: str = "left") -> None:
    """Apply standard data cell styling via a cached NamedStyle.

    Uses NamedStyle on first call per workbook to avoid per-attribute
    hash-chain overhead in openpyxl's IndexedList (the dominant hotspot
    for large sheets with many styled cells).
    """
    wb = cell.parent.parent
    _ensure_data_named_styles(wb)
    if align == "left":
        cell.style = "data_left"
    elif align == "center":
        cell.style = "data_center"
    else:
        cell.style = "data_right"


def apply_zebra_stripes(
    ws, start_row: int, end_row: int, max_col: int, skip_cols: set[int] | None = None
) -> None:
    """Apply alternating row shading, optionally skipping columns with semantic coloring.

    No-ops on hidden sheets: zebra fill is purely cosmetic and the ZEBRA_FILL
    PatternFill hash chain adds measurable overhead at scale (see S2.0 profile).
    """
    if ws.sheet_state == "hidden":
        return
    skip = skip_cols or set()
    for row in range(start_row, end_row + 1):
        if row % 2 == 0:
            for col in range(1, max_col + 1):
                if col not in skip:
                    ws.cell(row=row, column=col).fill = ZEBRA_FILL


def auto_width(ws, max_col: int, max_row: int, min_width: int = 8, max_width: int = 55) -> None:
    """Auto-size columns based on content.

    No-ops on hidden sheets: column sizing is cosmetically irrelevant when the
    sheet is not visible, and scanning cell values adds measurable overhead at
    scale (see S2.0 profile findings).
    """
    if ws.sheet_state == "hidden":
        return
    from openpyxl.utils import get_column_letter

    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(1, min(max_row + 1, 150)):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = max(
            min_width, min(max_len + 3, max_width)
        )


def write_kpi_card(
    ws,
    row: int,
    col: int,
    label: str,
    value: str | int | float,
    color: str | None = None,
    hyperlink: str | None = None,
) -> None:
    """Write a 2-row KPI card at the given position."""
    # Label cell
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = KPI_LABEL_FONT
    label_cell.fill = KPI_CARD_FILL
    label_cell.border = CARD_BORDER
    label_cell.alignment = CENTER

    # Value cell
    value_cell = ws.cell(row=row + 1, column=col, value=value)
    value_cell.font = Font("Calibri", 28, bold=True, color=color or NAVY)
    value_cell.fill = KPI_CARD_FILL
    value_cell.border = CARD_BORDER
    value_cell.alignment = CENTER

    if hyperlink:
        if hyperlink.startswith("#"):
            value_cell.hyperlink = Hyperlink(
                ref=value_cell.coordinate,
                location=hyperlink[1:],
                display=str(value),
            )
        else:
            value_cell.hyperlink = hyperlink

    # Merge 2 cols wide for each card
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)


def color_grade_cell(cell, grade: str) -> None:
    """Apply grade-specific coloring to a cell."""
    if grade in GRADE_FILLS:
        cell.fill = GRADE_FILLS[grade]
        cell.font = Font("Calibri", 10, bold=True, color=WHITE)


def color_tier_cell(cell, tier: str) -> None:
    """Apply tier-specific coloring to a cell."""
    if tier in TIER_FILLS:
        cell.fill = TIER_FILLS[tier]
        cell.font = TIER_FONT


def color_pattern_cell(cell, pattern: str) -> None:
    """Apply commit pattern coloring to a cell."""
    if pattern in PATTERN_FILLS:
        cell.fill = PATTERN_FILLS[pattern]
        cell.font = Font("Calibri", 10, bold=True, color=WHITE)
