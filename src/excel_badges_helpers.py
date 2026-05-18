"""Helpers for Badges workbook content."""

from __future__ import annotations

from collections import Counter
from collections.abc import Callable
from typing import Any

from openpyxl.chart import BarChart, Reference

BADGE_DISTRIBUTION_HEADERS = ["Badge", "Repos Earned", "% of Portfolio"]


def build_badges_sheet_content(audits: list[dict[str, Any]]) -> dict[str, Any]:
    badge_counts: Counter[str] = Counter()
    repo_badge_counts: list[tuple[str, int]] = []

    for audit in audits:
        badges = audit.get("badges", [])
        for badge in badges:
            badge_counts[badge] += 1
        repo_badge_counts.append((audit["metadata"]["name"], len(badges)))

    total_repos = len(audits)
    distribution_rows = [
        {
            "badge": badge,
            "count": count,
            "portfolio_percent": f"{count / total_repos * 100:.0f}%" if total_repos else "0%",
        }
        for badge, count in badge_counts.most_common()
    ]

    return {
        "total_repos": total_repos,
        "total_badges": sum(badge_counts.values()),
        "distribution_rows": distribution_rows,
        "achievement_rows": sorted(repo_badge_counts, key=lambda item: item[1], reverse=True)[:20],
    }


def build_badge_distribution_chart(ws: Any, *, start_row: int, end_row: int) -> BarChart:
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Badge Distribution"
    chart.style = 10
    chart_data = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
    chart_categories = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
    chart.add_data(chart_data, titles_from_data=False)
    chart.set_categories(chart_categories)
    chart.series[0].graphicalProperties.solidFill = "7C3AED"
    chart.width = 20
    chart.height = max(8, (end_row - start_row + 1) * 0.6)
    return chart


def write_badge_distribution_rows(
    ws: Any,
    rows: list[dict[str, Any]],
    *,
    start_row: int,
    style_data_cell: Callable[[Any], None],
) -> int:
    for row_number, row in enumerate(rows, start_row):
        ws.cell(row=row_number, column=1, value=row["badge"])
        ws.cell(row=row_number, column=2, value=row["count"])
        ws.cell(row=row_number, column=3, value=row["portfolio_percent"])
        for col in range(1, len(BADGE_DISTRIBUTION_HEADERS) + 1):
            style_data_cell(ws.cell(row=row_number, column=col))

    return start_row + len(rows) - 1


def write_badge_achievement_board(
    ws: Any,
    achievement_rows: list[tuple[str, int]],
    *,
    start_row: int,
    subheader_font: Any,
    subheader_fill: Any,
    border: Any,
) -> int:
    ws.cell(row=start_row, column=1, value="Repo").font = subheader_font
    ws.cell(row=start_row, column=2, value="Badges").font = subheader_font
    ws.cell(row=start_row, column=1).fill = subheader_fill
    ws.cell(row=start_row, column=2).fill = subheader_fill

    for row_number, (name, count) in enumerate(achievement_rows, start_row + 1):
        ws.cell(row=row_number, column=1, value=name).border = border
        ws.cell(row=row_number, column=2, value=count).border = border

    return start_row + len(achievement_rows)


def write_badges_sheet(
    ws: Any,
    content: dict[str, Any],
    *,
    section_font: Any,
    subtitle_font: Any,
    subheader_font: Any,
    subheader_fill: Any,
    border: Any,
    style_header_row: Callable[[Any, int, int], None],
    style_data_cell: Callable[[Any], None],
    build_badge_distribution_chart: Callable[..., Any],
) -> int:
    ws.merge_cells("A1:E1")
    ws["A1"].value = "Badge Dashboard"
    ws["A1"].font = section_font

    ws.cell(
        row=2,
        column=1,
        value=(
            f"Total badges earned: {content['total_badges']} across "
            f"{content['total_repos']} repos"
        ),
    ).font = subtitle_font

    for col, header in enumerate(BADGE_DISTRIBUTION_HEADERS, 1):
        ws.cell(row=4, column=col, value=header)
    style_header_row(ws, 4, len(BADGE_DISTRIBUTION_HEADERS))
    ws.freeze_panes = "B2"

    badge_end = write_badge_distribution_rows(
        ws,
        content["distribution_rows"],
        start_row=5,
        style_data_cell=style_data_cell,
    )

    if content["distribution_rows"]:
        ws.add_chart(build_badge_distribution_chart(ws, start_row=5, end_row=badge_end), "E4")

    board_row = badge_end + 3
    ws.cell(row=board_row, column=1, value="Achievement Board").font = section_font
    return write_badge_achievement_board(
        ws,
        content["achievement_rows"],
        start_row=board_row + 1,
        subheader_font=subheader_font,
        subheader_fill=subheader_fill,
        border=border,
    )


def build_badges_sheet(
    wb,
    data: dict[str, Any],
    *,
    get_or_create_sheet,
    configure_sheet_view,
    build_badges_sheet_content_fn,
    write_badges_sheet_fn,
    section_font,
    subtitle_font,
    subheader_font,
    subheader_fill,
    border,
    style_header_row,
    style_data_cell,
    build_badge_distribution_chart,
    auto_width,
) -> None:
    ws = get_or_create_sheet(wb, "Badges")
    ws.sheet_properties.tabColor = "7C3AED"
    configure_sheet_view(ws, zoom=110, show_grid_lines=True)
    content = build_badges_sheet_content_fn(data.get("audits", []))
    final_row = write_badges_sheet_fn(
        ws,
        content,
        section_font=section_font,
        subtitle_font=subtitle_font,
        subheader_font=subheader_font,
        subheader_fill=subheader_fill,
        border=border,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
        build_badge_distribution_chart=build_badge_distribution_chart,
    )

    auto_width(ws, 3, final_row)
