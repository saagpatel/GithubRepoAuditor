"""Helpers for repo-detail workbook sections."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

from src.excel_detail_helpers import repo_detail_lookup_formula, set_internal_hyperlink

REPO_DETAIL_DIMENSION_HEADERS = ["Rank", "Dimension", "Score", "Summary"]


def repo_detail_summary_fields() -> list[tuple[str, int, int, int, str]]:
    return [
        ("Repo", 1, 1, 1, "No repo selected."),
        ("Language", 3, 1, 4, "Unknown"),
        ("Badges", 9, 1, 7, "None"),
        ("Overall Score", 5, 2, 1, ""),
        ("Interest Score", 6, 2, 4, ""),
        ("Flags", 10, 2, 7, "None"),
        ("Grade", 7, 3, 1, "—"),
        ("Tier", 8, 3, 4, "—"),
        ("Collections", 11, 3, 7, "—"),
        ("Security", 12, 4, 1, "unknown"),
        ("Security Score", 13, 4, 4, ""),
        ("Trend", 18, 4, 7, "No trend history yet."),
    ]


def repo_detail_current_state_rows() -> list[tuple[str, int]]:
    return [
        ("Ship Readiness", 14),
        ("Momentum", 15),
        ("Security", 16),
        ("Portfolio Fit", 17),
    ]


def repo_detail_explanation_rows() -> list[tuple[str, int]]:
    return [
        ("Strongest Drivers", 19),
        ("Biggest Drags", 20),
        ("Last Movement", 30),
        ("Next Tier Gap", 21),
        ("Next Best Action", 22),
        ("Why This Action", 23),
    ]


def repo_detail_implementation_rows() -> list[tuple[str, int, str]]:
    return [
        ("Summary", 66, "No meaningful implementation hotspot is currently surfaced."),
        ("Implementation Hotspot 1", 67, "No implementation hotspot is currently surfaced."),
        (
            "Implementation Hotspot 2",
            68,
            "No second implementation hotspot is currently surfaced.",
        ),
        (
            "Implementation Hotspot 3",
            69,
            "No third implementation hotspot is currently surfaced.",
        ),
    ]


def build_repo_detail_sheet_content(
    repo_names: list[str],
    existing_selection: str | None,
) -> dict[str, object]:
    default_repo = (
        existing_selection
        if existing_selection in repo_names
        else (repo_names[0] if repo_names else "")
    )
    validation_formula = (
        f"=Data_RepoDetail!$A$2:$A${len(repo_names) + 1}" if repo_names else None
    )
    return {
        "default_repo": default_repo,
        "validation_formula": validation_formula,
        "summary_rows": [
            {
                "row": row_base + 5,
                "label_col": col_base,
                "value_col": col_base + 1,
                "label": label,
                "formula": repo_detail_lookup_formula(
                    column_index, fallback, allow_blank=column_index in {5, 6, 13}
                ),
            }
            for label, column_index, row_base, col_base, fallback in repo_detail_summary_fields()
        ],
        "description_formula": repo_detail_lookup_formula(4, "No description recorded yet."),
        "current_state_rows": [
            {
                "row": 13 + offset,
                "label": label,
                "formula": repo_detail_lookup_formula(column_index, "No summary recorded yet."),
            }
            for offset, (label, column_index) in enumerate(repo_detail_current_state_rows(), 1)
        ],
        "explanation_rows": [
            {
                "row": 13 + offset,
                "label": label,
                "formula": repo_detail_lookup_formula(
                    column_index, "No briefing detail recorded yet."
                ),
            }
            for offset, (label, column_index) in enumerate(repo_detail_explanation_rows(), 1)
        ],
        "dimension_rows": [
            {
                "row": 21 + rank,
                "rank": rank,
                "dimension_formula": (
                    f'=IFERROR(VLOOKUP($B$4&"::{rank}",Data_RepoDimensionRollups!$A:$F,4,FALSE),"")'
                ),
                "score_formula": (
                    f'=IFERROR(VLOOKUP($B$4&"::{rank}",Data_RepoDimensionRollups!$A:$F,5,FALSE),"")'
                ),
                "summary_formula": (
                    f'=IFERROR(VLOOKUP($B$4&"::{rank}",Data_RepoDimensionRollups!$A:$F,6,FALSE),"")'
                ),
            }
            for rank in range(1, 7)
        ],
        "hotspot_rows": [
            {
                "row": 20 + idx,
                "label": f"Hotspot {idx}",
                "formula": repo_detail_lookup_formula(23 + idx, "No hotspot recorded yet."),
            }
            for idx in range(1, 4)
        ],
        "implementation_rows": [
            {
                "row": 28 + offset,
                "label": label,
                "formula": repo_detail_lookup_formula(column_index, fallback),
            }
            for offset, (label, column_index, fallback) in enumerate(
                repo_detail_implementation_rows(), 1
            )
        ],
    }


def write_repo_detail_summary_rows(ws, summary_rows, *, style_data_cell, subheader_font) -> None:
    for item in summary_rows:
        ws.cell(row=item["row"], column=item["label_col"], value=item["label"]).font = subheader_font
        style_data_cell(
            ws.cell(row=item["row"], column=item["value_col"], value=item["formula"]),
            "left",
        )


def write_repo_detail_labeled_rows(
    ws,
    rows,
    *,
    value_column: int,
    style_data_cell,
    subheader_font,
) -> None:
    for item in rows:
        row = item["row"]
        ws.cell(row=row, column=value_column - 1, value=item["label"]).font = subheader_font
        style_data_cell(ws.cell(row=row, column=value_column, value=item["formula"]), "left")


def write_repo_detail_dimension_table(
    ws,
    rows,
    *,
    header_row: int,
    style_header_row,
    style_data_cell,
) -> None:
    for col, header in enumerate(REPO_DETAIL_DIMENSION_HEADERS, 1):
        ws.cell(row=header_row, column=col, value=header)
    style_header_row(ws, header_row, len(REPO_DETAIL_DIMENSION_HEADERS))
    for item in rows:
        row = item["row"]
        ws.cell(row=row, column=1, value=item["rank"])
        ws.cell(row=row, column=2, value=item["dimension_formula"])
        ws.cell(row=row, column=3, value=item["score_formula"])
        ws.cell(row=row, column=4, value=item["summary_formula"])
        for col in range(1, len(REPO_DETAIL_DIMENSION_HEADERS) + 1):
            style_data_cell(ws.cell(row=row, column=col), "center" if col in {1, 3} else "left")


def write_repo_detail_handoff_section(ws) -> int:
    ws["F25"] = "What To Do Next"
    handoff_rows = [
        ("Recommended Action", 22, "No clear next action is recorded yet."),
        ("Why This Action", 23, "No action rationale is recorded yet."),
        ("Follow-Through Status", 32, "Unknown"),
        ("Follow-Through Summary", 33, "No follow-through evidence is recorded yet."),
        ("Checkpoint Timing", 34, "Unknown"),
        ("Escalation", 35, "Unknown"),
        ("Escalation Summary", 36, "No stronger follow-through escalation is currently surfaced."),
        ("Recovery / Retirement", 37, "None"),
        (
            "Recovery Summary",
            38,
            "No follow-through recovery or escalation-retirement signal is currently surfaced.",
        ),
        ("Recovery Persistence", 39, "None"),
        (
            "Recovery Persistence Summary",
            40,
            "No follow-through recovery persistence signal is currently surfaced.",
        ),
        ("Relapse Churn", 41, "None"),
        ("Relapse Churn Summary", 42, "No relapse churn is currently surfaced."),
        ("Recovery Freshness", 43, "None"),
        (
            "Recovery Freshness Summary",
            44,
            "No follow-through recovery freshness signal is currently surfaced.",
        ),
        ("Recovery Memory Reset", 45, "None"),
        (
            "Recovery Memory Reset Summary",
            46,
            "No follow-through recovery memory reset signal is currently surfaced.",
        ),
        ("Recovery Rebuild Strength", 47, "None"),
        (
            "Recovery Rebuild Strength Summary",
            48,
            "No follow-through recovery rebuild-strength signal is currently surfaced.",
        ),
        ("Recovery Reacquisition", 49, "None"),
        (
            "Recovery Reacquisition Summary",
            50,
            "No follow-through recovery reacquisition signal is currently surfaced.",
        ),
        ("Reacquisition Durability", 51, "None"),
        (
            "Reacquisition Durability Summary",
            52,
            "No follow-through reacquisition durability signal is currently surfaced.",
        ),
        ("Reacquisition Confidence", 53, "None"),
        (
            "Reacquisition Confidence Summary",
            54,
            "No follow-through reacquisition confidence-consolidation signal is currently surfaced.",
        ),
        ("Reacquisition Softening Decay", 55, "None"),
        (
            "Reacquisition Softening Decay Summary",
            56,
            "No reacquisition softening-decay signal is currently surfaced.",
        ),
        ("Reacquisition Confidence Retirement", 57, "None"),
        (
            "Reacquisition Confidence Retirement Summary",
            58,
            "No reacquisition confidence-retirement signal is currently surfaced.",
        ),
        ("Revalidation Recovery", 59, "None"),
        (
            "Revalidation Recovery Summary",
            60,
            "No post-revalidation recovery or confidence re-earning signal is currently surfaced.",
        ),
        (
            "Progress Checkpoint",
            61,
            "Use the next run or linked artifact to confirm whether the recommendation moved.",
        ),
        ("Portfolio Catalog", 62, "No portfolio catalog contract is recorded yet."),
        (
            "Operating Path",
            66,
            "Operating Path: Unspecified (legacy confidence) — No operating-path rationale is recorded yet.",
        ),
        (
            "Intent Alignment",
            63,
            "missing-contract: Intent alignment cannot be judged until a portfolio catalog contract exists.",
        ),
        ("Scorecard", 64, "Scorecard: No maturity scorecard is recorded yet."),
        ("Maturity Gap", 65, "No maturity gap summary is recorded yet."),
        ("Action Candidate 2", 28, "No second action candidate recorded yet."),
        ("Action Candidate 3", 29, "No third action candidate recorded yet."),
    ]
    return len(handoff_rows)


def repo_detail_handoff_rows() -> list[tuple[str, int, str]]:
    return [
        ("Recommended Action", 22, "No clear next action is recorded yet."),
        ("Why This Action", 23, "No action rationale is recorded yet."),
        ("Follow-Through Status", 32, "Unknown"),
        ("Follow-Through Summary", 33, "No follow-through evidence is recorded yet."),
        ("Checkpoint Timing", 34, "Unknown"),
        ("Escalation", 35, "Unknown"),
        ("Escalation Summary", 36, "No stronger follow-through escalation is currently surfaced."),
        ("Recovery / Retirement", 37, "None"),
        (
            "Recovery Summary",
            38,
            "No follow-through recovery or escalation-retirement signal is currently surfaced.",
        ),
        ("Recovery Persistence", 39, "None"),
        (
            "Recovery Persistence Summary",
            40,
            "No follow-through recovery persistence signal is currently surfaced.",
        ),
        ("Relapse Churn", 41, "None"),
        ("Relapse Churn Summary", 42, "No relapse churn is currently surfaced."),
        ("Recovery Freshness", 43, "None"),
        (
            "Recovery Freshness Summary",
            44,
            "No follow-through recovery freshness signal is currently surfaced.",
        ),
        ("Recovery Memory Reset", 45, "None"),
        (
            "Recovery Memory Reset Summary",
            46,
            "No follow-through recovery memory reset signal is currently surfaced.",
        ),
        ("Recovery Rebuild Strength", 47, "None"),
        (
            "Recovery Rebuild Strength Summary",
            48,
            "No follow-through recovery rebuild-strength signal is currently surfaced.",
        ),
        ("Recovery Reacquisition", 49, "None"),
        (
            "Recovery Reacquisition Summary",
            50,
            "No follow-through recovery reacquisition signal is currently surfaced.",
        ),
        ("Reacquisition Durability", 51, "None"),
        (
            "Reacquisition Durability Summary",
            52,
            "No follow-through reacquisition durability signal is currently surfaced.",
        ),
        ("Reacquisition Confidence", 53, "None"),
        (
            "Reacquisition Confidence Summary",
            54,
            "No follow-through reacquisition confidence-consolidation signal is currently surfaced.",
        ),
        ("Reacquisition Softening Decay", 55, "None"),
        (
            "Reacquisition Softening Decay Summary",
            56,
            "No reacquisition softening-decay signal is currently surfaced.",
        ),
        ("Reacquisition Confidence Retirement", 57, "None"),
        (
            "Reacquisition Confidence Retirement Summary",
            58,
            "No reacquisition confidence-retirement signal is currently surfaced.",
        ),
        ("Revalidation Recovery", 59, "None"),
        (
            "Revalidation Recovery Summary",
            60,
            "No post-revalidation recovery or confidence re-earning signal is currently surfaced.",
        ),
        (
            "Progress Checkpoint",
            61,
            "Use the next run or linked artifact to confirm whether the recommendation moved.",
        ),
        ("Portfolio Catalog", 62, "No portfolio catalog contract is recorded yet."),
        (
            "Operating Path",
            66,
            "Operating Path: Unspecified (legacy confidence) — No operating-path rationale is recorded yet.",
        ),
        (
            "Intent Alignment",
            63,
            "missing-contract: Intent alignment cannot be judged until a portfolio catalog contract exists.",
        ),
        ("Scorecard", 64, "Scorecard: No maturity scorecard is recorded yet."),
        ("Maturity Gap", 65, "No maturity gap summary is recorded yet."),
        ("Action Candidate 2", 28, "No second action candidate recorded yet."),
        ("Action Candidate 3", 29, "No third action candidate recorded yet."),
    ]


def populate_repo_detail_handoff_section(ws, style_data_cell, subheader_font) -> int:
    ws["F25"] = "What To Do Next"
    ws["F25"].font = subheader_font
    handoff_rows = repo_detail_handoff_rows()
    for offset, (label, column_index, fallback) in enumerate(handoff_rows, 1):
        row = 25 + offset
        ws.cell(row=row, column=6, value=label).font = subheader_font
        style_data_cell(
            ws.cell(row=row, column=7, value=repo_detail_lookup_formula(column_index, fallback)),
            "left",
        )
    nav_row = 26 + len(handoff_rows)
    ws.cell(row=nav_row, column=1, value="Use Score Explainer")
    set_internal_hyperlink(
        ws.cell(row=nav_row, column=2), "Score Explainer", display="Open Score Explainer"
    )
    ws.cell(row=nav_row, column=4, value="Go To Explorer")
    set_internal_hyperlink(
        ws.cell(row=nav_row, column=5), "Portfolio Explorer", display="Open Explorer"
    )
    ws.cell(row=nav_row, column=6, value="Go To Queue")
    set_internal_hyperlink(
        ws.cell(row=nav_row, column=7), "Review Queue", display="Open Review Queue"
    )
    return nav_row


def build_repo_detail_sheet(
    wb: Workbook,
    data: dict,
    *,
    get_or_create_sheet,
    configure_sheet_view,
    set_sheet_header,
    write_instruction_banner,
    style_header_row,
    style_data_cell,
    auto_width,
    subheader_font,
    section_font,
    wrap_alignment,
    teal: str,
) -> None:
    existing_selection = ""
    if "Repo Detail" in wb.sheetnames:
        existing_selection = str(wb["Repo Detail"]["B4"].value or "").strip()
    ws = get_or_create_sheet(wb, "Repo Detail")
    ws.sheet_properties.tabColor = "1E40AF"
    configure_sheet_view(ws, zoom=115, show_grid_lines=False)
    set_sheet_header(
        ws,
        "Repo Detail",
        "Pick one repo, then use this page as the fastest single-repo briefing: score, tier, trend, risks, and next move.",
        width=10,
    )
    write_instruction_banner(
        ws,
        3,
        8,
        "Choose one repo, skim the briefing blocks top to bottom, then jump to Run Changes or Review Queue if you need more context.",
    )
    ws.freeze_panes = "A6"

    repo_names = sorted(
        audit.get("metadata", {}).get("name", "")
        for audit in data.get("audits", [])
        if audit.get("metadata", {}).get("name")
    )
    content = build_repo_detail_sheet_content(repo_names, existing_selection)
    default_repo = content["default_repo"]
    ws["A4"] = "Select Repo"
    ws["A4"].font = subheader_font
    ws["B4"] = default_repo
    if content["validation_formula"]:
        dv = DataValidation(
            type="list",
            formula1=content["validation_formula"],
            allow_blank=False,
            promptTitle="Repo Detail",
            prompt="Choose a repo to refresh this briefing page.",
        )
        dv.add("B4")
        ws.add_data_validation(dv)
    ws["D4"] = "GitHub"
    ws["D4"].font = subheader_font
    ws["E4"] = (
        '=IFERROR(IF(VLOOKUP($B$4,Data_RepoDetail!$A:$BZ,2,FALSE)="","Repo URL unavailable",HYPERLINK(VLOOKUP($B$4,Data_RepoDetail!$A:$BZ,2,FALSE),"Open Repo")),"Repo URL unavailable")'
    )
    ws["E4"].font = Font("Calibri", 10, bold=True, color=teal, underline="single")

    write_repo_detail_summary_rows(
        ws,
        content["summary_rows"],
        style_data_cell=style_data_cell,
        subheader_font=subheader_font,
    )

    ws["A11"] = "Description"
    ws["A11"].font = subheader_font
    ws.merge_cells("B11:H11")
    ws["B11"] = content["description_formula"]
    ws["B11"].alignment = wrap_alignment

    ws["A13"] = "Current State"
    ws["A13"].font = section_font
    write_repo_detail_labeled_rows(
        ws,
        content["current_state_rows"],
        value_column=2,
        style_data_cell=style_data_cell,
        subheader_font=subheader_font,
    )

    ws["E13"] = "Why This Repo Looks This Way"
    ws["E13"].font = section_font
    write_repo_detail_labeled_rows(
        ws,
        content["explanation_rows"],
        value_column=6,
        style_data_cell=style_data_cell,
        subheader_font=subheader_font,
    )

    ws["A20"] = "Dimension Breakdown"
    ws["A20"].font = section_font
    write_repo_detail_dimension_table(
        ws,
        content["dimension_rows"],
        header_row=21,
        style_header_row=style_header_row,
        style_data_cell=style_data_cell,
    )

    ws["F20"] = "What Changed"
    ws["F20"].font = section_font
    write_repo_detail_labeled_rows(
        ws,
        content["hotspot_rows"],
        value_column=7,
        style_data_cell=style_data_cell,
        subheader_font=subheader_font,
    )

    ws["A28"] = "Where To Start"
    ws["A28"].font = section_font
    write_repo_detail_labeled_rows(
        ws,
        content["implementation_rows"],
        value_column=2,
        style_data_cell=style_data_cell,
        subheader_font=subheader_font,
    )

    nav_row = populate_repo_detail_handoff_section(ws, style_data_cell, subheader_font)
    auto_width(ws, 8, nav_row)
