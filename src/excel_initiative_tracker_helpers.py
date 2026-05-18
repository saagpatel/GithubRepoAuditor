"""Excel helper: "Initiative Tracker" sheet (Arc G Sprint 7A.4).

Renders open tier-upgrade initiatives grouped by status:
  on-track / at-risk / overdue / met (not yet closed)

Gracefully degrades to an empty-state message when no initiatives exist.
"""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl.workbook import Workbook

from src.initiatives import (
    Initiative,
    derive_status,
    initiatives_path,
    load_initiatives,
)
from src.maturity_tiers import TierGap, compute_tier, tier_gap, tier_name

logger = logging.getLogger(__name__)

# Tab colours ─────────────────────────────────────────────────────────────────
_TAB_COLOR_GREEN = "15803D"  # on-track
_TAB_COLOR_AMBER = "B45309"  # at-risk  (fall back if mixed, prefer amber)
_TAB_COLOR_RED = "DC2626"  # overdue
_TAB_COLOR_BLUE = "0369A1"  # met / neutral

# Column layout ───────────────────────────────────────────────────────────────
_HEADERS = [
    "REPO",
    "CURRENT TIER",
    "TARGET TIER",
    "DEADLINE",
    "SET AT",
    "DAYS LEFT",
    "MISSING REQUIREMENTS",
]
_COL_WIDTHS = [22, 14, 13, 12, 12, 10, 60]

# Status ordering ─────────────────────────────────────────────────────────────
_LANE_ORDER: list[str] = ["overdue", "at-risk", "on-track", "met"]
_LANE_LABELS: dict[str, str] = {
    "overdue": "OVERDUE",
    "at-risk": "AT RISK",
    "on-track": "ON TRACK",
    "met": "MET (not yet closed)",
}


# ── Data helpers ──────────────────────────────────────────────────────────────


def _days_left(deadline: str, today: date | None = None) -> str:
    """Return a human-readable days-left string; negative = overdue."""
    t = today or date.today()
    try:
        d = date.fromisoformat(deadline)
        delta = (d - t).days
        if delta < 0:
            return f"{abs(delta)}d overdue"
        if delta == 0:
            return "today"
        return f"{delta}d"
    except (ValueError, TypeError):
        return "—"


def _format_missing_requirements(gap: TierGap) -> str:
    """Join missing requirements with an optional '(approx.)' suffix per source.

    Each requirement whose parallel ``requirement_sources`` entry is ``"proxy"``
    is annotated with ' (approx.)' to signal it was inferred, not directly
    verified.  When ``requirement_sources`` is empty (pre-Sprint-8.3 callers),
    all requirements are treated as proxy — consistent with the legacy behaviour
    of that code path.
    """
    if not gap.missing_requirements:
        return "—"
    # Fallback: treat everything as proxy when sources list is missing (legacy).
    sources = gap.requirement_sources or ["proxy"] * len(gap.missing_requirements)
    parts = []
    for req, src in zip(gap.missing_requirements, sources):
        if src == "proxy":
            parts.append(f"{req} (approx.)")
        else:
            parts.append(req)
    return "; ".join(parts)


def _missing_reqs_str(initiative: Initiative, repo: dict) -> str:
    """Return a semicolon-separated list of missing requirements, or '—'."""
    gap = tier_gap(repo, initiative.target_tier)
    return _format_missing_requirements(gap)


def _build_rows(
    initiatives: list[Initiative],
    projects: list[dict],
    today: date,
) -> dict[str, list[tuple[str, str, str, str, str, str, str]]]:
    """Build a dict[status → list[row tuples]]."""
    repo_map: dict[str, dict] = {p.get("identity", {}).get("display_name", ""): p for p in projects}

    lanes: dict[str, list[tuple[str, str, str, str, str, str, str]]] = {s: [] for s in _LANE_ORDER}

    for initiative in initiatives:
        # Skip closed initiatives
        if initiative.closed_at is not None:
            continue
        repo = repo_map.get(initiative.repo_name, {})
        status = derive_status(initiative, repo, today=today.isoformat())
        current = compute_tier(repo)
        row = (
            initiative.repo_name,
            f"{current} — {tier_name(current)}",
            f"{initiative.target_tier} — {tier_name(initiative.target_tier)}",
            initiative.deadline,
            initiative.set_at[:10] if initiative.set_at else "—",
            _days_left(initiative.deadline, today),
            _missing_reqs_str(initiative, repo),
        )
        lanes[status].append(row)

    return lanes


# ── Public API ────────────────────────────────────────────────────────────────


def build_initiative_tracker_sheet(
    wb: Workbook,
    data: dict[str, Any],
    *,
    # Style helpers (injected by excel_export.py to avoid re-importing)
    get_or_create_sheet,
    configure_sheet_view,
    set_sheet_header,
    style_header_row,
    auto_width,
    wrap_alignment,
    subtitle_font,
    # Data source override (for tests)
    output_dir: Path | None = None,
) -> None:
    """Build the 'Initiative Tracker' worksheet inside *wb*.

    Loads initiatives from ``output_dir/initiatives.json`` (falls back to
    ``Path("output")/initiatives.json`` if *output_dir* is None).
    Degrades gracefully if the file is missing or empty.
    """
    ws = get_or_create_sheet(wb, "Initiative Tracker")
    ws.sheet_properties.tabColor = _TAB_COLOR_BLUE
    configure_sheet_view(ws, zoom=115, show_grid_lines=False)

    set_sheet_header(
        ws,
        "Initiative Tracker",
        "Tier-upgrade initiatives grouped by status: overdue · at-risk · on-track · met.",
        width=len(_HEADERS),
    )

    # Apply column widths
    from openpyxl.utils import get_column_letter

    for col_idx, width in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Load data
    resolved_dir = output_dir or Path("output")
    initiatives = load_initiatives(initiatives_path(resolved_dir))

    if not initiatives:
        _write_empty_state(ws, wrap_alignment, subtitle_font)
        return

    # Extract projects from data (portfolio-truth-style project list)
    projects: list[dict] = data.get("projects", [])
    # Fallback: data may carry top-level 'audits' keyed repos
    if not projects:
        projects = [
            {
                "identity": {"display_name": a.get("metadata", {}).get("name", "")},
                "derived": a.get("derived", {}),
                "risk": a.get("risk", {}),
            }
            for a in data.get("audits", [])
        ]

    today = date.today()
    lanes = _build_rows(initiatives, projects, today)

    # Determine tab colour based on worst status present
    all_statuses = {s for s, rows in lanes.items() if rows}
    if "overdue" in all_statuses:
        ws.sheet_properties.tabColor = _TAB_COLOR_RED
    elif "at-risk" in all_statuses:
        ws.sheet_properties.tabColor = _TAB_COLOR_AMBER
    elif "on-track" in all_statuses:
        ws.sheet_properties.tabColor = _TAB_COLOR_GREEN

    row_num = 3  # header rows already consumed by set_sheet_header
    for status in _LANE_ORDER:
        rows = lanes[status]
        if not rows:
            continue

        # Lane header row
        lane_label = _LANE_LABELS[status]
        ws.cell(row=row_num, column=1, value=lane_label)
        ws.cell(row=row_num, column=1).font = subtitle_font
        row_num += 1

        # Column header row
        style_header_row(ws, row_num, len(_HEADERS))
        for col_idx, header in enumerate(_HEADERS, 1):
            ws.cell(row=row_num, column=col_idx, value=header)
        row_num += 1

        for row_data in rows:
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.alignment = wrap_alignment
            row_num += 1

        # Blank separator row
        row_num += 1


def _write_empty_state(ws, wrap_alignment, subtitle_font) -> None:
    """Write a single 'no initiatives yet' message row."""
    msg = (
        "No initiatives yet. "
        "Use `audit triage --set-initiative REPO --target-tier N --deadline YYYY-MM-DD`."
    )
    cell = ws.cell(row=3, column=1, value=msg)
    cell.alignment = wrap_alignment
    cell.font = subtitle_font
