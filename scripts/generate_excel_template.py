from __future__ import annotations

from _bootstrap import ensure_project_root
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName

ROOT = ensure_project_root()


def _load_template_constants() -> tuple[object, object, object]:
    from src.excel_template import DEFAULT_TEMPLATE_PATH, TEMPLATE_INFO_SHEET, TEMPLATE_SHEETS

    return DEFAULT_TEMPLATE_PATH, TEMPLATE_INFO_SHEET, TEMPLATE_SHEETS


DEFAULT_TEMPLATE_PATH, TEMPLATE_INFO_SHEET, TEMPLATE_SHEETS = _load_template_constants()


SECTION_FILL = PatternFill("solid", fgColor="E2E8F0")
HEADER_FILL = PatternFill("solid", fgColor="0F172A")


def _add_named_range(wb: Workbook, name: str, ref: str) -> None:
    wb.defined_names.add(DefinedName(name, attr_text=ref))


def _style_sheet(ws, title: str, description: str) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = description
    ws["A2"].font = Font(size=10, italic=True)
    ws["A4"] = "This sheet is hydrated from hidden workbook facts in template mode."
    ws["A4"].fill = SECTION_FILL
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20


def main() -> None:
    output_path = DEFAULT_TEMPLATE_PATH
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    first = wb.active
    first.title = TEMPLATE_SHEETS[0]
    for sheet_name in TEMPLATE_SHEETS[1:]:
        wb.create_sheet(sheet_name)

    descriptions = {
        "Index": "Workbook navigation grouped by persona.",
        "Dashboard": "Operator landing page and workbook KPI summary.",
        "Portfolio Explorer": "Profile-aware ranking and portfolio context.",
        "By Lens": "Ranking by decision lens.",
        "By Collection": "Collection-level rollup and leaders.",
        "Trend Summary": "Portfolio trend history and repo sparkline views.",
        "Scenario Planner": "Scenario lift and projected promotions.",
        "Review Queue": "Current material review targets.",
        "Review History": "Recurring review lifecycle history.",
        "Campaigns": "Campaign preview and status.",
        "Writeback Audit": "Writeback and external sync outcomes.",
        "Governance Controls": "Governed security-control preview.",
        "Governance Audit": "Governance readiness and drift summary.",
        "Executive Summary": "Compact analyst and executive summary.",
        "Print Pack": "Print-oriented workbook summary.",
    }

    for sheet_name in TEMPLATE_SHEETS:
        ws = wb[sheet_name]
        if sheet_name == TEMPLATE_INFO_SHEET:
            ws.sheet_state = "hidden"
            labels = [
                "Workbook Mode",
                "Review Open Count",
                "Review Deferred Count",
                "Review Resolved Count",
                "Campaign Action Count",
                "Campaign Repo Count",
                "Governance Ready Count",
                "Governance Drift Count",
                "Latest Portfolio Grade",
                "Latest Average Score",
                "Latest Review State",
                "Selected Profile",
                "Selected Collection",
            ]
            ws["A1"] = "Key"
            ws["B1"] = "Value"
            ws["A1"].fill = HEADER_FILL
            ws["B1"].fill = HEADER_FILL
            ws["A1"].font = Font(bold=True, color="FFFFFF")
            ws["B1"].font = Font(bold=True, color="FFFFFF")
            for idx, label in enumerate(labels, 2):
                ws.cell(row=idx, column=1, value=label).fill = SECTION_FILL
        else:
            _style_sheet(ws, sheet_name, descriptions[sheet_name])

    _add_named_range(wb, "nrGeneratedAt", "'Dashboard'!$A$2")
    _add_named_range(wb, "nrReviewOpenCount", f"'{TEMPLATE_INFO_SHEET}'!$B$2")
    _add_named_range(wb, "nrReviewDeferredCount", f"'{TEMPLATE_INFO_SHEET}'!$B$3")
    _add_named_range(wb, "nrReviewResolvedCount", f"'{TEMPLATE_INFO_SHEET}'!$B$4")
    _add_named_range(wb, "nrCampaignActionCount", f"'{TEMPLATE_INFO_SHEET}'!$B$5")
    _add_named_range(wb, "nrCampaignRepoCount", f"'{TEMPLATE_INFO_SHEET}'!$B$6")
    _add_named_range(wb, "nrGovernanceReadyCount", f"'{TEMPLATE_INFO_SHEET}'!$B$7")
    _add_named_range(wb, "nrGovernanceDriftCount", f"'{TEMPLATE_INFO_SHEET}'!$B$8")
    _add_named_range(wb, "nrPortfolioGrade", f"'{TEMPLATE_INFO_SHEET}'!$B$9")
    _add_named_range(wb, "nrAverageScore", f"'{TEMPLATE_INFO_SHEET}'!$B$10")
    _add_named_range(wb, "nrLatestReviewState", f"'{TEMPLATE_INFO_SHEET}'!$B$11")
    _add_named_range(wb, "nrSelectedProfileLabel", f"'{TEMPLATE_INFO_SHEET}'!$B$12")
    _add_named_range(wb, "nrSelectedCollectionLabel", f"'{TEMPLATE_INFO_SHEET}'!$B$13")

    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
